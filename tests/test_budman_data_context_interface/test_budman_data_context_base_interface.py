# ---------------------------------------------------------------------------- +
# test_budman_data_context_base_interface.py
# ---------------------------------------------------------------------------- +
#region imports
# python standard libraries
import pytest, os
from abc import ABC, abstractmethod
# third-party modules and packages
from openpyxl import Workbook
# local modules and packages
import logging, p3_utils as p3u, p3logging as p3l
from src.budman_namespace import design_language_namespace as bdmns
from budman_data_context import BudManDataContextBaseInterface
#endregion imports
# ---------------------------------------------------------------------------- +
#region Globals
logger = logging.getLogger(__name__)
#endregion Globals
# ---------------------------------------------------------------------------- +
class testBudManDC(BudManDataContextBaseInterface):
    """A test subclass of BudManDataContextBaseInterface."""
    def __init__(self):
        super().__init__()
        self.test_property = "test_value"
    #region Concrete Properties
    @property
    def dc_INITIALIZED(self) -> bool:
        """Indicates whether the data context has been initialized."""
        print("Getting INITIALIZED, True")
        return True
    @dc_INITIALIZED.setter
    def dc_INITIALIZED(self, value: bool) -> None:
        """Set the initialized state of the data context."""
        print(f"Setting INITIALIZED to {value}")
        return
    @property
    def dc_FI_KEY(self) -> str:
        """Return the FI_KEY for the financial institution."""
        print("Getting dc_FI_KEY, 'test_fi_key'")
        return "test_fi_key"
    @dc_FI_KEY.setter
    def dc_FI_KEY(self, value: str) -> None:
        """Set the FI_KEY for the financial institution."""
        print(f"Setting dc_FI_KEY to {value}")
        return
    @property
    def dc_WF_KEY(self) -> str:
        """Return the WF_KEY for the workflow."""
        print("Getting dc_WF_KEY, 'test_wf_key'")
        return "test_wf_key"
    @dc_WF_KEY.setter
    def dc_WF_KEY(self, value: str) -> None:
        """Set the WF_KEY for the workflow."""
        print(f"Setting dc_WF_KEY to {value}")
        return
    @property
    def dc_WB_TYPE(self) -> str:
        """Return the WB_TYPE workbook type."""
        print("Getting dc_WB_TYPE, 'test_wb_type'")
        return "test_wb_type"
    @dc_WB_TYPE.setter
    def dc_WB_TYPE(self, value: str) -> None:
        """Set the WB_TYPE workbook type."""
        print(f"Setting dc_WB_TYPE to {value}")
        return
    @property
    def dc_WB_NAME(self) -> str:
        """Return the WB_NAME workbook name."""
        print("Getting dc_WB_NAME, 'test_wb_name'")
        return "test_wb_name"
    @dc_WB_NAME.setter
    def dc_WB_NAME(self, value: str) -> None:
        """Set the WB_NAME workbook name."""
        print(f"Setting dc_WB_NAME to {value}")
        return
    @property
    def dc_BUDMAN_STORE(self) -> str:
        """Return the BUDMAN_STORE jsonc definition."""
        print("Getting dc_BUDMAN_STORE, 'test_budman_store.jsonc'")
        return "test_budman_store.jsonc"
    @dc_BUDMAN_STORE.setter
    def dc_BUDMAN_STORE(self, value: str) -> None:
        """Set the BUDMAN_STORE jsonc definition."""
        print(f"Setting dc_BUDMAN_STORE to {value}")
        return
    @property
    def dc_WORKBOOKS(self) -> bdmns.WORKBOOK_LIST:
        """Return the list of workbooks in the DC.
        This is a List of tuples, where each tuple contains the workbook name
        and its absolute path.
        """
        print("Getting dc_WORKBOOKS, [('test_wb1', '/path/to/test_wb1.xlsx'), ('test_wb2', '/path/to/test_wb2.xlsx')]")
        return [('test_wb1', '/path/to/test_wb1.xlsx'), ('test_wb2', '/path/to/test_wb2.xlsx')]
    @dc_WORKBOOKS.setter
    def dc_WORKBOOKS(self, value: bdmns.WORKBOOK_LIST) -> None:
        """Set the list of workbooks in the DC."""
        print(f"Setting dc_WORKBOOKS to {value}")
        return
    @property
    def dc_LOADED_WORKBOOKS(self) -> bdmns.LOADED_WORKBOOK_LIST:
        """Return the list of loaded workbooks in the DC.
        This is a List of tuples, where each tuple contains the workbook name
        and its Workbook object.
        """
        print("Getting dc_LOADED_WORKBOOKS, [('test_wb1', <Workbook>), ('test_wb2', <Workbook>)]")
        return [('test_wb1', Workbook()), ('test_wb2', Workbook())]
    @dc_LOADED_WORKBOOKS.setter
    def dc_LOADED_WORKBOOKS(self, value: bdmns.LOADED_WORKBOOK_LIST) -> None:
        """Set the list of loaded workbooks in the DC."""
        print(f"Setting dc_LOADED_WORKBOOKS to {value}")
        return
    #endregion Concrete Properties
    # ------------------------------------------------------------------------ +
    #region Concrete Methods
    def dc_initialize(self) -> None:
        """Initialize the data context."""
        print("Initializing the data context.")
        self.dc_INITIALIZED = True
        return
    def dc_FI_KEY_validate(self, fi_key: str) -> bool:
        """Validate the provided FI_KEY."""
        print(f"Validating FI_KEY: {fi_key}")
        return isinstance(fi_key, str) and len(fi_key) > 0
    
    def dc_WF_KEY_validate(self, wf_key: str) -> bool:
        """Validate the provided WF_KEY."""
        print(f"Validating WF_KEY: {wf_key}")
        return isinstance(wf_key, str) and len(wf_key) > 0
    def dc_WB_TYPE_validate(self, wb_type: str) -> bool:
        """Validate the provided WB_TYPE."""
        print(f"Validating WB_TYPE: {wb_type}")
        return isinstance(wb_type, str) and len(wb_type) > 0
    def dc_WB_NAME_validate(self, wb_name: str) -> bool:
        """Validate the provided WB_NAME."""
        print(f"Validating WB_NAME: {wb_name}")
        return isinstance(wb_name, str) and len(wb_name) > 0
    def dc_WB_REF_validate(self, wb_ref: str) -> bool:
        """Validate the provided workbook reference."""
        print(f"Validating WB_REF: {wb_ref}")
        return isinstance(wb_ref, str) and len(wb_ref) > 0
    def dc_WORKBOOK_loaded(self, wb_name: str) -> bool:
        """Indicates whether the named workbook is loaded."""
        print(f"Checking if workbook '{wb_name}' is loaded.")
        # For testing, we assume the workbook is loaded if it exists in dc_LOADED_WORKBOOKS.
        return None
    def dc_WORKBOOK_load(self, wb_name: str) -> Workbook:
        """Load the specified workbook by name."""
        print(f"Loading workbook '{wb_name}'.")
        return Workbook()  # Return a new Workbook instance for testing.
    def dc_WORKBOOK_save(self, wb_name: str, wb: Workbook) -> None:
        """Save the specified workbook by name."""
        print(f"Saving workbook '{wb_name}'.")
        return None
    def dc_WORKBOOK_remove(self, wb_name: str) -> None:
        """Remove the specified workbook by name."""
        print(f"Removing workbook '{wb_name}'.")
        return None
    def dc_WORKBOOK_add(self, wb_name: str, wb: Workbook) -> None:
        """Add a new workbook to the data context."""
        print(f"Adding workbook '{wb_name}'.")
        return None
    def dc_BUDMAN_STORE_load(self, file_path: str) -> None:
        """Load the BUDMAN_STORE from the specified file path."""
        print(f"Loading BUDMAN_STORE from {file_path}.")
        return None
    def dc_BUDMAN_STORE_save(self, file_path: str) -> None:
        """Save the BUDMAN_STORE to the specified file path."""
        print(f"Saving BUDMAN_STORE to {file_path}.")
        return None
    #endregion Concrete Methods
# ---------------------------------------------------------------------------- +
class TestBudManDataContextBaseInterface():
    def test_budman_data_context_base_interface(self):
        """Test the BudManDataContextBaseInterface ABC TypeError."""
        try:
            logger.info(self.test_budman_data_context_base_interface.__doc__)
            # Try to instantiate the abstract base class, should raise TypeError.
            with pytest.raises(TypeError) as excinfo:
                bdmdc = BudManDataContextBaseInterface(), \
                f"BudManDataContextBaseInterface() should not be instantiated directly, got: {type(bdmdc)}"
        except Exception as e:
            pytest.fail(f"BudManDataContextBaseInterface() raised an exception: {str(e)}")
    # ------------------------------------------------------------------------ +
    def test_budman_data_context_base_interface_subclass__init__(self,capsys):
        """Test concrete implementation of BudManDataContextBaseInterface class."""
        try:
            logger.info(self.test_budman_data_context_base_interface_subclass__init__.__doc__)
            # Try to instantiate the abstract base class, should raise TypeError.
            bdmdc = testBudManDC()
            assert bdmdc is not None, "testBudManDC() should not be None"
            assert isinstance(bdmdc, BudManDataContextBaseInterface), "Expected testBudManDC to be a subclass of BudManDataContextBaseInterface, got: " + str(type(bdmdc))
            assert bdmdc.dc_INITIALIZED is True, "Expected INITIALIZED to be True after instantiation"
            assert "Getting INITIALIZED" in capsys.readouterr().out, "Expected 'Getting INITIALIZED' in output, got: " + capsys.readouterr().out
            bdmdc.dc_INITIALIZED = True
            assert "Setting INITIALIZED to True" in capsys.readouterr().out, "Expected 'Setting INITIALIZED to True' in output, got: " + capsys.readouterr().out
            assert bdmdc.dc_FI_KEY == "test_fi_key", "Expected dc_FI_KEY to be 'test_fi_key'"
            assert bdmdc.dc_WF_KEY == "test_wf_key", "Expected dc_WF_KEY to be 'test_wf_key'"
            assert bdmdc.dc_WB_TYPE == "test_wb_type", "Expected dc_WB_TYPE to be 'test_wb_type'"
            assert bdmdc.dc_WB_NAME == "test_wb_name", "Expected dc_WB_NAME to be 'test_wb_name'"
            assert bdmdc.dc_BUDMAN_STORE == "test_budman_store.jsonc", "Expected dc_BUDMAN_STORE to be 'test_budman_store.jsonc'"
            bdmdc.dc_BUDMAN_STORE = "new_budman_store.jsonc"
            assert "Setting dc_BUDMAN_STORE to new_budman_store.jsonc" in capsys.readouterr().out, "Expected 'Setting dc_BUDMAN_STORE to new_budman_store.jsonc' in output, got: " + capsys.readouterr().out
            assert isinstance(bdmdc.dc_WORKBOOKS, list), "Expected dc_WORKBOOKS to be a list"
            assert isinstance(bdmdc.dc_LOADED_WORKBOOKS, list), "Expected dc_LOADED_WORKBOOKS to be a list"
            assert bdmdc.dc_FI_KEY_validate("test_fi_key") is True, "Expected dc_FI_KEY_validate to return True"
            assert bdmdc.dc_WF_KEY_validate("test_wf_key") is True, "Expected dc_WF_KEY_validate to return True"
            assert bdmdc.dc_WB_TYPE_validate("test_wb_type") is True, "Expected dc_WB_TYPE_validate to return True"
            assert bdmdc.dc_WB_NAME_validate("test_wb_name") is True, "Expected dc_WB_NAME_validate to return True"
            assert bdmdc.dc_WB_REF_validate("test_wb_ref") is True, "Expected dc_WB_REF_validate to return True"
            assert bdmdc.dc_WORKBOOK_loaded("test_wb1") is None, "Expected WORKBOOK_loaded to return None for test_wb1"
            wb = bdmdc.dc_WORKBOOK_load("test_wb1")
            assert isinstance(wb, Workbook), "Expected WORKBOOK_load to return a Workbook instance"
            assert "Loading workbook 'test_wb1'" in capsys.readouterr().out, "Expected 'Loading workbook 'test_wb1'' in output, got: " + capsys.readouterr().out
            bdmdc.dc_WORKBOOK_save("test_wb1", wb)
            assert "Saving workbook 'test_wb1'" in capsys.readouterr().out, "Expected 'Saving workbook 'test_wb1'' in output, got: " + capsys.readouterr().out
            bdmdc.dc_WORKBOOK_remove("test_wb1")
            assert "Removing workbook 'test_wb1'" in capsys.readouterr().out, "Expected 'Removing workbook 'test_wb1'' in output, got: " + capsys.readouterr().out
            bdmdc.dc_WORKBOOK_add("test_wb1", wb)
            assert "Adding workbook 'test_wb1'" in capsys.readouterr().out, "Expected 'Adding workbook 'test_wb1'' in output, got: " + capsys.readouterr().out
            bdmdc.dc_BUDMAN_STORE_load("test_budman_store.jsonc")
            assert "Loading BUDMAN_STORE from test_budman_store.jsonc" in capsys.readouterr().out, "Expected 'Loading BUDMAN_STORE from test_budman_store.jsonc' in output, got: " + capsys.readouterr().out
            bdmdc.dc_BUDMAN_STORE_save("test_budman_store.jsonc")
            assert "Saving BUDMAN_STORE to test_budman_store.jsonc" in capsys.readouterr().out, "Expected 'Saving BUDMAN_STORE to test_budman_store.jsonc' in output, got: " + capsys.readouterr().out
        except Exception as e:
            pytest.fail(f"BudManDataContextBaseInterface() raised an exception: {str(e)}")
    # ------------------------------------------------------------------------ +

    