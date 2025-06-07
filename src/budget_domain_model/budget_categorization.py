# ---------------------------------------------------------------------------- +
#region p3_budget_categorization.py module
""" Financial Budget Workflow: "categorization" of transaction workbooks.

    Workflow: categorization
    Input Folder: Financial Institution (FI) Incoming Folder (IF)
    Output Folder: Financial Institution (FI) Categorized Folder (CF)
    FI transaction workbooks are typically excel files. 

    Workflow Pattern: Apply a workflow_process (function) to each item in the 
    input folder, placing items in the output folder as appropriate to the 
    configured function. Each WorkFLow instance in the config applies one 
    function to the input with resulting output.

    TODO: Consider xlwings package for Excel integration.
"""
#endregion p3_execl_budget.p3_banking_transactions budget_transactions.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import re, pathlib as Path, logging, time, hashlib, datetime
from dataclasses import dataclass

# third-party modules and packages
import p3logging as p3l, p3_utils as p3u
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# local modules and packages
from budman_namespace import *
from budget_domain_model import (BudgetDomainModel, map_category, 
                                 category_map_count)
# from data.p3_fi_transactions.budget_model import BudgetModel
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)

# Note: Python lists are 0-based. With openpyxl, data is often returned in a
# list. In excel, worksheet columns are 1-based. So, we need to adjust the indices.
# A list of cells from a worksheet row is 0-based, with cell(0) being the value
# from column 1, or column 'A'.

# BOA workbooks arrive with these columns, beginning with "Status". 
BOA_WB_COLUMNS = [
    "Status", 
    "Date", 
    "Original Description",
    "Split Type", 
    "Category", 
    "Currency", 
    "Amount", 
    "User Description", 
    "Memo", 
    "Classification", 
    "Account Name",
    "Simple Description"
    ]

# BudMan insert 5 additional columns prior to processing transactions. These
# columns are filled by BudMan workflows, such as categorization.
BUDGET_CATEGORY_COL_NAME = "Budget Category"  # Added by BudMan.
ACCOUNT_CODE_COL_NAME = "Account Code" # Added by BudMan.
LEVEL_1_COL_NAME = "Level1" # Added by BudMan.
LEVEL_2_COL_NAME = "Level2" # Added by BudMan.
LEVEL_3_COL_NAME = "Level3" # Added by BudMan.


# A BudMan workbook will then have the following columns.
# BudMan users these columns.
DATE_COL_NAME = "Date" 
ORIGINAL_DESCRIPTION_COL_NAME = "Original Description"
CURRENCY_COL_NAME = "Currency"
AMOUNT_COL_NAME = "Amount" 
ACCOUNT_NAME_COL_NAME = "Account Name"  

BUDMAN_WB_COLUMNS = [
    "Status", 
    DATE_COL_NAME, 
    ORIGINAL_DESCRIPTION_COL_NAME,
    "Split Type", 
    "Category", 
    CURRENCY_COL_NAME, 
    AMOUNT_COL_NAME, 
    "User Description", 
    "Memo", 
    "Classification", 
    ACCOUNT_NAME_COL_NAME,
    "Simple Description",
    BUDGET_CATEGORY_COL_NAME,
    ACCOUNT_CODE_COL_NAME,
    LEVEL_1_COL_NAME,
    LEVEL_2_COL_NAME,
    LEVEL_3_COL_NAME
    ]

# Column indices for a list of cells from a row in the BOA workbook.
DATE_COL_INDEX = 1
ORIGINAL_DESCRIPTION_COL_INDEX = 2  
CURRENCY_COL_INDEX = 4  
AMOUNT_COL_INDEX = 6  
ACCOUNT_NAME_COL_INDEX = 10  
BUDGET_CATEGORY_COL_INDEX = 12  

BOA_WB_COL_DIMENSIONS = {
    DATE_COL_NAME: 12,
    ORIGINAL_DESCRIPTION_COL_NAME: 95,
    CURRENCY_COL_NAME: 9,
    AMOUNT_COL_NAME: 14,
    ACCOUNT_NAME_COL_NAME: 68,
    BUDGET_CATEGORY_COL_NAME: 40,
    LEVEL_1_COL_NAME: 20,
    LEVEL_2_COL_NAME: 20,
    LEVEL_3_COL_NAME: 20
}
BUDMAN_SHEET_NAME = "TransactionData"

BUDMAN_REQUIRED_COLUMNS = [
    DATE_COL_NAME, ORIGINAL_DESCRIPTION_COL_NAME, CURRENCY_COL_NAME,
    AMOUNT_COL_NAME, ACCOUNT_CODE_COL_NAME, BUDGET_CATEGORY_COL_NAME,
    LEVEL_1_COL_NAME, LEVEL_2_COL_NAME, LEVEL_3_COL_NAME
]

#endregion Globals and Constants
#region dataclasses
@dataclass
class TransactionData:
    """Data class to hold transaction data."""
    tid: str = None   # Transaction ID.
    date: datetime.date = None  # Transaction date - ISO 8601 format.
    description: str = None  # Transaction description.
    currency: str = None  # Transaction currency.
    account: str = None  # Account code.
    amount: float = 0.0  # Transaction amount.
    category: str = None  # Transaction budget category.

    def data_str(self) -> str:
        """Return a string representation of the transaction data."""
        ret =  f"{self.tid:12}|{self.date.strftime("%m/%d/%Y")}|"
        ret += f"{self.account:26}|" 
        ret += f"{self.amount:>+12.2f}|" 
        ret += f"({len(self.description):03}){self.description:102}|->" 
        ret += f"|({len(self.category):03})|{self.category:40}|"
        return ret
    
#endregion dataclasses
# ---------------------------------------------------------------------------- +
#region generate_hash_key(text:str) -> str
def generate_hash_key(text: str, length:int=12) -> str:
    """Generate a hash key for the given text.
    
    The hash key is generated by removing all non-alphanumeric characters 
    from the text, converting it to lowercase, and then hashing it using 
    SHA-256. The resulting hash is then converted to a hexadecimal string.

    Args:
        text (str): The input text to generate a hash key for.

    Returns:
        str: The generated hash key.
    """
    try:
        if not isinstance(text, str):
            raise TypeError(f"Expected 'text' arg to be a str, got {type(text)}")
        # Remove non-alphanumeric characters and convert to lowercase.
        # cleaned_text = re.sub(r'\W+', '', text).lower()
        # Generate the SHA-256 hash of the cleaned text.
        hash_object = hashlib.sha256(text.encode())
        # Convert the hash to a hexadecimal string.
        return hash_object.hexdigest()[:length]  #.[:HASH_KEY_LENGTH]
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion generate_hash_key(text:str) -> str
# ---------------------------------------------------------------------------- +
#region check_sheet_columns() function
def check_sheet_columns(sheet: Worksheet, add_columns: bool = True) -> bool:
    """Check that the sheet is ready to process transactions.
    
    BudMan uses these columns to process transactions in a sheet:
    1. 'Date' - the date of the transaction. It is a datetime.date object.
    2. 'Original Description' - the description of the transaction. It is a str.
    3. 'Amount' - the amount of the transaction. It is a float.
    4. 'Account Code' - the short-name of the account for the transaction. It is a str.
    5. 'Budget Category' - the budget category for the transaction. It is a str.
    6. 'Level1' - the first level of the budget category. It is a str.
    7. 'Level2' - the second level of the budget category. It is a str.
    8. 'Level3' - the third level of the budget category. It is a str.

    Args:
        sheet (openpyxl.worksheet): The worksheet to check.
        add_columns (bool): Whether to add missing columns or not.

    Returns:
        bool: True if all required columns are present, False otherwise.
    """
    try:
        logger.info("Check worksheet for required columns.")
        # Index the header row 1, it has the column names
        col_names = [cell.value for cell in sheet[1]]
        logger.debug(f"Column names in sheet('{sheet.title}'): {col_names}")
        
        # Check if all required columns are present
        missing_columns = [col for col in BUDMAN_REQUIRED_COLUMNS if col not in col_names]
        
        if not missing_columns:
            logger.info("All required columns are present.")
            return True
        
        if add_columns:
            # Add missing columns to the sheet to the right.
            for col_name in missing_columns:
                i = sheet.max_column + 1 # insert before column after last column
                sheet.insert_cols(i)
                sheet.cell(row=1, column=i).value = col_name
                # Set column width based on predefined dimensions
                width = BOA_WB_COL_DIMENSIONS.get(col_name, 20)
                sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width
                logger.info(f"Adding column '{col_name}' at index = {i}, "
                            f"column_letter = '{sheet.cell(row =1, column=i).column_letter}'")
            logger.info(f"Added missing columns: {', '.join(missing_columns)}")
        else:
            logger.error(f"Missing required columns: {', '.join(missing_columns)}")
            return False
        logger.info("Completed checks for required columns.")
        return True
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion check_sheet_columns() function
# ---------------------------------------------------------------------------- +
#region check_sheet_schema() function
def check_sheet_schema(wb: Workbook, correct: bool = False) -> bool:
    """Check that the sheet is ready to process transactions.
    
    Steps to check the active sheet of an excel workbook for BudMan processing:
    1. SHould be just 1 worksheet, the active worksheet.
    2. Title must be 'TransactionData'. BOA_SHEET_NAME
    3. Column names must match spelling an order from BOA_WB_COLUMNS.
    Args:
        wb (openpyxl.workbook): The workbook to check.
        correct (bool): Whether to correct the sheet schema or not.

    Returns:
        bool: True if all required columns are present, False otherwise.
    """
    try:
        logger.info("Check worksheet for schema structure.")
        # Check the active worksheet.
        ws = wb.active  # Get the active worksheet.
        good_schema = True  # Assume the schema is good.
        rule1 = rule2 = rule3 = good_schema
        # 1. Check if there is only one worksheet.
        sheet_names = wb.sheetnames 
        sheet_count = len(sheet_names)
        budman_sheet_index = -1
        if sheet_count > 1:
            logger.error(f"Workbook has {sheet_count} sheets, expected 1. "
                         f"Sheet names: {', '.join(sheet_names)}")
            good_schema = rule1 = False
        # 2. Check if the sheet title is BUDMAN_SHEET_NAME.
        if BUDMAN_SHEET_NAME not in sheet_names:
            logger.error(f"Workbook does not have a sheet named '{BUDMAN_SHEET_NAME}'. "
                         f"Sheet names: {', '.join(sheet_names)}")
            good_schema = rule2 = False
        else:
            # Get the index of the BudMan sheet.
            budman_sheet_index = sheet_names.index(BUDMAN_SHEET_NAME)
            ws = wb[sheet_names[budman_sheet_index]] 
        # 3. Check the column names are correct spelling and order.
        col_names = [cell.value for cell in ws[1]]
        # Check if all required columns are present
        missing_columns = [col for col in BUDMAN_REQUIRED_COLUMNS if col not in col_names]
        if len(missing_columns) > 0:
            logger.error(f"Missing required columns: {', '.join(missing_columns)}")
            good_schema = rule3 = False
        # If just one sheet with wrong title, rename it to BUDMAN_SHEET_NAME.
        if rule1 and rule3 and not rule2:
            logger.error(f"BudMan sheet '{BUDMAN_SHEET_NAME}' not found in workbook.")
            previous_title = ws.title  # Save the previous title.
            ws.title = BUDMAN_SHEET_NAME  # Rename the active sheet.
            logger.info(f"Renamed active sheet from '{previous_title}' to '{BUDMAN_SHEET_NAME}'.")
            good_schema = True
        # If more than one sheet, active sheet with wrong title, but is has 
        # the correct columns, rename the active sheet to BUDMAN_SHEET_NAME.
        if not rule1 and rule3 and not rule2:
            logger.error(f"Active sheet '{BUDMAN_SHEET_NAME}' has wrong name "
                         f"but has the correct columns. " 
                         f"Renaming it to '{BUDMAN_SHEET_NAME}'.")
            previous_title = ws.title  # Save the previous title.
            ws.title = BUDMAN_SHEET_NAME  # Rename the active sheet.
            logger.info(f"Renamed active sheet from '{previous_title}' to '{BUDMAN_SHEET_NAME}'.")
            good_schema = True
        return good_schema
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion check_sheet_schema() function
# ---------------------------------------------------------------------------- +
#region check_budget_category() function
def check_budget_category(sheet:Worksheet,add_columns : bool = True) -> bool:
    """Check that the sheet is ready to process budget category.
    
    BudMan uses 4 columns to categorize transactions in a sheet:
    1. 'Budget Category' - the budget category for the transaction. It is a str
        with up to 3 dotted levels of budget categories, e.g. "Housing.Improvements.Flooring".
    2. 'Level1' - the first level of the budget category, e.g. "Housing".
    3. 'Level2' - the second level of the budget category, e.g. "Improvements".
    4. 'Level3' - the third level of the budget category, e.g. "Flooring".
    A column 'Budget Category' is added to the sheet if it does not exist.
    3 columns are added to the sheet if they do not exist: 'Level1', 'Level2', 
    'Level3', adjacent to the 'Budget Category' column.

    Args:
        sheet (openpyxl.worksheet): The worksheet to map.
    """
    try:
        logger.info("Check worksheet for budget category columns.")
        # Index row 1 as the header row, it has the column names
        col_names = {}
        for cnum, cell in enumerate(sheet[1], start=1):
            col_names[cell.value] = cnum
        logger.debug(f"Column names in sheet('{sheet.title}'): {list(col_names.keys())}")
        # Is BUDGET_CATEGORY_COL_NAME in the sheet?
        if BUDGET_CATEGORY_COL_NAME in col_names:
            logger.info(f"Column '{BUDGET_CATEGORY_COL_NAME}' already exists in sheet.")
        elif add_columns:
            # Add the BUDGET_CATEGORY_COL_NAME, LEVEL_1_COL, LEVEL_2_COL, and 
            # LEVEL_3_COL columns to the sheet.
            i = sheet.max_column + 1
            sheet.insert_cols(i)
            sheet.cell(row=1, column=i).value = BUDGET_CATEGORY_COL_NAME
            col_names[BUDGET_CATEGORY_COL_NAME] = i  # Update the col_names dict.
            # Set the column width to 20.
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 20
            logger.info(f"Adding column '{BUDGET_CATEGORY_COL_NAME}' at index = {i}, "
                        f"column_letter = '{sheet.cell(row=1, column=i).column_letter}'")
            logger.info(f"Completed checks for budget category.")
        return True
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise    
#endregion check_budget_category() function
# ---------------------------------------------------------------------------- +
#region WORKSHEET_data(ws:Worksheet) -> List[TransactionData]
def WORKSHEET_data(ws:Worksheet, just_values:bool=False) -> list[TransactionData]:
    """Extract transaction data from a worksheet.

    Args:
        ws (Worksheet): The worksheet to extract data from.

    Returns:
        list[TransactionData]: A list of TransactionData objects.
    """
    try:
        if not isinstance(ws, Worksheet):
            raise TypeError(f"Expected 'ws' arg to be a Worksheet, got {type(ws)}")
        transactions = []
        for row in ws.iter_rows(min_row=2, values_only=just_values):
            transaction = WORKSHEET_row_data(row)  # Extract transaction data from the row.
            transactions.append(transaction)
        return transactions
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKSHEET_data(ws:Worksheet) -> List[TransactionData]
# ---------------------------------------------------------------------------- +
#region WORKSHEET_row_data(row:list) -> TransactionData
def WORKSHEET_row_data(row:tuple,hdr:list) -> TransactionData:
    """Extract transaction data from a worksheet row.

    Args:
        row (list): Array of cell values.
        hdr (list): The header list used to index values from the row.

    Returns:
        list[TransactionData]: A list of TransactionData objects.
    """
    try:
        p3u.is_not_obj_of_type("row", row, tuple, raise_TypeError=True)
        p3u.is_not_obj_of_type("hdr", hdr, list, raise_TypeError=True)
        # TODO: Verify and obtain the necessary index values from the header.
        # If the hdr is incorrect, we are screwed
        date_i = hdr.index(DATE_COL_NAME) if DATE_COL_NAME in hdr else -1
        desc_i = hdr.index(ORIGINAL_DESCRIPTION_COL_NAME) if ORIGINAL_DESCRIPTION_COL_NAME in hdr else -1
        currency_i = hdr.index(CURRENCY_COL_NAME) if CURRENCY_COL_NAME in hdr else -1
        amount_i = hdr.index(AMOUNT_COL_NAME) if AMOUNT_COL_NAME in hdr else -1
        account_i = hdr.index(ACCOUNT_NAME_COL_NAME) if ACCOUNT_NAME_COL_NAME in hdr else -1
        acct_code_i = hdr.index(ACCOUNT_CODE_COL_NAME) if ACCOUNT_CODE_COL_NAME in hdr else -1
        budget_cat_i = hdr.index(BUDGET_CATEGORY_COL_NAME) if BUDGET_CATEGORY_COL_NAME in hdr else -1
        if -1 in (date_i, desc_i, currency_i, amount_i, account_i, 
                  acct_code_i,budget_cat_i):
            raise ValueError(f"Missing required columns in header: {hdr}. "
                             f"Expected columns: {BUDMAN_REQUIRED_COLUMNS[1:]}")
        if isinstance(row[0], Cell):
            # This is Tuple of Cell objects.
            t_date_str = p3u.iso_date_only_string(row[date_i].value)  # Date is in the second column.
            t_desc = row[desc_i].value  # Description is in the third column.
            t_currency = row[currency_i].value  # Currency is in the fifth column.
            t_amt_str = str(row[amount_i].value)
            t_acct_str = row[account_i].value 
        else:
            # This is Tuple of Cell values-only.
            t_date_str = p3u.iso_date_only_string(row[date_i])  # Date is in the second column.
            t_desc = row[desc_i]  # Description is in the third column.
            t_currency = row[currency_i]  # Currency is in the fifth column.
            t_amt_str = str(row[amount_i]) 
            t_acct_str = row[account_i] 
        t_all_str = t_date_str + t_desc + t_currency + t_amt_str + t_acct_str
        t_acct_code = t_acct_str.split('-')[-1].strip()  # Get the last part of the account name.
        t_id = generate_hash_key(t_all_str)  # Generate a unique ID for the transaction.
        if isinstance(row[0], Cell):
            # Update the mapped acct_code in the cell.
            row[acct_code_i].value = t_acct_code  # Account code is in the 'Account Code' column.
            transaction = TransactionData(
                tid=t_id,
                date=row[date_i].value,
                description=row[desc_i].value,
                currency=row[currency_i].value,
                amount=row[amount_i].value,
                account=t_acct_code,  #row[account_i].value,
                category=row[budget_cat_i].value  # Category will be set later.
            )
        else:
            transaction = TransactionData(
                tid=t_id,
                date=p3u.iso_date_only_string(row[date_i]),
                description=row[desc_i],
                currency=row[currency_i],
                amount=float(row[amount_i]),
                account=t_acct_code,  #row[account_i],
                category=row[budget_cat_i]  # Category will be set later.
            )
        return transaction
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKSHEET_row_data(row:list) -> TransactionData
# ---------------------------------------------------------------------------- +
#region split_budget_category() -> tuple function
def split_budget_category(budget_category: str) -> tuple[str, str, str]:
    """Split a budget category string into three levels.
    
    The budget category is expected to be in the format "Level1.Level2.Level3".
    If the budget category does not have all three levels, the missing levels 
    will be set to an empty string.

    Args:
        budget_category (str): The budget category string to split.

    Returns:
        tuple[str, str, str]: A tuple containing Level1, Level2, and Level3.
    """
    try:
        if not isinstance(budget_category, str):
            raise TypeError(f"Expected 'budget_category' to be a str, got {type(budget_category)}")
        l1 = l2 = l3 = ""
        c = budget_category.count('.')
        if c >= 2:
            # Split the budget category by '.' and ensure we have 3 parts.
            l1, l2, l3 = budget_category.split('.',3)
        elif c == 1:
            # Split the budget category by '.' and ensure we have 2 parts.
            l1, l2 = budget_category.split('.',2)
        else:
            # If no '.' is present, treat the whole string as Level1.
            l1 = budget_category
        return l1, l2, l3
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion split_budget_category() function
# ---------------------------------------------------------------------------- +
#region map_budget_category() function
def map_budget_category(sheet:Worksheet,src,dst) -> None:
    """Map a src column to budget category putting result in dst column.
    
    The sheet has banking transaction data in rows and columns. 
    Column 'src' has the text presumed to be about the transaction.
    Column 'dst' will be assigned a mapped budget category. Append
    column 'Budget Category' if it is not already in the sheet.
    TODO: pass in a mapper function to map the src text to a budget category.

    Args:
        sheet (openpyxl.worksheet): The worksheet to map.
        src (str): The source column to map from.
        dst (str): The destination column to map to. 
    """
    try:
        _ = p3u.is_str_or_none("src", src, raise_TypeError=True)
        _ = p3u.is_str_or_none("dst", dst, raise_TypeError=True)
        rules_count = category_map_count()
        logger.info(f"Applying '{rules_count}' budget category mappings "
                    f"to {sheet.max_row-1} rows in sheet: '{sheet.title}' ")
        # transactions = WORKSHEET_data(sheet)
        # A row is a tuple of the Cell objects in the row. Tuples are 0-based
        # header_row is a list, also 0-based. So, using the index(name) will 
        # give the cell from a row tuple matching the column name in header_row.
        header_row = [cell.value for cell in sheet[1]] 
        # header_row.insert(0, "ignore item 0")  # Add an 'ignore' column at index 0.

        if src in header_row:
            src_col_index = header_row.index(src)
        else:
            logger.error(f"Source column '{src}' not found in header row.")
            return
        if dst in header_row:
            dst_col_index = header_row.index(dst)
        else:
            logger.error(f"Destination column '{dst}' not found in header row.")
            return
        
        # TODO: need to refactor this to do replacements by col_name or something.
        # This is specific to the Budget Category mapping, which now is to be
        # split into 3 levels: Level1, Level2, Level3.
        l1_i = header_row.index(LEVEL_1_COL_NAME) if LEVEL_1_COL_NAME in header_row else -1
        l2_i = header_row.index(LEVEL_2_COL_NAME) if LEVEL_2_COL_NAME in header_row else -1
        l3_i = header_row.index(LEVEL_3_COL_NAME) if LEVEL_3_COL_NAME in header_row else -1

        logger.info(f"Mapping '{src}'({src_col_index}) to "
                    f"'{dst}'({dst_col_index})")
        num_rows = sheet.max_row # or set a smaller limit
        other_count = 0
        for row in sheet.iter_rows(min_row=2):
            # row is a type 'tuple' of Cell objects.
            row_idx = row[0].row  # Get the row index.
            src_cell = row[src_col_index]
            dst_cell = row[dst_col_index]
            src_value = row[src_col_index].value 
            dst_value = map_category(src_value)
            row[dst_col_index].value = dst_value 
            dst_cell.value = dst_value 
            l1, l2, l3 = split_budget_category(dst_value)
            row[l1_i].value = l1 if l1_i != -1 else None
            row[l2_i].value = l2 if l2_i != -1 else None
            row[l3_i].value = l3 if l3_i != -1 else None
            transaction = WORKSHEET_row_data(row,header_row) 
            trans_str = transaction.data_str()
            del transaction  # Clean up the transaction object.
            if dst_value == 'Other':
                other_count += 1
                logger.debug(f"{row_idx:04}:{trans_str}" )
        logger.info(f"Completed budget category mapping for '{num_rows}' rows. "
                    f"Other count: '{other_count}'.")
        return None
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise    
#endregion map_budget_category() function
# ---------------------------------------------------------------------------- +
#region def execute_worklow_categorization(bm : BudgetModel, fi_key: str) -> None:
def execute_worklow_categorization(bm : BudgetDomainModel, fi_key: str, wf_key:str) -> None:
    """Process categorization wf_key for Financial Institution's 
    transaction workbooks.

    Execute the categorization wf_key to examine all fi transaction 
    workbooks presently in the IF (Incoming Folder) for the indicated 
    FI (Financial Institution). Each workbook file is opened and the 
    transactions are categorized and saved to the CF (Categorized Folder) 
    for the indicated FI.

    Args:
        bm (BudgetModel): The BudgetModel instance to use for processing.
        fi_key (str): The key for the financial institution.
    """
    # TODO: add logs directory to the budget folder.
    st = p3u.start_timer()
    wb_name = "BOAChecking2025.xlsx"
    cp = "Budget Model Categorization:"
    loaded_workbooks = {}
    # Execute a workflow for a specific financial institution (fi_key).
    #   The pattern is to apply a function based on wf_key to each item in 
    #   the input folder based on fi_key and wf_key.
    #   Execute steps:
    #     1: Load the input items from storage.
    #     2: Apply the workflow function to each input item.
    #     3: Save the output items to storage. 
    try:
        logger.info(f"{cp} Start: workflow: '{wf_key}' for FI('{fi_key}') ...")
        wb_type = WF_INPUT # input workbooks
        wb_c = bm.bdm_FI_WF_WORKBOOK_LIST_count(fi_key, wf_key, wb_type)
        # workbooks_dict = bm.bsm_WF_INPUT(fi_key, BDM_WF_INTAKE)
        # if workbooks_dict is None or len(workbooks_dict) == 0:
        if wb_c is None or wb_c == 0:
            logger.info(f"{cp}    No workbooks for input.")
            return
        logger.info(f"{cp}    {wb_c} workbooks for input.")
        # Now process each input workbook.
        # for wb_name, wb_ap in reversed(workbooks_dict.items()):
        # Step 1: Load the workbooks sequentially.
        for wb_name, wb in bm.bsm_FI_WF_WORKBOOKS_generate(fi_key, wf_key, wb_type):
            logger.info(f"{cp}    Workbook({wb_name})")
                
            # Step 2: Process the workbooks applying the workflow function
            try: 
                logger.info(f"{cp}    Workbook({wb_name})")
                sheet = wb.active
                # Check for budget category column, add it if not present.
                check_budget_category(sheet)
                # Map the 'Original Description' column to the 'Budget Category' column.
                map_budget_category(sheet, "Original Description", BUDGET_CATEGORY_COL_NAME)
            except Exception as e:
                logger.error(f"{cp}    Error processing workbook: {wb_name}: {e}")
                continue

            # Step 3: Save the output items to output storage.
            try:
                bm.bsm_FI_WF_WORKBOOK_save(wb, wb_name, fi_key, wf_key, WF_OUTPUT)
            except Exception as e:
                logger.error(f"{cp}    Error saving workbook: {wb_name}: {e}")
                continue
        logger.info(f"{cp} Complete: wf_key: '{wf_key}' {p3u.stop_timer(st)}")
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion execute_worklow_categorization() function
# ---------------------------------------------------------------------------- +