# ---------------------------------------------------------------------------- +
#region budget_domain_model.py module
""" budget_domain_model.py implements the class BudgetDomainModel.

    Following a rough MVVM pattern, the BudgetDomainModel class is acting as 
    both the cohesive Model, representing a budget model domain in
    memory as object API. 

    At present, the BudgetDomainModel class is a singleton class that manages the 
    mapping of excel "workbooks" stored in the filesystem to the budget model.
    The BudgetDomainModel class presents properties and methods to the outside world.
    Methods are separated into DomainModel-ish methods for the 
    Budget Domain Model (BDM), which is the in-memory data model,
    and StorageModel-ish methods for the Budget Storage Model (BSM), 
    which is the filesystem model structure.

    In the Budget Domain Model, a data pipeline pattern is used, anticipating 
    "raw data" will be introduced from financial institutions (FI) and and 
    processed through a series of workflow tasks. Raw data is a "workbook", 
    often an excel file, or a .cvs file.
    
    A "folder" concept is aligned with the stages of a pipeline as a 
    series of "workflows" applied to the data. Roughly, workflow task works 
    on data in its associated "work" folder, processing data inplace locally 
    or as modifications or output to workbooks in other folders. Configuration 
    is used to map the BDM folders to the BSM folders. But, a simple workflow 
    scheme is for a task to have a designated input, work and output folder in 
    BDM. Incoming data appears in the input folder. Tasks own the responsibility
    to move it into their work folder, and if and when to move it to the
    output folder. The workflow pipeline is roughly:

    Workflow: a function grouping of tasks to be applied to input data.
    workflowTaskFunc(InputFolder, Workfolder) -> OutputFolder

    Workflows are functional units of work with clearly defined concerns applied
    to input data and producing work and outout data. 

    Key Concepts:
    -------------
    - Folders: containers of workbooks associated with a workflow stage.
    - Workflows: a defined set of process functions applied to data.
    - Workbooks: a container of financial transaction data, e.g., excel file.
    - Raw Data: original data from a financial institution (FI), read-only.
    - Financial Budget: the finalized output, composed of workbooks,
        representing time-series transactions categorized by payments and
        deposits (debits and credits).
    - Financial Institution (FI): a bank or brokerage financial institution.
        
    Perhaps the primary domain should be "Financial Budget" instead of
    "Budget Model". The FB Domain and the "Storage Domain" or Sub-Domain. 

    The Budget Domain Model has the concern of presenting an API that is 
    independent of where the workbook data is sourced and stored. The 
    Budget Storage Model has the concern of managing sourcing and storing 
    workbooks in a storage system. Currently, it just maps the BDM structure to 
    filesystem folders and files.

    Assumptions:
    - The FI transaction data resides in a folder specified in the budget_config.
    - Banking transaction files are typically excel spreadsheets, referred to 
      as workbooks. 
    - A workbook's data content starts in cell A1.
    - Row 1 contains column headers. All subsequent rows are data with no blank
      rows until the end of the data.
"""
#endregion budget_domain_model.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
from abc import ABCMeta
import logging, os, getpass, time, copy
from pathlib import Path
from typing import List, Optional, Type, Generator, Dict, Tuple, Any, TYPE_CHECKING
# third-party modules and packages
import p3_utils as p3u, pyjson5, p3logging as p3l, p3_mvvm as p3m
from openpyxl import Workbook, load_workbook
# local modules and packages
from budman_namespace import *
from .budget_domain_model_config import BDMConfig
from budget_storage_model import (
    BSMFileTree,
    bsm_verify_folder, 
    bsm_BDM_STORE_url_put,
    bsm_get_workbook_names,
    bsm_BDM_STORE_to_json
    )                              
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)
# ---------------------------------------------------------------------------- +
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +
class BudgetDomainModel(p3m.Model_Base,metaclass=BDMSingletonMeta):
    # ======================================================================== +
    #region BudgetDomainModel class intrinsics
    # ======================================================================== +
    """BudgetDomainModel class implementing Singleton pattern.
    
        A singleton class to manage the BudgetDomainModel for the application.
        This class is used to store and manage the budget data, including
        the budget folder, institutions, workbooks, and options.

        When instantiated, a BDMConfig object is provided to the constructor.
    """
    # ------------------------------------------------------------------------ +
    #region    BudgetDomainModel class constructor __init__()
    def __init__(self, bdm_config : BDMConfig = None) -> None:
        """Constructor for the BudgetDomainModel class.

        This initial constructor is kept simple, returning an instance with
        empty initial values, with one critical exception: a valid BDMConfig
        object is required, created by one of the methods on the BDMConfig
        class. The bdm_config object is assigned as the initial value of
        the BDM_STORE_OBJECT property. The rest of the BDM content is hydrated
        later during initialization.

        Args:
            bdm_config (BDMConfig): A valid BDMConfig object used later
            to initialize and rehydrate the model state.
        """
        logger.debug("Start:")
        # If a valid BDMConfig object is not provided, raise an error.
        p3u.is_not_obj_of_type("bdm_config",bdm_config,BDMConfig,raise_error=True)
        setattr(self, BDM_ID, None)
        setattr(self, BDM_STORE_OBJECT, bdm_config) # CRITICAL to later initialization.
        setattr(self, BDM_INITIALIZED, False)
        setattr(self, BDM_FILENAME, None)
        setattr(self, BDM_FILETYPE, None)
        setattr(self, BDM_FOLDER, None)  
        setattr(self, BDM_URL, None)  
        setattr(self, BDM_FI_COLLECTION, {})
        setattr(self, BDM_WF_COLLECTION, {}) 
        setattr(self, BDM_OPTIONS, {})
        setattr(self, BDM_CREATED_DATE, p3u.now_iso_date_string()) 
        setattr(self, BDM_LAST_MODIFIED_DATE, self._created_date)
        setattr(self, BDM_LAST_MODIFIED_BY, getpass.getuser())
        setattr(self, BDM_DATA_CONTEXT, {})  
        setattr(self, BSM_FILE_TREE, None)  
        logger.debug("Complete:")
    #endregion BudgetDomainModel class constructor __init__()
    # ------------------------------------------------------------------------ +
    #region    BudgetDomainModel internal class methods
    def to_dict(self):
        '''Return BudgetDomainModelTemplate dictionary object. Used for serialization.'''
        ret = {
            BDM_ID: self.bdm_id,
            BDM_STORE_OBJECT: self.bdm_store_object,
            BDM_INITIALIZED: self.bdm_initialized,
            BDM_FILENAME: self.bdm_filename,
            BDM_FILETYPE: self.bdm_filetype,
            BDM_FOLDER: self.bdm_folder,
            BDM_URL: self.bdm_url,
            BDM_FI_COLLECTION: self.bdm_fi_collection,
            BDM_WF_COLLECTION: self.bdm_wf_collection,
            BDM_OPTIONS: self.bdm_options,
            BDM_CREATED_DATE: self.bdm_created_date,
            BDM_LAST_MODIFIED_DATE: self.bdm_last_modified_date,
            BDM_LAST_MODIFIED_BY: self.bdm_last_modified_by,
            BDM_DATA_CONTEXT: self.bdm_data_context,
        }
        return ret
    def __getitm__(self, key: str) -> Any:
        """Get an item from the BudgetDomainModel by key."""
        if hasattr(self, key):
            return getattr(self, key)
        raise KeyError(f"Key '{key}' not found in BudgetDomainModel.")
    def __setitm__(self, key: str, value: Any) -> None:
        """Set an item in the BudgetDomainModel by key."""
        if hasattr(self, key):
            setattr(self, key, value)
        else:
            raise KeyError(f"Key '{key}' not found in BudgetDomainModel.")
    def __delitm__(self, key: str) -> None:
        """Delete an item from the BudgetDomainModel by key."""
        if hasattr(self, key):
            delattr(self, key)
        else:
            raise KeyError(f"Key '{key}' not found in BudgetDomainModel.")
    def __repr__(self) -> str:
        ''' Return a str representation of the BudgetDomainModel object '''
        ret = f"{{ "
        ret += f"'{BDM_ID}': {self.bdm_id}, "
        ret += f"'{BDM_STORE_OBJECT}': {self.bdm_store_object}, "
        ret += f"'{BDM_INITIALIZED}': {self.bdm_initialized}, "
        ret += f"'{BDM_FOLDER}': '{self.bdm_folder}', "
        ret += f"'{BDM_FI_COLLECTION}': '{self.bdm_fi_collection}', "
        ret += f"'{BDM_URL}': '{self.bdm_url}', "
        ret += f"'{BDM_WF_COLLECTION}': '{self.bdm_wf_collection}', "
        ret += f"'{BDM_OPTIONS}': '{self.bdm_options}', "
        ret += f"'{BDM_CREATED_DATE}': '{self.bdm_created_date}', "
        ret += f"'{BDM_LAST_MODIFIED_DATE}': '{self.bdm_last_modified_date}', "
        ret += f"'{BDM_LAST_MODIFIED_BY}': '{self.bdm_last_modified_by}', "
        return ret
    def __str__(self) -> str:
        ''' Return a str representation of the BudgetDomainModel object '''
        ret = f"{self.__class__.__name__}({str(self.bdm_folder)}): "
        ret += f"{BDM_ID} = {self.bdm_id}, "
        ret += f"{BDM_STORE_OBJECT} = {str(self.bdm_store_object)}, "
        ret += f"{BDM_INITIALIZED} = {str(self.bdm_initialized)}, "
        ret += f"{BDM_FOLDER} = '{str(self.bdm_folder)}', "
        ret += f"{BDM_FI_COLLECTION} = [{', '.join([repr(fi_key) for fi_key in self.bdm_fi_collection.keys()])}], "
        ret += f"{BDM_URL} = '{self.bdm_url}' "
        ret += f"{BDM_WF_COLLECTION} = '{self.bdm_wf_collection}' "
        ret += f"{BDM_OPTIONS} = '{self.bdm_options}' "
        ret += f"{BDM_CREATED_DATE} = '{self.bdm_created_date}', "
        ret += f"{BDM_LAST_MODIFIED_DATE} = '{self.bdm_last_modified_date}', "
        ret += f"{BDM_LAST_MODIFIED_BY} = '{self.bdm_last_modified_by}', "
        return ret
    #endregion BudgetDomainModel internal class methods
    # ------------------------------------------------------------------------ +
    #region    BudgetDomainModel (BDM) properties
    @property
    def bdm_initialized(self) -> bool:
        """The initialized value."""
        return self._initialized
    @bdm_initialized.setter
    def bdm_initialized(self, value )-> None:
        """Set the initialized value."""
        self._initialized = value

    @property
    def bdm_id(self) -> str:
        """The budget model ID."""
        return getattr(self, BDM_ID)
    @bdm_id.setter
    def bdm_id(self, value: str) -> None:
        """Set the budget model ID."""
        if not isinstance(value, str):
            raise ValueError(f"bm_id must be a string: {value}")
        setattr(self, BDM_ID, value)

    @property
    def bdm_store_object(self) -> object:
        """The budget model configuration object."""
        return getattr(self, BDM_STORE_OBJECT)
    @bdm_store_object.setter
    def bdm_store_object(self, value: object) -> None:
        """Set the budget model configuration object."""
        if not isinstance(value, object):
            raise ValueError(f"bm_config_object must be an object: {value}")
        setattr(self, BDM_STORE_OBJECT, value)

    @property
    def bdm_filename(self) -> str:
        """The bdm_store filename is a string, e.g., 'bdm_store"""
        return self._bdm_filename
    @bdm_filename.setter
    def bdm_filename(self, value: str) -> None:
        """Set the bdm_store filename."""
        self._bdm_filename = value

    @property
    def bdm_filetype(self) -> str:
        """The bdm_store filetype, e.g., '.jsonc"""
        return self._bdm_filetype
    @bdm_filetype.setter
    def bdm_filetype(self, value: str) -> None:
        """Set the bdm_store filetype."""
        self._bdm_filetype = value

    @property
    def bdm_folder(self) -> str:
        """The budget folder path is a string, e.g., '~/OneDrive/."""
        return self._budget_folder
    @bdm_folder.setter
    def bdm_folder(self, value: str) -> None:
        """Set the budget folder path."""
        self._budget_folder = value

    @property
    def bdm_url(self) -> str:
        """The budget manager store file url."""
        return self._bdm_url
    @bdm_url.setter
    def bdm_url(self, value: str) -> None:
        """Set the budget manager store file url."""
        self._bdm_url = value

    @property
    def bdm_fi_collection(self) -> FI_COLLECTION_TYPE:
        """The financial institutions collection."""
        return self._financial_institutions
    @bdm_fi_collection.setter
    def bdm_fi_collection(self, value: FI_COLLECTION_TYPE) -> None:
        """Set the financial institutions collection."""
        self._financial_institutions = value

    @property
    def bdm_fi_count(self) -> int:
        """Return the number of Financial Institutions in the collection."""
        if self._financial_institutions is not None:
            return len(self._financial_institutions)
        return 0

    @property
    def bdm_wf_collection(self) -> WF_COLLECTION_TYPE:
        """The workflow collection."""
        return self._workflows
    @bdm_wf_collection.setter
    def bdm_wf_collection(self, value: WF_COLLECTION_TYPE) -> None:
        """Set the workflows collection."""
        self._workflows = value

    @property
    def bdm_wf_count(self) -> int:
        """Return the number of workflows in the collection."""
        if self._workflows is not None:
            return len(self._workflows)
        return 0

    @property
    def bdm_options(self) -> BDM_OPTIONS_TYPE:
        """The budget model options dictionary."""
        return self._options
    @bdm_options.setter
    def bdm_options(self, value: BDM_OPTIONS_TYPE) -> None:
        """Set the budget model options dictionary."""
        self._options = value

    @property
    def bdm_created_date(self) -> str:
        """The created date."""
        return self._created_date
    @bdm_created_date.setter
    def bdm_created_date(self, value: str) -> None:  
        """Set the created date."""
        self._created_date = value

    @property
    def bdm_last_modified_date(self) -> str:
        """The last modified date."""
        return self._last_modified_date
    @bdm_last_modified_date.setter
    def bdm_last_modified_date(self, value: str) -> None:
        """Set the last modified date."""
        self._last_modified_date = value

    @property
    def bdm_last_modified_by(self) -> str:
        """The last modified by."""
        return self._last_modified_by
    @bdm_last_modified_by.setter
    def bdm_last_modified_by(self, value: str) -> None:
        """Set the last modified by."""
        self._last_modified_by = value
    
    @property
    def bdm_data_context(self) -> DATA_CONTEXT_TYPE:
        """The budget domain model data context values."""
        self._data_context = {} if self._data_context is None else self._data_context
        return self._data_context
    @bdm_data_context.setter
    def bdm_data_context(self, value: DATA_CONTEXT_TYPE) -> None:
        """Set the budget domain model working data."""
        self._data_context = value

    @property
    def bsm_file_tree(self) -> BSMFileTree:
        """The BSMFileTree object representing the file tree of the budget folder."""
        return self._bsm_file_tree
    @bsm_file_tree.setter
    def bsm_file_tree(self, value: BSMFileTree) -> None:
        """Set the BSMFileTree object."""
        if not (value is None or isinstance(value, BSMFileTree)):
            raise ValueError(f"bsm_file_tree must be a BSMFileTree or None: {value}")
        self._bsm_file_tree = value

    #BDMBaseInterface properties
    @property
    def model_id(self) -> str:
        """Return the model ID."""
        return __class__.__name__ + "-" + self.bdm_id

    #endregion BudgetDomainModel (BDM)) properties
    # ------------------------------------------------------------------------ +
    #endregion BudgetDomainModel class intrinsics
    # ======================================================================== +

    # ======================================================================== +
    #region    BDM - Budget Domain Model methods
    # ------------------------------------------------------------------------ +    
    #region    BDM Design Notes
    """ Budget Model Domain Model (BDM) Documentation.

    Budget model Domain Model (BDM) is the conceptual model used to track and 
    analyze transactions from financial institutions (FI) overtime for the 
    purpose of following a budget. As a Domain Model, the implementation is 
    the python data structures used to represent the domain data.

    Class BudgetDomainModel is the base class for the budget model domain. As a 
    convention, while designing the structure, I adopted the use of constants
    for names of the various types and fields used for th class properties,
    default values and methods. Also, I adopted a pattern to use the
    BudgetDomainModelTemplate class, a subclass of BudgetDomainModel, as a means
    for storing the template for a budget model used for configuration and
    other conveniences. Each BudgetDomainModel created by a user, may have an 
    associated Budget Manager (BudMan) store. This is referred to as the 
    BDM_STORE and can be used as the configuration template.

    All BDM data storage is the concern of the BSM. BSM works with Path 
    objects, filenames, folders, relative, absolute path names and actual 
    filesystem operations. In the configuration data, the BDM data is stored 
    as strings. Internally, Path objects are used.

    In the BDM, all data is associated with a financial institution (FI). Data 
    is processed by workflow (WF) tasks that are configured for an FI. 
    Each FI has a unique fi_key which the api uses to keep data separated.

    Throughout, the following identifiers are used in reference to parts of
    the BDM.

    BDM_FOLDER - refers to the root folder for the budget model. BDM_FOLDER is 
    a path element, which is a str that will end up as a part of a Path str, 
    which is the BSM concern. See bsm_BDM_FOLDER methods for that.

    FI_FOLDER - refers to the root folder configured for a specific 
    Financial Institutions (FI). FI_FOLDER is a path element, which is a 
    str that will end up as a part of a Path str, which is the BSM concern.
    See bsm_FI_FOLDER methods for that. There is only one FI_FOLDER for
    each FI. The FI_FOLDER is a subfolder of the BDM_FOLDER.

    WF_INPUT_FOLDER - refers to the input folder for a specific workflow, again, 
    always in the context of a specific FI through the fi_key. WF_INPUT_FOLDER is
    a Path with various str reprs.

    WF_INPUT - is a WF_PURPOSE, indicating used for input to a workflow. 
    Likely, workbooks of this type are referenced in the collection of 
    workbooks currently in the WF_INPUT_FOLDER folder.

    WF_WORKING_FOLDER - refers to the folder for a specific workflow. It is
    a Path with various str reprs.

    WF_WORKING - is a WF_PURPOSE, indicating used to work on (input and output) 
    data for a workflow. Likely, workbooks of this type are referenced in the 
    collection of workbooks currently in the WF_WORKING_FOLDER folder. It is
    typical that working data is opened for read and write.

    WF_OUTPUT_FOLDER - refers to the output folder for a specific workflow. It is
    a Path with various str reprs.

    WF_OUTPUT - is a WF_PURPOSE, indicating used for output from a workflow. 
    Likely, workbooks of this type are referenced in the collection of 
    workbooks currently in the WF_OUTPUT_FOLDER folder. A common pattern is
    for a workflow to just copy workbooks to an output folder.

    From the configuration data, the values of BDM_FOLDER, FI_FOLDER and
    WF_FOLDER are used to construct folder and file names for the file system
    and the BSM. The BDM methods do not use Path objects. The overlap seen in 
    the WB_INPUT and WB_OUTPUT for a purpose. The workbooks are
    common in the two layers.
     
    The following naming conventions apply to the purpose of the BDM and BSM 
    methods to  separate the concerns for partial path names, full path names 
    and absolute pathname values..

    Naming Conventions:
    -------------------
    - bm_ - BudgetDomainModel class properties, e.g., bm_folder, bm_fi, etc.
    - fi_ - Related to financial institution, e.g., fi_key, fi_name, etc.
    - wf_ - Related to workflow, e.g., wf_key, wf_name, etc.
    - bdm_ - BudgetDomainModel Domain methods, concerning the in memory data model.
    - bsm_ - BudgetDomainModel Storage Model methods, folders/files stored the filesystem.
    - _path_str - is the simplest string for a path name or component of. These
    methods return str values and do not manipulate with Path objects. Some
    folder values in the BDM, for example, are simply the name of a folder, 
    not a complete path name.
    - _path - constructs a Path object using _path_str values, 
    invokes .expanduser(). 
    _abs_path - invokes .resolve() on _path results, return Path object.
    - Path objects are never serialized to .jsonc files, only _path_str values.

    Abbrevs used in method names: 
        bdm - Budget Model Domain, the domain model of the budget model.
        bms - Budget Storage Model, the filesystem storage model.
        bm - BudgetDomainModel class instance, parent of the Budget Model data structure.
        bf - budget folder - attribute, root folder, used in path objects.
        fi - financial institution dictionary, attribute of bdm.
             fi_key = int_key, short name of FI, e.g., "boa", "chase", etc.
             fi_value = dict of info about the FI.
        wf - workflow
    """
    #endregion BDM Design Notes
    # ======================================================================== +
    #region    bdm_initialize(self, bsm_init, ...)
    def bdm_initialize(self, 
                 create_missing_folders : bool = True,
                 raise_errors : bool = True
                 ) -> "BudgetDomainModel":
        """Hydrate the BDM from a BDMConfig object.

        Currently, as a singleton class, BudgetDomainModel is initialized from
        BDMConfig object, which must be the current value of the 
        self.bdm_store_object property. Applications can support different 
        techniques to create the BDMConfig using a subclass. Having a valid
        bdm_store_object property is important, lest not much can be initialized.
        This method will apply the BDM_STORE_OBJECT values
        and set the BDM state to match.

        Args:
            self.bdm_store_object property value (dict): A valid BDM_CONFIG object.
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
        """
        st = p3u.start_timer()
        logger.info(f"BizEVENT: Model setup for BudgetDomainModel (BDM)")
        try:
            # If a valid BDMConfig object is not provided, raise an error.
            p3u.is_not_obj_of_type("bdm_store_object property", 
                                   self.bdm_store_object, BDMConfig, 
                                   raise_error=True)
            bdm_config = self.bdm_store_object
            # Hydrate the BDM from the BDMConfig content.
            # No defaults here, the BDMConfig must have all attributevalues.
            setattr(self, BDM_STORE_OBJECT, bdm_config)
            setattr(self, BDM_ID, bdm_config[BDM_ID])
            setattr(self, BDM_STORE_OBJECT, bdm_config)
            setattr(self, BDM_INITIALIZED, bdm_config[BDM_INITIALIZED])
            setattr(self, BDM_FILENAME, bdm_config[BDM_FILENAME])
            setattr(self, BDM_FILETYPE, bdm_config[BDM_FILETYPE])
            setattr(self, BDM_FOLDER, bdm_config[BDM_FOLDER])  
            setattr(self, BDM_URL, bdm_config[BDM_URL])  
            setattr(self, BDM_FI_COLLECTION, copy.deepcopy(bdm_config[BDM_FI_COLLECTION]))
            setattr(self, BDM_WF_COLLECTION, copy.deepcopy(bdm_config[BDM_WF_COLLECTION])) 
            setattr(self, BDM_OPTIONS, copy.deepcopy(bdm_config[BDM_OPTIONS]))
            setattr(self, BDM_CREATED_DATE, bdm_config[BDM_CREATED_DATE]) 
            setattr(self, BDM_LAST_MODIFIED_DATE, bdm_config[BDM_LAST_MODIFIED_DATE])
            setattr(self, BDM_LAST_MODIFIED_BY, bdm_config[BDM_LAST_MODIFIED_BY])
            setattr(self, BDM_DATA_CONTEXT, bdm_config[BDM_DATA_CONTEXT])
            # Initialize the Budget Storage Model (BSM).
            # This resolves dependencies for the storage systems supported by 
            # the BSM, mapping folders and files according to the supported 
            # storage system, e.g., filesystem, database, etc. Within the BDM,
            # all folders and files are referenced by URL.
            self.bsm_initialize(create_missing_folders, raise_errors)
            # When the BDM is constructed and then initialized, the assumption 
            # is that data was marshalled from a storage format such as json.
            # bdm_rehydrate() reinstates any native class objects based from 
            # the persisted storage format.
            self.bdm_rehydrate()
            # Load the BSM_FILE_TREE
            self.bsm_file_tree = BSMFileTree(self.bsm_BDM_FOLDER_url())
            # Initialization complete.
            self.bdm_initialized = True
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return self
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdm_initialize(self, bsm_init, ...) 
    # ------------------------------------------------------------------------ +
    #region    bdm_rehydrate() method
    def bdm_rehydrate(self) -> None:
        """Rehydrate dicts loaded from the BDM_STORE into class instances."""
        try:
            logger.debug("Start:  ...")
            # Focus on the BDM_FI_COLLECTION.
            if (self.bdm_fi_collection is None or
                len(self.bdm_fi_collection) == 0):
                logger.debug("FI_COLLECTION is empty.")
                return None
            for fi_key, fi_object in self.bdm_fi_collection.items():
                if (fi_object[FI_WORKBOOK_DATA_COLLECTION] is None or
                    not isinstance(fi_object[FI_WORKBOOK_DATA_COLLECTION], dict) or
                    len(fi_object[FI_WORKBOOK_DATA_COLLECTION]) == 0):
                    logger.debug(f"FI_KEY('{fi_key}') has no WORKBOOK_DATA_COLLECTION.")
                    continue
                fi_folder_abs_path: Path = self.bsm_FI_FOLDER_abs_path(fi_key)
                for wb_id, wb_data in fi_object[FI_WORKBOOK_DATA_COLLECTION].items():
                    if not isinstance(wb_data, dict):
                        # During this initialization, expecting the wb_data as
                        # a dict, not a BDMWorkbook object or some other type.
                        logger.warning(f"FI_KEY('{fi_key}') wb_id('{wb_id}') "
                                        f"wb_data is not a dict, skipping "
                                        f"type: {type(wb_data)}")
                        continue
                    # Check if the wb_data.wb_url still exists. If not,
                    # remove it from the collection.
                    if (WB_URL in wb_data and
                        wb_data[WB_URL] is not None and
                        isinstance(wb_data[WB_URL], str)):
                        wb_url = wb_data[WB_URL]
                        try:
                            _ = p3u.verify_url_file_path(wb_url)
                        except Exception as e: 
                            m = p3u.exc_err_msg(e)
                            logger.error(f"Error verifying WORKBOOK URL "
                                         f"'{wb_url}': {m} ")
                            logger.error(f"Skipping: FI_KEY('{fi_key}') WB_ID('{wb_id}')")
                            continue
                    # Check the wb_data content for schema compliance with
                    # the BDMWorkbook class. If not compliant, skip it.
                    wb_data_2: dict = BDMWorkbook.check_schema(wb_data)
                    # Convert the WORKBOOK_ITEM to a WORKBOOK_OBJECT.
                    wb_object = BDMWorkbook(**wb_data_2)
                    # Get the wf_folder_url expected for this workbook.
                    wb_folder_url = self.bdm_FI_WF_FOLDER_CONFIG_ATTRIBUTE(
                        fi_key=wb_object.fi_key, wf_key=wb_object.wf_key,
                        wf_purpose=wb_object.wf_purpose, attribute=WF_FOLDER_URL,
                        raise_errors=True)
                    if wb_folder_url is None:
                        # wb_folder_url not found in the fi_wf_folder_config_collection
                        # which means the fi_workbook_data_collection is out of
                        # sync with the fi_wf_folder_config_collection.
                        wf_folder_abs_path: Path = fi_folder_abs_path / wb_object.wf_folder
                        logger.warning(f"FI_KEY('{fi_key}') WF_KEY('{wb_object.wf_key}') "
                                     f"WF_PURPOSE('{wb_object.wf_purpose}'): "
                                     f"FI_WF_FOLDER_CONFIG_COLLECTION is missing "
                                     f"BDMWorkbook(wb_id='{wb_object.wb_id}').wf_folder "
                                     f"'{wb_object.wf_folder}'.")
                    # Set the wb_folder_url in the WORKBOOK_OBJECT.
                    wb_object.wb_folder_url = wb_folder_url
                    # Replace the DATA_OBJECT with the WORKBOOK_OBJECT.
                    fi_object[FI_WORKBOOK_DATA_COLLECTION][wb_id] = wb_object
                # Sort the FI_WORKBOOK_DATA_COLLECTION by wb_id, for
                # WB_INDEX order from here on.
                sorted_wdc = dict(
                    sorted(fi_object[FI_WORKBOOK_DATA_COLLECTION].items()))
                fi_object[FI_WORKBOOK_DATA_COLLECTION] = sorted_wdc
            logger.debug(f"Complete:")
            return None
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdm_rehydrate() method
    # ------------------------------------------------------------------------ +
    #region    bdm_dehydrate() method
    def bdm_dehydrate(self) -> BDM_STORE_TYPE:
        """Return a serializable BDM_STORE dict with the current BDM content."""
        try:
            logger.debug("Start:  ...")
            bdm_store: BDM_STORE_TYPE = self.to_dict()
            # Traverse the BDM_STORE and remove non-serializable objects.
            # Replace objects with known non-serializable attributes with a
            # a serializable dict copy.
            if self.bdm_fi_count == 0:
                logger.debug("FI_COLLECTION is empty.")
                return bdm_store
            for fi_key, fi_object in bdm_store[BDM_FI_COLLECTION].items():
                if not isinstance(fi_object, dict):
                    logger.warning(f"FI_KEY('{fi_key}') is not a dict, "
                                    f"skipping type: {type(fi_object)}")
                    continue
                wdc: WORKBOOK_DATA_COLLECTION_TYPE = fi_object[FI_WORKBOOK_DATA_COLLECTION]
                # If the FI_WORKBOOK_DATA_COLLECTION is empty, skip it.
                if (wdc is None or len(wdc) == 0):
                    logger.debug(f"FI_KEY('{fi_key}') WORKBOOK_DATA_COLLECTION is empty.")
                    continue
                for wb_id, bdm_wb in wdc.items():
                    if isinstance(bdm_wb, BDMWorkbook):
                        # Convert the BDMWorkbook object to a dict.
                        # Don't modify the BDMWorkbook objects
                        bdm_wb_dict = bdm_wb.to_dict()
                    elif isinstance(bdm_wb, dict):
                        bdm_wb_dict = bdm_wb
                    else:
                        logger.warning(f"FI_KEY('{fi_key}') WB_ID('{wb_id}') "
                                       f"type:({type(bdm_wb).__name__}).")
                        continue
                    # A bdm_wb dict may have an object for wb_content
                    if bdm_wb_dict[WB_CONTENT] is not None:
                        # Never serialize the wb_content, so set it to None.
                        wbc_type = type(bdm_wb_dict[WB_CONTENT]).__name__
                        logger.debug(f" Dehydrating BDMWorkbook({wb_id}): "
                                        f"wb_content type: '{wbc_type}'")
                        bdm_wb_dict[WB_CONTENT] = None
                        bdm_wb_dict[WB_LOADED] = False 
                    # Replace the bdm_wb in fi_object[FI_WORKBOOK_DATA_COLLECTION]
                    fi_object[FI_WORKBOOK_DATA_COLLECTION][wb_id] = bdm_wb_dict
            logger.debug(f"Complete:")   
            return bdm_store
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdm_dehydrate() method
    # ------------------------------------------------------------------------ +
    #region    bdm_save_model() method                       +
    def bdm_save_model(self) -> None:
        """Save the model for this view_model to the BDM_STORE."""
        try:
            st = p3u.start_timer()
            logger.info(f"Start: ...")
            # Get a Dict of the BudgetModel to store.
            bdm_dict = self.bdm_dehydrate()
            # Save the BDM_STORE file to storage.
            bsm_BDM_STORE_url_put(bdm_dict, self.bdm_url)
            logger.info(f"Saved BDM_STORE url: {self.bdm_url}")
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return None
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion bdm_save_model() method                       +
    # ------------------------------------------------------------------------ +
    #region    bdm_BDM_STORE_json() method
    def bdm_BDM_STORE_json(self) -> str:
        """Return the BDM_STORE as a JSON string."""
        try:
            bdm_store: BDM_STORE_TYPE = self.bdm_dehydrate()
            json_str: str = bsm_BDM_STORE_to_json(bdm_store)
            return json_str
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise ValueError(m)
    #endregion bdm_BDM_STORE_json() method
    # ------------------------------------------------------------------------ +
    #region    bdm_FI_OBJECT_TYPE methods
    def bdm_FI_OBJECT(self, fi_key:str) -> FI_OBJECT_TYPE:
        """Return the FI_OBJECT for fi_key."""
        self.bdm_FI_KEY_validate(fi_key)
        return self.bdm_fi_collection[fi_key]
    
    def bdm_FI_KEY(self, fi_key:str) -> str:
        """Return the FI_KEY value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_KEY]

    def bdm_FI_NAME(self, fi_key:str) -> str:
        """Return the FI_NAME value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_NAME]
    
    def bdm_FI_TYPE(self, fi_key:str) -> str:
        """Return the FI_TYPE value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_TYPE]

    def bdm_FI_FOLDER(self, fi_key:str) -> str:
        """Return the FI_FOLDER value of the FI_OBJECT for fi_key."""
        return self.bdm_FI_OBJECT(fi_key)[FI_FOLDER]
    
    def bdm_FI_WF_FOLDER_CONFIG_COLLECTION(self, 
                                    fi_key:str) -> FI_WF_FOLDER_CONFIG_COLLECTION_TYPE:
        """Return the FI_WF_FOLDER_CONFIG_COLLECTION object of the FI_OBJECT 
        for fi_key."""
        try:
            fi_obj: FI_OBJECT_TYPE = self.bdm_FI_OBJECT(fi_key)
            if (fi_obj[FI_WF_FOLDER_CONFIG_COLLECTION] is None or
                len(fi_obj[FI_WF_FOLDER_CONFIG_COLLECTION]) == 0):
                return {}
            return self.bdm_FI_OBJECT(fi_key)[FI_WF_FOLDER_CONFIG_COLLECTION]
        except ValueError as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bdm_FI_WF_FOLDER_CONFIG_COLLECTION_append(self, 
                fi_key:str, wf_key:str, wf_folder_config:dict) -> None:
        """Append a new WF_FOLDER_CONFIG to the FI_WF_FOLDER_CONFIG_COLLECTION for fi_key."""
        try:
            fi_wf_fldr_cfg_coll: FI_WF_FOLDER_CONFIG_COLLECTION_TYPE = None
            fi_wf_fldr_cfg_coll = self.bdm_FI_WF_FOLDER_CONFIG_COLLECTION(fi_key)
            if fi_wf_fldr_cfg_coll is None:
                fi_wf_fldr_cfg_coll = {}
            if wf_key not in fi_wf_fldr_cfg_coll:
                fi_wf_fldr_cfg_coll[wf_key] = []
            fi_wf_fldr_cfg_coll[wf_key].append(wf_folder_config)
            return None
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bdm_FI_WORKBOOK_DATA_COLLECTION(self, fi_key:str) -> WORKBOOK_DATA_COLLECTION_TYPE:
        """Return the FI_WORKBOOK_COLLECTION object of the FI_OBJECT for fi_key."""
        try:
            fi_obj: FI_OBJECT_TYPE = self.bdm_FI_OBJECT(fi_key)
            if (fi_obj[FI_WORKBOOK_DATA_COLLECTION] is None or
                len(fi_obj[FI_WORKBOOK_DATA_COLLECTION]) == 0):
                return {}
            return self.bdm_FI_OBJECT(fi_key)[FI_WORKBOOK_DATA_COLLECTION]
        except ValueError as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bdm_FI_WORKBOOK_DATA_COLLECTION_set(self, fi_key:str, value:DATA_COLLECTION_TYPE) -> DATA_COLLECTION_TYPE:
        """Set the FI_WORKBOOK_DATA_COLLECTION object of the FI_OBJECT for fi_key."""
        try:
            self.bdm_FI_OBJECT(fi_key)[FI_WORKBOOK_DATA_COLLECTION] = value
            return
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise ValueError(m)

    def bdm_FI_KEY_validate(self, fi_key:str) -> bool:
        """Validate the financial institution key as a member of the
        BDM_FI_COLLECTION and referencing a valid FI_OBJECT dict."""
        p3u.is_not_non_empty_str(fi_key, "fi_key",raise_error=True)
        if (self.bdm_fi_collection is None or
            len(self.bdm_fi_collection) == 0):
            m = f"Cannot validate FI_KEY('{fi_key}'), "
            m += f"BDM_FI_COLLECTION is None or empty."
            raise ValueError(m)
        if fi_key not in self.bdm_fi_collection:
            m = f"Financial Institution '{fi_key}' not found in "
            m += f"BDM_FI_COLLECTION: {list(self.bdm_fi_collection.keys())}"
            logger.error(m)
            raise ValueError(m)
        if (self.bdm_fi_collection[fi_key] is None or 
            not isinstance(self.bdm_fi_collection[fi_key], dict) or
            not FI_KEY in self.bdm_fi_collection[fi_key] or
            self.bdm_fi_collection[fi_key][FI_KEY] != fi_key):
            m = f"FI_KEY('{fi_key}') does not reference a valid FI_OBJECT "
            m += f"or does not match one BDM_FI_COLLECTION: "
            m += f"{list(self.bdm_fi_collection.keys())}"
            logger.error(m)
            raise ValueError(m)
        return True
    #endregion bdm_FI_OBJECT_TYPE methods
    # ------------------------------------------------------------------------ +
    #region    bdm_FI_WF_FOLDER_CONFIG & _ATTRIBUTE methods:
    def bdm_FI_WF_FOLDER_CONFIG(self, fi_key:str, 
                             wf_key:str, wf_purpose:str) -> Optional[str]:
        """Return the FI_WF_FOLDER_CONFIG for a given fi_key, wf_key and wf_purpose."""
        try:
            fi_wf_fldr_cfg_coll: FI_WF_FOLDER_CONFIG_COLLECTION_TYPE = None
            fi_wf_fldr_cfg_coll = self.bdm_FI_WF_FOLDER_CONFIG_COLLECTION(fi_key)
            if (fi_wf_fldr_cfg_coll is None or
                len(fi_wf_fldr_cfg_coll) == 0):
                m = f"FI_WF_FOLDER_CONFIG_COLLECTION for fi_key('{fi_key}') "
                m += "is empty or None."
                logger.debug(m)
                return None
            self.bdm_WF_KEY_validate(wf_key)
            self.bdm_WF_PURPOSE_validate(wf_purpose)
            if (wf_key not in fi_wf_fldr_cfg_coll):
                m = f"Workflow key '{wf_key}' not found in "
                m += f"FI_WF_FOLDER_CONFIG_COLLECTION for fi_key('{fi_key}'): "
                m += f"{list(fi_wf_fldr_cfg_coll.keys())}"
                logger.debug(m)
                return None
            wf_fldr_cfg_list : WF_FOLDER_CONFIG_LIST_TYPE = fi_wf_fldr_cfg_coll[wf_key]
            if (wf_fldr_cfg_list is None or
                len(wf_fldr_cfg_list) == 0):
                m = f"Workflow folder config list for FI_KEY('{fi_key}') "
                m += f"WF_KEY('{wf_key}') is empty or None."
                logger.debug(m)
                return None
            for wf_folder_config in wf_fldr_cfg_list:
                if (wf_folder_config[WF_PURPOSE] == wf_purpose):
                    return wf_folder_config
            return None
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise ValueError(m)
        
    def bdm_FI_WF_FOLDER_CONFIG_ATTRIBUTE(self, 
                                          fi_key:str, 
                                          wf_key:str, 
                                          wf_purpose:str,
                                          attribute:str,
                             raise_errors: bool =  True) -> Optional[str]:
        """Return an attribute value from the workflow folder config."""
        try:
            wf_folder_config: Optional[WF_FOLDER_CONFIG_TYPE] = None
            wf_folder_config = self.bdm_FI_WF_FOLDER_CONFIG(fi_key, wf_key, wf_purpose)
            if wf_folder_config is None:
                m = f"Workflow folder config for FI_KEY('{fi_key}'), "
                m += f"WF_KEY('{wf_key}'), WF_PURPOSE('{wf_purpose}') is None."
                logger.warning(m)
                if raise_errors:
                    raise ValueError(m)
                return None
            if attribute not in VALID_WF_FOLDER_CONFIG_OBJECT_ATTR_KEYS:
                m = f"Attribute '{attribute}' is not a valid workflow folder config attribute."
                logger.warning(m)
                if raise_errors:
                    raise ValueError(m)
                return None
            attribute_value: str = wf_folder_config.get(attribute, None)

            if attribute_value is None:
                m = f"Workflow folder config for FI_KEY('{fi_key}'), "
                m += f"WF_KEY('{wf_key}'), WF_PURPOSE('{wf_purpose}') has no value "
                m += f"for attribute '{attribute}'."
                logger.warning(m)
                if raise_errors:
                    raise ValueError(m)
            return attribute_value
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise ValueError(m)
    #endregion bdm_FI_WF_FOLDER_CONFIG methods
    # ------------------------------------------------------------------------ +
    #region    bdm_WF_OBJECT_TYPE Dict attribute getter methods
    def bdm_WF_OBJECT(self, wf_key:str) -> WF_OBJECT_TYPE:
        """Return the WF_OBJECT specified wf_key."""
        self.bdm_WF_KEY_validate(wf_key)
        return self.bdm_wf_collection[wf_key]
    
    def bdm_WF_OBJECT_count(self) -> int:
        """Return a count of WF_OBJECTS in the BDM_WF_COLLECTION."""
        if self.bdm_wf_collection is None:
            return 0
        if not isinstance(self.bdm_wf_collection, dict):
            m = f"BDM_WF_COLLECTION is not a dict, "
            m+= "but type: {type(self.bdm_wf_collection)}"
            logger.error(m)
            raise ValueError(m)
        return len(self.bdm_wf_collection)
    
    def bdm_WF_KEY(self, wf_key:str) -> str:
        """Return the WF_KEY value the WF_OBJECT for wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_KEY]

    def bdm_WF_NAME(self, wf_key:str) -> str:
        """Return the WF_NAME value for wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_NAME]
    
    def bdm_WF_FOLDER_CONFIG_LIST(self, wf_key:str) -> WF_FOLDER_CONFIG_LIST_TYPE:
        """Return the WF_FOLDER_CONFIG_LIST value for wf_key."""
        if (self.bdm_WF_OBJECT(wf_key)[WF_FOLDER_CONFIG_LIST] is None or
            len(self.bdm_WF_OBJECT(wf_key)[WF_FOLDER_CONFIG_LIST]) == 0):
            return []
        return self.bdm_WF_OBJECT(wf_key)[WF_FOLDER_CONFIG_LIST]
    
    def bdm_WF_PURPOSE_FOLDER_CONFIG(self, wf_key:str, 
                                     wf_purpose:str) -> Optional[WF_FOLDER_CONFIG_TYPE]:
        """Return the WF_FOLDER_CONFIG value for a given wf_key and wf_purpose."""
        self.bdm_WF_PURPOSE_validate(wf_purpose)
        for wf_folder_config_object in self.bdm_WF_FOLDER_CONFIG_LIST(wf_key):
            if wf_folder_config_object[WF_PURPOSE] == wf_purpose:
                return wf_folder_config_object
        return None

    def bdm_WF_KEY_validate(self, wf_key:str) -> bool:
        """Validate the workflow key."""
        workflows = self.bdm_wf_collection
        if ((wf_key is None or len(wf_key) == 0) or
            workflows is None or 
            (wf_key not in workflows)):
            m = f"Invalid wf_key('{wf_key}') is not found in "
            m += f"BDM_WF_COLLECTION: {str(list(self.bdm_wf_collection.keys()))}"
            logger.error(m)
            raise ValueError(m)
        return True

    def bdm_WF_PURPOSE_validate(self, wf_purpose:str) -> bool:
        """Validate the workflow purpose."""
        if wf_purpose not in VALID_WF_PURPOSE_VALUES:
            m = f"Invalid wf_purpose('{wf_purpose}') is not found in "
            m += f"valid purposes: {str(VALID_WF_PURPOSE_VALUES)}"
            logger.error(m)
            raise ValueError(m)
        return True

    #endregion bdm_WF_OBJECT_TYPE pseudo-Object properties
    # ------------------------------------------------------------------------ +
    #endregion BDM - Budget Domain Model methods
    # ======================================================================== +

    # ======================================================================== +
    #region    BSM - Budget Storage Model methods
    # ------------------------------------------------------------------------ +    
    #region    BSM Design Notes
    """ Budget model Storage Model (BSM) Documentation.

    All BDM data is stored in the filesystem by the BSM. BSM works with Path 
    objects, filenames, folders, relative, absolute path names and actual 
    filesystem operations. In the configuration data, the BDM data is stored 
    as strings. Internally, Path objects are used.

    In the BDM, all data is associated with a financial institution (FI). Data 
    is processed by workflows (WF) that are configured for an FI. Hence,
    in all cases the BSM requires an fi_key to identify which FI's data is being
    processed. The wf_key, identifying a particular workflow, is also used
    in the BSM methods, because naming conventions for folder and files are 
    based on the workflow settings.

    Throughout, the following identifiers are used as a convention:
    
    FI_FOLDER - refers to the root folder configured for a specific 
    Financial Institutions (FI). FI_FOLDER is a Path with various str reprs.

    WF_INPUT_FOLDER - refers to the input folder for a specific workflow, again, 
    always in the context of a specific FI through the fi_key. WF_INPUT_FOLDER is
    a Path with various str reprs.

    WB_INPUT - refers to the type of workbooks currently in the 
    WF_INPUT_FOLDER folder.

    WF_OUTPUT_FOLDER - refers to the output folder for a specific workflow. It is
    a Path with various str reprs.

    WB_OUTPUT - refers to the type of workbooks currently in 
    the WF_OUTPUT_FOLDER folder.

    The configuration data and stored budget model data contains str values 
    that are utilized in path names for folders and files. The following 
    naming conventions apply to the purpose of the BDM and BSM methods to 
    separate the concerns for partial path names, full path names and absolute
    pathname values..

    Naming Conventions:
    -------------------
    - bm_ - BudgetDomainModel class properties, e.g., bm_folder, bm_fi, etc.
    - fi_ - Related to financial institution, e.g., fi_key, fi_name, etc.
    - wf_ - Related to workflow, e.g., wf_key, wf_name, etc.
    - bdm_ - BudgetDomainModel methods, concerning the in memory data model.
    - bsm_ - BudgetDomainModel Storage Model methods, folders/files stored the filesystem.
    - _path_str - is the simplest string for a path name or component of. These
    methods return str values and do not manipulate with Path objects. Some
    folder values in the BDM, for example, are simply the name of a folder, 
    not a complete path name.
    - _path - constructs a Path object using _path_str values, 
    invokes .expanduser(). 
    _abs_path - invokes .resolve() on _path results, return Path object.
    - Path objects are never serialized to .jsonc files, only _path_str values.

    Abbrevs used in method names: 
        bdm - Budget Model Domain, the domain model of the budget model.
        bms - Budget Storage Model, the filesystem storage model.
        bm - BudgetDomainModel class instance, parent of the Budget Model data structure.
        bf - budget folder - attribute, root folder, used in path objects.
        fi - financial institution dictionary, attribute of bdm.
             fi_key = int_key, short name of FI, e.g., "boa", "chase", etc.
             fi_value = dict of info about the FI.
        wf - workflow

    A WORKBOOK_DATA_COLLECTION(Dict) is a DATA_OBJECT(Dict) with key/value pairs specific
    to data for a workflow. A wb_data_collection is retrieved with the 
    bdm_FI_WF_DATA_OBJECT() method which could return other types of 
    DATA_OBJECTs in the future. These methods are BSM-related.

    """
    #endregion BSM Design Notes
    # ======================================================================== +
    #region bsm_initialize() method
    def bsm_initialize(self, 
                        create_missing_folders:bool=True,
                        raise_errors:bool=True) -> "BudgetDomainModel":
        """Initialize BSM aspects of the BDM.
        
        Examine elements of self, the BudgetDomainModel class, as initialized
        so far from configuration (BDM_STORE, or Template). Validate the 
        BDM data dependent on folders and files in the filesystem. Flags control 
        whether to create the filesystem structure if it is not present. Also 
        scan the workflow folders for the presence of workbook excel files and 
        resolve their references from the BDM as appropriate.

        During resolution, differences between the workbooks referenced in the
        BDM thus far from the configuration and the actual workbooks in the
        filesystem are detected. The BDM might reference a workbook which is
        no longer present in the BSM. The BSM may discover a workbook in the
        filesystem which is not referenced in the BDM. 

        Args:
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
        """
        st = p3u.start_timer()
        # Plan: validate filesystem folders: bf, fi, wf
        #       update the BDM workbook mappings to BSM 
        try:
            logger.info("BizEVENT: Model setup for BudgetStorageModel (BSM)")
            self.bsm_BDM_FOLDER_resolve(create_missing_folders, raise_errors)
            # Initialize the bdm_fi_collection of financial institutions.
            _ = [self.bsm_FI_initialize(fi_key, create_missing_folders, raise_errors) 
                   for fi_key, fi_object in self.bdm_fi_collection.items()]                
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return self
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_initialize() method
    # ------------------------------------------------------------------------ +    
    #region bsm_BDM_FOLDER Path methods
    def bsm_BDM_FOLDER_validate(self) -> bool:
        """Validate the bm_folder property setting.
        
        Raise a ValueError if the bm_folder property is not set or is not
        usable as part of a valid path string.
        """
        if self.bdm_folder is None or len(self.bdm_folder) == 0:
            m = f"Budget folder path is not set. "
            m += f"Set the BDM_FOLDER('{BDM_FOLDER}') property to valid path value."
            logger.error(m)
            raise ValueError(m)
        return True
    def bsm_BDM_FOLDER_path_str(self) -> str:
        """str version of the BDM_FOLDER value as a Path."""
        # In the BSM, the bm_folder property must be a valid setting that will
        # result in a valid Path. Raise a ValueError if not.
        self.bsm_BDM_FOLDER_validate() # Raises ValueError if not valid
        return str(Path(self.bdm_folder))
    def bsm_BDM_FOLDER_path(self) -> Path:
        """Path of self.bsm_BDM_FOLDER_path_str().expanduser()."""
        return Path(self.bsm_BDM_FOLDER_path_str()).expanduser()
    def bsm_BDM_FOLDER_abs_path(self) -> Path:
        """Path of self.bsm_BDM_FOLDER_path().resolve()."""
        return self.bsm_BDM_FOLDER_path().resolve()
    def bsm_BDM_FOLDER_abs_path_str(self) -> str:
        """str of self.bsm_BDM_FOLDER_abs_path()."""
        return str(self.bsm_BDM_FOLDER_abs_path())
    def bsm_BDM_FOLDER_url(self) -> str:
        """URL str of self.bsm_BDM_FOLDER_abs_path()."""
        return self.bsm_BDM_FOLDER_abs_path().as_uri()
    def bsm_BDM_FOLDER_resolve(self, 
                              create_missing_folders : bool=True,
                              raise_errors : bool=True) -> None:
        """Resolve the BDM_FOLDER path and create it if it does not exist."""
        try:
            logger.debug(f"Checking BDM_FOLDER path: '{self.bdm_folder}'")
            if self.bdm_folder is None:
                m = f"Budget folder path is not set. "
                m += f"Set the '{BDM_FOLDER}' property to a valid path."
                logger.error(m)
                raise ValueError(m)
            # Resolve the BDM_FOLDER path.
            bf_ap = self.bsm_BDM_FOLDER_abs_path()
            if bsm_verify_folder(bf_ap, create_missing_folders, raise_errors):
                wb_paths = bsm_get_workbook_names(bf_ap)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_BDM_FOLDER Path methods
    # ------------------------------------------------------------------------ +
    #region bsm_BDM_URL Path methods
    def bsm_BDM_URL_validate(self) -> bool:
        """Validate the BDM_URL property setting.
        
        Raise a ValueError if the BDM_URL property is not set or is not
        usable as part of a valid path string.
        """
        if self.bdm_url is None or not isinstance(self.bdm_url,str) or len(self.bdm_url) == 0:
            m = f"BDM_URL value is not set to a non-zero length str. "
            m += f"Set the BDM_URL('{BDM_URL}') property to valid path str."
            logger.error(m)
            raise ValueError(m)
        return True
    def bsm_BDM_URL_path_str(self) -> str:
        """str version of the BDM_URL validated as a Path."""
        # In the BSM, the BDM_URL property must be a valid setting that will
        # result in a valid Path. Raise a ValueError if not.
        self.bsm_BDM_URL_validate() # Raises ValueError if not valid
        return str(Path(self.bdm_url))
    def bsm_BDM_URL_path(self) -> Path:
        """Path of self.bsm_BDM_URL_path_str().expanduser()."""
        return Path(self.bsm_BDM_URL_path_str()).expanduser()
    def bsm_BDM_URL_abs_path(self) -> Path:
        """Path of self.bsm_BDM_URL_path().resolve()."""
        return self.bsm_BDM_URL_path().resolve()
    def bsm_BDM_URL_abs_path_str(self) -> str:
        """str of self.bsm_BDM_URL_abs_path()."""
        return str(self.bsm_BDM_URL_abs_path())
    def bsm_BDM_URL_resolve(self, 
                              create_missing_folders : bool=True,
                              raise_errors : bool=True) -> None:
        """Resolve the BDM_URL path and create it if it does not exist."""
        try:
            logger.debug(f"Checking BDM_URL path: '{self.bdm_url}'")
            if self.bdm_url is None:
                m = f"Budget folder path is not set. "
                m += f"Set the '{BDM_URL}' property to a valid path."
                logger.error(m)
                raise ValueError(m)
            # Resolve the BDM_URL path.
            bf_ap = self.bsm_BDM_URL_abs_path()
            bsm_verify_folder(bf_ap, create_missing_folders, raise_errors)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BDM_URL Path methods
    # ------------------------------------------------------------------------ +
    #region bsm_FI_initialize() method
    def bsm_FI_initialize(self, fi_key : str, 
                        create_missing_folders:bool=True,
                        raise_errors:bool=True) -> bool:
        """Initialize BSM aspects for a financial institution.
        
        Examine elements of self for the financial institution indicated by
        fi_key. Validate the mapping of the BDM data dependent on mapping to 
        folders and files in the storage system. Flags control whether to 
        create the storage system structure if it is not present. 

        Args:
            fi_key (str): The financial institution key.
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
            
        Returns:
            bool: True if successful, False otherwise.

        Raises:
            TypeError: If the financial institution key is not a string.
            ValueError: If the financial institution key is not valid.
            Exception: If there is an error during initialization.
        """
        st = p3u.start_timer()
        # Plan: validate filesystem folders: bf, fi, wf
        #       update and refresh any storage system URL references. 
        try:
            logger.debug("Start: ...")
            p3u.str_empty(fi_key, raise_error=True) # Raises TypeError, ValueError
            # Resolve FI_FOLDER path.
            self.bsm_FI_FOLDER_resolve(fi_key, 
                                        create_missing_folders, raise_errors)
            # Resolve/initialize the FI_WF_FOLDER_CONFIG_COLLECTION.
            self.bsm_FI_WF_FOLDER_CONFIG_resolve(fi_key, 
                                        create_missing_folders, raise_errors)
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return True
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_FI_initialize() method
    # ------------------------------------------------------------------------ +    
    #region bsm_FI_WORKBOOK_DATA_COLLECTION_resolve() method
    def bsm_FI_WORKBOOK_DATA_COLLECTION_resolve(self, fi_key:str,
                                                reconcile:bool = False) -> Tuple[WORKBOOK_DATA_COLLECTION_TYPE, str]:
        """Discover all actual WORKBOOKS in storage for the FI. Return a 
        WORKBOOK_DATA_COLLECTION of BDMWorkbook objects, an inventory of
        what is actually in storage.

        Use the configured Workflows for the storage structure in terms of
        folders associated with the fi_key. All configured workflows apply to 
        each financial institution and have WORKFLOW_DATA_FOLDERS in storage
        that might contain workbooks. The goal is to find all workbooks and
        based on their location, will in the metadata about WF_KEY, WF_FOLDER,
        WF_PURPOSE, and WF_FOLDER_ID.

        A WORKFLOW_DATA_FOLDER pathname is constructed from the path elements
        in the workflow configuration.
        """
        discovered_wdc: WORKBOOK_DATA_COLLECTION_TYPE = {}
        wf_object: WF_OBJECT_TYPE = None
        r_msg: str = ""
        fi_folder: Path = None
        wf_name: str = ""
        fi_name: str = ""
        wfdf_name: str = ""
        wfdf_abs_path: Path = None
        try:
            _ = self.bdm_FI_KEY_validate(fi_key)
            fi_folder = self.bsm_FI_FOLDER_abs_path(fi_key)
            fi_name = self.bdm_FI_NAME(fi_key)
            m = (f"Start: FI_KEY('{fi_key}') "
                            f"WORKBOOK_DATA_COLLECTION storage discovery "
                            f"FI_FOLDER('{str(fi_folder)}')")
            r_msg += f"{P2}{m}\n"
            logger.debug(m)
            # Traverse the FI_OBJECT's FI_WF_FOLDER_CONFIG_COLLECTION structure, 
            # to compile a list of WORKBOOKS as BDMWorkbook objects with 
            # populated metadata. 
            # Then reconcile list with the FI_WORKBOOK_DATA_COLLECTION.
            for wf_key, wf_object in self.bdm_wf_collection.items():
                wf_name = wf_object[WF_NAME]

                for wf_purpose in VALID_WF_PURPOSE_VALUES:
                    # For each possible wf_purpose, see if a folder is configured.
                    folder_id = self.bdm_WF_PURPOSE_FOLDER_MAP(wf_key, wf_purpose)
                    if folder_id is None or len(folder_id) == 0:
                        m = (f"Workflow: '{wf_name}') has no configured folder "
                             f" for wf_purpose: '{wf_purpose}', skipping.")
                        logger.debug(m)
                        r_msg += f"{P4}{m}\n"
                        continue
                    # wfdf - WORKFLOW_DATA_FOLDER
                    wfdf_name = self.bdm_WF_FOLDER(wf_key, folder_id)
                    if wfdf_name is None or len(wfdf_name) == 0:
                        m = (f"Workflow: '{wf_name}') has no configured folder "
                             f" for wf_purpose: '{wf_purpose}', skipping.")
                        logger.debug(m)
                        r_msg += f"{P4}{m}\n"
                        continue
                    bdm_wb_paths = []
                    id = f"<{fi_key}:{wf_key}:{wf_purpose}:{folder_id}>"
                    wfdf_abs_path = self.bsm_FI_WORKFLOW_DATA_FOLDER_abs_path(fi_key, wf_key, folder_id)
                    if wfdf_abs_path is None: 
                        m = f"{id} wfdf_abs_path path is None.",
                        logger.debug(m)
                        r_msg += f"{P4}{m}\n"
                        continue
                    if not wfdf_abs_path.exists():
                        m = f"{id} wfdf_abs_path does not exist: {wfdf_abs_path}"
                        logger.debug(m)
                        r_msg += f"{P4}{m}\n"
                        continue
                    # This is where the bsm scans a folder for actual workbook files.
                    bdm_wb_paths = bsm_get_workbook_names(wfdf_abs_path)
                    if len(bdm_wb_paths) == 0:
                        m = f"{id} wfdf_abs_path has no workbooks: {wfdf_abs_path}"
                        logger.debug(m)
                        r_msg += f"{P4}{m}\n"
                        continue
                    m = f"{id} WORKFLOW_DATA_FOLDER('{wfdf_abs_path}') "
                    m += f"found {len(bdm_wb_paths)} workbooks."
                    logger.debug(m)
                    r_msg += f"{P4}{m}\n"
                    for wb_path in bdm_wb_paths:
                        # Create a BDMWorkbook object for each workbook.
                        wb_filename = wb_path.stem
                        wb_filetype = wb_path.suffix.lower()
                        wb_name = wb_path.name
                        wb_url = wb_path.as_uri()
                        bdm_wb = BDMWorkbook(
                            wb_name = wb_name, 
                            wb_filename =  wb_filename,
                            wb_filetype = wb_filetype,
                            wb_url = wb_url,
                            fi_key= fi_key,
                            wf_key = wf_key,
                            wf_purpose = wf_purpose,
                            wf_folder_id = folder_id,
                            wf_folder = wfdf_name,
                            wb_loaded = False,
                            wb_content = None
                            )
                        wb_id = bdm_wb.wb_id
                        bdm_wb.determine_wb_type() 
                        discovered_wdc[wb_id] = bdm_wb
                        m = bdm_wb.wb_info_display_str()
                        logger.debug(f"Collected workbook: {m}")
                        r_msg += f"{P6}workbook: {m}\n"
            # Now have scanned all of the BSM storage to capture all BDMWorkbooks
            # from the FI_WORKFLOW structure mapped to storage.
            # Now reconcile the list of BDMWorkbook objects with the 
            # FI_WORKBOOK_DATA_COLLECTION.
            m = f"FI_KEY('{fi_key}') discovered {len(discovered_wdc)} workbooks."
            logger.debug(m)
            r_msg += f"{P4}{m}\n"
            if reconcile:
                self.bsm_FI_WORKBOOK_DATA_COLLECTION_reconcile(fi_key, discovered_wdc)
            m = f"Complete: FI_KEY('{fi_key}') WORKBOOK storage discovery"
            logger.debug(m)
            r_msg += f"{P2}{m}\n"
            return discovered_wdc, r_msg
        except Exception as e:
                m = p3u.exc_err_msg(e)
                logger.error(m)
                r_msg += f"{P2}{m}\n"
                return discovered_wdc, r_msg
    #endregion bsm_FI_WORKBOOK_DATA_COLLECTION_resolve() method
    # ------------------------------------------------------------------------ +   
    #region bsm_FI_WORKBOOK_DATA_COLLECTION_reconcile() method
    def bsm_FI_WORKBOOK_DATA_COLLECTION_reconcile(self, fi_key:str,
                                                  disc_wdc:WORKBOOK_DATA_COLLECTION_TYPE) -> None:
        """Reconcile the FI_WORKBOOK_DATA_COLLECTION with the discovered workbooks.
        
        This method reconciles the discovered workbooks with the existing 
        FI_WORKBOOK_DATA_COLLECTION. It updates the BDM to reflect the current
        state of the workbooks in the storage system.
        
        Args:
            fi_key (str): The financial institution key.
            disc_wdc (Dict[str, BDMWorkbook]): The discovered workbooks.
        """
        try:
            logger.debug(f"Start: FI_KEY('{fi_key}') reconcile WORKBOOK_DATA_COLLECTION")
            wdc = self.bdm_FI_WORKBOOK_DATA_COLLECTION(fi_key)
            if len(wdc) == 0:
                m = f"FI_KEY('{fi_key}') has no workflow data."
                logger.debug(m)
                return
            # Enumerate the discoverd WORKBOOK_DATA_COLLECTION (disc_wdc),
            # and test the FI WORKBOOK_DATA_COLLECTION for the presence of each
            # discovered BDMWorkbook. If the BDMWorkbook is not present, add it.
            added : List[str] = []
            for wb_id, wb in disc_wdc.items():
                if wb_id not in wdc:
                    # Add a new discovered workbook.
                    logger.debug(f"FI_KEY('{fi_key}') adding discovered "
                                 f"BDMWorkbook '{wb_id}' to FI_WORKBOOK_DATA_COLLECTION.")
                    wdc[wb_id] = wb
                    added.append(wb_id)
            logger.debug(f"Complete: FI_KEY('{fi_key}') reconciled, added "
                         f"{len(added)} new BDMWorkbook(s) {str(added)}")
        except Exception as e:
                m = p3u.exc_err_msg(e)
                logger.error(m)
                raise
    #endregion bsm_FI_WORKBOOK_DATA_COLLECTION_reconcile() method
    # ------------------------------------------------------------------------ + 
    #region bsm_FI_FOLDER Path methods
    def bsm_FI_FOLDER_path_str(self, fi_key: str) -> str:
        """str version of the FI_FOLDER value as a Path.
        
        Raises ValueError for invalid settings for any of: bm_folder property,
        fi_key, of FI_FOLDER value of FI dictionary."""
        bf_p_s = self.bsm_BDM_FOLDER_path_str() # ValueError on BDM_FOLDER property
        self.bdm_FI_KEY_validate(fi_key) # ValueError fi_key
        fi_p_s = str(Path(self.bdm_FI_FOLDER(fi_key))) 
        if fi_p_s is None or len(fi_p_s) == 0:
            m = f"FI_FOLDER value is not set for FI_KEY('{fi_key}'). "
            m += f"In the BudgetDomainModel configuration, correct FI_FOLDER setting."
            logger.error(m)
            raise ValueError(m)   
        return str(Path(bf_p_s) / fi_p_s)
    def bsm_FI_FOLDER_path(self, fi_key: str) -> Path:
        """Path of self.bsm_FI_FOLDER_path_str().expanduser()."""
        return Path(self.bsm_FI_FOLDER_path_str(fi_key)).expanduser()
    def bsm_FI_FOLDER_abs_path(self, fi_key: str) -> Path:
        """Path of self.bsm_FI_FOLDER_path().resolve()."""
        return Path(self.bsm_FI_FOLDER_path(fi_key)).resolve()
    def bsm_FI_FOLDER_abs_path_str(self, fi_key: str) -> str:
        """str of self.bsm_FI_FOLDER_abs_path()."""
        return str(self.bsm_FI_FOLDER_abs_path(fi_key))
    def bsm_FI_FOLDER_resolve(self, fi_key : str,
                              create_missing_folders : bool=True,
                                raise_errors : bool=True) -> None: 
        """Resolve the FI_FOLDER path and create it if it does not exist."""
        fi_ap = self.bsm_FI_FOLDER_abs_path(fi_key)
        logger.debug(f"FI_KEY('{fi_key}') Checking FI_FOLDER('{fi_ap}')")
        _ = bsm_verify_folder(fi_ap, create_missing_folders, raise_errors)
    #endregion bsm_FI_FOLDER Path methods
    # ------------------------------------------------------------------------ + 
    #region bsm_FI_WF_FOLDER_CONFIG_resolve() method
    def bsm_FI_WF_FOLDER_CONFIG_resolve(self, 
                fi_key:str, create_missing_folders:bool=True, 
                raise_errors:bool=True) -> None:
        """Resolve all WF_FOLDER_CONFIG settings an FI_KEY.
        
        A BDM_STORE provides configurations for workflow processes in the
        application in the BDM_WF_COLLECTION attribute. Each defined workflow 
        provides a WF_FOLDER_CONFIG_LIST with 0 or more WF_FOLDER_CONFIG 
        dictionary objects. A WF_FOLDER_CONFIG dict has four attributes: 
        WF_FOLDER, WF_PURPOSE, WF_PREFIX, and WF_FOLDER_URL. At initialization
        time, when the BDM_STORE is loaded to initialize the BDM, these settings
        are resolved. 

        The first part of this resolve step is to validate that each FI_KEY 
        in the BDM_FI_COLLECTION is configured with the same 
        WF_FOLDER_CONFIG_LISTs, using the BDM_WF_COLLECTION definitions.
        
        The second part is to resolve each FI_KEY FI_WF_FOLDER_CONFIG_COLLECTION
        to create missing folders and update the WF_FOLDER_URL attributes.
        """
        try:
            logger.debug(f"FI_KEY('{fi_key}') FI_WF_FOLDER_CONFIG_COLLECTION.")
            fi_folder_abs_path: Path = self.bsm_FI_FOLDER_abs_path(fi_key)

            # Step 1. Validate the FI_WF_FOLDER_CONFIG_COLLECTION has all the
            # WF_FOLDER_CONFIG_LISTs prescribed in the BDM_WF_COLLECTION.
            if (self.bdm_wf_collection is None or len(self.bdm_wf_collection) == 0):
                m = f"BDM_WF_COLLECTION is empty for FI_KEY('{fi_key}'). "
                m += "Cannot resolve workflow folder configurations."
                logger.error(m)
                raise ValueError(m)
            # Iteration each workflow's wf_folder_config_list.
            for wf_key, wf_object in self.bdm_wf_collection.items():
                wf_folder_config_list: WF_FOLDER_CONFIG_LIST_TYPE = None
                wf_folder_config_list = wf_object[WF_FOLDER_CONFIG_LIST]
                for wf_folder_config in wf_folder_config_list:
                    # Lookup wf_folder_config in the fi_wf_folder_config_list.
                    fi_wf_folder_config: WF_FOLDER_CONFIG_TYPE = None
                    fi_wf_folder_config = self.bdm_FI_WF_FOLDER_CONFIG(
                        fi_key, wf_key, wf_folder_config[WF_PURPOSE])
                    if fi_wf_folder_config is None:
                        # Not found, so add it to the fi_wf_folder_config_list.
                        self.bdm_FI_WF_FOLDER_CONFIG_COLLECTION_append(
                            fi_key, wf_key, wf_folder_config)
                        logger.debug(f"FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
                                      f"added wf_folder_config: ({wf_folder_config})")
            # Step 2. Resolve the FI_WF_FOLDER_CONFIG_COLLECTION.
            fi_wf_fldr_cfg: FI_WF_FOLDER_CONFIG_COLLECTION_TYPE = None
            fi_wf_fldr_cfg = self.bdm_FI_WF_FOLDER_CONFIG_COLLECTION(fi_key)
            for wf_key, wf_config_list in fi_wf_fldr_cfg.items():
                for wf_folder in wf_config_list:
                    wf_folder_abs_path: Path = fi_folder_abs_path / wf_folder[WF_FOLDER]
                    if wf_folder_abs_path is not None:
                        m = f"FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
                        m += f"WF_PURPOSE('{wf_folder[WF_PURPOSE]}') "
                        m += f"Checking WF_FOLDER('{str(wf_folder_abs_path)}')"
                        logger.debug(m)
                        bsm_verify_folder(wf_folder_abs_path, 
                                          create_missing_folders, 
                                          raise_errors)
                        wf_folder[WF_FOLDER_URL] = wf_folder_abs_path.as_uri()
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_FI_WF_FOLDER_CONFIG_resolve() method
    # ------------------------------------------------------------------------ + 
    #region bsm_FI_WF_FOLDER Path methods TODO: Are these still needed?
    def bsm_FI_WORKFLOW_DATA_FOLDER_path_str(self, fi_key : str, wf_key : str,
                               folder_id : str) -> str:
        """str version of the WF_FOLDER value for fi_key/wf_key/folder_id."""
        if folder_id not in VALID_WF_PURPOSE_VALUES:
            m = f"Invalid folder_id '{folder_id}' for FI_KEY('{fi_key}') "
            m += f"and WF_KEY('{wf_key}')"
            logger.error(m)
            raise ValueError(m)
        fi_p_s = self.bsm_FI_FOLDER_path_str(fi_key) # FI_FOLDER Path component
        wf_p_s = self.bdm_WF_FOLDER(wf_key, folder_id)
        return str(Path(fi_p_s) / wf_p_s) if wf_p_s is not None else None
    def bsm_FI_WORKFLOW_DATA_FOLDER_path(self, fi_key : str, wf_key : str, 
                           folder_id:str) -> Path:
        """Path of self.bsm_wf_folder_path_str().expanduser()."""
        p_s = self.bsm_FI_WORKFLOW_DATA_FOLDER_path_str(fi_key, wf_key, folder_id)
        return Path(p_s).expanduser() if p_s is not None else None
    def bsm_FI_WORKFLOW_DATA_FOLDER_abs_path(self, fi_key : str, wf_key : str,
                               folder_id : str) -> Path:
        """Path of self.bsm_wf_folder_path().resolve()."""
        p = self.bsm_FI_WORKFLOW_DATA_FOLDER_path(fi_key, wf_key, folder_id)
        return p.resolve() if p is not None else None
    def bsm_FI_WORKFLOW_DATA_FOLDER_abs_path_str(self, fi_key : str, wf_key : str,
                                   folder_id : str) -> str:
        """str of self.bsm_wf_folder_abs_path()."""
        p = self.bsm_FI_WORKFLOW_DATA_FOLDER_abs_path(fi_key, wf_key, folder_id)
        return str(p) if p is not None else None
    #endregion bsm_FI_WF_FOLDER_CONFIG_resolve Path methods
    # ------------------------------------------------------------------------ +
    #endregion BSM - Budget Storage Model methods
    # ======================================================================== +
