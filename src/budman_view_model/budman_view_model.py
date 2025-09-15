# ---------------------------------------------------------------------------- +
#region budman_view_model.py module
""" budman_view_model.py implements the BudManViewModel class.

    In MVVM, the ViewModel is the interface between the View and the Model. As
    such, it provides the inherent View Model behavior, a CommandProcessor
    Pattern, and a Data Context Pattern for the View. It also consumes the 
    services from one or more Model objects, in this case, the Budget Domain Model.

    Typically, an mvvm application implements the user interface with the View 
    and data models and storage with the Model. A ViewModel is a middle ground
    where requests from a user or client module are submitted by the View for
    processing. A ViewModel directs the actions to fulfill the request and
    accesses the Model for retrieval and storage of data.

    Our mvvm provides a CommandProcessor as an abstract to represent the requests
    submitted by the View for processing. A command a format for such a request
    and will contain a command action or verb, perhaps a subcommand noun for 
    an object or other information used in the the request along with various
    parameters and option flags to more clearly specify the intent and scope of
    the request.

    Lastly this mvvm design supports a Data Context as a data model shareable 
    between the View and ViewModel, typically application-specific in nature, but
    designed as a simple abstraction of the application data model and state.

    This ViewModel inherits classes implementing a CommandProcessor, DataContext, 
    and Model. The classes used are subclasses of the basic View, ViewModel and 
    Model abstractions from the p3_mvvm package.

    Inherited Class Interfaces:
    ---------------------------
       1. DataContext - BudManAAppDataContext_Binding.
       2. CommandProcessor - p3m.CommandProcessor.
       3. Model - p3m.Model_Binding.

    p3m is an abbreviation for the p3_mvvm package, indicating that classes
    from that base package are subclassed directly. BudManAppDataContext_Binding
    is a class within the BudMan application which specializes other super 
    classes.

    Each of the classes mentioned is a concrete implementation of one or more
    abstract base class (ABC) interface. Some objects implement the ABC interface
    directly as a concrete object inherited by the BudManViewModel class. Others,
    indicated by the _Binding suffix, provide a concrete implementation of an ABC
    interface which is bound at runtime to an instance of a class that implements
    the necessary ABC interface. This is known as Dependency Injection enabling 
    a late binding of a configured concrete class to provide a dependent 
    interface, but constructed at runtime during setup and initialization.

    Basic Application Startup
    -------------------------

    When the BudMan Application is started, the BudManViewModel is instantiated
    first based on a provided path to a configuration file and an application
    settings file. Both files are loaded into the ViewModel. Since the 
    BudManViewModel class inherits the p3m.CommandProcessor class, it will serve
    as the CommandProcessor for the mvvm pattern. During the setup process, 
    Dependency Injection is implemented by steps to instantiate the
    Model and "bind" it to the ViewModel to initialize the p3m.Model_Binding. 
    Likewise, the BDMDataContext class is instantiated and bound to the ViewModel
    to initialize the p3m.DataContext_Binding. Lastly, the View is instantiated
    as the BudManCLIView class with the BudManViewModel instance bound as the 
    CommandProcessor for the View. The same DataContext intance is also bound to
    the View, completing the Dependency Injection setup of the application.

    Basic Command Flow
    ------------------

    As a CLI user interface, the View class, BudManCLIView, takes input from the
    user and parses it into p3m.CMD_OBJECT_TYPE objects. Each CMD_OBJECT is
    configured in the View CLI parsing structure to ensure only allowed commands
    and parameters are entered. When a user submits a command line, the View
    passes the command object to the CommandProcessor for validation and
    execution. Command validation covers all parameter values and options to be
    validated and looks up the configured execution function to process the
    command using a command map configured in the CommandProcessor. Finally,
    the CommandProcessor invokes the command execution function using the ViewModel
    to route the command to various modules configured to process commands.

    The general pattern for command execution is to have a module associated with
    a command that decomposes the command into one or more execution Tasks and
    dispatches the sequence of Tasks for execution. As output is provided, it is
    captured and returned in a p3m.CMD_RESULT_TYPE object to the CommandProcessor
    and subsequently returned back to the View for handling output to the user.

    DataContext Usage
    -----------------

    When a command is dispatched for execution, the CMD_OBJECT and THe 
    DataContext object are provided to the top-level command execution function.
    That way, a common data model state is maintained with application-specific
    support as needed. A DataContext may be Model-Aware or not, based on the
    designers approach. A Model-Aware DC can access the Model for data
    retrieval and storage while hiding that complexity from the ViewModel and
    View. In other cases, command execution functions are free to directly 
    access the Model.

    Throughout the BudgetModel (budman) application, a design language is
    used as a convention for naming within the code-base. 

    Convention: <Domain_Model_Object_Name>_<action_verb>()
                <Domain_Model_Object_Name>_<action_verb>_<dc_item>()
                dc_item: a reference to data in the Data Context

    Budget Domain Model Design Language (namespace) 
    -----------------------------------------------

    Budget Domain Concepts
    ---------------------
        
    - Budget - a means of tracking financial transactions over time.

    - Budget Model (BM) - a functional model of budget processes. It is composed
    of worklows that process transactions from financial institutions, income 
    and expenses. Transactions are categorized by scanning input data and
    producing output data.

    - Budget Domain (BD) and Budget Domain Model (BDM) - the conceptual model 
    of the budget top-level concept.

    Budget Domain Objects
    ---------------------

    Objects are considered things, pieces of data that commands do something to.
    An app will execute commands to take action on an object. Object names are 
    nouns of the design language.

    - Workbook - typically an excel workbook, stored in a file on the user's
    file system. It contains data in worksheets, and is often received from a 
    bank or brokerage firm

    - Financial Institution (FI) - a bank or brokerage firm that provides
    financial services to the user. The FI is the source of transactions for the
    budget model.

    - Loaded Workbook - a workbook that has been loaded into the budget model
    for processing. It is a workbook that has been opened by the app and is in 
    the current session context, available to have commands executed on it.

    - Workflow (WF) - a sequence of steps to process data. A Budget Model workflow 
    has a purpose, and applies to workbooks. As workflows are executed,
    the workbooks are processed and the data is transformed. Workflows have 
    folders in the filesystem where workbooks are accessed. In general, a 
    workflow has both an input and output folder. The input folder is where the
    workbooks are loaded from, and the output folder is where the processed, or 
    changed workbooks are saved.

    - Commands - a command is an action to be taken on an object. Commands are
    typically implemented as methods in the code. A command is named, has a
    source and a target, and in MVVM packages, a binding to configure how
    Views bind to the commands of a ViewModel.

    For the sake of the BudgetModel design language, the command source is 
    some class serving as a View. The command target is a workbook (WB), 
    a loaded workbook, a financial institution (FI), or a workflow (WF). The
    command method name will include a noun (probably abbreviated) from the
    design language for objects and one of the a verb for the action: init, load, 
    save, show, delete, add, etc., suffixed by '_cmd'. Additional method 
    naming conventions are described below.

    MVVM Design Pattern
    -------------------

    MVVM - Model View ViewModel - a design pattern applied throughout for 
    separation of concerns in the Budget Domain.

    View (v) - code implementing access to a View Model, either by a 
    user-facing experience (UX) or an API interface.

    View Model (vm) - code implementing the domain interface to the Model. 
    View Models separate the concerns of the Model from those anticipated
    by the View.

    Model (m) - code implementing the domain model data. The Model is caboose
    in terms of the application data models, persistence, etc.
    
    These View Model commands have a verb, an noun (object reference), 
    and supporting parameter arguments. The verb is the action to be taken,
    the noun is the object to be acted upon, and the parameters are the
    supporting data for the action. The command methods are typically
    implemented as a single method. A naming convention reveals the verb and
    noun. The method name is a concatenation of the verb and noun, with an  
    underscore between them.

    Character case convention:
    - Class names are in CamelCase.
    - Method names are in prefix_snake_Camel_UPPER_lower_case.
    - Constants are in UPPER_SNAKE_CASE.
    - Seeing UPPER case in a mixed name implies the UPPER case part is also
    a constant defined in the module or class and used to represent a single
    value or a list of values both in the design language and the code.

    Method Naming convention: <scope>_<noun>_<verb>()
    where scope is the domain or sub-domain of the method (function), verb is the action to be taken, 
    and noun is the object to be acted upon. The noun is broken into two parts:

    <noun>: <object>[_<property>]
    <object>: <class_name> | <class_abbrev> | <pseudo_class_abbrev>
    <class_abbrev>: "bdm" | "bmc" |"bdmwd" | "bdm_vm"
    "bdm" - Budget Domain Model (BudgetDomainModel class, budget_domain_model pkg)
    "bmc" - Budget Domain Model Config (BDMConfig class, budget_domain_model pkg)
    "bsm" - Budget Storage Model (budget_storage_model package)

    bdm_vm - scope is the concern of the Budget Domain Model View Model
    bdmwd - scope is the concern of the Budget Domain Model Working Data

    The term "Design Language" is abbreviated as "DL" or "dl" in the code. All
    constants are defined in a design_language.py module for the intended scope. 
    A package will have a design_language.py module for the package. If deemed
    helpful, a specific module will have a module_name_dl.py module for the
    design language for that module. Names mean something, and acronyms are used
    prolifically in code, as are fully descriptive names. 

    Maintaining consistency with the design language is critical for good
    design and keeping cohesion continuous. Here, python constants are used
    heavily to represent the design language. Think of constants as "tags" 
    from the design language word cloud. Searching the code-base for a constant
    will reveal the design language and the code that implements it. The best
    API design with high cohesion and low coupling reflect a consistent design
    language, and ruthless refactoring to keep it consistent and documented.

    Also, adopting the practice of documenting the design language in the code
    is proving to be beneficial for GPT coding assistance. In this case, 
    GitHub Copilot is mind-blowingly adept at applying the design language in
    all parts of a large application, not just the neighboring code. Thank
    God for code-folding though.

    Command Methods: To be a clean ViewModel, command methods are provided to
    enable the View to call the command methods in a consistent and 
    loosely-coupled manner. The command methods are typically implemented as a
    single method. A naming convention reveals the verb and noun. In the MVVM
    pattern, a DataContext is used to bind the ViewModel to the View. For 
    the view, Commands are methods on the DataContext, without actually knowing
    the type of the DataContext. Since python is dynamically typed, we just
    use it to dispatch the command to the appropriate method.
"""
#endregion budman_view_model.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import logging, os, sys, getpass, time, copy, importlib
from pathlib import Path
from typing import List, Type, Optional, Dict, Tuple, Any, Callable
import importlib

# third-party modules and packages
import p3_utils as p3u, p3logging as p3l, p3_mvvm as p3m
from openpyxl import Workbook, load_workbook

# local modules and packages
# from budman_command_processor.cp_utils import CMD_RESULT_OBJECT
import budman_workflows
import budman_command_processor as cp
import budman_settings as bdms
from budman_namespace.design_language_namespace import *
from budman_namespace.bdm_workbook_class import BDMWorkbook
from budman_workflows import (
    BDMTXNCategoryManager, TXNCategoryCatalog,
    check_sheet_schema, check_sheet_columns, 
    validate_budget_categories, process_budget_category,
    output_category_tree,
    WORKFLOW_CMD_process,
    budget_category_mapping
    )
from budget_domain_model import (
    BudgetDomainModel, 
    BDMConfig
    )
from budget_storage_model import (
    bsm_BDM_STORE_url_put,
    bsm_BDM_STORE_url_get
    )
from budman_data_context.budman_app_data_context_binding_class import BudManAppDataContext_Binding
from budman_data_context.budget_domain_model_data_context import BDMDataContext
from budman_cli_view import budman_cli_parser, budman_cli_view
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +
class BudManViewModel(BudManAppDataContext_Binding, p3m.CommandProcessor, 
                      p3m.Model_Binding): 
    # ======================================================================== +
    #region BudManViewModel_Base class intrinsics                              +
    # ======================================================================== +
    #                                                                          +
    # ------------------------------------------------------------------------ +
    #region BudManViewModel class doc string                                   +
    """BudManViewModel - A Budget Manager View Model providing 
    CommandProcessor & Data Context.
    
    This ViewModel, BudManViewModel, depends on some primary MVVM design 
    pattern abstract behaviors:
        1. ViewModel_Base - the behavior of a View Model in MVVM.
        2. ViewModelCommandProcessor_Base (CP), Command Processor behavior,
        3. ViewModelDataContext_Base (DC), Data Context behavior, and
        4. Model_Base - the Model behavior.
         
    BudManViewModel is a concrete implementation of the ViewModel_Base and
    ViewModelCommandProcessor_Base. It applies the ViewModelDataContext_Binding
    and Model_Binding to access concrete Model and Data Context implementations.

    Each pattern is backed by defined interfaces for properties and methods. 
    CP is the means for a View to take actions performing commands against data 
    in the DC. From an MVVM pattern perspective, DC is the only medium of data 
    exchange between the View and the ViewModel using properties, methods and
    eventing. Commands are sent from the View, to be processed by the ViewModel 
    in the "context" of the data state in the DC. Command implementation methods 
    access the DC properties and invoke methods on the View Model DC interface 
    to perform the requested actions. A View Model may publish events as well, 
    which a View may subscribe to.
    """
    #endregion BudManViewModel class doc string                                +
    # ------------------------------------------------------------------------ +
    #region    BudManViewModel Class __init__() constructor method             +
    def __init__(self, bdms_url : str = None, settings : bdms.BudManSettings = None) -> None:
        super().__init__()
        self._initialized : bool = False
        self._bdm_store_url : str = bdms_url
        self._settings = settings
        self._initialized : bool = False
        self._BDM_STORE_loaded : bool = False
        # Dependency Injection: p3m.Model_Binding model() method backing value
        self._budget_domain_model : BudgetDomainModel = None
    #endregion __init__() constructor method
    # ------------------------------------------------------------------------ +
    #region    BudManViewModel Class initialize() method                       +
    def initialize(self) -> "BudManViewModel":
        """Initialize the command view_model."""
        try:
            st = p3u.start_timer()
            logger.info(f"BizEVENT: View Model setup for '{self.app_name}'")
            self._initialized = False
            logger.debug(f"Complete: {p3u.stop_timer(st)}")
            return self
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion BudManViewModel Class initialize() method                       +
    # ------------------------------------------------------------------------ +
    #region    BudManViewModel Class initialize_model() method                 +
    def initialize_model(self, bdms_url : str) -> BudgetDomainModel:
        """Create a model using the bdms_url location for a valid BDM_STORE.
            The BDM_STORE object provides the model configuration and state.
        """
        try:
            st = p3u.start_timer()
            logger.debug(f"Start: ...")
            # if a bdms_url is provided, load the BDM_STORE file.
            if p3u.str_notempty(self.bdms_url):
                # Load the BDM_STORE file from the URL, initializing 
                # a BDMConfig object, to use for initialization tasks.
                bdmc : BDMConfig = BDMConfig.BDM_STORE_url_get(self.bdms_url)
                if bdmc is None:
                    m = f"Failed to load BDM_STORE from URL: {self.bdms_url}"
                    logger.error(m)
                    raise ValueError(m)
                # Use the loaded BDM_STORE file as a config_object 
                self._BDM_STORE_loaded = True
            else:
                # Use the builtin default template as a config_object.
                bdm_config = BDMConfig.BDM_CONFIG_default()
                # Use the default BDM_CONFIG object as a config_object 
                self._BDM_STORE_loaded = False
            # Now to create the model and initialize it.
            model : BudgetDomainModel = BudgetDomainModel(bdmc).bdm_initialize()
            if not model.bdm_initialized: 
                raise ValueError("BudgetModel is not initialized.")
            logger.debug(f"Complete: {p3u.stop_timer(st)}")
            return model
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion BudManViewModel Class initialize_model() method                   +
    # ------------------------------------------------------------------------ +
    #region    BudManViewModel Class save_model() method                       +
    def save_model(self) -> None:
        """Save the model for this view_model."""
        try:
            st = p3u.start_timer()
            logger.info(f"Start: ...")
            self.model.bdm_save_model()
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return None
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion BudManViewModel Class save_model() method                       +
    # ------------------------------------------------------------------------ +
    #region    BudManViewModel Class Properties
    @property
    def app_name(self) -> str:
        """Return the application name."""
        if self._settings is None:
            raise ValueError("Settings not configured.")
        return self._settings.get(bdms.APP_NAME, "BudManViewModel")

    @property
    def bdms_url(self) -> str:
        """Return the BDM_STORE URL."""
        return self._bdm_store_url
    @bdms_url.setter
    def bdms_url(self, url: str) -> None:
        """Set the BDM_STORE URL."""
        if not isinstance(url, str):
            raise TypeError("bdms_url must be a string")
        if not url.startswith("file://") and not url.startswith("http://"):
            raise ValueError("bdms_url must be a valid file or http URL")
        self._bdm_store_url = url

    @property
    def settings(self) -> bdms.BudManSettings:
        """Return the application settings."""
        return self._settings
    @settings.setter
    def settings(self, settings: bdms.BudManSettings) -> None:
        """Set the application settings."""
        if not isinstance(settings, bdms.BudManSettings):
            raise TypeError("settings must be a BudManSettings instance")
        self._settings = settings

    @property
    def initialized(self) -> bool:
        """Return True if the ViewModel is initialized."""
        return self._initialized
    @initialized.setter
    def initialized(self, value: bool) -> None:
        """Set the initialized property."""
        if not isinstance(value, bool):
            raise ValueError("initialized must be a boolean value.")
        self._initialized = value

    @property
    def budget_domain_model(self) -> BudgetDomainModel:
        """Dependency Injection: Return the BudgetDomainModel instance."""
        return self._budget_domain_model
    @budget_domain_model.setter
    def budget_domain_model(self, value: BudgetDomainModel) -> None:
        """Dependency Injection: Set the BudgetModel binding."""
        if not isinstance(value, BudgetDomainModel):
            raise ValueError("budget_model must be a BudgetModel instance.")
        self._budget_domain_model = value

    #endregion BudManViewModel Class Properties                                +
    # ------------------------------------------------------------------------ +
    #endregion BudManViewModel class intrinsics
    # ======================================================================== +

    # ======================================================================== +
    #region    BudManViewModel p3m.CommandProcessor super class Override methods          +
    # ======================================================================== +
    #region    Design Notes                                                    +
    """ ViewModelCommandProcessor Design Notes (future ABC)

    A Command Pattern supports a means to represent a command as an object,
    in our case, a dictionary. The command is a request to perform an action
    and includes arguments. Both the command and the arguments are validated
    before executing the command. The command is a dictionary with a
    command key and a an optional sub-command value. The command key includes
    the name of the requested action, often, but not limited to a verb, 
    e.g., init, show, save, etc. That part is the cmd_name. So, here are some
    rules:
    - A cmd_name is a string of characters (which cannot include '_cmd').
    - A cmd_key is also a string composed of the cmd_name and the
      '_cmd' suffix.
    - A cmd_key becomes an actual 'key' in the cmd dict object with an 
      Optional[str] value. It is one identifier to use for binding to an
      execution function in the cmd_map.
    - A sub-command is given as a value to the cmd_key, but may be None.
    - A second key in the cmd dict is 'command_name' with the value being the 
      command name string, e.g., 'init', 'load', 'save', etc.
    - A third key used in a valid cmd dict object is 'subcmd_key' with a 
      string value constructed as" <cmd_key>_<subcmd_name>".

      Examples:
      - cli:  init fin_inst -> cmd_key = init_cmd, subcmd_key = init_cmd_fin_inst  
      - cli:  workflow categorization 2 -> cmd_key = workflow_cmd, subcmd_key = workflow_cmd_categorization   
      - cli:  app log rollover -> cmd_key = app_cmd, subcmd_key = app_cmd_log_subcmd

    When binding to an execution function, the subcmd_key is searched first, and
    used when present. If no subcmd_key binding exists in the cmd_map, then
    the cmd_key is sought and used. If no cmd_key binding found, then the cmd 
    is auto-bound to the UNKNOWN_cmd() exec function.

    The command key is a string, which is the name of the command.  
    The characters '_cmd' are appended to the command key. The sub-command 
    value is optional, but most common. It is a noun indicating the type
    of object to be acted upon. The command key is combined with the subcommand
    value to form a full command key. The full command key is used to look
    up the command method in the command map. The command map is a dictionary
    with the full command key as the key and the command method as the value.

    An example command object content is:

    {'init_cmd': 'fin_inst', 'fi_key': 'boa', 'wf_key': None, 'wb_name': 'all'}

    The command key is 'init_cmd' and the sub-command value is 'fin_inst'. The 
    resulting full command key is 'init_cmd_fin_inst'. 

    The other key/values are arguments: fi_key = 'boa', wf_key = None, and
    wb_name = 'all'. Some arguments are included with default values and will
    raise an error if their value is None.

    - fi_key: The key for the financial institution, cannot be None. The 
      reserved fi_key = 'all' indicates to apply the action to all FIs.
    - wf_key: The key for the workflow, cannot be None. The reserved 
      wf_key = 'all' indicates to apply the action to all workflows.
    - wb_name: The name of a workbook, None or the reserved name of 'all'
      which indicates to apply the action to all workbooks in the scope.
    
    """
    #endregion design notes                                                    +
    # ------------------------------------------------------------------------ +
    #region    cp_initialize_cmd_map() method+
    def cp_initialize_cmd_map(self) -> None:
        """Override: BudManApp-Specific Initialize the cmd_map dictionary."""
        try:
            # Use the following cmd_map to dispatch the command for execution.
            self.cp_cmd_map = {
                cp.CV_LIST_CMD_KEY: self.LIST_cmd,
                cp.CV_LOAD_BDM_STORE_SUBCMD_KEY: self.BDM_STORE_load_cmd,
                cp.CV_SAVE_BDM_STORE_SUBCMD_KEY: self.BDM_STORE_save_cmd,
                cp.CV_LOAD_WORKBOOKS_SUBCMD_KEY: self.WORKBOOKS_load_cmd,
                cp.CV_SAVE_WORKBOOKS_SUBCMD_KEY: self.WORKBOOKS_save_cmd,
                cp.CV_CHANGE_WORKBOOKS_SUBCMD_KEY: self.CHANGE_cmd,
                cp.CV_CATEGORIZATION_SUBCMD_KEY: self.WORKFLOW_categorization_cmd,
                cp.CV_WORKFLOW_CMD_KEY: self.WORKFLOW_cmd,
                cp.CV_APPLY_SUBCMD_KEY: self.WORKFLOW_apply_cmd,
                cp.CV_CHECK_SUBCMD_KEY: self.WORKFLOW_check_cmd,
                # cp.CV_WORKFLOW_INTAKE_SUBCMD_KEY: self.WORKFLOW_intake_cmd,
                cp.CV_SHOW_CMD_KEY: self.SHOW_cmd,
                cp.CV_CHANGE_CMD_KEY: self.CHANGE_cmd,
                cp.CV_APP_CMD_KEY: self.APP_cmd,
            }
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion cp_initialize_cmd_map() method
    #region    cp_validate_cmd() Command Processor method
    def cp_validate_cmd(self, cmd : p3m.CMD_OBJECT_TYPE = None,
                        validate_all : bool = False) -> p3m.CMD_RESULT_TYPE:
        """BudMan-App-specific: Validate the cmd object for cmd_key and parameters.

        Extract a valid, known full_cmd_key/cmd_key to succeed.
        Consider values to common arguments which can be validated with
        the data context.

        Arguments:
            cmd (Dict): A candidate Command object to validate.
            validate_all (bool): If True, validate all parts of the cmd, 
                before returning a result. If False, return on first discovery
                of an error.

        returns:
            Tuple[bool, str]: A tuple containing a boolean indicating if the
                cmd and arguments are valid.
                True returns the full cmd_key value as result.
                False, the message will contain an error message.
        """
        try:
            self.cp_validate_cmd_object(cmd, raise_error = True)
            # If cp_validate_cmd_object() returns, CMD_OBJECT has 
            # content to examine and validate.
            # Setup a CMD_RESULT object to return.
            cmd_result : p3m.CMD_RESULT_TYPE = p3m.create_CMD_RESULT_OBJECT(
                cmd_result_status=False, 
                result_content_type=p3m.CMD_STRING_OUTPUT,
                result_content="Command validation failed.",
                cmd_object=cmd
            )
            # For a few args, apply the DC values if no value given in the cmd.
            if self.cp_cmd_attr_get(cmd, cp.CK_FI_KEY) is None:
                self.cp_cmd_attr_set(cmd, cp.CK_FI_KEY,
                                    self.dc_FI_KEY)
            if self.cp_cmd_attr_get(cmd, cp.CK_WF_KEY) is None:
                self.cp_cmd_attr_set(cmd, cp.CK_WF_KEY,
                                    self.dc_WF_KEY)
            # If validate_all, check all cmd args before return. 
            validate_all: bool = self.cp_cmd_attr_get(cmd, cp.CK_VALIDATE_ONLY)
            all_results : str = "All Results:\n" if validate_all else ""
            result = "It's all good." 
            if validate_all:
                all_results = f"Command validation info: \n{P2}cmd: {str(cmd)}\n"
            success:bool = True
            # Validate the cmd arguments.
            for key, value in cmd.items():
                if key == cmd[cp.p3m.CK_CMD_KEY]: 
                    continue
                elif key == cp.CK_WB_LIST:
                    for wb_index in cmd[cp.CK_WB_LIST]:
                        if not self.dc_WB_INDEX_validate(wb_index):
                            result = f"Invalid wb_list value: '{wb_index}'."
                            success = False 
                            logger.error(result)
                        continue
                    continue
                elif key == cp.CK_FI_KEY:
                    if not self.dc_FI_KEY_validate(value):
                        result = f"Invalid fi_key value: '{value}'."
                        success = False 
                        logger.error(result)
                    continue
                elif key == cp.CK_WB_NAME:
                    continue
                elif (key == cp.CK_WF_KEY or
                      key == cp.CK_SRC_WF_KEY or
                      key == cp.CK_DST_WF_KEY):
                    if value is None:
                        # No further validation if value is None
                        continue
                    if not self.dc_WF_KEY_validate(value):
                        result = f"Invalid wf_key value: '{value}'."
                        success = False 
                        logger.error(result)
                    continue
                elif (key == cp.CK_WF_PURPOSE or
                      key == cp.CK_SRC_WF_PURPOSE or
                      key == cp.CK_DST_WF_PURPOSE):
                    if value is None:
                        # No further validation if value is None
                        continue
                    if self.cp_cmd_attr_get(cmd, key) in VALID_WF_PURPOSE_CHOICES:
                        # Map the choices value to actual value
                        value = VALID_WF_PURPOSE_MAP[value]
                    if not self.dc_WF_PURPOSE_validate(value):
                        result = f"Invalid wf_purpose value: '{value}'."
                        success = False 
                        logger.error(result)
                    self.cp_cmd_attr_set(cmd,key,value)
                    continue
                elif key == cp.CK_WB_ID:
                    if not self.dc_WB_ID_validate(value):
                        result = f"Invalid wb_id level: '{value}'."
                        success = False 
                        logger.error(result)
                elif key == cp.CK_PARSE_ONLY: 
                    po = cmd.get(cp.CK_PARSE_ONLY, False)
                    continue
                elif key == cp.CK_VALIDATE_ONLY: 
                    vo = cmd.get(cp.CK_VALIDATE_ONLY, False)
                    continue
                elif key == cp.CK_WHAT_IF: 
                    what_if = cmd.get(cp.CK_WHAT_IF, False)
                    continue
                elif key == cp.CK_CHECK_REGISTER:
                    continue
                else:
                    if key not in cp.BUDMAN_VALID_CK_ATTRS:
                        result = f"Unchecked argument key: '{key}': '{value}'."
                        logger.debug(result)
                # If not validate_all and success is False, return. Else, 
                # continue validating all cmd args.
                if not validate_all and not success:
                    # Without validate_all, if error, return, else continue
                    cmd_result[p3m.CMD_RESULT_STATUS] = success
                    cmd_result[p3m.CMD_RESULT_CONTENT] = result
                    return cmd_result
                else:
                    # If validate_all, accumulate results.
                    all_results += f"{P2}{result}\n"
            # Argument check is complete
            if success:
                m = (f"Command validated - cmd_key: '{cmd[cp.p3m.CK_CMD_KEY]}' "
                            f"subcmd_key: {str(cmd[cp.p3m.CK_SUBCMD_KEY])}")
                logger.info(m)
                cmd_result[p3m.CMD_RESULT_STATUS] = success
                cmd_result[p3m.CMD_RESULT_CONTENT] = m
                return cmd_result # The happy path return 
            if validate_all:
                cmd_result[p3m.CMD_RESULT_STATUS] = success
                cmd_result[p3m.CMD_RESULT_CONTENT] = all_results
                return cmd_result # The happy path return 
            cmd_result[p3m.CMD_RESULT_STATUS] = success
            cmd_result[p3m.CMD_RESULT_CONTENT] = result
            return cmd_result
        except Exception as e:
            m = f"Error validating command: {str(cmd)}: {p3u.exc_err_msg(e)}"
            logger.error(m)
            return False, m
    #endregion cp_validate_cmd() Command Processor method
    # ------------------------------------------------------------------------ +
    #                                                                          +
    #endregion BudManViewModel p3mCommandProcessor super class Override methods          +
    # ======================================================================== +
 
    # ======================================================================== +
    #region    Model_Binding super class overrides for BudgetDomainModel                              + 
    # ======================================================================== +
    #                                                                          +
    # ------------------------------------------------------------------------ +
    #region    Model_Binding Properties                                        +
    @property
    def model(self) -> BudgetDomainModel:
        """Override: Return the model object reference from Dependency Injection."""
        return self._budget_domain_model
    @model.setter
    def model(self, bdm: BudgetDomainModel) -> None:
        """Override: Dependency Injection: Set the model object binding reference."""
        if not isinstance(bdm, BudgetDomainModel):
            raise TypeError("model must be a BudgetDomainModel instance")
        self._budget_domain_model = bdm
    #endregion Model_Binding Properties                                        +
    #endregion Model_Binding super class overrides for BudgetDomainModel               +
    # ======================================================================== +

    # ======================================================================== +
    #region    Command Execution Methods                                       +
    # ======================================================================== +
    #region LIST_cmd() execution method
    def LIST_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> p3m.CMD_RESULT_TYPE:
        """List requested info from Budget Manager Data Context.

        Arguments:
            cmd (CMD_OBJECT_TYPE): 
                A validated CommandProcessor CMD_OBJECT_TYPE. Contains
                the command attributes and parameters to execute.

        Returns:
            p3m.CMD_RESULT_TYPE:
            The outcome of the command execution. 
        """
        try:
            st = p3u.start_timer()
            logger.debug(f"Start: ...{P2}")
            # Should be called only for list cmd.
            cmd_result : p3m.CMD_RESULT_TYPE = cp.verify_cmd_key(cmd, cp.CV_LIST_CMD_KEY)
            if not cmd_result[p3m.CMD_RESULT_STATUS]: return cmd_result
            # Process BudMan command tasks.
            cmd_result = cp.BUDMAN_CMD_process(cmd, self.DC)
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return cmd_result
        except Exception as e:
            return p3m.create_CMD_RESULT_EXCEPTION(cmd, e)
    #endregion SHOW_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region BDM_STORE_save_cmd() execution method
    def BDM_STORE_save_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> None:
        """Save the Budget Manager store (BDM_STORE) file with the BSM.

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains the 
            BDM_STORE dict after updating DC_BDM_STORE in the DATA_CONTEXT.
            If False, a description of the error.
            
        Raises:
            RuntimeError: If the BDM_STORE file cannot be saved.
        """
        try:
            st = p3u.start_timer()
            logger.info(f"Start: ...")
            # Save the BDM_STORE file with the BSM.
            # Get a Dict of the BudgetModel to store.
            bdm_dict = self.budget_domain_model.bdm_dehydrate()
            # Save the BDM_STORE file.
            bdm_url = self.dc_BDM_STORE[BDM_URL]
            bsm_BDM_STORE_url_put(bdm_dict, bdm_url)
            self.dc_BDM_STORE_changed = False
            logger.info(f"Saved BDM_STORE url: {bdm_url}")
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return True, bdm_dict
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion BDM_STORE_save_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region BDM_STORE_load_cmd() execution method
    def BDM_STORE_load_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> Tuple[bool, str]:
        """Load the Budget Manager store (BDM_STORE) file from the BSM.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. For this
            command, must contain load_cmd = 'BDM_STORE' resulting in
            a full command key of 'load_cmd_BDM_STORE'.

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            st = p3u.start_timer()
            logger.info(f"Start: ...")
            # Load the BDM_STORE file with the BSM.
            # Use the BDM_STORE configured in BUDMAN_SETTINGS.
            bdm_url = self.dc_BDM_STORE[BDM_URL]
            # Load the BDM_STORE file.
            budman_store_dict = bsm_BDM_STORE_url_get(bdm_url)
            self.dc_BDM_STORE = budman_store_dict
            # Now refresh the model state to reflect the BDM_STORE content.
            model_refresh: BudgetDomainModel = self.model.bdm_initialize()
            view_model_refresh: BudManViewModel = self.dc_initialize()
            self._BDM_STORE_loaded = True
            self.dc_BDM_STORE_changed = False
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return True, budman_store_dict
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion BDM_STORE_load_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region SHOW_cmd() execution method
    def SHOW_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE:
        """Show requested info from Budget Manager Data Context.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. 
    
        Required cmd object attributes:
            cmd_key: 'show_cmd' 
        Optional cmd object attributes:
            cmd_name: 'show'
            subcmd_name: CV_BUDGET_CATEGORIES_SUBCMD
            subcmd_key: 'show_cmd_BUDGET_CATEGORIES'
            CK_CAT_LIST: A list of budget categories to include, len()==0 means All. 

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            st = p3u.start_timer()
            logger.debug(f"Start: ...")
            # Should be called only for show cmd.
            cmd_result : p3m.CMD_RESULT_TYPE = cp.verify_cmd_key(cmd, cp.CV_SHOW_CMD_KEY)
            if not cmd_result[p3m.CMD_RESULT_STATUS]: return cmd_result
            # Process BudMan command tasks.
            cmd_result = cp.BUDMAN_CMD_process(cmd, self.DC)
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return cmd_result
        except Exception as e:
            m = (f"Error executing cmd: {cmd[cp.p3m.CK_CMD_NAME]} {cmd[cp.p3m.CK_SUBCMD_NAME]}: "
                 f"{p3u.exc_err_msg(e)}")
            logger.error(m)
            return False, m
    #endregion SHOW_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region WORKBOOKS_load_cmd() execution method
    def WORKBOOKS_load_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE:
        """Model-aware: Load workbook content for one or more WORKBOOKS in the DC.

        A load_cmd_workbooks command will use the wb_ref value in the cmd. 
        Value is a number, a wb_name or ALL.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. For this
            command, must contain load_cmd = 'workbooks' resulting in
            a full command key of 'load_cmd_workbooks'.

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            st = p3u.start_timer()
            logger.debug(f"Start: ...")
            selected_bdm_wb_list : List[BDMWorkbook] = None
            selected_bdm_wb_list = self.process_selected_workbook_input(cmd)
            bdm_wb : Optional[BDMWorkbook] = None
            r = f"\nBudget Manager Workbooks({len(selected_bdm_wb_list)}):"
            for bdm_wb in selected_bdm_wb_list:
                # Select the current workbook in the Data Context.
                self.dc_WORKBOOK = bdm_wb
                success, result = self.dc_WORKBOOK_load(bdm_wb) 
                if not success:
                    m = f"Error loading wb_id: '{bdm_wb.wb_id}': {result}"
                    logger.error(m)
                    r += f"\n{P2}Error wb_index: {self.dc_WB_INDEX:>4} wb_id: '{bdm_wb.wb_id:<40}' Reason:{m}"
                    continue
                # Cmd output string
                r_str = bdm_wb.wb_index_display_str(self.dc_WB_INDEX)
                r += f"\n{P2}Loaded {r_str}"
                continue
            logger.debug(f"Complete Command: 'Load' {p3u.stop_timer(st)}")   
            return True, r
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion WORKBOOKS_load_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region WORKBOOKS_save_cmd() execution method
    def WORKBOOKS_save_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE: 
        """Model-Aware: Execute save command for one WB_INDEX or ALL_WBS.
        In BudMan, loaded workbook content is maintained in the 
        dc_LOADED_WORKBOOKS collection. This command will cause that content
        to be saved to its storage location.        

        Raises:
            RuntimeError: For exceptions.
        """
        try:
            r : str = ""
            m : str = ""
            st = p3u.start_timer()  
            logger.debug(f"Start: {str(cmd)}")
            selected_bdm_wb_list : List[BDMWorkbook] = None
            selected_bdm_wb_list = self.process_selected_workbook_input(cmd)
            bdm_wb : Optional[BDMWorkbook] = None
            r = f"\nBudget Manager Workbooks({len(selected_bdm_wb_list)}):"
            bdm_wb : Optional[BDMWorkbook] = None
            for bdm_wb in selected_bdm_wb_list:
                wb_index = self.dc_WORKBOOK_index(bdm_wb.wb_id)
                if bdm_wb.wb_content is None:
                    m = f"Workbook wb_id: '{bdm_wb.wb_id}' has no loaded content."
                    logger.error(m)
                    r += f"\n{P2}Error wb_index: {wb_index:>4} wb_id: '{bdm_wb.wb_id:<40}' Reason:{m}"
                    continue
                # Save the workbook content.
                success, result = self.dc_WORKBOOK_save(bdm_wb)
                if not success:
                    m = f"Error saving wb_id: '{bdm_wb.wb_id}': {result}"
                    logger.error(m)
                    r += f"\n{P2}Error wb_index: {wb_index:>2} wb_id: '{bdm_wb.wb_id:<40}' Reason:{m}"
                    continue
                r_str = bdm_wb.wb_index_display_str(wb_index)
                r += f"\n{P2}Saved {result!r}"
            logger.info(f"Complete Command: 'Save' {p3u.stop_timer(st)}")   
            return True, r
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion WORKBOOKS_save_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region CHANGE_cmd() execution method
    def CHANGE_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> Tuple[bool, str]:
        """Change aspects of the Data Context.

        A CHANGE_cmd command uses the wb_ref arg parameter.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. 
    
        Required cmd object attributes:
            cmd_key: 'change_cmd' 
        Optional cmd object attributes:
            cmd_name: 'change'
            subcmd_name: CV_BUDGET_CATEGORIES_SUBCMD
            subcmd_key: 'change_cmd_BUDGET_CATEGORIES'
            CK_CAT_LIST: A list of budget categories to include, len()==0 means All. 

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            logger.info(f"Start: ...")
            if cmd[cp.p3m.CK_SUBCMD_NAME] == cp.CV_WORKBOOKS_SUBCMD_NAME:
                selected_bdm_wb_list : List[BDMWorkbook] = None
                selected_bdm_wb_list = self.process_selected_workbook_input(cmd)
                result: str = f"Changing {len(selected_bdm_wb_list)} workbooks:"
                r: str = ""
                for bdm_wb in selected_bdm_wb_list:
                    # Select the current workbook in the Data Context.
                    self.dc_WORKBOOK = bdm_wb
                    result += f"\n{P2}workbook: {str(self.dc_WB_INDEX):>4} '{bdm_wb.wb_id:<40}'"
                    # Apply the include argument switches to the selected workbook.
                    new_wb_type = self.cp_cmd_attr_get(cmd, cp.CK_CMDLINE_WB_TYPE, None)
                    if new_wb_type is not None:
                        bdm_wb.wb_type = new_wb_type
                        result += (f"\n{P4}Changed wb_type: '{new_wb_type}' for "
                                f"wb_index: {str(self.dc_WB_INDEX):>4}' wb_id: '{bdm_wb.wb_id}'")
                        self.dc_BDM_STORE_changed = True
                    new_wf_key = self.cp_cmd_attr_get(cmd, cp.CK_CMDLINE_WF_KEY, None)
                    if new_wf_key is not None:
                        bdm_wb.wf_key = new_wf_key
                        result += (f"\n{P4}Changed wf_key: '{new_wf_key}' for "
                                f"wb_index: '{str(self.dc_WB_INDEX):>4}' wb_id: '{bdm_wb.wb_id}'")
                        self.dc_BDM_STORE_changed = True
                    new_wf_purpose = self.cp_cmd_attr_get(cmd, cp.CK_CMDLINE_WF_PURPOSE, None)
                    if new_wf_purpose is not None:
                        wf_key = new_wf_key or bdm_wb.wf_key
                        # DEPRECATED, refactor to change wf_purpose without folder_id
                        # folder_id = self.dc_WF_PURPOSE__FOLDER_MAP(wf_key, new_wf_purpose)
                        # if folder_id != bdm_wb.wf_folder_id:
                        #     bdm_wb.wf_folder_id = folder_id
                        #     m = f", wf_folder_id: '{folder_id}' "
                        # else:
                        #     m = ""
                        m = "fix me"
                        bdm_wb.wf_purpose = new_wf_purpose
                        result += (f"\n{P4}Changed wf_purpose: '{new_wf_purpose}'"
                                   f" {m} for wb_index: "
                                f"'{str(self.dc_WB_INDEX):>4}' wb_id: '{bdm_wb.wb_id}'")
                        self.dc_BDM_STORE_changed = True
                    self.dc_WORKBOOK = bdm_wb
                return True, result
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion CHANGE_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region APP_cmd() execution method
    def APP_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE:
        """App commands to manipulate app values and settings.

        The APP_cmd command can use a variety of other command line arguments.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. 
    
        Required cmd object attributes:
            cmd_key: 'app_cmd' 
        Optional cmd object attributes:
            cmd_name: 'app'
            subcmd_key: 'app_cmd_delete'
            subcmd_name: CV_DELETE_SUBCMD
            subcmd_key: 'app_cmd_reload'
            subcmd_name: CV_RELOAD_SUBCMD
            subcmd_key: 'app_cmd_log'
            subcmd_name: CV_LOG_SUBCMD

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            r_msg: str = "Start:"
            m:Optional[str] = None
            logger.debug(r_msg)
            success: bool = False
            result: Any = None
            subcmd_name = cmd[cp.p3m.CK_SUBCMD_NAME]
            if subcmd_name == cp.CV_LOG_SUBCMD_NAME:
                # Show the current log level.
                return True, "App Log cmd."
            elif subcmd_name == cp.CV_RELOAD_SUBCMD_NAME:
                try:
                    reload_target = self.cp_cmd_attr_get(cmd, cp.CK_RELOAD_TARGET, None)
                    if reload_target is None:
                        m = f"reload_target is None, no action taken."
                        logger.error(m)
                        return False, m
                    if reload_target == cp.CV_CATEGORY_MAP:
                        catman: BDMTXNCategoryManager = BDMTXNCategoryManager() #self.WF_CATEGORY_MANAGER
                        category_catalog: TXNCategoryCatalog = None
                        if catman :
                            category_catalog = catman.catalogs[self.dc_FI_KEY]
                            category_catalog.CATEGORY_MAP_WORKBOOK_import()
                            mod = category_catalog.category_map_module
                            if mod:
                                cat_count = len(category_catalog.category_collection)
                                rule_count = len(category_catalog.category_map)
                                task = "CATEGORY_MAP_WORKBOOK_import()"
                                m = (f"{P2}Task: {task:30} {rule_count:>3} "
                                     f"rules, {cat_count:>3} categories.")
                                logger.debug(m)
                                return True, m
                            else:
                                return False, "Failed to reload category_map_module"
                    if reload_target == cp.CV_FI_WORKBOOK_DATA_COLLECTION:
                        m = "deprecated"
                        # wdc: WORKBOOK_DATA_COLLECTION_TYPE = None
                        # wdc, m = self.model.bsm_FI_WORKBOOK_DATA_COLLECTION_resolve(self.dc_FI_KEY)
                        return True, m
                    if reload_target == cp.CV_WORKFLOWS_MODULE:
                        importlib.reload(budman_workflows.workflow_utils)
                        return True, "Reloaded workflows module."
                    return True, r_msg
                except Exception as e:
                    m = f"Error reloading target: {reload_target}: {p3u.exc_err_msg(e)}"
                    logger.error(m)
                    return False, m
            elif subcmd_name == cp.CV_DELETE_SUBCMD_NAME:
                try:
                    delete_target = self.cp_cmd_attr_get(cmd, cp.CK_DELETE_TARGET, -1)
                    if self.dc_WB_INDEX_validate(delete_target):
                        bdm_wb: BDMWorkbook = self.dc_WORKBOOK_remove(delete_target)
                        return True, f"Deleted workbook: {bdm_wb.wb_id}"
                    return False, f"Invalid wb_index: '{delete_target}'"
                except Exception as e:
                    m = f"Error deleting workbook: {delete_target}: {p3u.exc_err_msg(e)}"
                    logger.error(m)
                    return False, m
            else:
                return False, f"Unknown app subcmd: {subcmd_name}"
        except Exception as e:
            m = f"Error executing cmd: {cmd[cp.p3m.CK_CMD_NAME]} {cmd[cp.p3m.CK_SUBCMD_NAME]}: "
            m += p3u.exc_err_msg(e)
            logger.error(m)
            return False, m
    #endregion APP_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region WORKFLOW_categorization_cmd() execution method
    def WORKFLOW_categorization_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE:
        """Apply categorization workflow to one or more WORKBOOKS in the DC.

        As a workflow process, the WORKFLOW_categorization_cmd method has the
        job to marshall the necessary input objects, implied by the command
        arguments and perhaps Data Context, and then invoke the appropriate 
        function or method to conduct the requested process, tasks, etc. It also
        catches the return status and result to return.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. 
    
        Required cmd object attributes:
            cmd_key: 'workflow_cmd' 
        Optional cmd object attributes:
            cmd_name: CV_WORKFLOW_CMD
            Valid subcommands:
                subcmd_key: 'workflow_cmd_categorization'
                subcmd_name: CV_DELETE_SUBCMD
            subcmd_key: 'app_cmd_reload'
            subcmd_name: CV_RELOAD_SUBCMD
            subcmd_key: 'app_cmd_log'
            subcmd_name: CV_LOG_SUBCMD

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            logger.info(f"Start: ...")
            selected_bdm_wb_list : List[BDMWorkbook] = None
            selected_bdm_wb_list = self.process_selected_workbook_input(cmd)
            fr: str = f"Categorizing {len(selected_bdm_wb_list)} workbooks:"
            success : bool = False
            r : str = ""
            m : str = ""
            log_all : bool = self.cp_cmd_attr_get(cmd, cp.CK_LOG_ALL, False)
            # Process the intended workbooks.
            for bdm_wb in selected_bdm_wb_list:
                # Select the current workbook in the Data Context.
                self.dc_WORKBOOK = bdm_wb
                bdm_wb_abs_path = bdm_wb.abs_path()
                fr += f"\n{P2}workbook: {str(self.dc_WB_INDEX):>4} '{self.dc_WB_ID:<40}'"
                bdm_wb_abs_path = bdm_wb.abs_path()
                if bdm_wb_abs_path is None:
                    m = f"Workbook path is not valid: {bdm_wb.wb_url}"
                    logger.error(m)
                    fr += f"\n{P4}Error: {m}"
                    continue
                # Check cmd needs loaded workbooks to check
                if not bdm_wb.wb_loaded:
                    m = f"wb_name '{bdm_wb.wb_name}' is not loaded, no action taken."
                    logger.error(m)
                    fr += f"\n{P4}Error: {m}"
                    continue
                # Now we have a valid bdm_wb to process.
                if bdm_wb.wb_type == WB_TYPE_EXCEL_TXNS:
                    task = "process_budget_category()"
                    m = (f"{P2}Task: {task:30} {str(self.dc_WB_INDEX):>4} "
                         f"'{self.dc_WB_ID:<40}'")
                    logger.debug(m)
                    fr += f"\n{P2}{m}"
                    success, r = process_budget_category(bdm_wb, self.DC, log_all)
                    if not success:
                        r = (f"{P4}Task Failed: process_budget_category() Workbook: "
                             f"'{self.dc_WB_ID}'\n{P8}Result: {r}")
                        logger.error(r)
                        fr += r + "\n"
                        continue
                    fr += f"\n{P8}Result: {r}"
                    task = "dc_WORKBOOK_save()"
                    m = (f"{P2}Task: {task:30} {str(self.dc_WB_INDEX):>4} "
                         f"'{self.dc_WB_ID:<40}'")
                    logger.debug(m)
                    fr += f"\n{P2}{m}"
                    success, r = self.dc_WORKBOOK_save(bdm_wb)
                    if not success:
                        m = (f"{P4}Task Failed: dc_WORKBOOK_save() Workbook: "
                             f"'{bdm_wb.wb_id}'\n{P8}Result: {m}")
                        logger.error(m)
                        fr += m + "\n"
                        continue
                    fr += f"\n{P8}Result: {r}"
            logger.info(m)
            return True, fr
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion WORKFLOW_categorization_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region WORKFLOW_cmd() execution method
    def WORKFLOW_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> p3m.CMD_RESULT_TYPE:
        """Execute a workflow task.

        Arguments:
            cmd (CMD_OBJECT_TYPE): 
                A validated CommandProcessor CMD_OBJECT_TYPE. Contains
                the command attributes and parameters to execute.

        Returns:
            p3m.CMD_RESULT_TYPE:
                The outcome of the command execution. 
        """
        try:
            st = p3u.start_timer()
            logger.debug(f"Start: ...")
            # Should be called only for workflow cmd.
            cmd_result : p3m.CMD_RESULT_TYPE = cp.verify_cmd_key(cmd, cp.CV_WORKFLOW_CMD_KEY)
            if not cmd_result[p3m.CMD_RESULT_STATUS]: return cmd_result
            # Process workflow command tasks.
            cmd_result = WORKFLOW_CMD_process(cmd, self.DC)
            logger.info(f"Complete: {p3u.stop_timer(st)}")
            return cmd_result
        except Exception as e:
            return p3m.create_CMD_RESULT_EXCEPTION(cmd, e)
    #endregion WORKFLOW_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region WORKFLOW_apply_cmd() execution method
    def WORKFLOW_apply_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE:
        """Apply workflow tasks to WORKBOOKS.

        A WORKFLOW_apply_cmd command will use the wb_ref value in the cmd. 
        Value is a number or a wb_name.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. For this
            command, must contain workflow_cmd = 'apply' resulting in
            a full command key of 'workflow_cmd_apply'.

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            logger.info(f"Start: ...")
            r_msg: str = "Start:"
            m:Optional[str] = None
            logger.debug(r_msg)
            success: bool = False
            result: Any = None
            subcmd_name = cmd[cp.p3m.CK_SUBCMD_NAME]
            if subcmd_name == cp.CV_APPLY_SUBCMD_NAME:
                # Update the txn_categories by apply the category_map.
                return True, "Applied category_map to txn_categories."
            return True, ""
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion WORKFLOW_apply_cmd() execution method
    # ------------------------------------------------------------------------ +
    #region WORKFLOW_check() execution method 
    def WORKFLOW_check_cmd(self, cmd : p3m.CMD_OBJECT_TYPE) -> BUDMAN_RESULT_TYPE:
        """Apply workflow to one or more WORKBOOKS in the DC.

        A WORKFLOW_categorization_cmd command will use the wb_ref value in the cmd. 
        Value is a number or a wb_name.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. For this
            command, must contain workflow_cmd = 'categorization' resulting in
            a full command key of 'workflow_cmd_categorization'.

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
        """
        try:
            logger.info(f"Start: ...")
            selected_bdm_wb_list : List[BDMWorkbook] = None
            selected_bdm_wb_list = self.process_selected_workbook_input(cmd)
            result: str = f"Checking {len(selected_bdm_wb_list)} workbooks:"
            r: str = ""
            for bdm_wb in selected_bdm_wb_list:
                # Select the current workbook in the Data Context.
                self.dc_WORKBOOK = bdm_wb
                bdm_wb_abs_path = bdm_wb.abs_path()
                result += f"\n{P2}workbook: {str(self.dc_WB_INDEX):>4} '{bdm_wb.wb_id:<40}'"
                # Check cmd needs loaded workbooks to check
                if not bdm_wb.wb_loaded:
                    m = f"wb_name '{bdm_wb.wb_name}' is not loaded, no action taken."
                    logger.error(m)
                    result += f"\n{P2}{m}"
                    continue
                # By default, check the sheet schema. But other cli switches
                # can added to check something else.
                if cmd[cp.CK_VALIDATE_CATEGORIES]:
                    # Validate the categories in the workbook.
                    task = "validate_budget_categories()"
                    m = (f"{P2}Task: {task:30} {str(self.dc_WB_INDEX):>4} "
                        f"'{bdm_wb.wb_id:<40}'")
                    logger.debug(m)
                    success, r = validate_budget_categories(bdm_wb, self.DC, P4)
                    result += f"\n{r}"
                    continue
                success: bool = check_sheet_schema(bdm_wb.wb_content)
                r = f"Task: check_sheet_schema workbook: Workbook: '{bdm_wb.wb_id}' "
                if success:
                    result += f"\n{P2}{r}"
                    continue
                if cmd[cp.CK_FIX_SWITCH]:
                    r = f"Task: check_sheet_columns workbook: Workbook: '{bdm_wb.wb_id}' "
                    ws = bdm_wb.wb_content.active
                    success = check_sheet_columns(ws, add_columns=True)
                    if success: 
                        bdm_wb.wb_content.save(bdm_wb_abs_path)
                result += f"\n{P2}{r}"
                continue
            return success, result
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion WORKFLOW_check_cmd() execution method
    # ------------------------------------------------------------------------ +
    #                                                                          +
    #endregion Command Execution Methods                                       +
    # ======================================================================== +

    # ======================================================================== +
    #region    helper methods                                                  +
    # ======================================================================== +
    #region process_workbook_input()
    def process_selected_workbook_input(self, cmd: p3m.CMD_OBJECT_TYPE) -> List[BDMWorkbook]:
        """Process the workbook input from the command, return a list of 
        BDMWorkbooks, which may be empty.

        Arguments:
            cmd (Dict): A p3m.CMD_OBJECT.

        Returns:
            List[BDMWorkbook]: A list of BDMWorkbook objects.
        """
        try:
            # Extract common command attributes to select workbooks for task action.
            wb_list : List[int] = cmd.get(cp.CK_WB_LIST, [])
            all_wbs : bool = self.cp_cmd_attr_get(cmd, cp.CK_ALL_WBS, self.dc_ALL_WBS)
            selected_bdm_wb_list : List[BDMWorkbook] = []
            load_workbook:bool = self.cp_cmd_attr_get(cmd, cp.CK_LOAD_WORKBOOK, False)
            if all_wbs:
                # If all_wbs is True, process all workbooks in the data context.
                selected_bdm_wb_list = list(self.dc_WORKBOOK_DATA_COLLECTION.values())
            elif len(wb_list) > 0:
                for wb_index in wb_list:
                    bdm_wb = self.dc_WORKBOOK_by_index(wb_index)
                    selected_bdm_wb_list.append(bdm_wb)
            else:
                # No workbooks selected by the command parameters.
                return selected_bdm_wb_list
            for bdm_wb in selected_bdm_wb_list:
                bdm_wb_abs_path = bdm_wb.abs_path()
                if bdm_wb_abs_path is None:
                    selected_bdm_wb_list.remove(bdm_wb)
                    m = f"Excluded workbook: '{bdm_wb.wb_id}', "
                    m += f"workbook url is not valid: {bdm_wb.wb_url}"   
                    logger.error(m)
                    continue
                if load_workbook and not bdm_wb.wb_loaded:
                    # Load the workbook content if it is not loaded.
                    success, result = self.dc_WORKBOOK_content_get(bdm_wb)
                    if not success:
                        selected_bdm_wb_list.remove(bdm_wb)
                        m = f"Excluded workbook: '{bdm_wb.wb_id}', "
                        m += f"failed to load: {result}"
                        logger.error(m)
                        continue
                    if not bdm_wb.wb_loaded:
                        logger.warning(f"Workbook '{bdm_wb.wb_id}' wb_loaded was False!")
                        bdm_wb.wb_loaded = True
                    # self.dc_LOADED_WORKBOOKS[bdm_wb.wb_id] = wb_content
            return selected_bdm_wb_list
        except Exception as e:
            m = f"Error processing workbook input: {p3u.exc_err_msg(e)}"
            logger.error(m)
            raise RuntimeError(m)
    #endregion process_workbook_input()
    # ------------------------------------------------------------------------ +
    #region    wb_ref_not_valid() method
    def recon_cmd_args_to_DC(self, cmd: dict) -> bool:
        """Reconcile.

        Arguments:
            cmd: reconcile the args in cmd with the DC state values

        Returns:
            bool: 
        """
        # Get the command arguments.
        wb_index : int = self.cp_cmd_attr_get(cmd, cp.CK_WB_INDEX, self.dc_WB_INDEX)
        all_wbs : bool = self.cp_cmd_attr_get(cmd, cp.CK_ALL_WBS, self.dc_ALL_WBS)
        wb_ref : str = self.cp_cmd_attr_get(cmd, cp.CK_WB_REF, self.dc_WB_ID)
        # Get the command arguments.
        fi_key = cmd.get(cp.CK_FI_KEY, None)
        wf_key = cmd.get(cp.CK_WF_KEY, EXAMPLE_BDM_WF_CATEGORIZATION)
        wf_purpose = cmd.get(cp.CK_WF_PURPOSE, WF_INPUT)
        wb_type = cmd.get(cp.CK_WB_TYPE, WB_TYPE_EXCEL_TXNS)
        wb_name = cmd.get("wb_name", None)
        if fi_key != self.dc_FI_KEY:
            logger.warning(f"fi_key: arg '{fi_key}' differs from "
                            f" current value '{self.dc_FI_KEY}'")
            fi_key = self.dc_FI_KEY
            logger.warning(f"fi_key: using current value '{self.dc_FI_KEY}'")
        if wf_key != self.dc_WF_KEY:
            logger.warning(f"wf_key: arg '{wf_key}' differs from "
                            f" current value '{self.dc_WF_KEY}'")
            wf_key = self.dc_WF_KEY
            logger.warning(f"wf_key: using current value '{self.dc_WF_KEY}'")
        if wb_name != self.dc_WB_NAME:
            logger.warning(f"wb_name: arg '{wb_name}' differs from "
                            f" current value '{self.dc_WB_NAME}'")
        if wb_type != self.dc_WB_TYPE:
            logger.warning(f"wb_type: arg '{wb_type}' differs from "
                            f" current value '{self.dc_WB_TYPE}'")
            wb_type = self.dc_WB_TYPE
            logger.warning(f"wb_type: using current value '{self.dc_WB_TYPE}'")
        return True
    #endregion wb_ref_not_valid() method
    # ------------------------------------------------------------------------ +
    #region    wb_ref_not_valid() method
    @classmethod
    def wb_ref_not_valid(cls, all_wbs: bool, wb_index: int, wb_name: str) -> bool:
        """Check if the workbook reference is not valid.

        Arguments:
            all_wbs (bool): True if all workbooks are to be processed.
            wb_index (int): The index of the workbook.
            wb_name (str): The name of the workbook.

        Returns:
            bool: True if the workbook reference is not valid, False otherwise.
        """
        if not all_wbs and wb_index == -1 and wb_name is None:
            m = f"wb_ref '{wb_ref}' is not valid."
            logger.error(m)
            return True
        return False
    #endregion wb_ref_not_valid() method
    # ------------------------------------------------------------------------ +
    #                                                                          +
    #endregion helper methods for command execution                            +
    # ======================================================================== +
