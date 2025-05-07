# ---------------------------------------------------------------------------- +
#region budget_model.py module
""" budget_model.py implements the class BudgetModel.

    Following a rough MVVM pattern, the BudgetModel class is acting as 
    both the cohesive Model, representing a budget model domain in
    memory as object API. 

    At present, the BudgetModel class is a singleton class that manages the 
    mapping of excel "workbooks" stored in the filesystem to the budget model.
    The BudgetModel class presents properties and methods to the outside world.
    Methods are separated into ViewModel-ish methods for the Budget Domain 
    and Model-ish methods for the Storage Domain, which is the filesystem.

    In the Budget Domain Model, a data pipeline pattern is used, anticipating 
    "raw data" will be introduced from finanancial institutions (FI) and and 
    proceed through a series of transformations to a "finalized", although
    updatable budget model. Raw data is a "workbook", often an excel file, 
    or a .cvs file.
    
    A "folder" concept is aligned with the stages of transformation as a 
    series of "workflows" applied to the data. Roughly, workflow stage works 
    on data in its associated folder and then may process data inplace locally 
    or as modifications or output to workbooks to another stage folder.

    [raw_data] -> [incoming] -> [categorized] -> [finalized]

    Workflows are functional units of work with clearly defined concerns applied
    to input data and producing outout data. 

    Key Concepts:
    -------------
    - Folders: containers of workbooks associated with a workflow stage.
    - Workflows: a defined set of process functions applied to data.
    - Workbooks: a container of financial transaction data, e.g., excel file.
    - Raw Data: original data from a financial institution (FI), read-only.
    - Financial Budget: the finalized output, composed of workbooks,
        representing time-series transactions categorized by payments and
        deposits (debits and credits).
    - Financial Institution (FI): a bank or brokerage financial institution.
        
    Perhaps the primary domain should be "Financial Budget" instead of
    "Budget Model". The FB Domain and the "Storage Domain" or Sub-Domain. 

    The Budget Domain Model has the concern of presenting an API that is 
    independent of where the workbook data is sourced and stored. The 
    Budget Storage Model has the concern of managing sourcing and storing 
    workbooks in the filesystem. It maps the Budget Domain structure to 
    filesystem folders and files.

    Assumptions:
    - The FI transactions are in a folder specified in the budget_config.
    - Banking transaction files are typical excel spreadsheets. 
    - Data content starts in cell A1.
    - Row 1 contains column headers. All subsequent rows are data.
"""
#endregion budget_model.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import logging, os, getpass, time, copy
from pathlib import Path
from typing import List, Type, Generator, Dict, Tuple, Any

# third-party modules and packages
import p3_utils as p3u, pyjson5, p3logging as p3l
from openpyxl import Workbook, load_workbook

# local modules and packages
from .budget_model_constants import *
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(THIS_APP_NAME)
# ---------------------------------------------------------------------------- +
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +
class SingletonMeta(type):
    """Metaclass for implementing the Singleton pattern for subclasses."""
    _instances = {}

    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            # Invokes cls.__init__()
            cls._instances[cls] = super().__call__(*args, **kwargs)
        return cls._instances[cls]
# ---------------------------------------------------------------------------- +
class BudgetModel(metaclass=SingletonMeta):
    # ======================================================================== +
    #region BudgetModel class intrinsics
    # ======================================================================== +
    """BudgetModel class implementing Singleton pattern.
    
        A singleton class to manage the budget model for the application.
        This class is used to store and manage the budget data, including
        the budget folder, institutions, and options. Uses only dict-friendly
        items.

        Properties:
        -----------
        - budget_folder_path_str: A pathname to a parent budget folder, 
            e.g., ~/OneDrive/budget.
        - budget_folder_abs_path: A Path object representing the absolute path
            to the budget folder.
        - institutions: A dictionary to store the financial institutions.
        - options: A dictionary to store the options for the budget model.
        - budget_data: A dictionary to store non-persistent the budget data.
    """
    config_template : object = None #
    # ------------------------------------------------------------------------ +
    def __init_subclass__(cls, **kwargs):
        # called at import time, not at runtime.
        print(f"BudgetModel.__init_subclass__({cls}): called")
    # ------------------------------------------------------------------------ +
    #region    BudgetModel class constructor __init__()
    def __init__(self,classname : str = "BudgetModel") -> None:
        """Constructor for the BudgetModel class."""
        self._classname = classname
        BudgetModel.config_template = self if classname == "BudgetModelTemplate" else None

        # Private attributes initialization, basic stuff only.
        # for serialization ease, always persist dates as str type.
        logger.debug("Start: BudgetModel().__init__() ...")
        setattr(self, BM_INITIALIZED, False)
        setattr(self, BM_FOLDER, None)  # budget folder path
        setattr(self, BM_STORE_URI, None)  # uri for budget model store
        setattr(self, BM_FI_COLLECTION, {})  # financial institutions
        setattr(self, BM_WF_COLLECTION, {}) 
        setattr(self, BM_OPTIONS, {})  # budget model options
        setattr(self, BM_CREATED_DATE, p3u.now_iso_date_string()) 
        setattr(self, BM_LAST_MODIFIED_DATE, self._created_date)
        setattr(self, BM_LAST_MODIFIED_BY, getpass.getuser())
        setattr(self, BM_WORKING_DATA, {})  # wd - budget model working data
        logger.debug("Complete: BudgetModel().__init__() ...")
    #endregion BudgetModel class constructor __init__()
    # ------------------------------------------------------------------------ +
    #region    BudgetModel internal class methods
    def to_dict(self):
        '''Return BudgetModelTemplate dictionary object. Used for serialization.'''
        ret = {
            BM_INITIALIZED: self.bm_initialized,
            BM_FOLDER: self.bm_folder,
            BM_FI_COLLECTION: self.bm_fi_collection,
            BM_STORE_URI: self.bm_store_uri,
            BM_WF_COLLECTION: self.bm_wf_collection,
            BM_OPTIONS: self.bm_options,
            BM_CREATED_DATE: self.bm_created_date,
            BM_LAST_MODIFIED_DATE: self.bm_last_modified_date,
            BM_LAST_MODIFIED_BY: self.bm_last_modified_by,
            BM_WORKING_DATA: self.bm_working_data
        }
        return ret
    def __repr__(self) -> str:
        ''' Return a str representation of the BudgetModel object '''
        ret = f"{{ "
        ret += f"'{BM_INITIALIZED}': {self.bm_initialized}, "
        ret += f"'{BM_FOLDER}': '{self.bm_folder}', "
        ret += f"'{BM_FI_COLLECTION}': '{self.bm_fi_collection}', "
        ret += f"'{BM_STORE_URI}': '{self.bm_store_uri}', "
        ret += f"'{BM_WF_COLLECTION}': '{self.bm_wf_collection}', "
        ret += f"'{BM_OPTIONS}': '{self.bm_options}', "
        ret += f"'{BM_CREATED_DATE}': '{self.bm_created_date}', "
        ret += f"'{BM_LAST_MODIFIED_DATE}': '{self.bm_last_modified_date}', "
        ret += f"'{BM_LAST_MODIFIED_BY}': '{self.bm_last_modified_by}', "
        ret += f"'{BM_WORKING_DATA}': '{self.bm_working_data}' }} "
        return ret
    def __str__(self) -> str:
        ''' Return a str representation of the BudgetModel object '''
        ret = f"{self.__class__.__name__}({str(self.bm_folder)}): "
        ret += f"{BM_INITIALIZED} = {str(self.bm_initialized)}, "
        ret += f"{BM_FOLDER} = '{str(self.bm_folder)}', "
        ret += f"{BM_FI_COLLECTION} = [{', '.join([repr(fi_key) for fi_key in self.bm_fi_collection.keys()])}], "
        ret += f"{BM_STORE_URI} = '{self.bm_store_uri}' "
        ret += f"{BM_WF_COLLECTION} = '{self.bm_wf_collection}' "
        ret += f"{BM_OPTIONS} = '{self.bm_options}' "
        ret += f"{BM_CREATED_DATE} = '{self.bm_created_date}', "
        ret += f"{BM_LAST_MODIFIED_DATE} = '{self.bm_last_modified_date}', "
        ret += f"{BM_LAST_MODIFIED_BY} = '{self.bm_last_modified_by}', "
        ret += f"{BM_WORKING_DATA} = {self.bm_working_data}"
        return ret
    #endregion BudgetModel internal class methods
    # ------------------------------------------------------------------------ +
    #region    BudgetModel public class properties
    @property
    def bm_initialized(self) -> bool:
        """The initialized value."""
        return self._initialized
    
    @bm_initialized.setter
    def bm_initialized(self, value )-> None:
        """Set the initialized value."""
        self._initialized = value

    @property
    def bm_folder(self) -> str:
        """The budget folder path is a string, e.g., '~/OneDrive/."""
        return self._budget_folder

    @bm_folder.setter
    def bm_folder(self, value: str) -> None:
        """Set the budget folder path."""
        self._budget_folder = value

    @property
    def bm_store_uri(self) -> str:
        """The budget model store URI."""
        return self._budget_model_store_uri
    
    @bm_store_uri.setter
    def bm_store_uri(self, value: str) -> None:
        """Set the budget model store URI."""
        self._budget_model_store_uri = value

    @property
    def bm_fi_collection(self) -> dict:
        """The financial institutions collection."""
        return self._financial_institutions
    
    @bm_fi_collection.setter
    def bm_fi_collection(self, value: dict) -> None:
        """Set the financial institutions collection."""
        self._financial_institutions = value

    @property
    def bm_wf_collection(self) -> dict:
        """The worklow collection."""
        return self._workflows
    
    @bm_wf_collection.setter
    def bm_wf_collection(self, value: dict) -> None:
        """Set the worklows collection."""
        self._workflows = value

    @property
    def bm_options(self) -> dict:
        """The budget model options dictionary."""
        return self._options
    
    @bm_options.setter
    def bm_options(self, value: dict) -> None:
        """Set the budget model options dictionary."""
        self._options = value

    @property
    def bm_created_date(self) -> str:
        """The created date."""
        return self._created_date
    
    @bm_created_date.setter
    def bm_created_date(self, value: str) -> None:  
        """Set the created date."""
        self._created_date = value

    @property
    def bm_last_modified_date(self) -> str:
        """The last modified date."""
        return self._last_modified_date
    
    @bm_last_modified_date.setter
    def bm_last_modified_date(self, value: str) -> None:
        """Set the last modified date."""
        self._last_modified_date = value

    @property
    def bm_last_modified_by(self) -> str:
        """The last modified by."""
        return self._last_modified_by
    
    @bm_last_modified_by.setter
    def bm_last_modified_by(self, value: str) -> None:
        """Set the last modified by."""
        self._last_modified_by = value
    
    @property
    def bm_working_data(self) -> dict:
        """The budget model working data."""
        return self._wd
    
    @bm_working_data.setter
    def bm_working_data(self, value: dict) -> None:
        """Set the budget model working data."""
        self._wd = value

    # budget_model_working_data is a dictionary to store dynamic, non-property data.
    def set_budget_model_working_data(self, key, value):
        self.budget_model_working_data[key] = value

    def get_budget_model_working_data(self, key):
        return self.budget_model_working_data.get(key, 0)

    #endregion BudgetModel public class properties
    # ------------------------------------------------------------------------ +
    #endregion BudgetModel class intrinsics
    # ======================================================================== +

    # ======================================================================== +
    #region    Budget model Domain Model (BDM) methods
    """ Budget Model Domain Model (BDM) Documentation.

    Budget model Domain Model (BDM) is the conceptual model used to track and 
    analyze transactions from financial institutions (FI) overtime for the 
    purpose of following a budget. As a Domain Model, the implementation is 
    the python data structures used to represent the domain data.

    Class BudgetModel is the base class for the budget model domain. As a 
    convention, will desigining the structure, I adopted the use of constants
    for names of the various types and fields used for th class properties,
    default values and methods. Also, I adopted a pattern to use the
    BudgetModelTemplate class, a subclass of BudgetModel, as a means
    for storing the template for a budget model used for configuration and
    other conveniences. 

    All BDM data storage is the cncern of the BSM. BSM works with Path 
    objects, filenames, folders, relative, absolute path names and actual 
    filesystem operations. In the configuration data, the BDM data is stored 
    as strings. Internally, Path objects are used.

    In the BDM, all data is associated with a financial institution (FI). Data 
    is procecessed by workflows (WF) that are configured for an FI. Each FI has 
    a unique fi_key which the api uses to keep data separated.

    Throughout, the following identifiers are used as a convention:
    
    BF_FOLDER - refers to the root folder for the budget model.

    FI_FOLDER - refers to the root folder configured for a specific 
    Financial Institutions (FI). FI_FOLDER is a Path with various str reprs.

    WF_FOLDER_IN - refers to the input folder for a specific workflow, again, 
    always in the context of a specific FI through the fi_key. WF_FOLDER_IN is
    a Path with various str reprs.

    WF_WORKBOOKS_IN - refers to the collection of workbooks currently in the 
    WF_FOLDER_IN folder.

    WF_FOLDER_OUT - refers to the output folder for a specific workflow. It is
    a Path with various str reprs.

    WF_WORKBOOKS_OUT - refers to the collection of workbooks currently in 
    the WF_FOLDER_OUT folder.

    From the configuration data, the values of BF_FOLDER, FI_FOLDER and
    WF_FOLDER are used to construct folder and file names for the file system
    and the BSM. The BDM methods do not use Path objects. The overlap seen in 
    the WF_WORKBOOKS_IN and WF_WORKBOOKS_OUT for a purpose. The workbooks are
    common in the two layers.
     
    The following naming conventions apply to the purpose of the BDM and BSM 
    methods to  separate the concerns for partial path names, full path names 
    and absolute pathname values..

    Naming Conventions:
    -------------------
    - bm_ - BudgetModel class properties, e.g., bm_folder, bm_fi, etc.
    - fi_ - Related to financial institution, e.g., fi_key, fi_name, etc.
    - wf_ - Related to workflow, e.g., wf_key, wf_name, etc.
    - bdm_ - BudgetModel Domain methods, concerning the in memory data model.
    - bsm_ - BudgetModel Storage Model methods, folders/files stored the filesystem.
    - _path_str - is the simplest string for a path name or component of. These
    methods return str values and do not manipulate with Path objects. Some
    folder values in the BDM, for example, are simply the name of a folder, 
    not a complete path name.
    - _path - constructs a Path object using _path_str values, 
    invokes .expanduser(). 
    _abs_path - invokes .resovle() on _path results, return Path object.
    - Path objects are never serialized to .jsonc files, only _path_str values.

    Abrevs used in method names: 
        bdm - Budget Model Domain, the domain model of the budget model.
        bms - Budget Storage Model, the filesystem storage model.
        bm - BudgetModel class instance, parent of the Budget Model data structure.
        bf - budget folder - attribute, root folder, used in path objects.
        fi - financial institution dictionary, attribute of bm.
             fi_key = int_key, short name of FI, e.g., "boa", "chase", etc.
             fi_value = dict of info about the FI.
        wf - workflow
    """
    # ======================================================================== +
    #region    BDM bdm_initialize(self, budget_node : dict = None) public 
    def bdm_initialize(self, 
                 bm_config_src:dict=None, 
                 bsm_init : bool = True,
                 create_missing_folders : bool = True,
                 raise_errors : bool = True
                 ) -> "BudgetModel":
        """Initialize BDM from config, either bm_config_src, or the template.

        Args:
            bm_config_src (dict): A BudgetModel object to configure from. If
                None, use the BudgetModelTemplate internally.
            bsm_init (bool): Initialize the BSM if True.
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
        """
        st = p3u.start_timer()
        bm_config : dict = None
        try:
            if bm_config_src is None:
                logger.debug("bm_config_src parameter is None, using builtin template.")
                bm_config = BudgetModel.config_template
                if bm_config is None:
                    from data.p3_budget_model.create_config_template import __create_config_template__
                    bm_config = __create_config_template__()
            else:
                bm_config = bm_config_src
            logger.debug(f"Start: ...")
            # Apply the configuration to the budget model (self)
            self.bm_initialized = bm_config.bm_initialized
            self.bm_folder = bm_config.bm_folder
            self.bm_store_uri = bm_config.bm_store_uri
            self.bm_fi_collection = copy.deepcopy(bm_config.bm_fi_collection) if bm_config.bm_fi_collection else {}
            self.bm_wf_collection = copy.deepcopy(bm_config.bm_wf_collection) if bm_config.bm_wf_collection else None
            self.bm_options = bm_config.bm_options.copy() if bm_config.bm_options else {}
            self.bm_created_date = bm_config.bm_created_date
            self.bm_last_modified_date = bm_config.bm_last_modified_date
            self.bm_last_modified_by = bm_config.bm_last_modified_by
            self.bm_working_data = copy.deepcopy(bm_config.bm_working_data) if bm_config.bm_working_data else {}
            if bsm_init:
                self.bsm_inititalize(create_missing_folders, raise_errors)
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return self
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BDM bdm_initialize(self, budget_node : dict = None) public 
    # ------------------------------------------------------------------------ +
    #region BDM FI_OBJECT pseudo-Object properties
    def bdm_validate_FI_KEY(self, fi_key:str) -> bool:
        """Validate the financial institution key."""
        if fi_key not in self.bm_fi_collection.keys():
            m = f"Financial Institution '{fi_key}' not found in "
            m += f"{self.bm_fi_collection.keys()}"
            logger.error(m)
            raise ValueError(m)
        return True

    def bdm_FI_OBJECT(self, fi_key:str) -> FI_OBJECT:
        """Return the FI_OBJECT for fi_key."""
        self.bdm_validate_FI_KEY(fi_key)
        return self.bm_fi_collection[fi_key]
    
    def bdm_FI_OBJECT_count(self) -> int:
        """Return a count of FI_OBJECTS in the FI_COLLECTION."""
        if self.bm_fi_collection is None:
            return 0
        if not isinstance(self.bm_fi_collection, dict):
            m = f"FI_COLLECTION is not a dict: {type(self.bm_fi_collection)}"
            logger.error(m)
            raise ValueError(m)
        return len(self.bm_fi_collection)
    
    def bdm_FI_KEY(self, fi_key:str) -> str:
        """Return the FI_KEY value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_KEY]

    def bdm_FI_NAME(self, fi_key:str) -> str:
        """Return the FI_NAME value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_NAME]
    
    def bdm_FI_TYPE(self, fi_key:str) -> str:
        """Return the FI_TYPE value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_TYPE]

    def bdm_FI_FOLDER(self, fi_key:str) -> str:
        """Return the FI_FOLDER value of the FI_OBJECT for fi_key."""
        return self.bdm_FI_OBJECT(fi_key)[FI_FOLDER]
    
    def bdm_FI_WORKFLOW_DATA(self, fi_key:str) -> WF_DATA_COLLECTION:
        """Return the WF_DATA_COLLECTION value of the FI_OBJECT for fi_key."""
        return self.bdm_FI_OBJECT(fi_key)[FI_WORKFLOW_DATA]

    def bdm_FI_WORKFLOW_DATA_count(self, fi_key:str) -> int:
        """Return a count of WF_DATA_OBJECTS in the WF_DATA_COLLECTION."""
        wf_dc = self.bdm_FI_WORKFLOW_DATA(fi_key)
        if wf_dc is None:
            return 0
        if not isinstance(wf_dc, dict):
            m = f"FI_WORKFLOW_DATA is not a dict: {type(wf_dc)}"
            logger.error(m)
            raise ValueError(m)
        return len(wf_dc)

    def bdm_FI_WF_DATA_OBJECT(self, fi_key:str, wf_key : str) -> WF_DATA_OBJECT:
        """Return the WF_DATA_OBJECT value of the FI_OBJECT for fi_key, wf_key."""
        if self.bdm_FI_WORKFLOW_DATA_count(fi_key) == 0:
            return None
        if self.bdm_FI_WORKFLOW_DATA(fi_key) is None:
            return None
        if wf_key not in self.bdm_FI_WORKFLOW_DATA(fi_key).keys():
            return None
        return self.bdm_FI_WORKFLOW_DATA(fi_key)[wf_key]
    
    def bdm_FI_WF_WORKBOOK_LIST(self, 
                                fi_key:str, wf_key:str, 
                                wb_type:str) -> WORKBOOK_LIST:
        """Return the WORKBOOK_LIST for the specified fi_key, wf_key, wb_type."""
        if not self.bdm_validate_FI_KEY(fi_key):
            m = f"Invalid financial institution key '{fi_key}'."
            logger.error(m)
            raise ValueError(m)
        if not self.bdm_WF_KEY_validate(wf_key):
            m = f"Invalid workflow key '{wf_key}'."
            logger.error(m)
            raise ValueError(m)
        if wb_type not in WF_DATA_OBJECT_KEYS:
            m = f"Invalid workbook type '{wb_type}' for workflow '{wf_key}'."
            logger.error(m)
            raise ValueError(m)
        wf_do = self.bdm_FI_WF_DATA_OBJECT(fi_key, wf_key)
        if wf_do is None:
            return None
        if wb_type not in wf_do.keys():
            return None
        wf_wbl = wf_do[wb_type]
        if wf_wbl is None:
            return None
        if not isinstance(wf_wbl, WORKBOOK_LIST):
            m = f"FI_WORKFLOW_DATA '{fi_key}' '{wf_key}' is not a WORKBOOK_LIST"
            m += f"{type(wf_wbl)}"
            logger.error(m)
            raise ValueError(m)
        return wf_wbl

    def bdm_FI_WF_WORKBOOK_LIST_count(self,
                                fi_key:str, wf_key:str, wb_type:str) -> int:
        """Return a count of WORKBOOK_LIST in the FI_OBJECT for fi_key, wf_key."""
        wf_wbl = self.bdm_FI_WF_WORKBOOK_LIST(fi_key, wf_key, wb_type)
        if wf_wbl is None:
            return 0
        if not isinstance(wf_wbl, WORKBOOK_LIST):
            m = f"FI_WORKFLOW_DATA '{fi_key}' '{wf_key}' is not a WORKBOOK_LIST"
            m += f"{type(wf_wbl)}"
            logger.error(m)
            raise ValueError(m)
        return len(wf_wbl)
    #endregion FI pseudo-Object properties
    # ------------------------------------------------------------------------ +
    #region BDM WF_OBJECT pseudo-Object properties
    def bdm_WF_KEY_validate(self, wf_key:str) -> bool:
        """Validate the workflow key."""
        supp_wf = self.bm_wf_collection
        if supp_wf is None or wf_key not in supp_wf:
            m = f"Workflow('{wf_key}') not found supported workflows "
            m += f"{self.bm_wf_collection}"
            logger.error(m)
            raise ValueError(m)
        return True

    def bdm_WF_OBJECT(self, wf_key:str) -> WF_OBJECT:
        """Return the WF dictionary specified by wf_key."""
        self.bdm_WF_KEY_validate(wf_key)
        return self.bm_wf_collection[wf_key]
    
    def bdm_WF_OBJECT_count(self) -> int:
        """Return a count of WF_OBJECTS in the WF_COLLECTION."""
        if self.bm_wf_collection is None:
            return 0
        if not isinstance(self.bm_wf_collection, dict):
            m = f"WF_COLLECTION is not a dict: {type(self.bm_wf_collection)}"
            logger.error(m)
            raise ValueError(m)
        return len(self.bm_wf_collection)
    
    def bdm_WF_KEY(self, wf_key:str) -> str:
        """Return the WF_KEY value the WF dictionary specified by wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_KEY]

    def bdm_WF_NAME(self, wf_key:str) -> str:
        """Return the WF_NAME value for the specified wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_NAME]
    
    def bdm_WF_FOLDER(self, wf_key:str, folder_id : str) -> str:
        """Return the WF_FOLDER value for the specified wf_key, and folder_id.
        Not a full path, just the configuration value."""
        if folder_id not in (WF_FOLDER_IN, WF_FOLDER_OUT):
            m = f"Invalid folder_id '{folder_id}' for workflow '{wf_key}'."
            logger.error(m)
            raise ValueError(m)
        return self.bdm_WF_OBJECT(wf_key)[folder_id]
    
    def bdm_WF_PREFIX_IN(self, wf_key:str) -> str:
        """Return the WF_PREFIX_IN value for the specified wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_PREFIX_IN]
    
    def bdm_WF_PREFIX_OUT(self, wf_key:str) -> str:
        """Return the WF_PREFIX_OUT value for the specified wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_PREFIX_OUT]    
    
    def bdm_WF_WORKBOOK_MAP(self, wf_key:str, wb_type:str=None) -> str|dict:
        """Return the WF_WORKBOOK_MAP or specific value for the specified wb_type."""
        if wb_type is None:
            return self.bdm_WF_OBJECT(wf_key)[WF_WORKBOOK_MAP]
        return self.bdm_WF_OBJECT(wf_key)[WF_WORKBOOK_MAP][wb_type]
    #endregion BDM WF_OBJECT pseudo-Object properties
    # ------------------------------------------------------------------------ +
    #endregion BudgetModel Domain methods
    # ======================================================================== +

    # ======================================================================== +
    #region Budget model Storage Model (BSM) methods
    """ Budget model Storage Model (BSM) Documentation.

    All BDM data is stored in the filesystem by the BSM. BSM works with Path 
    objects, filenames, folders, relative, absolute path names and actual 
    filesystem operations. In the configuration data, the BDM data is stored 
    as strings. Internally, Path objects are used.

    In the BDM, all data is associated with a financial institution (FI). Data 
    is procecessed by workflows (WF) that are configured for an FI. Hence,
    in all cases the BSM requries an fi_key to identify which FI's data is being
    processed. The wf_key, identifiying a particular workflow, is also used
    in the BSM methods, because naming conventions for folder and files are 
    based on the workflow settings.

    Throughout, the following identifiers are used as a convention:
    
    FI_FOLDER - refers to the root folder configured for a specific 
    Financial Institutions (FI). FI_FOLDER is a Path with various str reprs.

    WF_FOLDER_IN - refers to the input folder for a specific workflow, again, 
    always in the context of a specific FI through the fi_key. WF_FOLDER_IN is
    a Path with various str reprs.

    WF_WORKBOOKS_IN - refers to the collection of workbooks currently in the 
    WF_FOLDER_IN folder.

    WF_FOLDER_OUT - refers to the output folder for a specific workflow. It is
    a Path with various str reprs.

    WF_WORKBOOKS_OUT - refers to the collection of workbooks currently in 
    the WF_FOLDER_OUT folder.

    The configuration data and stored budget model data contains str values 
    that are utilized in path names for folders and files. The following 
    naming conventions applyr to the purpose of the BDM and BSM methods to 
    separate the concerns for partial path names, full path names and absolute
    pathname values..

    Naming Conventions:
    -------------------
    - bm_ - BudgetModel class properties, e.g., bm_folder, bm_fi, etc.
    - fi_ - Related to financial institution, e.g., fi_key, fi_name, etc.
    - wf_ - Related to workflow, e.g., wf_key, wf_name, etc.
    - bdm_ - BudgetModel Domain methods, concerning the in memory data model.
    - bsm_ - BudgetModel Storage Model methods, folders/files stored the filesystem.
    - _path_str - is the simplest string for a path name or component of. These
    methods return str values and do not manipulate with Path objects. Some
    folder values in the BDM, for example, are simply the name of a folder, 
    not a complete path name.
    - _path - constructs a Path object using _path_str values, 
    invokes .expanduser(). 
    _abs_path - invokes .resovle() on _path results, return Path object.
    - Path objects are never serialized to .jsonc files, only _path_str values.

    Abrevs used in method names: 
        bdm - Budget Model Domain, the domain model of the budget model.
        bms - Budget Storage Model, the filesystem storage model.
        bm - BudgetModel class instance, parent of the Budget Model data structure.
        bf - budget folder - attribute, root folder, used in path objects.
        fi - financial institution dictionary, attribute of bm.
             fi_key = int_key, short name of FI, e.g., "boa", "chase", etc.
             fi_value = dict of info about the FI.
        wf - workflow
    """
    # ======================================================================== +
    #region bsm_inititalize() method
    def bsm_inititalize(self, 
                        create_missing_folders:bool=True,
                        raise_errors:bool=True) -> "BudgetModel":
        """Initialize BSM aspects of the BDM.
        
        Examine elements of self, the BudgetModel class, representing the BDM.
        Validate the mapping of the BDM data dependent on mapping to folders 
        and files in the filesystem. Flags control whether to create the
        filesystem structure if it is not present. Also scan the workflow 
        folders for the presence of workbook excel files and load their 
        references into the BDM as appropriate.

        Args:
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
        """
        st = p3u.start_timer()
        # Plan: validate filesystem folders: bf, fi, wf
        #       update the BDM workbook mappings to BSM 
        try:
            logger.debug("Start: ...")
            self.bsm_BM_FOLDER_resolve(create_missing_folders, raise_errors)
            # Enumerate the financial institutions.
            for fi_key, fi_object in self.bm_fi_collection.items():
                # Resolve FI_FOLDER path.
                self.bsm_FI_FOLDER_resolve(fi_key, 
                                           create_missing_folders, raise_errors)
                # Resolve WF_FOLDER paths for the workflows.
                self.bsm_WF_FOLDER_resolve(fi_key, 
                                           create_missing_folders, raise_errors)
                # Resolve the FI_WORKFLOW_DATA collection, refresh actual
                # data from folders in BSM.
                self.bsm_FI_WORKFLOW_DATA_resolve(fi_key)
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return self
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_inititalize() method
    # ------------------------------------------------------------------------ +    
    #region BM_FOLDER Path methods
    def bsm_BM_FOLDER_validate(self) -> bool:
        """Validate the bm_folder property setting.
        
        Raise a ValueError if the bm_folder property is not set or is not
        usable as part of a valid path string.
        """
        # TODO: expand and resolve flags?
        if self.bm_folder is None or len(self.bm_folder) == 0:
            m = f"Budget folder path is not set. "
            m += f"Set the BM_FOLDER('{BM_FOLDER}') property to valid path value."
            logger.error(m)
            raise ValueError(m)
        return True
    def bsm_BM_FOLDER_path_str(self) -> str:
        """str version of the BM_FOLDER value as a Path."""
        # In the BSM, the bm_folder property must be a valid setting that will
        # result in a valid Path. Raise a ValueError if not.
        self.bsm_BM_FOLDER_validate() # Raises ValueError if not valid
        return str(Path(self.bm_folder))
    def bsm_BM_FOLDER_path(self) -> Path:
        """Path of self.bsm_BM_FOLDER_path_str().expanduser()."""
        return Path(self.bsm_BM_FOLDER_path_str()).expanduser()
    def bsm_BM_FOLDER_abs_path(self) -> Path:
        """Path of self.bsm_BM_FOLDER_path().resolve()."""
        return self.bsm_BM_FOLDER_path().resolve()
    def bsm_BM_FOLDER_abs_path_str(self) -> str:
        """str of self.bsm_BM_FOLDER_abs_path()."""
        return str(self.bsm_BM_FOLDER_abs_path())
    
    def bsm_BM_FOLDER_resolve(self, 
                              create_missing_folders : bool=True,
                              raise_errors : bool=True) -> None:
        """Resolve the BM_FOLDER path and create it if it does not exist."""
        try:
            logger.info(f"Checking BM_FOLDER path: '{self.bm_folder}'")
            if self.bm_folder is None:
                m = f"Budget folder path is not set. "
                m += f"Set the '{BM_FOLDER}' property to a valid path."
                logger.error(m)
                raise ValueError(m)
            # Resolve the BM_FOLDER path.
            bf_ap = self.bsm_BM_FOLDER_abs_path()
            bsm_verify_folder(bf_ap, create_missing_folders, raise_errors)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BM_FOLDER Path methods
    # ------------------------------------------------------------------------ +
    #region FI_OBJECT pseudo-Object methods
    # ------------------------------------------------------------------------ +   
    #region FI_OBJECT FI_FOLDER Path methods
    def bsm_FI_FOLDER_path_str(self, fi_key: str) -> str:
        """str version of the FI_FOLDER value as a Path.
        
        Raises ValueError for invalid settings for any of: bm_folder property,
        fi_key, of FI_FOLDER value of FI dictionary."""
        bf_p_s = self.bsm_BM_FOLDER_path_str() # ValueError on BM_FOLDER property
        self.bdm_validate_FI_KEY(fi_key) # ValueError fi_key
        fi_p_s = str(Path(self.bdm_FI_FOLDER(fi_key))) 
        if fi_p_s is None or len(fi_p_s) == 0:
            m = f"FI_FOLDER value is not set for FI_KEY('{fi_key}'). "
            m += f"In the BudgetModel configuration, correct FI_FOLDER setting."
            logger.error(m)
            raise ValueError(m)   
        return str(Path(bf_p_s) / fi_p_s)
    def bsm_FI_FOLDER_path(self, fi_key: str) -> Path:
        """Path of self.bsm_FI_FOLDER_path_str().expanduser()."""
        return Path(self.bsm_FI_FOLDER_path_str(fi_key)).expanduser()
    def bsm_FI_FOLDER_abs_path(self, fi_key: str) -> Path:
        """Path of self.bsm_FI_FOLDER_path().resolve()."""
        return Path(self.bsm_FI_FOLDER_path(fi_key)).resolve()
    def bsm_FI_FOLDER_abs_path_str(self, fi_key: str) -> str:
        """str of self.bsm_FI_FOLDER_abs_path()."""
        return str(self.bsm_FI_FOLDER_abs_path(fi_key))
    def bsm_FI_FOLDER_resolve(self, fi_key : str,
                              create_missing_folders : bool=True,
                                raise_errors : bool=True) -> None: 
        """Resolve the FI_FOLDER path and create it if it does not exist."""
        fi_ap = self.bsm_FI_FOLDER_abs_path(fi_key)
        logger.info(f"FI_KEY('{fi_key}') Checking FI_FOLDER('{fi_ap}')")
        bsm_verify_folder(fi_ap, create_missing_folders, raise_errors)
    #endregion FI_OBJECT FI_FOLDER Path methods
    # ------------------------------------------------------------------------ +   
    #region FI_OBJECT FI_WORKFLOW_DATA pseudo-Object methods
    def bsm_FI_WORKFLOW_DATA_resolve(self, fi_key:str) -> None:
        """Resolve the FI_WORKFLOW_DATA for the specified fi_key and wf_key."""
        try:
            if self.bdm_FI_WORKFLOW_DATA_count(fi_key) == 0:
                m = f"FI_KEY('{fi_key}') has no workflow data."
                logger.debug(m)
                return
            # Enumerate the FI_OBJECT WF_DATA_COLLECTION.
            wf_dc : WF_DATA_COLLECTION = self.bdm_FI_WORKFLOW_DATA(fi_key)
            for wf_key, wf_data_object in wf_dc.items():
                # Resolve each WF_DATA_OBJECT in the collection.
                self.bms_WF_DATA_OBJECT_resolve(wf_data_object,
                                                fi_key, wf_key)
        except Exception as e:
                m = p3u.exc_err_msg(e)
                logger.error(m)
                raise
    #endregion FI_OBJECT FI_WORKFLOW_DATA pseudo-Object methods
    # ------------------------------------------------------------------------ +   
    #endregion FI FI_OBJECT pseudo-Object methods 
    # ------------------------------------------------------------------------ +   
    #region WF_OBJECT WF_FOLDER Path methods
    """WF_FOLDER refers to a str element used to construct a folder path name.
    It is an actual Path value or a string that is used
    to initialize a Path object. These Path mehtods are for that purpose.
    """
    def bsm_WF_FOLDER_resolve(self, fi_key:str, 
                              create_missing_folders:bool=True, 
                              raise_errors:bool=True) -> None:
        """Resolve any WF-related folders for the fi_key, create if requested."""
        try:
            logger.debug(f"FI_KEY('{fi_key}') scan all WF_FOLDERs.")
            for wf_key, wf_object in self.bm_wf_collection.items():
                # Resolve the WF_FOLDER path elements.
                for f_id in WF_FOLDER_PATH_ELEMENTS:
                    wf_in_p = self.bsm_WF_FOLDER_path(fi_key, wf_key, f_id)
                    if wf_in_p is not None:
                        logger.info(f"FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
                                    f"Checking WF_FOLDER('{f_id}')")
                        bsm_verify_folder(wf_in_p, create_missing_folders, 
                                         raise_errors)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bsm_WF_FOLDER_path_str(self, fi_key : str, wf_key : str,
                               folder_id : str) -> str:
        """str version of the WF_FOLDER value for fi_key/wf_key/folder_id."""
        if folder_id not in [WF_FOLDER_IN, WF_FOLDER_OUT]:
            m = f"Invalid folder_id '{folder_id}' for FI_KEY('{fi_key}') "
            m += f"and WF_KEY('{wf_key}')"
            logger.error(m)
            raise ValueError(m)
        fi_p_s = self.bsm_FI_FOLDER_path_str(fi_key) # FI_FOLDER Path component
        wf_p_s = self.bdm_WF_FOLDER(wf_key, folder_id)
        return str(Path(fi_p_s) / wf_p_s) if wf_p_s is not None else None
    def bsm_WF_FOLDER_path(self, fi_key : str, wf_key : str, 
                           folder_id:str) -> Path:
        """Path of self.bsm_wf_folder_path_str().expanduser()."""
        p_s = self.bsm_WF_FOLDER_path_str(fi_key, wf_key, folder_id)
        return Path(p_s).expanduser() if p_s is not None else None
    def bsm_WF_FOLDER_abs_path(self, fi_key : str, wf_key : str,
                               folder_id : str) -> Path:
        """Path of self.bsm_wf_folder_path().resolve()."""
        p = self.bsm_WF_FOLDER_path(fi_key, wf_key, folder_id)
        return p.resolve() if p is not None else None
    def bsm_WF_FOLDER_abs_path_str(self, fi_key : str, wf_key : str,
                                   folder_id : str) -> str:
        """str of self.bsm_wf_folder_abs_path()."""
        p = self.bsm_WF_FOLDER_abs_path(fi_key, wf_key, folder_id)
        return str(p) if p is not None else None
    #endregion WF_OBJECT WF_FOLDER Path methods
    # ------------------------------------------------------------------------ +
    #region WF_OBJECT WF_DATA_OBJECT (WF_DO) pseudo-property methods
    """
    Witht the BSM, the properties relating to the storage model are 
    WF_FOLDER_IN, WF_WORKBOOKS_IN, WF_FOLDER_OUT and WF_WORKBOOKS_OUT. These 
    concern actual access to data in the filesystem. The BDM methods are used to
    access the data in the BDM.

    One key overlap, is bsm_WF_WORKBOOKS_IN and bsm_WF_WORKBOOKS_OUT. These are 
    the same methods in both the BDM and BSM. Each are dictionaries with tuples
    of each workbook in the corresponding WF_FOLDER_IN and WF_FOLDER_OUT folders.
    The key is the filename, or the workbook name, the value is the full path 
    to the file.
    """
    def bms_WF_DATA_OBJECT_resolve(self, wf_do: WF_DATA_OBJECT,
                                   fi_key : str, wf_key : str):
        """Resolve the WF_DATA_OBJECT based on the keys and values present."""
        try:
            logger.debug(f"FI_KEY('{fi_key}') WF_KEY('{wf_key}')")
            if wf_do is None or len(wf_do) == 0:
                logger.debug(f"  WF_DATA_OBJECT is empty.")
                return
            logger.debug(f"  WF_DATA_OBJECT({len(wf_do)} keys): {str(wf_do.keys())}")
            # Resolve all keys in the WF_DATA_OBJECT.
            did_workbooks = False
            for wf_do_key, wf_do_value in wf_do.items():
                if wf_do_key not in WF_DATA_OBJECT_VALID_KEYS:
                    m = f"Invalid WF_DATA_OBJECT key '{wf_do_key}' "
                    m += f"for FI_KEY('{fi_key}') and WF_KEY('{wf_key}')"
                    logger.error(m)
                    raise ValueError(m)
                if wf_do_key in WF_WORKBOOK_TYPES:
                    if not did_workbooks:
                        # Resolve all the WORKBOOK_LIST for WF_FOLDERs, just once.
                        self.bsm_WF_WORKBOOKS_resolve(wf_do, fi_key, wf_key)
                        did_workbooks = True
                    else:
                        continue
                else:
                    raise NotImplementedError(f"WF_DATA_OBJECT key '{wf_do_key}' ")
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
 
    def bsm_WF_WORKBOOKS_resolve(self, wf_do: WF_DATA_OBJECT, fi_key: str, wf_key : str):
        """Resolve all WORKBOOK_LISTs for WF_FOLDERS(fi_key, wf_key) in the WF_DATA_OBJECT. 

        Workbooks are contained in folders in the filesystem. The folder_id 
        determines the absolute path to the folder, as per the following 
        mapping:
        - WF_WORKBOOKS_IN -> WF_FOLDER_IN - input folder for the workflow.
        - WF_WORKBOOKS_OUT -> WF_FOLDER_OUT - output folder for the workflow.

        Args:
            fi_key (str): The key of the institution to get the workbooks for.
            wf_key (str): The workflow to get the workbooks for.
            folder_id (str): The folder id to get the workbooks for. Must be
            listed in BM_VALID_PATH_ELEMENTS.

        Returns:
            dict { str: str, ... }: A dict of workbook file names and their paths.
        """
        try:
            # TODO: validate parameters.
            for wf_do_key in wf_do.keys():
                # Only handle WF_WORKBOOK scope for the workflow.
                # A wf_do may have other keys, but we only care about the workbooks.
                if wf_do_key not in WF_WORKBOOK_TYPES:
                    logger.debubg(f"  Skipping WF_DATA_OBJECT key: '{wf_do_key}'")
                    continue
                # Only interested in the WF_WORKBOOK_TYPES keys. Each of them
                # is configured to one WF_FOLDER_PATH_ELEMENT. Get the
                # appropriate folder_id for the WF_WORKBOOK_TYPE.
                f_id = self.bdm_WF_WORKBOOK_MAP(wf_key, wf_do_key)
                if f_id is None:
                    logger.debug(f"  FI('{fi_key}') WF('{wf_key}') "
                                f"WF_DO['{wf_do_key}'] is None")
                    continue
                # WORKBOOKS are found in WF_FOLDERS, scan them all.
                # Resolve all folders in this workflow for fi_key, wf_key.
                wb_list = []
                wf_f_ap = self.bsm_WF_FOLDER_abs_path(fi_key, wf_key, f_id)
                wb_files = list(wf_f_ap.glob("*.xlsx"))
                for wb_p in wb_files:
                    wb_name = wb_p.name
                    wb_item = (wb_name, str(wb_p))
                    wb_list.append(wb_item)
                wf_do[wf_do_key] = wb_list # save wb_list to the wf_do
                logger.debug(f"  FI('{fi_key}') WF('{wf_key}') "
                            f"WF_DO['{wf_do_key}'] maps to FOLDER('{f_id}') "
                            f"with {len(wb_list)} workbooks)")
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise

    def bsm_FI_WF_WORKBOOK_generator(self,
                fi_key : str, 
                wf_key : str, 
                wb_type : str) -> Generator[Tuple[str, Workbook], None, None]: 
        """For fi_key,wf_key,wb_type, yield (wb_name, loaded_Workbook).
        
        Yields:
            Tuple[str, Workbook]: a tuple containing the file name the 
            loaded workbook object.
        """
        try:
            wbl = self.bdm_FI_WF_WORKBOOK_LIST(fi_key, wf_key, wb_type)
            if wbl is None:
                return
            for wb_name, wb_path in wbl:
                wb = load_workbook(filename=wb_path)
                yield (wb_name, wb)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bsm_FI_WF_WORKBOOK_save(self, wb : Workbook, wb_name: str, 
                               fi_key:str, wf_key:str, wb_type : str):
        """Save workbook output to storage associated (fi_key,wf_key,wb_type).
        
        Map the fi_key and wf_key to the appropriate WF_FOLDER_OUT folder in 
        the filesystem.
        """
        try:
            f_id = self.bdm_WF_WORKBOOK_MAP(wf_key, wb_type)
            if f_id is None:
                m = f"Invalid folder_id '{f_id}' for FI_KEY('{fi_key}') "
                m += f"and WF_KEY('{wf_key}')"
                logger.error(m)
                raise ValueError(m)
            fi_wf_ap = self.bsm_WF_FOLDER_abs_path(fi_key, wf_key, f_id)
            # TODO: strip the in_prefix if it is there.
            # Prepend the out_prefix to the workbook name.
            wb_name = f"{self.bdm_WF_PREFIX_OUT(wf_key)}{wb_name}"
            wb_path = fi_wf_ap / wb_name
            wb.save(wb_path)
            logger.info(f"Saved workbook '{wb_name}' to '{str(fi_wf_ap)}'")
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion WF_OBJECT WF_DATA_OBJECT (WF_DO) pseudo-property methods
    # ------------------------------------------------------------------------ +   
    #region load_fi_transactions() function
    def load_fi_transactions(self, fi_key:str, process_folder: str, 
                             trans_file : str = None) -> Workbook:
        """Get FI transactions from an excel file into an openpyxl.Workbook.
        
        Transactions downloaded from an FI are in a folder. 
        The file is assumed to be in the folder specified in the budget_config. 
        A folder is specified in the budget_config.json file. That folder is scanned 
        for transactions in excel files if a particular file is not specified.

        TODO: If trans_file is not specified, then scan the folder for files. 
        Return a dictionary of workbooks with the file name as the key.

        Args:
            trans_file (str): The name of the transaction file to load.

        Returns:
            Workbook: The workbook containing the FI transactions.

        """
        me = self.load_fi_transactions
        st = time.time()
        try:
            trans_path = self.fi_abs_path / trans_file
            logger.info("Loading FI transactions...")
            wb = load_workbook(filename=trans_path)
            logger.info(f"Loaded FI transactions from {str(trans_path)}")
            delta = f"{time.time() - st:.3f} seconds."
            logger.info(f"Complete: {delta}")
            return wb
        except Exception as e:
            logger.error(p3u.exc_msg(me, e))
            raise    
    #endregion load_fi_transactions() function
    # ------------------------------------------------------------------------ +
    #region fi_load_workbook(self, fi_key:str, process_folder:str, file_name:str) function
    def fi_load_workbook(self, fi_key:str, workflow:str, 
                              workbook_name:str) -> Workbook:
        """Load a transaction file for a Financial Institution Workflow.

        ViewModel: This is a ViewModel function, mapping budget domain model 
        to how budget model data is stored in filesystem.

        Args:
            fi_key (str): The key of the institution to load the transaction file for.
            workflow (str): The workflow to load the transaction file from.
            workbook_name (str): The name of the workbook file to load.

        Returns:
            Workbook: The loaded transaction workbook.
        """
        me = self.fi_load_workbook
        try:
            # Budget Folder Financial Institution Workflow Folder absolute path
            bffiwfap = self.bsm_FI_FOLDER_wf_abs_path(fi_key, workflow) 
            wbap = bffiwfap / workbook_name # workbook absolute path
            m = f"BDM: Loading FI('{fi_key}') workflow('{workflow}') "
            m += f"workbook('{workbook_name}'): abs_path: '{str(wbap)}'"
            logger.debug(m)
            wb = self.bsm_load_workbook(wbap)
            return wb
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion fi_load_workbook(self, fi_key:str, process_folder:str) function
    # ------------------------------------------------------------------------ +
    #region bsm_load_workbook(self, workbook_path:Path) function
    def bsm_load_workbook(self, workbook_path:Path) -> Workbook:
        """Load a transaction file for a Financial Institution Workflow.

        Storage Model: This is a Model function, loading an excel workbook
        file into memory.

        Args:
            workbook_path (Path): The path of the workbook file to load.

        Returns:
            Workbook: The loaded transaction workbook.
        """
        me = self.bsm_load_workbook
        try:
            logger.debug(f"BSM: Loading workbook file: '{workbook_path}'")
            wb = load_workbook(filename=workbook_path)
            return wb
        except Exception as e:
            logger.error(p3u.exc_msg(me, e))
            raise
    #endregion bsm_load_workbook(self, fi_key:str, process_folder:str) function
    # ------------------------------------------------------------------------ +
    #region bsm_save_workbook() function
    def bsm_save_workbook(self, workbook : Workbook = None, output_path:str=None) -> None:
        """Save the FI transactions workbook to the filesystem.
        
        The file is assumed to be in the folder specified in the budget_config. 
        Use the folder specified in the budget_config.json file. 
        the budget_config may have an output_prefix specified which will be
        prepended to the output_path name.

        TODO: If output_path is not specified, then scan the folder for files. 
        Return a dictionary of workbooks with the file name as the key.

        Args:
            output_path (str): The absolute path of the transaction file to save.

        Returns:
            None

        """
        me = self.save_fi_transactions
        st = time.time()
        try:
            # TODO: add logic to for workbook open in excel, work around.
            if (budget_config["output_prefix"] is not None and 
                isinstance(budget_config["output_prefix"], str) and
                len(budget_config["output_prefix"]) > 0):
                file_path = budget_config["output_prefix"] + output_path
            else:
                file_path = output_path
            trans_path = Path(budget_config["bank_transactions_folder"]) / file_path
            logger.info("Saving FI transactions...")
            workbook.save(filename=trans_path)
            logger.info(f"Saved FI transactions to '{str(trans_path)}'")
            return
        except Exception as e:
            logger.error(p3u.exc_msg(me, e))
            raise    
    #endregion bsm_save_workbook() function
    # ------------------------------------------------------------------------ +
    #endregion BudgetModel Storage Model methods
    # ======================================================================== +
# ---------------------------------------------------------------------------- +
#region log_BDM_info() function
def log_BDM_info(bm : BudgetModel) -> None:  
    try:
        if bm is None:
            logger.warning("bm parameter is None.")
            return None
        logger.debug("BDM Content:")
        logger.debug(f"{P2}BM_INITIALIZED['{BM_INITIALIZED}']: "
                     f"{bm.bm_initialized}")
        logger.debug(f"{P2}BM_FOLDER['{BM_FOLDER}']: '{bm.bm_folder}'")
        logger.debug(f"{P2}BM_STORE_URI['{BM_STORE_URI}]: '{bm.bm_store_uri}'")
        # Enumerate the financial institutions in the budget model
        c = bm.bdm_FI_OBJECT_count()
        logger.debug(
            f"{P2}BM_FI_COLLECTION['{BM_FI_COLLECTION}']({c}): "
            f"{str(list(bm.bm_fi_collection.keys()))}")
        for fi_key in bm.bm_fi_collection.keys():
            logger.debug(f"{P4}Financial Institution: "
                         f"{bm.bdm_FI_KEY(fi_key)}:{bm.bdm_FI_NAME(fi_key)}:"
                         f"{bm.bdm_FI_TYPE(fi_key)}: '{bm.bdm_FI_FOLDER(fi_key)}'")
        # Enumerate workflows in the budget model
        c = bm.bdm_WF_OBJECT_count()
        logger.debug(
            f"{P2}BM_WF_COLLECTION['{BM_WF_COLLECTION}']({c}): "
            f"{str(list(bm.bm_wf_collection.keys()))}")
        for wf_key in bm.bm_wf_collection.keys():
            logger.debug(f"{P4}Workflow:({bm.bdm_WF_KEY(wf_key)}:{bm.bdm_WF_NAME(wf_key)}: ")
            logger.debug(f"{P6}WF_WORKBOOKS_IN: '{bm.bdm_WF_FOLDER(wf_key,WF_FOLDER_IN)}'")
            logger.debug(f"{P6}WF_FOLDER_OUT: '{bm.bdm_WF_FOLDER(wf_key,WF_FOLDER_OUT)}'")
            logger.debug(f"{P6}WF_PREFIX_IN: '{bm.bdm_WF_PREFIX_IN(wf_key)}' "
                         f"WF_PREFIX_OUT: '{bm.bdm_WF_PREFIX_OUT(wf_key)}'")
            logger.debug(f"{P6}WF_WORKBOOK_MAP: {str(bm.bdm_WF_WORKBOOK_MAP(wf_key))}")
        # Enumerate Budget Model Options
        bmoc = len(bm.bm_options)
        logger.debug(f"{P2}BM_OPTION['{BM_OPTIONS}']({bmoc})")
        for opt_key, opt in bm.bm_options.items():
            logger.debug(f"{P4}Option('{opt_key}') = '{opt}'")

        # And the rest
        logger.debug(f"{P2}BM_CREATED_DATE['{BM_CREATED_DATE}']: "
                        f"{bm.bm_created_date}")
        logger.debug(f"{P2}BM_LAST_MODIFIED_DATE['{BM_LAST_MODIFIED_DATE}']: "
                        f"{bm.bm_last_modified_date}")
        logger.debug(f"{P2}BM_LAST_MODIFIED_BY['{BM_LAST_MODIFIED_BY}']: "
                        f"'{bm.bm_last_modified_by}'")
        logger.debug(f"{P2}BM_WORKING_DATA('{BM_WORKING_DATA}'): "
                        f"'{bm.bm_working_data}'")
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion log_BDM_info() function
# ---------------------------------------------------------------------------- +
#region log_BSM_info() function
def log_BSM_info(bm : BudgetModel) -> None:  
    try:
        if bm is None:
            logger.warning("bm parameter is None.")
            return None
        logger.debug("BSM Content:")
        bf_p = bm.bsm_BM_FOLDER_path() # budget folder path
        bf_p_exists = "exists." if bf_p.exists() else "does not exist!"
        bf_ap = bm.bsm_BM_FOLDER_abs_path() # budget folder path
        bf_ap_exists = "exists." if bf_ap.exists() else "does not exist!"
        logger.debug(f"{P2}BM_FOLDER['{BM_FOLDER}']: '{bm.bm_folder}' {bf_ap_exists}")
        logger.debug(f"{P4}bsm_BM_FOLDER_path(): '{str(bf_p)}' {bf_p_exists}")
        logger.debug(f"{P4}bsm_BM_FOLDER_abs_path(): '{str(bf_ap)}' {bf_ap_exists}")
        bmc_p = bf_p / BSM_DEFAULT_BUDGET_MODEL_FILE_NAME # bmc: BM config file
        bmc_p_exists = "exists." if bmc_p.exists() else "does not exist!"
        logger.debug(
            f"{P2}BM_STORE_URI['{BM_STORE_URI}]: '{bm.bm_store_uri}' "
            f"{bmc_p_exists}")
        # Enumerate the financial institutions in the budget model
        c = bm.bdm_FI_OBJECT_count()
        logger.debug(
            f"{P2}BM_FI_COLLECTION['{BM_FI_COLLECTION}']({c}): "
            f"{str(list(bm.bm_fi_collection.keys()))}")
        for fi_key in bm.bm_fi_collection.keys():
            logger.debug(f"{P4}Financial Institution: "
                         f"{bm.bdm_FI_KEY(fi_key)}:{bm.bdm_FI_NAME(fi_key)}:"
                         f"{bm.bdm_FI_TYPE(fi_key)}: '{bm.bdm_FI_FOLDER(fi_key)}'")
            c = bm.bdm_FI_WORKFLOW_DATA_count(fi_key)
            fi_wd = bm.bdm_FI_WORKFLOW_DATA(fi_key)
            m = str(list(fi_wd.keys())) if fi_wd is not None else "'None'"
            logger.debug(f"{P6}Workflow Data({c}): {m}")
        # Enumerate workflows in the budget model
        c = bm.bdm_WF_OBJECT_count()
        logger.debug(
            f"{P2}BM_WF_COLLECTION['{BM_WF_COLLECTION}']({c}): "
            f"{str(list(bm.bm_wf_collection.keys()))}")
        for wf_key in bm.bm_wf_collection.keys():
            logger.debug(f"{P4}Workflow:({bm.bdm_WF_KEY(wf_key)}:{bm.bdm_WF_NAME(wf_key)}: ")
            logger.debug(f"{P6}WF_FOLDER_IN: '{bm.bdm_WF_FOLDER(wf_key,WF_FOLDER_IN)}'")
            logger.debug(f"{P6}WF_FOLDER_OUT: '{bm.bdm_WF_FOLDER(wf_key,WF_FOLDER_OUT)}'")
            logger.debug(f"{P6}WF_WORKBOOK_MAP: {str(bm.bdm_WF_WORKBOOK_MAP(wf_key))}")
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion log_BSM_info() function
# ---------------------------------------------------------------------------- +
#region check_budget_model() -> bool
def check_budget_model() -> bool:   
    """Quick check of the budget model schema is valid.

    Returns:
        bool: True if the budget model schema is valid, raise otherwise.
    """
    me = check_budget_model
    global budget_config, budget_model, budget_config_expected_keys
    global institution_expected_keys, options_expected_keys
    try:
        if (budget_model is None or 
            not isinstance(budget_model, dict) or
            len(budget_model) == 0):
            logger.warning("Budget model is not initialized.")
            return False
        # Check if the budget model contains expected keys
        for key in budget_config_expected_keys:
            if key not in budget_model:
                m = f"Budget model missing key: '{key}'"
                logger.error(m)
                raise ValueError(m)
        # Check if the institutions are valid
        for institution_key in BudgetModel.valid_institutions_keys:
            if institution_key not in budget_config["institutions"].keys():
                m = f"Budget model has invalid institution key: '{institution_key}'"
                logger.error(m)
                raise ValueError(m)
            for key in institution_expected_keys:
                if key not in budget_model["institutions"][institution_key]:
                    m = f"Budget model institution[{institution_key}] missing key: '{key}'"
                    logger.error(m)
                    raise ValueError(m)
        for key in options_expected_keys:
            if key not in budget_model["options"]:
                m = f"Budget model missing options key: '{key}'"
                logger.error(m)
                raise ValueError(m)
        return True
    except Exception as e:
        logger.error(p3u.exc_msg(me, e))
        return False
#endregion check_budget_model() -> bool
# ---------------------------------------------------------------------------- +
#region verify_folder(ap: Path, create:bool=True, raise_errors:bool=True) -> bool
def bsm_verify_folder(ap: Path, create:bool=True, raise_errors:bool=True) -> bool:
    """Verify the folder exists, create it if it does not exist.

    Args:
        ap (Path): The absolute path to the folder to verify.
        create (bool): Create the folder if it does not exist.
        raise_errors (bool): Raise errors if True.
    """
    try:
        if not ap.is_absolute():
            m = f"Path is not absolute: '{str(ap)}'"
            logger.error(m)
            raise ValueError(m)
        if ap.exists() and ap.is_dir():
            logger.debug(f"Folder exists: '{str(ap)}'")
            return True
        if not ap.exists():
            m = f"Folder does not exist: '{str(ap)}'"
            logger.error(m)
            if create:
                logger.info(f"Creating folder: '{str(ap)}'")
                ap.mkdir(parents=True, exist_ok=True)
            else:
                raise ValueError(m)
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion verify_folder(ap: Path, create:bool=True, raise_errors:bool=True) -> bool
# ---------------------------------------------------------------------------- +
#region Local __main__ stand-alone
if __name__ == "__main__":
    try:
        # this_app_name = os.path.basename(__file__)
        # Configure logging
        logger_name = THIS_APP_NAME
        log_config = "budget_model_logging_config.jsonc"
        # Set the log filename for this application.
        # filenames = {"file": "logs/p3ExcelBudget.log"}
        _ = p3l.setup_logging(
            logger_name = logger_name,
            config_file = log_config
        )
        p3l.set_log_flag(p3l.LOG_FLAG_PRINT_CONFIG_ERRORS, True)
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)
        logger.info("+ ----------------------------------------------------- +")
        logger.info(f"+ Running {THIS_APP_NAME} ...")
        logger.info("+ ----------------------------------------------------- +")
        logger.debug(f"Start: {THIS_APP_NAME}...")

        bm = BudgetModel()
        bm.inititailize()
        bms = str(bm)
        bmr = repr(bm)
        bdm = bm.to_dict()

        logger.debug(f"Budget Model: str() = '{bms}'")
        logger.debug(f"Budget Model: repr() = '{bmr}'")
        logger.debug(f"Budget Model: to_dict() = '{bdm}'")
        logger.info(f"Budget Model: {str(bm)}")
        _ = "pause"
    except Exception as e:
        m = p3u.exc_msg("__main__", e)
        logger.error(m)
    logger.info(f"Exiting {THIS_APP_NAME}...")
    exit(1)
#endregion
# ---------------------------------------------------------------------------- +
