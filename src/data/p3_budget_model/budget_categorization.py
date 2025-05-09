# ---------------------------------------------------------------------------- +
#region p3_budget_categorization.py module
""" Financial Budget Workflow: "categorization" of transaction workbooks.

    Workflow: cantegorization
    Input Folder: Financial Institution (FI) Incoming Folder (IF)
    Output Folder: Financial Institution (FI) Categorized Folder (CF)
    FI transaction workbooks are typically excel files. 

    Workflow Pathern: Apply a workflow_proccess (function) to each item in the 
    input folder, placing items in the output folder as appropiate to the 
    configured function. Each WorkFLow instance in the config applies one 
    function to the input with resulting output.
"""
#endregion p3_execl_budget.p3_banking_transactions budget_transactions.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import re, pathlib as Path, logging, time

# third-party modules and packages
import p3logging as p3l, p3_utils as p3u
from openpyxl import Workbook, load_workbook

# local modules and packages
from .budget_model_constants import  *
from .category_mapping import (
    map_category, category_map_count)
from .budget_model import BudgetModel
# from data.p3_fi_transactions.budget_model import BudgetModel
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(THIS_APP_NAME)

# ---------------------------------------------------------------------------- +
#region check_budget_category() function
def check_budget_category(sheet) -> bool:
    """Check that the sheet is ready to budget category.
    
    A column 'Budget Category' is added to the sheet if it does not exist.

    Args:
        sheet (openpyxl.worksheet): The worksheet to map.
    """
    try:
        me = check_budget_category
        logger.info("Check sheet for budget category.")
        # Is BUDGET_CATEGORY_COL in the sheet?
        if BUDGET_CATEGORY_COL not in sheet.columns:
            # Add the column to the sheet.
            i = sheet.max_column + 1
            sheet.insert_cols(i)
            sheet.cell(row=1, column=i).value = BUDGET_CATEGORY_COL
            # Set the column width to 20.
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 20
            logger.info(f"Adding column '{BUDGET_CATEGORY_COL}' at index = {i}, "
                        f"column_letter = '{sheet.cell(row=1, column=i).column_letter}'")
        else:
            logger.info(f"Column '{BUDGET_CATEGORY_COL}' already exists in sheet.")

        logger.info(f"Completed checks for budget category.")
        return True
    except Exception as e:
        logger.error(p3u.exc_msg(me, e))
        raise    
#endregion check_budget_category() function
# ---------------------------------------------------------------------------- +
#region map_budget_category() function
def map_budget_category(sheet,src,dst) -> None:
    """Map a src column to budget category putting result in dst column.
    
    The sheet has banking transaction data in rows and columns. 
    Column 'src' has the text presumed to be about the transaction.
    Column 'dst' will be assigned a mapped budget category. Append
    column 'Budget Category' if it is not already in the sheet.
    TODO: pass in a mapper function to map the src text to a budget category.

    Args:
        sheet (openpyxl.worksheet): The worksheet to map.
        src (str): The source column to map from.
        dst (str): The destination column to map to. 
    """
    me = map_budget_category
    try:
        rules_count = category_map_count()
        logger.info(f"Applying '{rules_count}' budget category mappings...")
        header_row = [cell.value for cell in sheet[1]] 
        if src in header_row:
            src_col_index = header_row.index(src) + 1
        else:
            logger.error(f"Source column '{src}' not found in header row.")
            return
        if dst in header_row:
            dst_col_index = header_row.index(dst) + 1
        else:
            logger.Error(f"Destination column '{dst}' not found in header row.")
            return

        logger.info(f"Mapping '{src}'({src_col_index}) to "
                    f"'{dst}'({dst_col_index})")
        num_rows = sheet.max_row # or set a smaller limit
        for row_idx, row in enumerate(sheet.iter_rows(min_col=src_col_index, 
                                                      max_col=src_col_index, 
                                                      min_row=2, 
                                                      max_row=num_rows), 
                                                      start=2):
            # Get the value from the source column.
            src_value = row[0].value
            # Map the value to a budget category.
            dst_value = map_category(src_value)
            # Set the value in the destination column.
            sheet.cell(row=row_idx, column=dst_col_index).value = dst_value
            # if dst_value != 'Other':
            logger.debug(f"Row {row_idx:04}: " 
                        f"({len(src_value):03})|{str(src_value):102}| -> " 
                        f"({len(dst_value):03})|{str(dst_value):40}|")

        logger.info(f"Completed budget category mapping for '{num_rows}' rows.")
        return None
    except Exception as e:
        logger.error(p3u.exc_err_msg(me, e))
        raise    
#endregion map_budget_category() function
# ---------------------------------------------------------------------------- +
#region def workfloexecute_worklow_categorizationw_categorization(bm : BudgetModel, fi_key: str) -> None:
def execute_worklow_categorization(bm : BudgetModel, fi_key: str, wf_key:str) -> None:
    """Process categorization wf_key for Financial Institution's 
    transaction workbooks.

    Excecute the categorization wf_key to examine all fi transaction 
    workbooks presently in the IF (Incoming Folder) for the indicated 
    FI (Financial Institution). Each workbook file is opened and the 
    transactions are categorized and saved to the CF (Categorized Folder) 
    for the indicated FI.

    Args:
        bm (BudgetModel): The BudgetModel instance to use for processing.
        fi_key (str): The key for the financial institution.
    """
    # TODO: add logs directory to the budget folder.
    st = p3u.start_timer()
    wb_name = "BOAChecking2025.xlsx"
    cp = "Budget Model Categorization:"
    loaded_workbooks = {}
    # Execute a workflow for a specific financial institution (fi_key).
    #   The pattern is to apply a function based on wf_key to each item in 
    #   the input folder based on fi_key and wf_key.
    #   Execute steps:
    #     1: Load the input items from storage.
    #     2: Apply the workflow function to each input item.
    #     3: Save the output items to storage. 
    try:
        logger.info(f"{cp} Start: workflow: '{wf_key}' for FI('{fi_key}') ...")
        wb_type = WF_WORKBOOKS_IN # input workbooks
        wb_c = bm.bdm_FI_WF_WORKBOOK_LIST_count(fi_key, wf_key, wb_type)
        # workbooks_dict = bm.bsm_WF_WORKBOOKS_IN(fi_key, BM_WF_INTAKE)
        # if workbooks_dict is None or len(workbooks_dict) == 0:
        if wb_c is None or wb_c == 0:
            logger.info(f"{cp}    No workbooks for input.")
            return
        logger.info(f"{cp}    {wb_c} workbooks for input.")
        # Now process each input workbook.
        # for wb_name, wb_ap in reversed(workbooks_dict.items()):
        # Step 1: Load the workbooks sequentially.
        for wb_name, wb in bm.bsm_FI_WF_WORKBOOK_generate(fi_key, wf_key, wb_type):
            logger.info(f"{cp}    Workbook({wb_name})")
                
            # Step 2: Process the workbooks applying the workflow function
            try: 
                logger.info(f"{cp}    Workbook({wb_name})")
                sheet = wb.active
                # Check for budget category column, add it if not present.
                check_budget_category(sheet)
                # Map the 'Original Description' column to the 'Budget Category' column.
                map_budget_category(sheet, "Original Description", BUDGET_CATEGORY_COL)
            except Exception as e:
                logger.error(f"{cp}    Error processing workbook: {wb_name}: {e}")
                continue

            # Step 3: Save the output items to output storage.
            try:
                bm.bsm_FI_WF_WORKBOOK_save(wb, wb_name, fi_key, wf_key, WF_WORKBOOKS_OUT)
            except Exception as e:
                logger.error(f"{cp}    Error saving workbook: {wb_name}: {e}")
                continue
        logger.info(f"{cp} Complete: wf_key: '{wf_key}' {p3u.stop_timer(st)}")
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion execute_worklow_categorization() function
# ---------------------------------------------------------------------------- +