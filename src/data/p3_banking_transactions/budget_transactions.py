# ---------------------------------------------------------------------------- +
#region p3_execl_budget.p3_banking_transactions budget_transactions.py module
""" Provide functions to apply a budget to banking transactions.

    Assumptions:
    - The banking transactions are in a folder specified in the user_config.
    - Banking transaction files are typical excel spreadsheets. 
    - Data content starts in cell A1.
    - Row 1 contains column headers. All subsequent rows are data.
"""
#endregion p3_execl_budget.p3_banking_transactions budget_transactions.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import re, pathlib as Path, logging

# third-party modules and packages
import p3logging as p3l
from openpyxl import Workbook, load_workbook

# local modules and packages
from p3_excel_budget_constants import  *
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(THIS_APP_NAME)

user_config = {
}

# Map values to column ['Category'] by re pattern-matching to 
# column ['Original Description'].
# This list of patterns will be quite long. 
# TODO: How to use data to train an LLM or ML model to do this?
category_mapping = {
    # Donations
    r'(?i)\bCAMPUS\s\bCRUSADE\b': 'Charity.Campus Crusade',
    r'(?i)\bTUNNELTOTOWERS\b': 'Charity.Tunnel to Towers',
    r'(?i)\bFOCUS\sON\sTHE\sFAMILY\b': 'Charity.Focus on the Family',
    # Housing, Utilities etc.
    r'(?i)\bReliant\sEnergy\b': 'Grape Cove.Utilities.Electric',
    r'(?i)\bPedernales_Elec\b': 'Castle Pines.Utilities.Electric',
    r'(?i)\bONE\sGAS\b': 'Grape Cove.Utilities.Natural Gas',
    r'(?i)\bATMOS\sENERGY\b': 'Grape Cove.Utilities.Natural Gas',
    r'(?i)\bService\sExperts\b': 'Grape Cove.Utilities.HVAC',
    r'(?i)\bBrushy\sCreek\sMUD\b': 'Grape Cove.Utilities.MUD',
    r'(?i)\bAT\&T\sU-Verse\b': 'Grape Cove.Utilities.Internet',
    r'(?i)\bGoogle\sFIBER\b': 'Castle Pines.Utilities.Internet',
    r'(?i)\bGoogle\s*\*FIBER\b': 'Castle Pines.Utilities.Internet',
    r'(?i)\bCity\sof\sAustin\b': 'Castle Pines.Utilities.City of Austin',
    r'(?i)\bGOOGLE\s\*FIBER\b': 'Castle Pines.Utilities.Internet',
    r'(?i)\bavery\W*.*?\branch\W*.*?\bHOA\W*.*?\bdues\b': 'Castle Pines.HOA',
    r'(?i)\bFREDERICK\sPEVA\b': 'Housing.Lawn Care',
    r'(?i)\bTRANSFER\sPAUL\sB\sPAINTER\sLAURA:john hogge\b': 'Housing.Mortgage',
    r'(?i)\bTHE\sHOME\sDEPOT\b': 'Housing.Maintenance',
    r"(?i)\bLOWE's\b": 'Housing.Maintenance',
    r'(?i)\bCASHWAY\sBLDG\sMATERIALS\b': 'Housing.Maintenance',
    r'(?i)\bCULLIGAN\b': 'Grape Cove.Culligan',
    r'(?i)\bCULLINGAN\b': 'Grape Cove.Culligan',
    r'(?i)\bRIGHTSPACE\sSTORAGE\b': 'Housing.Storage Unit',
    r'(?i)\bATT\b': 'Telecom.Cellular',
    # Auto
    r'(?i)\bDISCOUNT-TIRE-CO\b': 'Auto.Tires',
    r"(?i)\bO'REILLY\s\bTOLL\b": "Auto.Maintenance.O'Reilly",
    r'(?i)\bCHEVRON\b': 'Auto.Gasoline.Chevron',
    r'(?i)\bGO\s\bCARWASH\b': 'Auto.Carwash.Go Carwash',
    r'(?i)\bHCTRA\sEFT\b': 'Auto.Tolls.HCTRA',
    r'(?i)\bdoxoPLUS\b': 'Auto.Tolls.DoxoPlus',
    # Groceries
    r'(?i)\bDOLLAR\sGENERAL\b': 'Groceries.Dollar General',
    r'(?i)\bCOSTCO\sWHSE\b': 'Groceries.CostCo',
    r'(?i)\bH-E-B\b': 'Groceries.H-E-B',
    # Insurance
    r'(?i)\bHP\s*\bDES:INS\sPREM\b': 'Insurance.Cobra Medical',
    r'(?i)\bState\s\bFARM\b': 'Insurance.State Farm',
    # Medical
    r'(?i)\bGEORGETOWN\sOB-GYN\b': 'Medical.Bio-T',
    r'(?i)\bHAROLD\sF\sADELMAN\sMD\b': 'Medical.Adelman',
    r'(?i)\bBSWHealth.*\b': 'Medical.BSW Health',
    r'(?i)\bJOHN\sF\sLANN\sDDS\b': 'Dental.Lann',
    r'(?i)\bCVS/PHARM\b': 'Medical.Pharmacy.CVS',
    r'(?i)\bCAREMARK\sMAIL\b': 'Medical.Pharmacy.CVS',
    r'(?i)\bWWW\.CAREMARK\.COM\b': 'Medical.Pharmacy.CVS',
    r'(?i)\bCVS.*PHARMACY\b': 'Medical.Pharmacy.CVS',
    r'(?i)\bMINUTECLINIC\b': 'Medical.Pharmacy.CVS Minute Clinic',
    # Health and Fitness, Coaching
    r'(?i)\bFINLEYS\sAVERY\sRANCH\b': 'Health and Fitness.Haircut Paul',
    r'(?i)\bCATHY\s\bDUNFORD\b': 'Health and Fitness.Coach Cathy',
    r'(?i)\bSNICOLA\s\bMCKERLIE\b': 'Health and Fitness.Nicola McKerlie',
    # Streaming and Subscriptions
    r'(?i)\bCINEMARK\b': 'Theater.Cinemark',
    r'(?i)\bAudible\*.*\b': 'Subscription.Audible',
    r'(?i)\bID:ANGIES\sLIST\b': 'Subscription.Angies List',
    r'(?i)\bID:GITHUB\sINC\b': 'Subscription.GitHub',
    r'(?i)\bID:HULU\b': 'Subscription.Hulu',
    r'(?i)\bID:PATREON\sMEMBERNETFLIX\b': 'Subscription.Patreon',
    r'(?i)\bPrime\s\bVideo\b': 'Subscription.Amazon Prime Video',
    r'(?i)\bTHIRTEEN\b': 'Subscription.Thirteen',
    r'(?i)\bID:NETFLIX\.COM\b': 'Subscription.Netflix',
    r'(?i)\bID:DROPBOX\b': 'Subscription.DropBox',
    # Restaurants, Door Dash, Eating Out.
    r"(?i)\b183\sPHIL's\b": "Restaurants.183 Phil's",
    r'(?i)\s\*SMOKEY\sMOS\sBBQ\b': "Restaurants.Smokey Mo's BBQ",
    r'(?i)\bTST\*LA\sMARGARITA\b': 'Restaurants.La Margarita',
    r'(?i)\bSURF\sAND\sTURF\b': 'Restaurants.Surf and Turf',
    r'(?i)\bCASA\sOLE\b': 'Restaurants.Casa Ole',
    r'(?i)\bID:DOORDASHINC\b': 'Restaurants.Door Dash',
    r'(?i)\s\*DOORDASH\b': 'Restaurants.Door Dash',
    r'(?i)\bDAIRY\sQUEEN\b': 'Restaurants.Dairy Queen',
    r'(?i)\bCHICK-FIL-A\b': 'Restaurants.Chick-Fil-A',
    r'(?i)\bMOD\sPIZZA\b': 'Restaurants.Mod Pizza',
    r"(?i)\bMCDONALD'S\b": "Restaurants.McDonald's",
    r"(?i)\bMANDOLAS\b": "Restaurants.McDonald's",
    r"(?i)\bTONY\sC'S\sCOAL\sFIRED\b": "Restaurants.Tony C's Coal Fired",
    # Shopping - Amazon, Apple, etc.
    r'(?i)\bMICROSOFT\b': 'Shopping.Microsoft',    
    r'(?i)\bETSY,\sINC\.': 'Shopping.Etsy',
    r'(?i)\bDIGISTORE\d\d\sINC\.': 'Shopping.DIGISTORE Unknown',
    r'(?i)\bBARNES\s&\sNOBLE\b': 'Shopping.Barnes & Noble',
    r'(?i)\bBASS\sPRO\sSTORE\b': 'Shopping.Bass Pro',
    r'(?i)\bAT\sHOME\b': 'Shopping.At Home',
    r'(?i)\bAMAZON\s.*?\bPRIME\b': 'Shopping.Amazon Prime',
    r'(?i)\bAMAZON\sRETA\*\s\b': 'Shopping.Amazon Retail',
    r'(?i)\bAMAZON\s\bMKTPL\*.*\b': 'Shopping.Amazon Marketplace',
    r'(?i)\bAMAZON\.com\*.*\b': 'Shopping.Amazon',
    r'(?i)\bAMAZON\s(MARKETPLA\s|MKTPLACE\s)\b': 'Shopping.Amazon Marketplace',
    r'(?i)\bAMAZON\s\b(DIGITAL|DIGI).*?\b(LINKEDIN|LINKWA)\b': 'Shopping.Amazon Digital',
    r'(?i)\b.*APPLE\.COM.*\b': 'Shopping.Apple',
    r'(?i)\bCIRCLE\sK\b': 'Shopping.Circle K',
    r'(?i)\b7-ELEVEN\s.*?MOBILE\sPURCHASE\b': 'vape',
    r'(?i)\b7-ELEVEN\b': 'Shoppint.7-Eleven',
    r'(?i)\bMICHAELS\sSTORES\b': 'Shopping.Michaels',
    # Pets
    r'(?i)\bGREAT\sOAKS\sANIMAL\b': 'Pets.Veternary',
    r'(?i)\bID:CHEWY\sINC\b': 'Pets.Dog Food',
    # Work-related Expenses
    r'(?i)\bEXECUTIVE\sCAREER\sUPGRA\b': 'Work-related.ECU Recruiting',
    r'(?i)\bOTTER\.AI\b': 'Work-related.OTTER-AI',
    r'(?i)\bID:LINKEDIN\b': 'Work-related.Linked In',
    # Income
    r'(?i)\bInterest\sEarned\b': 'Income.Interest',
    r'(?i)\bGerson\sLehrman\sG\b': 'Income.Consulting.GL Group',
    r'(?i)\bHP\sINC*?\bPAYROLL\b': 'Income.HP Inc.',
    # Banking, Finance and Taxes
    r'(?i)\bPMT\*WILCO\sTAX\b': 'Taxes.Williamson County',
    r'(?i)\bLATE\sFEE\sFOR\sPAYMENT\sDUE\b': 'Banking.Late Fee',
    r'(?i)\bINTEREST\sCHARGED\sON\sPURCHASES\b': 'Banking.Interest',
    r'(?i)\bFRAUD\sDISPUTE\b': 'Banking.Fraud Dispute',
    r'(?i)\bBG\sREVERS(AL|ED)\b': 'Banking.Fraud Disput',
    r'(?i)\bFOREIGN\sTRANSACTION\sFEE\b': 'Banking.Foreign Transaction Fee',
    r'(?i)\bExperian\*\sCredit\sReport\b': 'Finance.Experian',
    r'(?i)\bBKOFAMERICA\sATM.*\bWITHDRWL\b': 'Banking.ATM',
    r'(?i)\bBKOFAMERICA\sMOBILE.*\bDEPOSIT\b': 'Banking.Mobile Deposit',
    # Credit Card Payments
    r'(?i)\bOnline\spayment\sfrom\sCHK\s1391\b': 'Credit Cards.Bank of America',
    r'(?i)\bPayment\s-\sTHANK\sYOU\b': 'Credit Cards.Bank of America',
    r'(?i)\bVisa\sBank\sOf\sAmerica\sBill\sPayment\b': 'Credit Cards.Bank of America',
    r'(?i)\bOnline\sBanking\sTRANSFER\sTO\sCRD\b': 'Credit Cards.Bank of America',
    r'(?i)\bCHASE\sCREDIT\sCRD\b': 'Credit Cards.Chase',
    r'(?i)\bBANK\sOF\sAMERICA\sCREDIT\sCARD\b': 'Credit Cards.Band of America',
    # Transfers
    r'(?i)\bForis\sUSA\sINC\sCF\b': 'Investment.Foris USA',
    r'(?i)\bFID\sBKG\sSVC\sLLC\b': 'Investment.Fidelity',
    r'(?i)\bAgent\sAssisted\stransfer\sfrom\b': 'Investment.Transfer',
    r'(?i)\bMobile\sTransfer\sfrom\sCHK\b': 'Investment.Transfer',    
    r'(?i)\bOnline\sTRANSFER\sTO\sCHK\b': 'Investment.Transfer',
    r'(?i)\bOnline\sBanking\sTRANSFER\sTO\sSAV\b': 'Investment.Transfer.ToSavings',
    r'(?i)\bOnline\sBanking\sTRANSFER\sTO\sINV\b': 'Investment.Transfer.ToInvestments',
    # Travel, trips, one-offs
    r'(?i)\bEXPEDIA\b': 'Travel.Expedia',
    r'(?i)\bATGPay\sonline\b': 'Unknown.ATGPay',
    r'(?i)\bGITKRAKEN\sSOFTWARE\b': 'Crypto.Gitkraken',
    r'(?i)\bGUNNISON\s*CO\b': 'Travel.ColoradoTrip',
    r'(?i)\bIRVING\s*TX\b': 'Travel.ColoradoTrip',
    r'(?i)\bEARLY\s*TX\b': 'Travel.ColoradoTrip',
    r'(?i)\bSAGUACHE\s*CO\b': 'Travel.ColoradoTrip',
    r'(?i)\bSNYDER\s*TX\b': 'Travel.ColoradoTrip',
    r'(?i)\bRATON\s*NM\b': 'Travel.ColoradoTrip',
    r'(?i)\bDALLAS\s*TX\b': 'Travel.ColoradoTrip',
    r'(?i)\bPOST\s*TX\b': 'Travel.ColoradoTrip',
    r'(?i)\bCAMPO\s*TX\b': 'Travel.ColoradoTrip',
    r'(?i)\bCAMPO\s*CO\b': 'Travel.ColoradoTrip',
    r'(?i)\bCHEYENNE\s*WELLCO\b': 'Travel.ColoradoTrip',
    r'(?i)\bDUMAS\s*TX\b': 'Travel.ColoradoTrip',
}

#endregion Globals and Constants
# ---------------------------------------------------------------------------- +
#region map_category() function
def map_category(description):
    try:
        for pattern, category in category_mapping.items():
            if re.search(pattern, str(description), re.IGNORECASE):
                return category
        return 'Other'  # Default category if no match is found
    except re.PatternError as e:
        logger.error(p3l.exc_msg(map_category, e))
        logger.error(f'Pattern error: category_mapping dict: ' 
                     f'{{ \"{e.pattern}\": \"{category}\" }}')
        raise
    except Exception as e:
        logger.error(p3l.exc_msg(map_category, e))
        raise
#endregion map_category() function
# ---------------------------------------------------------------------------- +
#region check_budget_category() function
def check_budget_category(sheet) -> bool:
    """Check that the sheet is ready to budget category.
    
    A column 'Budget Category' is added to the sheet if it does not exist.

    Args:
        sheet (openpyxl.worksheet): The worksheet to map.
    """
    try:
        me = check_budget_category
        logger.info("Check sheet for budget category.")
        # Is BUDGET_CATEGORY_COL in the sheet?
        if BUDGET_CATEGORY_COL not in sheet.columns:
            # Add the column to the sheet.
            i = sheet.max_column + 1
            sheet.insert_cols(i)
            sheet.cell(row=1, column=i).value = BUDGET_CATEGORY_COL
            # Set the column width to 20.
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 20
            logger.info(f"Adding column '{BUDGET_CATEGORY_COL}' at index = {i}, "
                        f"column_letter = '{sheet.cell(row=1, column=i).column_letter}'")
        else:
            logger.info(f"Column '{BUDGET_CATEGORY_COL}' already exists in sheet.")

        logger.info(f"Completed checks for budget category.")
        return True
    except Exception as e:
        logger.error(p3l.exc_msg(me, e))
        raise    
#endregion check_budget_category() function
# ---------------------------------------------------------------------------- +
#region load_banking_transactions() function
def map_budget_category(sheet,src,dst) -> None:
    """Map a src column to budget category putting result in dst column.
    
    The sheet has banking transaction data in rows and columns. 
    Column 'src' has the text presumed to be about the transaction.
    Column 'dst' will be assigned a mapped budget category. Append
    column 'Budget Category' if it is not already in the sheet.
    TODO: pass in a mapper function to map the src text to a budget category.

    Args:
        sheet (openpyxl.worksheet): The worksheet to map.
        src (str): The source column to map from.
        dst (str): The destination column to map to. 
    """
    me = map_budget_category
    try:
        logger.info("Applying budget category mapping...")
        header_row = [cell.value for cell in sheet[1]] 
        if src in header_row:
            src_col_index = header_row.index(src) + 1
        else:
            logger.Error(f"Source column '{src}' not found in header row.")
            return
        if dst in header_row:
            dst_col_index = header_row.index(dst) + 1
        else:
            logger.Error(f"Destination column '{dst}' not found in header row.")
            return

        logger.info(f"Mapping '{src}'({src_col_index}) to "
                    f"'{dst}'({dst_col_index})")
        num_rows = sheet.max_row # or set a smaller limit
        for row_idx, row in enumerate(sheet.iter_rows(min_col=src_col_index, 
                                                      max_col=src_col_index, 
                                                      min_row=2, 
                                                      max_row=num_rows), 
                                                      start=2):
            # Get the value from the source column.
            src_value = row[0].value
            # Map the value to a budget category.
            dst_value = map_category(src_value)
            # Set the value in the destination column.
            sheet.cell(row=row_idx, column=dst_col_index).value = dst_value
            # if dst_value != 'Other':
            logger.info(f"Row {row_idx:04}: " 
                        f"({len(src_value):03})|{str(src_value):102}| -> " 
                        f"({len(dst_value):03})|{str(dst_value):40}|")

        logger.info(f"Completed budget category mapping")
        return None
    except Exception as e:
        logger.error(p3l.exc_msg(me, e))
        raise    
#endregion load_banking_transactions() function
# ---------------------------------------------------------------------------- +
