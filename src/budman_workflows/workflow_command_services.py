# ---------------------------------------------------------------------------- +
#region workflow_command_services.py module
""" workflow_command_services.py implements functions for BudMan app
workflow-related commands.

In general, services should return either data objects or command result objects.
Leave it to the caller, such as a View, or ViewMModel to handle additional
services, such as a pipeline, or to perform output functions.
"""
#endregion workflow_command_services.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import logging, shutil
from pathlib import Path
from typing import List, Type, Optional, Dict, Tuple, Any, Callable
from treelib import Tree
from datetime import datetime as dt

# third-party modules and packages
import p3_utils as p3u, pyjson5, p3logging as p3l, p3_mvvm as p3m
from openpyxl import Workbook, load_workbook

# local modules and packages
import budman_command_processor as cp
import budman_namespace as bdm
from budman_namespace import P2, P4
from budman_namespace import BDMWorkbook
from budget_domain_model import BudgetDomainModel
from budman_data_context import BudManAppDataContext_Base
from budget_storage_model import (BSMFile, BSMFileTree,
                                  csv_DATA_LIST_url_get) 
from .budget_intake import *
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_process() function
def WORKFLOW_CMD_process(cmd: p3m.CMD_OBJECT_TYPE, 
                       bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Process workflow tasks.

    This function processes various worklow tasks as requested in the validated
    CMD_OBJECT.

    Args:
        cmd (CMD_OBJECT_TYPE): 
            A validated CommandProcessor CMD_OBJECT_TYPE. Contains
            the command attributes and parameters to execute.
        bdm_DC (BudManAppDataContext_Base): 
            The data context for the BudMan application.
    """
    try:
        # If bdm_DC is bad, just raise an error.
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        # Should be called only for workflow cmd.
        cmd_result : p3m.CMD_RESULT_TYPE = cp.verify_cmd_key(cmd, cp.CV_WORKFLOW_CMD_KEY)
        if not cmd_result[p3m.CMD_RESULT_STATUS]: return cmd_result
        # Assuming the cmd parameters have been validated before reaching this point.
        # Process the CMD_OBJECT based on its CK_SUBCMD_KEY.
        # workflow transfer subcmd
        if cmd[p3m.CK_SUBCMD_KEY] == cp.CV_WORKFLOW_TRANSFER_SUBCMD_KEY:
            transfer_files: bool = cmd.get(cp.CK_TRANSFER_FILES, False)
            if transfer_files:
                return WORKFLOW_TASK_transfer_files(cmd, bdm_DC)
            transfer_workbooks: bool = cmd.get(cp.CK_TRANSFER_WORKBOOKS, False)
            if transfer_workbooks:
                return WORKFLOW_TASK_transfer_workbooks(cmd, bdm_DC)
            return p3m.unknown_CMD_RESULT_ERROR(cmd)
        # workflow intake subcmd
        elif cmd[p3m.CK_SUBCMD_KEY] == cp.CV_WORKFLOW_INTAKE_SUBCMD_KEY:
            return INTAKE_TASK_process(cmd, bdm_DC)
        # workflow set subcmd
        elif cmd[p3m.CK_SUBCMD_KEY] == cp.CV_SET_SUBCMD_KEY:
            # Process the set_value task.
            return WORKFLOW_TASK_set_value(cmd, bdm_DC)
        # workflow task sync subcmd
        elif (cmd[p3m.CK_SUBCMD_KEY] == cp.CV_TASK_SUBCMD_KEY and
              cmd[cp.CK_TASK_NAME] == cp.CV_SYNC):
            # Process the wf sync task.
            recon: bool = cmd.get(cp.CK_RECONCILE, False)
            return WORKFLOW_TASK_sync_wdc(recon, bdm_DC)
        # workflow unknown subcmd
        else:
            return p3m.unknown_CMD_RESULT_ERROR(cmd)
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_CMD_process() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_transfer_workbooks() function
def WORKFLOW_TASK_transfer_workbooks(cmd: p3m.CMD_OBJECT_TYPE, 
                    bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_TRANSFER_subcmd: Transfer data workbooks between workflows.

    Tasks required to transfer workbooks to a workflow for a specified purpose.
    Processing requirements vary based on the specific workflow, purpose and 
    workbook types.

    Arguments:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            cp.CK_WB_LIST - list of src wb_index values from the DC workbook_data_collection
            cp.CK_WF_KEY - wf_folder wf_key
            cp.CK_WF_PURPOSE - wf_folder wf_purpose
            cp.CK_WB_TYPE - dst workflow workbook type

    Workbooks are files are in wf_folders. A list of wb_index values references 
    workbooks to transfer from a source (src) workflow folder to a destination (dst) 
    wf_folder. Depending on the dst wf_key and wf_purpose, naming conventions 
    applied to the workbooks being transferred. The dst wb_type also impacts the 
    dst workbook naming and possible conversion, or transformation tasks 
    applied during the transfer operation.

    At present, the workflow processes are fixed: intake, categorize_transactions,
    and budget. Each workflow may have up to three wf_folders configured for
    the wf_purpose values of: wf_input, wf_working, and wf_output. Also, a 
    wf_prefix value may be applied to workbooks arriving anew in a wf_folder.

    Hence, a dst full_file_name could have the following structure:
        <wf_prefix><filename><wb_type>.<extension>

    The src_filename is parsed for the four components, wf_prefix may be None and
    wb_type may be None for a file that is not yet a workbook.

    The dst_filename is constructed considering the dst_wf_prefix and dst_wb_type.
    The extension is determined by the wb_type.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.
    """
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cp.validate_cmd_arguments(
            cmd=cmd, 
            bdm_DC=bdm_DC,
            cmd_key=cp.CV_WORKFLOW_CMD_KEY, 
            subcmd_key=cp.CV_WORKFLOW_TRANSFER_SUBCMD_KEY,
            model_aware=True,
            required_args=[
                cp.CK_TRANSFER_WORKBOOKS,
                cp.CK_WB_LIST,
                cp.CK_WF_KEY,
                cp.CK_WF_PURPOSE,
                cp.CK_WB_TYPE
            ]
        )
        logger.info(f"Start: ...")
        # Extract and validate required parameters from the command.
        model: BudgetDomainModel = bdm_DC.model
        valid_prefixes: List[str] = model.bdm_valid_WF_PREFIX_values()
        dst_wf_key = cmd_args.get(cp.CK_WF_KEY)
        dst_wf_purpose = cmd_args.get(cp.CK_WF_PURPOSE)
        dst_wb_type = cmd_args.get(cp.CK_WB_TYPE)
        fi_key: str = bdm_DC.dc_FI_KEY
        fr: str = ""  # final report string
        m: str = ""
        success: bool = False
        result: Any = None
        selected_bdm_wb_list : List[BDMWorkbook] = None
        selected_bdm_wb_list = process_selected_workbook_input(cmd, bdm_DC)
        log_all : bool = cmd.get(cp.CK_LOG_ALL, False)
        # Process the intended workbooks.
        src_wb: BDMWorkbook = None
        dst_wb: BDMWorkbook = None
        src_wb_type: str = ''
        for src_wb in selected_bdm_wb_list:
            # Select the current workbook in the Data Context.
            bdm_DC.dc_WORKBOOK = src_wb
            wb_index: int = bdm_DC.dc_WB_INDEX
            bdm_wb_abs_path = src_wb.abs_path()
            fr += f"\n{P2}workbook: {str(bdm_DC.dc_WB_INDEX):>4} '{bdm_DC.dc_WB_ID:<40}'"
            bdm_wb_abs_path = src_wb.abs_path()
            if bdm_wb_abs_path is None:
                m = f"Workbook path is not valid: {src_wb.wb_url}"
                logger.error(m)
                fr += f"\n{P4}Error: {m}"
                continue
            # Transfer cmd needs loaded workbooks
            if not src_wb.wb_loaded:
                # Load the workbook content if it is not loaded.
                success, result = bdm_DC.dc_WORKBOOK_content_get(src_wb)
                if not success:
                    selected_bdm_wb_list.remove(src_wb)
                    m = f"Excluded workbook: '{src_wb.wb_id}', "
                    m += f"failed to load: {result}"
                    logger.error(m)
                    continue
            src_wb_type = src_wb.wb_type
            # Supportted transfer cases.
            if src_wb_type == bdm.WB_TYPE_CSV_TXNS and dst_wb_type == bdm.WB_TYPE_EXCEL_TXNS:
                # Transfer a csv_txns workbook by converting to a excel_txns workbook.
                # Create a file_url for the new workbook being transferred.
                src_wb_bsm_file: BSMFile = BSMFile(
                    file_url=src_wb.wb_url,
                    valid_prefixes=valid_prefixes,
                    valid_wb_types=bdm.VALID_WB_TYPE_VALUES
                )
                success, result = WORKFLOW_TASK_construct_bdm_workbook(
                    src_filename=src_wb_bsm_file.filename,
                    wb_type=dst_wb_type,
                    fi_key=fi_key,
                    wf_key=dst_wf_key,
                    wf_purpose=dst_wf_purpose,
                    bdm_DC=bdm_DC
                )
                if not success:
                    msg = (f"Failed to construct file URL for workbook: "
                            f"'{wb_index:2}:{src_wb.wb_name}' "
                            f" Error: {result}")
                    logger.error(msg)
                    result_content += msg + "\n"
                    continue
                dst_wb = result
                WORKFLOW_TASK_convert_csv_txns_to_excel_txns(src_wb, dst_wb)
                # Add the new workbook to the wdc.
                wdc = bdm_DC.dc_WORKBOOK_DATA_COLLECTION
                wb_id = dst_wb.wb_id
                wdc[wb_id] = dst_wb
            elif src_wb_type == bdm.WB_TYPE_EXCEL_TXNS and dst_wb_type == bdm.WB_TYPE_EXCEL_TXNS:
                # Transfer a excel_txns workbook to another workflow wf_folder.
                # Create a file_url for the new workbook file being transferred.
                src_wb_bsm_file: BSMFile = BSMFile(
                    file_url=src_wb.wb_url,
                    valid_prefixes=valid_prefixes,
                    valid_wb_types=bdm.VALID_WB_TYPE_VALUES
                )
                # Create a BDMWorkbook for the dst workbook.
                success, result = WORKFLOW_TASK_construct_bdm_workbook(
                    src_filename=src_wb_bsm_file.filename,
                    wb_type=dst_wb_type,
                    fi_key=fi_key,
                    wf_key=dst_wf_key,
                    wf_purpose=dst_wf_purpose,
                    bdm_DC=bdm_DC
                )
                if not success:
                    msg = (f"Failed to create workbook: "
                            f"'{wb_index:2}:{src_wb.wb_name}' "
                            f" Error: {result}")
                    logger.error(msg)
                    result_content += msg + "\n"
                    continue
                dst_wb = result
                WORKFLOW_TASK_copy_workbook(src_wb, dst_wb)
                # Add the new workbook to the wdc.
                wdc = bdm_DC.dc_WORKBOOK_DATA_COLLECTION
                wb_id = dst_wb.wb_id
                wdc[wb_id] = dst_wb
            else:
                m = (f"Unsupported transfer from src wb_type '{src_wb_type}' "
                     f"to dst wb_type '{dst_wb_type}'")
                logger.error(m)
                fr += f"\n{P4}Error: {m}"
                continue
        return p3m.create_CMD_RESULT_OBJECT(
            cmd_object=cmd,
            cmd_result_status=True,
            result_content_type=p3m.CMD_DICT_OUTPUT,
            result_content="all done"
        )
    except p3m.CMDValidationException as e:
        logger.error(e.message)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_transfer: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.create_CMD_RESULT_ERROR(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
#endregion WORKFLOW_TASK_transfer_workbooks() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_transfer_files() function
def WORKFLOW_TASK_transfer_files(cmd: p3m.CMD_OBJECT_TYPE, 
                    bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_TRANSFER_subcmd: Transfer data files into workflow workbooks.

    Tasks required to transfer files to a workflow for a specified purpose.
    This is how files from a file_list are transformed into workbooks in the
    model. Processing requirements vary based on the specific workflow, 
    purpose and workbook types.

    Arguments:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            cp.CK_FILE_LIST - list of src file_index values from the file_list
            cp.CK_WF_KEY - wf_folder wf_key of destination.
            cp.CK_WF_PURPOSE - wf_folder wf_purpose of destination.
            cp.CK_WB_TYPE - workbook type of destination.

    Files are in wf_folders. A list of file_index values references files to 
    transfer from a source (src) workflow folder to a destination (dst) 
    wf_folder. Depending on the dst wf_key and wf_purpose, naming conventions 
    applied to the files being transferred. The dst wb_type also impacts the 
    dst filename and possible conversion, or transformation tasks applied during
    the transfer operation.

    At present, the workflow processes are fixed: intake, categorize_transactions,
    and budget. Each workflow may have up to three wf_folders configured for
    the wf_purpose values of: wf_input, wf_working, and wf_output. Also, a 
    wf_prefix value may be applied to files arriving anew in a wf_folder.

    Hence, a dst full_file_name could have the following structure:
        <wf_prefix><filename><wb_type>.<extension>

    The src_filename is parsed for the four components, wf_prefix may be None and
    wb_type may be None for a file that is not yet a workbook.

    The dst_filename is constructed considering the dst_wf_prefix and dst_wb_type.
    The extension is determined by the wb_type.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.
    """
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cp.validate_cmd_arguments(
            cmd=cmd, 
            bdm_DC=bdm_DC,
            cmd_key=cp.CV_WORKFLOW_CMD_KEY, 
            subcmd_key=cp.CV_WORKFLOW_TRANSFER_SUBCMD_KEY,
            model_aware=True,
            required_args=[
                cp.CK_TRANSFER_FILES,
                cp.CK_FILE_LIST,
                cp.CK_WF_KEY,
                cp.CK_WF_PURPOSE,
                cp.CK_WB_TYPE
            ]
        )
        model: BudgetDomainModel = bdm_DC.model
        bsm_file_tree : BSMFileTree = model.bsm_file_tree
        # Extract and validate required parameters from the command.
        wf_key = cmd_args.get(cp.CK_WF_KEY)
        wf_purpose = cmd_args.get(cp.CK_WF_PURPOSE)
        src_file_list = cmd_args.get(cp.CK_FILE_LIST)
        wb_type = cmd_args.get(cp.CK_WB_TYPE)
        fi_key: str = bdm_DC.dc_FI_KEY
        bsm_files: List[cp.BSMFile] = []
        bsm_files = bsm_file_tree.validate_file_list(src_file_list)
        result_content: str = ""
        msg: str = ""
        # Supported cases:
        # 1. transfer .csv files from file_list to a .csv_txns workbooks
        success: bool = False
        result: str = ""
        bsm_file: BSMFile = None
        cvs_wb: BDMWorkbook = None
        for bsm_file in bsm_files:
            # Process for supported transfer dst wb_types.
            if wb_type == bdm.WB_TYPE_CSV_TXNS:
                # Transfer a csv file to a csv_txns workbook.
                # Input file must have .csv extension.
                if bsm_file.extension == bdm.WB_FILETYPE_CSV:
                    # Transfer a .csv file to a .csv_txns workbook.
                    # Create a file_url for the new file being transferred.
                    success, result = WORKFLOW_TASK_construct_bdm_workbook(
                        src_filename=bsm_file.filename,
                        wb_type=wb_type,
                        fi_key=fi_key,
                        wf_key=wf_key,
                        wf_purpose=wf_purpose,
                        bdm_DC=bdm_DC
                    )
                    if not success:
                        msg = (f"Failed to construct file URL for file: "
                               f"'{bsm_file.file_index:2}:{bsm_file.full_filename}'"
                               f" Error: {result}")
                        logger.error(msg)
                        result_content += msg + "\n"
                        continue
                    csv_wb = result
                    success, result = WORKFLOW_TASK_transfer_csv_file_to_workbook(
                        src_file_url=bsm_file.file_url,
                        dst_wb=csv_wb,
                        wb_type=wb_type
                    )
                    if not success:
                        msg = (f"Failed to transfer file: "
                               f"'{bsm_file.file_index:2}:{bsm_file.full_filename}' "
                               f" to .csv_txns workbook. Error: {result}")
                        logger.error(msg)
                        result_content += msg + "\n"
                        continue
                else:
                    # Unsupported file type for transfer.
                    msg = (f"Unsupported source file type file "
                            f"'{bsm_file.file_index:2}:{bsm_file.full_filename}'"
                            f" must be .csv file.")
                    logger.error(msg)
                    result_content += msg + "\n"
                    continue
                # Add the new workbook to the wdc.
                wdc = bdm_DC.dc_WORKBOOK_DATA_COLLECTION
                wb_id = csv_wb.wb_id
                wdc[wb_id] = csv_wb
            else:
                # Unsupported dst wb_type for transfer.
                msg = (f"Unsupported destination workbook type for transfer: {wb_type}")
                logger.error(msg)
                result_content += msg + "\n"
                continue
        return p3m.create_CMD_RESULT_OBJECT(
            cmd_object=cmd,
            cmd_result_status=True,
            result_content_type=p3m.CMD_DICT_OUTPUT,
            result_content="all done"
        )
    except p3m.CMDValidationException as e:
        logger.error(e.message)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_transfer: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.create_CMD_RESULT_ERROR(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
#endregion WORKFLOW_TASK_transfer_files() function
# ---------------------------------------------------------------------------- +
#region process_txn_intake() function
def WORKFLOW_TASK_convert_csv_txns_to_excel_txns(csv_wb: BDMWorkbook, 
                                                 excel_wb: BDMWorkbook ) -> bdm.BUDMAN_RESULT_TYPE:
    """Convert CSV transactions to Excel transactions.

        Args:
        csv_wb (BDMWorkbook): The source CSV workbook.
        excel_wb (BDMWorkbook): The destination Excel workbook.
    """
    try:
        st = p3u.start_timer()
        fr: str = "WORKFLOW_TASK_convert_csv_txns_to_excel_txns:"
        logger.debug(f"{fr} ")
        csv_txns = csv_wb.wb_content
        headers = list(csv_txns[0].keys())  # Get headers from the first row
        
        excel_wb.wb_content = Workbook() # Create a new Excel workbook.
        ws: Worksheet = excel_wb.wb_content.active
        #TODO: get the worksheet title from settings or config.
        ws.title = "TransactionData" # Set the name of the first worksheet.

        ws.append(headers)  # Write headers to the first row
        # Need to convert date strings to datetime objects, and
        # convert amount strings to float.
        for row in csv_txns:
            for key, value in row.items():
                if key.lower() == "date":
                    value = datetime.datetime.strptime(value, "%m/%d/%Y").date()
                elif key.lower() == "amount":
                    cleaned = re.sub(r'[^\d.-]', '', value)  # Remove non-numeric characters
                    value = float(cleaned)
                row[key] = value
            ws.append(list(row.values()))
        # Task 3: Save the excel_txns workbook
        excel_wb.wb_content.save(excel_wb.abs_path())
        logger.debug(f"Saved excel_txns to {excel_wb.abs_path() }")
        return True, f"Converted CSV to Excel in {p3u.stop_timer(st)} "
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
                    # row[key].number_format = '#,##0.00'  # Set currency format in Excel
                    # row[key].number_format = 'DD/MM/YYYY'  # Set date format in Excel
#endregion process_txn_intake() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_transfer_csv_file()
def WORKFLOW_TASK_transfer_csv_file_to_workbook(src_file_url: str,
                                    dst_wb: BDMWorkbook,
                                    wb_type: str) -> bdm.BUDMAN_RESULT_TYPE:
    """Transfer a CSV file to the specified wb_type.

    Args:
        src_file_url (str): The source file URL.
        dst_wb (BDMWorkbook): The destination workbook object.
        wb_type (str): The desired workbook type for the destination workbook.

    Returns:
        bdm.BUDMAN_RESULT_TYPE: The result of the transfer operation.
    """
    try:
        if wb_type == bdm.WB_TYPE_CSV_TXNS:
            # bsm_file must be a .csv file. Transfer to .csv_txns workbook
            # is just a copy to the new file_url.
            src: Path = Path.from_uri(src_file_url)
            dst: Path = Path.from_uri(dst_wb.wb_url)
            shutil.copyfile(src, dst)
            return True, f"Transferred .csv to .csv_txns workbook: {dst_wb.wb_url}"
        else:
            err_msg = f"Unsupported wb_type for CSV transfer: {wb_type}"
            logger.error(err_msg)
            return False, err_msg
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception transfer_csv_file_to_workbook(): {m}")
        logger.error(err_msg)
        return False, err_msg
#endregion WORKFLOW_TASK_transfer_csv_file()
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_copy_workbook()
def WORKFLOW_TASK_copy_workbook(src_wb: BDMWorkbook,
                                dst_wb: BDMWorkbook) -> bdm.BUDMAN_RESULT_TYPE:
    """Copy a workbook to a new wf_folder location.

    Args:
        src_wb (BDMWorkbook): The source workbook object.
        dst_wb (BDMWorkbook): The destination workbook object.

    Returns:
        bdm.BUDMAN_RESULT_TYPE: The result of the transfer operation.
    """
    try:
        p3u.is_not_obj_of_type("src_wb", src_wb, BDMWorkbook, raise_error=True)
        p3u.is_not_obj_of_type("dst_wb", dst_wb, BDMWorkbook, raise_error=True)
        src: Path = Path.from_uri(src_wb.wb_url)
        dst: Path = Path.from_uri(dst_wb.wb_url)
        shutil.copyfile(src, dst)
        return True, f"Copied workbook from {src_wb.wb_url} to {dst_wb.wb_url}"
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception WORKFLOW_TASK_copy_workbook(): {m}")
        logger.error(err_msg)
        return False, err_msg
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception transfer_csv_file_to_workbook(): {m}")
        logger.error(err_msg)
        return False, err_msg
#endregion WORKFLOW_TASK_transfer_csv_file()
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_set_value() function
def WORKFLOW_TASK_set_value(cmd: Dict[str, Any], 
                       bdm_DC: BudManAppDataContext_Base) -> bdm.BUDMAN_RESULT_TYPE:
    """WORKFLOW_SET_subcmd: Set values in the DC for workflow tasks.

    This function sets values in the data context for workflow tasks based on the
    provided command object.

    Args:
        cmd (Dict[str, Any]): A valid BudMan View Model Command object.
        bdm_DC (BudManAppDataContext_Base): The data context for the 
            BudMan application.
        Associate cmd object keys for the set subcmd:
            cp.CK_CMDLINE_WF_KEY
            cp.CK_CMDLINE_WF_PURPOSE
    """
    try:
        # Assuming the cmd parameters have been validated before reaching this point.
        wf_key: str = cmd.get(cp.CK_CMDLINE_WF_KEY, None)
        wf_purpose: str = cmd.get(cp.CK_CMDLINE_WF_PURPOSE, None)
        msg: str = ""
        if wf_key:
            bdm_DC.dc_WF_KEY = wf_key
            msg += f"Workflow key set to: '{wf_key}'"
        sep = ", " if wf_key else ""
        if wf_purpose:
            bdm_DC.dc_WF_PURPOSE = wf_purpose
            msg += f"{sep}Workflow purpose set to: '{wf_purpose}' "
        return True, msg
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_TASK_set_value() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_sync_wdc() function
def WORKFLOW_TASK_sync_wdc(reconcile:bool, 
                           bdm_DC: BudManAppDataContext_Base) -> bdm.BUDMAN_RESULT_TYPE:
    """Model-Aware: Sync the WORKBOOK_DATA_COLLECTION for the current FI_KEY.

    For the current DC FI_KEY, synchronize the WORKBOOK_DATA_COLLECTION with 
    files in storage. Use the BSM functions.

    Args:
        bdm_DC (BudManAppDataContext_Base): The data context for the 
            BudMan application.

    Returns:
        treelib.Tree: The folder tree for the specified workflow.
    """
    try:
        st = p3u.start_timer()
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        fi_key: str = bdm_DC.dc_FI_KEY
        if not fi_key:
            m = "No FI_KEY set in the DC."
            logger.error(m)
            return False, m
        model:BudgetDomainModel = bdm_DC.model
        if not model:
            m = "No BudgetDomainModel binding in the DC."
            logger.error(m)
            return False, m
        task_name: str = "sync_wdc()"
        msg: str = f"DEPRECATED: Syncing WORKBOOK_DATA_COLLECTION for FI_KEY: '{fi_key}'"
        logger.debug(f"Start Task: {task_name} {msg}")
        r_msg: str = ""
        # discovered_wdc: bdm.WORKBOOK_DATA_COLLECTION_TYPE = None
        # discovered_wdc, r_msg = model.bsm_FI_WORKBOOK_DATA_COLLECTION_resolve(fi_key)
        return True, msg
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_TASK_sync_wdc() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_contstruct_wb_file_url() function
def WORKFLOW_TASK_construct_bdm_workbook(src_filename: str,
                                       wb_type: str,
                                       fi_key: str,
                                       wf_key: str,
                                       wf_purpose: str,
                                       bdm_DC: BudManAppDataContext_Base) -> bdm.BUDMAN_RESULT_TYPE:
    """Construct a workbook file URL based on the provided parameters.
    This function targets a workbook file url for for filename using:
        fi_key, wf_key, wf_purpose and wb_type.

    Args:
        filename (str): The base filename without extension.
        wb_type (str): The workbook type (e.g., 'excel_txns').
        fi_key (str): The financial institution key.
        wf_key (str): The workflow key.
        wf_purpose (str): The workflow purpose.

    Returns:
        bdm.BUDMAN_RESULT_TYPE: A tuple containing boolean indicating success, 
        and the constructed workbook file URL or an error message.
    """
    try:
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        model:BudgetDomainModel = bdm_DC.model
        if not model:
            m = "No BudgetDomainModel binding in the DC."
            logger.error(m)
            return False, m
        # Validate the provided parameters.
        if p3u.str_empty(src_filename):
            m = "Empty filename provided."
            logger.error(m)
            return False, m
        if not model.bdm_FI_KEY_validate(fi_key):
            m = f"Invalid fi_key: '{fi_key}'"
            logger.error(m)
            return False, m
        if not model.bdm_WF_KEY_validate(wf_key):
            m = f"Invalid wf_key: '{wf_key}'"
            logger.error(m)
            return False, m
        if not model.bdm_WF_PURPOSE_validate(wf_purpose):
            m = f"Invalid wf_purpose: '{wf_purpose}'"
            logger.error(m)
            return False, m
        if wb_type not in bdm.VALID_WB_TYPE_VALUES:
            m = f"Invalid wb_type: '{wb_type}'"
            logger.error(m)
            return False, m
        # Construct the workbook file URL.
        wb_extension: str = bdm.WB_FILETYPE_MAP[wb_type]
        wb_prefix: str = model.bdm_FI_WF_FOLDER_CONFIG_ATTRIBUTE(
            fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose,
            attribute=bdm.WF_PREFIX, raise_errors=True)
        wf_folder_url: str = model.bdm_FI_WF_FOLDER_CONFIG_ATTRIBUTE(
            fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose,
            attribute=bdm.WF_FOLDER_URL, raise_errors=True)
        wf_folder: str = model.bdm_FI_WF_FOLDER_CONFIG_ATTRIBUTE(
            fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose,
            attribute=bdm.WF_FOLDER, raise_errors=True)
        filename = f"{wb_prefix}{src_filename}{wb_type}"
        full_filename: str = f"{filename}{wb_extension}"
        folder_path = Path.from_uri(wf_folder_url) 
        file_path = folder_path / full_filename
        # Check if the dst file already exists.
        if file_path.exists():
            m = (f"Destination file already exists: "
                 f"'{file_path.as_uri()}'")
            logger.error(m)
        file_url: str = file_path.as_uri()
        new_wb: BDMWorkbook = BDMWorkbook()
        new_wb.wb_type = wb_type
        new_wb.wb_name = full_filename
        new_wb.wb_filename = filename
        new_wb.wb_filetype = wb_extension
        new_wb.wb_url = file_url
        new_wb.fi_key = fi_key
        new_wb.wf_key = wf_key
        new_wb.wf_purpose = wf_purpose
        new_wb.wf_folder_url = wf_folder_url
        new_wb.wf_folder = wf_folder
        return True, new_wb
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_construct_wb_file_url: {m}")
        logger.error(err_msg)
        return False, err_msg
#endregion WORKFLOW_TASK_contstruct_wb_file_url() function
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK helper functions
# ---------------------------------------------------------------------------- +
#region process_selected_workbook_input() function
def process_selected_workbook_input(cmd: p3m.CMD_OBJECT_TYPE, 
                    bdm_DC: BudManAppDataContext_Base) -> List[BDMWorkbook]:
    """Process the workbook input from the command, return a list of 
    BDMWorkbooks, which may be empty.

    Arguments:
        cmd (Dict): A p3m.CMD_OBJECT.

    Returns:
        List[BDMWorkbook]: A list of BDMWorkbook objects.
    """
    try:
        # TODO: this is duplicated code from a view_model method, needed until
        # the workflow command processing is refactored.
        # Extract common command attributes to select workbooks for task action.
        wb_list : List[int] = cmd.get(cp.CK_WB_LIST, [])
        all_wbs : bool = cmd.get(cp.CK_ALL_WBS, bdm_DC.dc_ALL_WBS)
        selected_bdm_wb_list : List[BDMWorkbook] = []
        load_workbook:bool = cmd.get(cp.CK_LOAD_WORKBOOK, False)
        if all_wbs:
            # If all_wbs is True, process all workbooks in the data context.
            selected_bdm_wb_list = list(bdm_DC.dc_WORKBOOK_DATA_COLLECTION.values())
        elif len(wb_list) > 0:
            for wb_index in wb_list:
                bdm_wb = bdm_DC.dc_WORKBOOK_by_index(wb_index)
                selected_bdm_wb_list.append(bdm_wb)
        else:
            # No workbooks selected by the command parameters.
            return selected_bdm_wb_list
        for bdm_wb in selected_bdm_wb_list:
            bdm_wb_abs_path = bdm_wb.abs_path()
            if bdm_wb_abs_path is None:
                selected_bdm_wb_list.remove(bdm_wb)
                m = f"Excluded workbook: '{bdm_wb.wb_id}', "
                m += f"workbook url is not valid: {bdm_wb.wb_url}"   
                logger.error(m)
                continue
            if load_workbook and not bdm_wb.wb_loaded:
                # Load the workbook content if it is not loaded.
                success, result = bdm_DC.dc_WORKBOOK_content_get(bdm_wb)
                if not success:
                    selected_bdm_wb_list.remove(bdm_wb)
                    m = f"Excluded workbook: '{bdm_wb.wb_id}', "
                    m += f"failed to load: {result}"
                    logger.error(m)
                    continue
                if not bdm_wb.wb_loaded:
                    logger.warning(f"Workbook '{bdm_wb.wb_id}' wb_loaded was False!")
                    bdm_wb.wb_loaded = True
                # self.dc_LOADED_WORKBOOKS[bdm_wb.wb_id] = wb_content
        return selected_bdm_wb_list
    except Exception as e:
        m = f"Error processing workbook input: {p3u.exc_err_msg(e)}"
        logger.error(m)
        raise RuntimeError(m)
#endregion process_selected_workbook_input() function
# ---------------------------------------------------------------------------- +
#endregion WORKFLOW_TASK helper functions
# ---------------------------------------------------------------------------- +
