# ---------------------------------------------------------------------------- +
#region budget_categorization.py module
""" Financial Budget Workflow: "categorization" of transaction workbooks.

    Workflow: categorization
    Input Folder: Financial Institution (FI) Incoming Folder (IF)
    Output Folder: Financial Institution (FI) Categorized Folder (CF)
    FI transaction workbooks are typically excel files. 

    Workflow Pattern: Apply a workflow_process (function) to each item in the 
    input folder, placing items in the output folder as appropriate to the 
    configured function. Each WorkFLow instance in the config applies one 
    function to the input with resulting output.

    TODO: Consider xlwings package for Excel integration.
"""
#endregion budget_categorization.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import re, logging, time, hashlib, datetime
from pathlib import Path
from dataclasses import dataclass, field

# third-party modules and packages
import p3logging as p3l, p3_utils as p3u
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from pyparsing import Optional

# local modules and packages
from budman_namespace.design_language_namespace import *
from budman_namespace.bdm_workbook_class import BDMWorkbook
from .budget_category_mapping import (
    get_category_map, check_register_map, get_category_histogram,
    clear_category_histogram
)
from .workflow_utils import (
    category_map_count, 
    check_register_map,
    get_category_histogram,
)
from .txn_category import (BDMTXNCategoryManager, TXNCategoryCatalog)
from budman_data_context import BudManAppDataContext_Base
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)

# Note: Python lists are 0-based. With openpyxl, data is often returned in a
# list. In excel, worksheet columns are 1-based. So, we need to adjust the indices.
# A list of cells from a worksheet row is 0-based, with cell(0) being the value
# from column 1, or column 'A'.

# Symbols for BOA .csv files.
BOA_ORIGINAL_DESCRIPTION_COL_NAME = "Original Description"
BOA_DATE_COL_NAME = "Date"
BOA_CURRENCY_COL_NAME = "Currency"
BOA_AMOUNT_COL_NAME = "Amount"
BOA_ACCOUNT_NAME_COL_NAME = "Account Name"

# BOA csv files originals contain these columns, beginning with "Status".
BOA_WB_COLUMNS = [
    "Status", 
    BOA_DATE_COL_NAME, 
    BOA_ORIGINAL_DESCRIPTION_COL_NAME,
    "Split Type", 
    "Category", 
    BOA_CURRENCY_COL_NAME, 
    BOA_AMOUNT_COL_NAME, 
    "User Description", 
    "Memo", 
    "Classification", 
    BOA_ACCOUNT_NAME_COL_NAME,
    "Simple Description"
    ]

# The following columns are not needed and can be removed
BOA_WB_COLUMNS_TO_REMOVE = [
    "Status", 
    "Split Type", 
    "Category", 
    "User Description", 
    "Memo", 
    "Classification"
    ]


# BudMan adds additional columns prior to processing transactions. These
# columns are filled by BudMan workflows, such as categorization.
BUDGET_CATEGORY_COL_NAME = "Budget Category"  
ACCOUNT_CODE_COL_NAME = "Account Code" 
LEVEL_1_COL_NAME = "Level1" 
LEVEL_2_COL_NAME = "Level2" 
LEVEL_3_COL_NAME = "Level3" 
DEBIT_CREDIT_COL_NAME = "DebitOrCredit"  
YEAR_MONTH_COL_NAME = "YearMonth"  
PAYEE_COL_NAME = "Payee"
ESSENTIAL_COL_NAME = "Essential" 
RULE_COL_NAME = "Rule"

# BudMan utilizes the following columns from the BOA side:
DATE_COL_NAME = BOA_DATE_COL_NAME 
TRANSACTION_DESCRIPTION_COL_NAME = BOA_ORIGINAL_DESCRIPTION_COL_NAME
CURRENCY_COL_NAME = BOA_CURRENCY_COL_NAME
AMOUNT_COL_NAME = BOA_AMOUNT_COL_NAME 
ACCOUNT_NAME_COL_NAME = BOA_ACCOUNT_NAME_COL_NAME  

# BudMan processes the original .csv file from BOA to produce an excel
# transactions workbook with the following columns:
BUDMAN_WB_COLUMNS = [
    "Status", 
    DATE_COL_NAME, 
    TRANSACTION_DESCRIPTION_COL_NAME,
    "Split Type", 
    "Category", 
    CURRENCY_COL_NAME, 
    AMOUNT_COL_NAME, 
    "User Description", 
    "Memo", 
    "Classification", 
    ACCOUNT_NAME_COL_NAME,
    "Simple Description",
    BUDGET_CATEGORY_COL_NAME,
    ACCOUNT_CODE_COL_NAME,
    LEVEL_1_COL_NAME,
    LEVEL_2_COL_NAME,
    LEVEL_3_COL_NAME,
    DEBIT_CREDIT_COL_NAME,
    YEAR_MONTH_COL_NAME,
    PAYEE_COL_NAME,
    ESSENTIAL_COL_NAME
    ]

# Column indices for a list of cells from a row in the BOA workbook.
DATE_COL_INDEX = 1
ORIGINAL_DESCRIPTION_COL_INDEX = 2  
CURRENCY_COL_INDEX = 4  
AMOUNT_COL_INDEX = 6  
ACCOUNT_NAME_COL_INDEX = 10  
BUDGET_CATEGORY_COL_INDEX = 12  

BOA_WB_COL_DIMENSIONS = {
    DATE_COL_NAME: 12,
    TRANSACTION_DESCRIPTION_COL_NAME: 95,
    CURRENCY_COL_NAME: 9,
    AMOUNT_COL_NAME: 14,
    ACCOUNT_NAME_COL_NAME: 68,
    BUDGET_CATEGORY_COL_NAME: 40,
    LEVEL_1_COL_NAME: 20,
    LEVEL_2_COL_NAME: 20,
    LEVEL_3_COL_NAME: 20,
    DEBIT_CREDIT_COL_NAME: 10,
    YEAR_MONTH_COL_NAME: 20,
    PAYEE_COL_NAME: 25,
    ESSENTIAL_COL_NAME: 10,
}
BUDMAN_SHEET_NAME = "TransactionData"

BUDMAN_REQUIRED_COLUMNS = [
    DATE_COL_NAME, 
    TRANSACTION_DESCRIPTION_COL_NAME, 
    CURRENCY_COL_NAME,
    AMOUNT_COL_NAME, 
    ACCOUNT_CODE_COL_NAME, 
    BUDGET_CATEGORY_COL_NAME,
    LEVEL_1_COL_NAME, 
    LEVEL_2_COL_NAME, 
    LEVEL_3_COL_NAME,
    DEBIT_CREDIT_COL_NAME,
    YEAR_MONTH_COL_NAME,
    PAYEE_COL_NAME,
    ESSENTIAL_COL_NAME,
    RULE_COL_NAME
]

#endregion Globals and Constants
# ---------------------------------------------------------------------------- +
#region dataclasses
TRANS_PARAMETERS = ["tid", "date", "description", "currency",
                    "account_code", "amount", "category",
                    "level1", "level2", "level3", 
                    "debit_credit", "year_month", "payee", "essential"]
@dataclass
class TransactionData:
    """Data class to hold transaction data."""
    tid: str = None   # Transaction ID.
    date: datetime.date = None
    description: str = None
    currency: str = None
    account_code: str = None
    amount: float = 0.0
    category: str = None
    level1: str = None
    level2: str = None
    level3: str = None
    debit_credit: str = None  # 'D' for debit, 'C' for credit.
    year_month: str = None  # Formant 'YYYY-MM-mmm' e.g., 2025-01-Jan
    essential: bool = False
    payee: str = field(default="")
    rule: int = field(default=-1)
    manual: bool = field(default=False)
    
    def data_str(self) -> str:
        """Return a string representation of the transaction data."""
        ret =  f"{self.tid:12}|{self.date.strftime("%m/%d/%Y")}|"
        ret += f"{self.year_month:11}|{self.account_code:26}|" 
        ret += f"{self.payee:25}|{self.amount:>+12.2f}|{self.debit_credit:1}|" 
        ret += f"({len(self.description):03}){self.description:102}|->" 
        ret += f"|({len(self.category):03})|{self.category:40}|"
        return ret
    
#endregion dataclasses
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_check_sheet_columns() function
def WORKFLOW_TASK_check_sheet_columns(sheet: Worksheet, 
                        trans_desc: str,
                        add_columns: bool = True) -> bool:
    """Check that the sheet is ready to process transactions.
    
    BudMan uses these columns to process transactions in a sheet:
    DATE_COL_NAME - date of the transaction.
    TRANSACTION_DESCRIPTION_COL_NAME - the description of the transaction.
    CURRENCY_COL_NAME - the unit of currency, USD, EUR, etc.
    AMOUNT_COL_NAME - the amount of the transaction, as a float.
    ACCOUNT_CODE_COL_NAME - the short-name of the account for the transaction. 
    BUDGET_CATEGORY_COL_NAME - the budget category for the transaction.
    LEVEL_1_COL_NAME - the first level of the budget category.
    LEVEL_2_COL_NAME - the second level of the budget category.
    LEVEL_3_COL_NAME - the third level of the budget category.
    DEBIT_CREDIT_COL_NAME - 'D' for debit, 'C' for credit.
    YEAR_MONTH_COL_NAME - trans date in format 'YYYY-MM-mmm', e.g. '2025-01-Jan'.
    PAYEE - parsed name of the org associated with the transaction
    ESSENTIAL - True/False the budget category is defined essential.

    Args:
        sheet (openpyxl.worksheet): The worksheet to check.
        add_columns (bool): Whether to add missing columns or not.

    Returns:
        bool: True if all required columns are present, False otherwise.
    """
    try:
        logger.info("Check worksheet for required columns.")
        # Index the header row 1, it has the column names
        col_names = [cell.value for cell in sheet[1]]
        logger.debug(f"Column names in sheet('{sheet.title}'): {col_names}")
        
        # Check if all required columns are present
        missing_columns = [col for col in BUDMAN_REQUIRED_COLUMNS if col not in col_names]
        
        if missing_columns is not None and len(missing_columns) > 0:
            # If the trans_desc column is missing, cannot continue.
            if trans_desc in missing_columns:
                m = f"Source transaction description column '{trans_desc}' not found in header row."
                logger.error(m)
                return False
            logger.error(f"Some required columns are present in sheet:('{sheet.title}')")
            logger.error(f"  Missing Columns: '{missing_columns}'")
        
        missing_count = len(missing_columns)
        added_count = 0
        if add_columns:
            # Add missing columns to the sheet to the right.
            for col_name in missing_columns:
                i = sheet.max_column + 1 # insert before column after last column
                sheet.insert_cols(i)
                sheet.cell(row=1, column=i).value = col_name
                # Set column width based on predefined dimensions
                width = BOA_WB_COL_DIMENSIONS.get(col_name, 20)
                sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width
                logger.debug(f"Column '{col_name}' added at index = {i}, "
                            f"column_letter = '{sheet.cell(row =1, column=i).column_letter}'")
                added_count += 1
            logger.info(f"Added '{added_count}' of '{missing_count}' missing columns from: {', '.join(missing_columns)}")
        if added_count < missing_count:
            logger.error(f"Sheet('{sheet.title}') missing required columns: {', '.join(missing_columns)}")
            return False
        logger.info("Completed checks for required columns.")
        return True
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_TASK_check_sheet_columns() function
# ---------------------------------------------------------------------------- +
#region check_sheet_schema() function
def check_sheet_schema(wb: Workbook, correct: bool = False) -> bool:
    """Check that the Workbook active sheet is ready to process transactions.
    
    Steps to check the active sheet of an excel workbook for BudMan processing:
    1. Should be just 1 worksheet, the active worksheet.
    2. Title must be 'TransactionData'. BOA_SHEET_NAME
    3. Column names must match spelling an order from BOA_WB_COLUMNS.
    Args:
        wb (openpyxl.workbook): The workbook to check.
        correct (bool): Whether to correct the sheet schema or not.

    Returns:
        bool: True if all required columns are present, False otherwise.
    """
    try:
        logger.info("Check worksheet for schema structure.")
        p3u.is_not_obj_of_type("wb", wb, Workbook, raise_error=True)
        # Check the active worksheet.
        ws = wb.active  # Get the active worksheet.
        good_schema = True  # Assume the schema is good.
        rule1 = rule2 = rule3 = good_schema
        # 1. Check if there is only one worksheet.
        sheet_names = wb.sheetnames 
        sheet_count = len(sheet_names)
        budman_sheet_index = -1
        if sheet_count > 1:
            logger.error(f"Workbook has {sheet_count} sheets, expected 1. "
                         f"Sheet names: {', '.join(sheet_names)}")
            good_schema = rule1 = False
        # 2. Check if the sheet title is BUDMAN_SHEET_NAME.
        if BUDMAN_SHEET_NAME not in sheet_names:
            logger.error(f"Workbook does not have a sheet named '{BUDMAN_SHEET_NAME}'. "
                         f"Sheet names: {', '.join(sheet_names)}")
            good_schema = rule2 = False
        else:
            # Get the index of the BudMan sheet.
            budman_sheet_index = sheet_names.index(BUDMAN_SHEET_NAME)
            ws = wb[sheet_names[budman_sheet_index]] 
        # 3. Check the column names are correct spelling and order.
        col_names = [cell.value for cell in ws[1]]
        # Check if all required columns are present
        missing_columns = [col for col in BUDMAN_REQUIRED_COLUMNS if col not in col_names]
        if len(missing_columns) > 0:
            logger.error(f"Missing required columns: {', '.join(missing_columns)}")
            good_schema = rule3 = False
        # If just one sheet with wrong title, rename it to BUDMAN_SHEET_NAME.
        if rule1 and rule3 and not rule2:
            logger.error(f"BudMan sheet '{BUDMAN_SHEET_NAME}' not found in workbook.")
            previous_title = ws.title  # Save the previous title.
            ws.title = BUDMAN_SHEET_NAME  # Rename the active sheet.
            logger.info(f"Renamed active sheet from '{previous_title}' to '{BUDMAN_SHEET_NAME}'.")
            good_schema = True
        # If more than one sheet, active sheet with wrong title, but is has 
        # the correct columns, rename the active sheet to BUDMAN_SHEET_NAME.
        if not rule1 and rule3 and not rule2:
            logger.error(f"Active sheet '{BUDMAN_SHEET_NAME}' has wrong name "
                         f"but has the correct columns. " 
                         f"Renaming it to '{BUDMAN_SHEET_NAME}'.")
            previous_title = ws.title  # Save the previous title.
            ws.title = BUDMAN_SHEET_NAME  # Rename the active sheet.
            logger.info(f"Renamed active sheet from '{previous_title}' to '{BUDMAN_SHEET_NAME}'.")
            good_schema = True
        return good_schema
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion check_sheet_schema() function
# ---------------------------------------------------------------------------- +
#region WORKSHEET_data(ws:Worksheet) -> List[TransactionData]
def WORKSHEET_data(ws:Worksheet, just_values:bool=False) -> list[TransactionData]:
    """Extract transaction data from a worksheet.

    Args:
        ws (Worksheet): The worksheet to extract data from.

    Returns:
        list[TransactionData]: A list of TransactionData objects.
    """
    try:
        if not isinstance(ws, Worksheet):
            raise TypeError(f"Expected 'ws' arg to be a Worksheet, got {type(ws)}")
        transactions = []
        headers = [cell.value for cell in ws[1]]  # Get the header row values.
        for row in ws.iter_rows(min_row=2, values_only=just_values):
            transaction = WORKSHEET_row_data(row,headers)  # Extract transaction data from the row.
            transactions.append(transaction)
        return transactions
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKSHEET_data(ws:Worksheet) -> List[TransactionData]
# ---------------------------------------------------------------------------- +
#region WORKSHEET_row_data(row:list) -> TransactionData
def WORKSHEET_row_data(row:tuple,hdr:list=BUDMAN_WB_COLUMNS) -> TransactionData:
    """Extract transaction data from a worksheet row.

    Args:
        row (list): List of either cells or just the cell values.
        hdr (list): A list of header column names used to index values from the row.

    Returns:
        list[TransactionData]: A list of TransactionData objects.
    """
    try:
        # Validation
        p3u.is_not_obj_of_type("row", row, tuple, raise_error=True)
        p3u.is_not_obj_of_type("hdr", hdr, list, raise_error=True)
        if len(row) == 0 or len(hdr) == 0 or len(row) != len(hdr):
            raise ValueError("Row and Hdr must be equal, non-zero length.")
        # Check if all required columns are present in the hdr.
        missing_columns = [col for col in BUDMAN_REQUIRED_COLUMNS if col not in hdr]
        if missing_columns:
            raise ValueError(f"Hdr is missing required columns: {missing_columns}")
        # First, determine if row contains Cells or value-only cell.value's and
        # extract the values accordingly.
        if isinstance(row[0], Cell):
            # This is a Tuple of Cell objects.
            row_values = [cell.value for cell in row]
        else:
            row_values = row 
        # Second, verify the hdr contains the required columns.
        row_dict = dict(zip(hdr, row_values))  # Create a dict from the header and row values.

        # Extract some of the values as strings to generate a hash from
        t_date_str = p3u.iso_date_only_string(row_dict[DATE_COL_NAME])
        t_desc = row_dict[TRANSACTION_DESCRIPTION_COL_NAME]
        t_currency = row_dict[CURRENCY_COL_NAME]
        t_amt_str = str(row_dict[AMOUNT_COL_NAME])
        t_acct_str = row_dict[ACCOUNT_NAME_COL_NAME]
        t_all_str = t_date_str + t_desc + t_currency + t_amt_str + t_acct_str
        t_id = p3u.gen_hash_key(t_all_str) 
        t_manual: bool = False
        t_rule: int = row_dict[RULE_COL_NAME] if RULE_COL_NAME in row_dict else -1
        if isinstance(t_rule, int) and t_rule > 20200000:
            t_manual = True

        transaction = TransactionData(
            tid=t_id,
            date=row_dict[DATE_COL_NAME],
            description=row_dict[TRANSACTION_DESCRIPTION_COL_NAME],
            currency=row_dict[CURRENCY_COL_NAME],
            amount=row_dict[AMOUNT_COL_NAME],
            account_code=row_dict[ACCOUNT_CODE_COL_NAME],
            category=row_dict[BUDGET_CATEGORY_COL_NAME],
            level1=row_dict[LEVEL_1_COL_NAME],
            level2=row_dict[LEVEL_2_COL_NAME],
            level3=row_dict[LEVEL_3_COL_NAME],
            debit_credit=row_dict[DEBIT_CREDIT_COL_NAME],
            year_month=row_dict[YEAR_MONTH_COL_NAME],
            rule=row_dict[RULE_COL_NAME],
            manual=t_manual
        )
        return transaction
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKSHEET_row_data(row:list) -> TransactionData
# ---------------------------------------------------------------------------- +
#region col_i() function
def col_i(col_name:str, hdr:list) -> int:
    """Get the 0-based index of a column name in a header list.
    
    Args:
        col_name (str): The name of the column to find.
        hdr (list): The header list of column names from row(1) to search in.

    Returns:
        int: The index of the column name in the header row, or -1 if not found.
    """
    try:
        if not isinstance(col_name, str):
            raise TypeError(f"Expected 'col_name' to be a str, got {type(col_name)}")
        if not isinstance(hdr, list):
            raise TypeError(f"Expected 'hdr' to be a list, got {type(hdr)}")
        return hdr.index(col_name) if col_name in hdr else -1
    except ValueError:
        return -1
#endregion col_i() function
# ---------------------------------------------------------------------------- +
#region year_month_str() function
def year_month_str(date:object) -> str:
    """Convert a date object to a year-month string in the format 'YYYY-MM-mmm'.
    
    Args:
        date (object): A date object or a string that can be converted to a date.

    Returns:
        str: The year-month string in the format 'YYYY-MM-mmm'.
    """
    try:
        if isinstance(date, str):
            date = datetime.datetime.strptime(date, "%m/%d/%Y").date()
        elif not isinstance(date, datetime.date):
            raise TypeError(f"Expected 'date' to be a date object or a str, got {type(date)}")
        return date.strftime("%Y-%m-%b")
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion year_month_str() function
# ---------------------------------------------------------------------------- +
#region validate_budget_categories() function
def validate_budget_categories(bdm_wb:BDMWorkbook, 
                               bdm_DC: BudManAppDataContext_Base,
                               pad:str='') -> BUDMAN_RESULT_TYPE:
    """Validate budget categories in the workbook.

    Args:
        bdm_wb (BDMWorkbook): The budget workbook to validate.
        bdm_DC (BudManAppDataContext_Base): The data context for the budget.

    Returns:
        bool: True if validation is successful, False otherwise.
    """
    try:
        # Validate the input parameters.
        _ = p3u.is_not_obj_of_type("bdm_wb", bdm_wb, BDMWorkbook, 
                                   raise_error=True)
        _ = p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base, 
                                   raise_error=True)

        # Check if the budget categories are valid.
        task = "validate_budget_categories()"
        result : str = f"{pad}{task}"
        unique_categories: Dict[str, int] = dict()  # To track unique budget categories.
        if bdm_wb.wb_type != WB_TYPE_EXCEL_TXNS:
            # This is not a transactions workbook, no action taken.
            m = (f"Workbook '{bdm_wb.wb_id}' is not wb_type: "
                 f"'{WB_TYPE_EXCEL_TXNS}', no action taken.")
            logger.error(m)
            result += f"\n{pad}{m}"
            return False, result
        if not bdm_wb.wb_loaded:
            # Load the workbook from the data context.
            m = f"Workbook '{bdm_wb.wb_id}' is not loaded, no action taken."
            logger.error(m)
            result += f"\n{pad}{m}"
            return False, result
        if p3u.is_not_obj_of_type("wb", bdm_wb.wb_content, Workbook):
            m = f"Error accessing wb_content for workbook: '{bdm_wb.wb_id}'."
            logger.error(m)
            result += f"\n{pad}{m}"
            return False, result
        ws : Worksheet = bdm_wb.wb_content.active  # Get the active worksheet.
        if not WORKFLOW_TASK_check_sheet_columns(ws, add_columns=False):
            m = (f"Sheet '{ws.title}' cannot be mapped due to "
                    f"missing required columns.")
            logger.error(m)
            result += f"\n{pad}{m}"
            return False, result
        # A row is a tuple of the Cell objects in the row. Tuples are 0-based
        # hdr is a list, also 0-based. So, using the index(name) will 
        # give the cell from a row tuple matching the column name in hdr.
        hdr = [cell.value for cell in ws[1]] 
        # These are values to validate.
        budget_cat_i = col_i(BUDGET_CATEGORY_COL_NAME,hdr)
        l1_i = col_i(LEVEL_1_COL_NAME,hdr)
        l2_i = col_i(LEVEL_2_COL_NAME,hdr)
        l3_i = col_i(LEVEL_3_COL_NAME,hdr)
        # Model-Aware: For the FI currently in the focus of the DC, we need
        # some FI-specific info, the name of the transaction workbook 
        # column used to map budget categories, the source.
        txn_desc_col_name = bdm_DC.dc_FI_OBJECT[FI_TRANSACTION_DESCRIPTION_COLUMN]
        txn_desc_i = col_i(txn_desc_col_name, hdr)
        errors = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # row is a 'tuple' of Cell objects, 0-based index
            # row_idx = row[0].row  # Get the row index, the row number, 1-based.
            # Validate the budget category.
            budget_category = row[budget_cat_i]
            txn_desc = row[txn_desc_i]
            if not p3u.is_non_empty_str(BUDGET_CATEGORY_COL_NAME, budget_category):
                errors += 1
                # BUDGET_CATEGORY_COL_NAME cell cannot be empty.
                m = (f"Row {row.index} has an invalid budget category: '{budget_category}'.")
                logger.error(m)
                result += f"\n{pad}{m}"
                continue
            unique_categories[budget_category] = unique_categories.get(budget_category, 0) + 1
            # Validate the levels.
            level1 = row[l1_i] 
            level2 = row[l2_i] or ''
            level3 = row[l3_i] or ''
            l1, l2, l3 = p3u.split_parts(budget_category)
            if not p3u.is_non_empty_str("level1", level1):
                # Level 1 cannot be empty.
                m = (f"Row {row_idx} has an invalid Level 1: '{level1}' "
                     f"for budget category: '{budget_category}'.")
                logger.error(m)
                result += f"\n{pad}{m}"
                continue
            if l1 != level1:
                m = (f"Row {row_idx} has a Level 1: '{level1}' that does not match "
                     f"the budget category: '{budget_category}' for description '{txn_desc}'.")
                logger.error(m)
                result += f"\n{pad}{m}"
                continue
            if l2 != level2:
                m = (f"Row {row_idx} has a Level 2: '{level2}' that does not match "
                     f"the budget category: '{budget_category}' for description '{txn_desc}'.")
                logger.error(m)
                result += f"\n{pad}{m}"
                continue
            if l3 != level3:
                m = (f"Row {row_idx} has a Level 3: '{level3}' that does not match "
                     f"the budget category: '{budget_category}' for description '{txn_desc}'.")
                logger.error(m)
                result += f"\n{pad}{m}"
                continue
        result += f"\n{pad}{P2}Workbook validation phase completed. unique categories: {len(unique_categories)} errors: {errors}"
        catman : BDMTXNCategoryManager = bdm_DC.WF_CATEGORY_MANAGER
        fi_txn_catalog : TXNCategoryCatalog = catman.catalogs[bdm_wb.fi_key]
        cat_collection_list = list(fi_txn_catalog.category_collection.keys())
        for key in unique_categories:
            if key not in cat_collection_list:
                errors += 1
                m = (f"Row {row_idx} has a budget category: '{key}' that is not in the "
                     f"category collection for FI: '{bdm_wb.fi_key}'.")
                logger.error(m)
                result += f"\n{pad}{m}"
        result += f"\n{pad}{P2}Validation with CATEGORY_COLLECTION phase completed. unique categories: {len(unique_categories)} errors: {errors}"
        return True, result
    except Exception as e:
        m = p3u.exc_err_msg(e)
        result = f"Error validating budget categories: {m}"
        logger.error(p3u.exc_err_msg(e))
        return False, result
#endregion validate_budget_categories() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_process_budget_category() function
def WORKFLOW_TASK_process_budget_category(bdm_wb:BDMWorkbook,
                            bdm_DC : BudManAppDataContext_Base,
                            log_all : bool,
                            clear_other : bool=False) -> BUDMAN_RESULT_TYPE:
    """Process budget categorization for the workbook.
    
    The sheet has banking transaction data in rows and columns. 
    Column 'src' has the text presumed to be about the transaction.
    Column 'dst' will be assigned a mapped budget category. Append
    column 'Budget Category' if it is not already in the sheet.
    TODO: pass in a mapper function to map the src text to a budget category.

    Args:
        bdm_wb (BDMWorkbook): The workbook with an excel workbook.
        bdm_DC (BudManAppDataContext_Base): The data context for the budget.
        log_all (bool): Whether to log all mappings or just unmapped.
        clear_other (bool): Whether to clear the Other category workbook content.
    """
    try:
        #region Validate all required information is accessible.
        #
        # Validatethe input parameters.
        _ = p3u.is_not_obj_of_type("wb_object", bdm_wb, BDMWorkbook,
                                   raise_error=True)
        _ = p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                                   raise_error=True)
        success : bool = False
        # Validate applicable WB_TYPE_EXCEL_TXNS workbooks
        if bdm_wb.wb_type != WB_TYPE_EXCEL_TXNS:
            # This is not a transactions workbook, no action taken.
            m = (f"Workbook '{bdm_wb.wb_id}' is not wb_type: "
                 f"'{WB_TYPE_EXCEL_TXNS}', no action taken.")
            logger.error(m)
            return False, m
        # Validate the workbook is loaded.
        if not bdm_wb.wb_loaded:
            # Load the workbook from the data context.
            m = f"Workbook '{bdm_wb.wb_id}' is not loaded, no action taken."
            logger.error(m)
            return False, m
        if p3u.is_not_obj_of_type("wb", bdm_wb.wb_content, Workbook):
            m = f"Error accessing wb_content for workbook: '{bdm_wb.wb_id}'."
            logger.error(m)
            return False, m
        #
        #endregion Validate all required information is accessible.

        #region FI-specific data needed for the workbook 
        #
        # Validate category manager to get to the appropriate category_map.
        # Use the bdm_wb fi_key to access the category manager for the FI in 
        # the data context. Obtain the compiled category map (ccm) for the 
        # FI by FI_KEY. Uses the current FI_OBJECT in the supplied DC.
        fi_key: str = bdm_wb.fi_key
        if not p3u.is_non_empty_str("fi_key", fi_key):
            m = f"Invalid fi_key: '{fi_key}' in workbook '{bdm_wb.wb_id}'."
            logger.error(m)
            return False, m
        fi_obj: dict = bdm_DC.dc_FI_OBJECT
        if fi_key != fi_obj[FI_KEY]:
            m = (f"FI_Key: '{fi_key}' from workbook does not match FI_Key "
                 f"from current DC FI_OBJECT: '{fi_obj[FI_KEY]}'.")
            logger.error(m)
            return False, m
        # Process the Category Managers with the fi_key
        catman : BDMTXNCategoryManager = bdm_DC.WF_CATEGORY_MANAGER
        if fi_key not in catman.catalogs:
            # No WB_TYPE_TXN_CATEGORIES dictionary for this FI, no action taken.
            m = (f"WB_TYPE_TXN_CATEGORIES dictionary not found for FI '{fi_key}', "
                 f"no action taken.")
            logger.error(m)
            return False, m
        fi_txn_catalog : TXNCategoryCatalog = catman.catalogs[fi_key]
        if not fi_txn_catalog:
            # No TXNCategoryCatalogItem for this FI, no action taken.
            m = (f"TXNCategoryCatalogItem not found for FI '{fi_key}', "
                 f"no action taken.")
            logger.error(m)
            return False, m
        ccm : Dict[str, re.Pattern] = fi_txn_catalog.compiled_category_map
        if not ccm:
            # No compiled category map for this FI, no action taken.
            m = (f"Compiled category map not found for FI '{fi_key}', "
                 f"no action taken.")
            logger.error(m)
            return False, m
        # For the FI_KEY, get the name of the workbook column used as the 
        # transaction description (trans_desc) which is mapped to a budget 
        # category (bud_cat). Map 'trans_desc' column to 'bud_cat' column in 
        # the workbook being category mapped.
        trans_desc = fi_obj[FI_TRANSACTION_DESCRIPTION_COLUMN]
        if not p3u.is_non_empty_str("src", trans_desc):
            m = f"Source transaction description column name '{str(trans_desc)}' is not valid."
            logger.error(m)
            return False, m
        bud_cat = fi_obj[FI_TRANSACTION_BUDGET_CATEGORY_COLUMN]
        if not p3u.is_non_empty_str("bud_cat", bud_cat):
            m = f"Destination budget category column name '{str(bud_cat)}' is not valid."
            logger.error(m)
            return False, m
        # Now get the name of the worksheet in the workbook containung the
        # transactions for catgegory mapping.
        ws_name = fi_obj[FI_TRANSACTION_WORKSHEET_NAME]
        if not p3u.is_non_empty_str("ws_name", ws_name):
            m = f"For FI_KEY: '{fi_key}', Transaction worksheet name '{str(ws_name)}' is not valid."
            logger.error(m)
            return False, m
        #
        # Access the named worksheet from the wb_content
        ws : Worksheet = bdm_wb.wb_content[ws_name] 
        if ws is None or not isinstance(ws, Worksheet):
            m = (f"Worksheet '{ws_name}' not found in workbook '{bdm_wb.wb_id}'.")
            logger.error(m)
            return False, m
        #
        #endregion FI-specific data needed for the workbook

        # The workbooks's transaction worksheet is now available as ws.
        # Check that the required columns are present.
        if not WORKFLOW_TASK_check_sheet_columns(ws, trans_desc, add_columns=False):
            m = (f"Sheet '{ws.title}' cannot be mapped due to "
                    f"missing required columns.")
            logger.error(m)
            return False, m

        # An openpyxl Worksheet is an array of tuples, 1-based index.
        # A row is a tuple of the Cell objects in the row. Tuples are 0-based.
        # The first row is the header row, hdr, a list, also 0-based. So, using 
        # the index of the column name from the hdr gives the cell from a 
        # row tuple matching the column name in hdr.
        hdr = [cell.value for cell in ws[1]] # Extract hdr col names. 

        # TODO: need to refactor this to do replacements by col_name or something.
        # This is specific to the Budget Category mapping, which now is to be
        # split into 3 levels: Level1, Level2, Level3.

        # Setup index values for each hdr column of interest for row access.
        trans_desc_i = col_i(trans_desc,hdr)
        bud_cat_i = col_i(bud_cat,hdr)
        date_i = col_i(DATE_COL_NAME,hdr)
        l1_i = col_i(LEVEL_1_COL_NAME,hdr)
        l2_i = col_i(LEVEL_2_COL_NAME,hdr)
        l3_i = col_i(LEVEL_3_COL_NAME,hdr)
        amt_i = col_i(AMOUNT_COL_NAME,hdr)
        dORc_i = col_i(DEBIT_CREDIT_COL_NAME,hdr)
        year_month_i = col_i(YEAR_MONTH_COL_NAME,hdr)
        acct_name_i = col_i(ACCOUNT_NAME_COL_NAME,hdr)
        acct_code_i = col_i(ACCOUNT_CODE_COL_NAME,hdr)
        acct_cell : Cell = ws.cell(row=1, column=acct_name_i + 1)
        payee_i = col_i(PAYEE_COL_NAME,hdr)
        essential_i = col_i(ESSENTIAL_COL_NAME,hdr)
        rule_i = col_i(RULE_COL_NAME,hdr)

        logger.debug(f"Mapping '{trans_desc}'({trans_desc_i}) to "
                    f"'{bud_cat}'({bud_cat_i})")
        num_rows = ws.max_row # or set a smaller limit
        other_count = 0
        ch = clear_category_histogram()  # Clear the category histogram.
        rules_count = category_map_count()
        # Open Other category workbook to save unmapped rows.
        other_wb: Workbook = None
        other_ws: Worksheet = None
        other_wb, other_ws = open_other_category_workbook(hdr, 
                                                          clear_content=clear_other)

        transactions: List[TransactionData] = []  # To hold all transactions.
        task_name = "WORKFLOW_TASK_process_budget_category()"
        logger.info(f"Start Task: {task_name}: Apply '{rules_count}' budget category mapping rules "
                    f"to {ws.max_row-1} rows in workbook: '{bdm_wb.wb_id}' "
                    f"worksheet: '{ws.title}'")
        st = p3u.start_timer()
        for row in ws.iter_rows(min_row=2):
            # row is a 'tuple' of Cell objects, 0-based index
            row_idx = row[0].row  # Get the row index, the row number, 1-based.
            transaction = WORKSHEET_row_data(row,hdr) 
            transactions.append(transaction)
            trans_str = f"Row({row_idx}): '{transaction.data_str()}'"
            rule_value = row[rule_i].value if rule_i != -1 else None
            # Do the mapping from trans_desc to bud_cat columns.
            if transaction.manual:
                # Skip manually modified transactions
                logger.debug(f"{P2}Skipping manual: {trans_str}")
                continue # skip due to manual category settings in workbook
            logger.debug(f"{P2}{trans_str}") if log_all else None
            bud_cat_value, payee, rule_index = WORKFLOW_TASK_categorize_transaction(
                transaction.description, 
                fi_txn_catalog, 
                log_all)
            row[bud_cat_i].value = bud_cat_value # Capture bud_cat mapping 
            # Modify the actual row with additional values for BudMan.
            date_val = row[date_i].value
            year_month: str = year_month_str(date_val) if date_val else None
            row[year_month_i].value = year_month
            l1, l2, l3 = p3u.split_parts(bud_cat_value)
            row[l1_i].value = l1 if l1_i != -1 else None
            row[l2_i].value = l2 if l2_i != -1 else None
            row[l3_i].value = l3 if l3_i != -1 else None
            row[dORc_i].value = 'C' if row[amt_i].value > 0 else 'D'
            acct_value = row[acct_name_i].value
            t_acct_code = acct_value.split('-')[-1].strip()
            row[acct_code_i].value = t_acct_code if acct_code_i != -1 else None
            if payee is not None:
                row[payee_i].value = payee if payee_i != -1 else None
            essential_value = False
            if (bud_cat_value != 'Other' and 
                bud_cat_value in fi_txn_catalog.category_collection):
                essential_value = fi_txn_catalog.category_collection[bud_cat_value].essential
            else:
                logger.warning(f"'{bud_cat_value}' not found in category collection.")
            row[essential_i].value = essential_value if essential_i != -1 else None
            row[rule_i].value = rule_index if rule_i != -1 else None
            # Capture 'Other' category transactions.
            if bud_cat_value == 'Other':
                copy_row_to_worksheet(row, other_ws)
                other_count += 1
                logger.debug(f"'Other': {trans_str}" )
        time_taken = p3u.stop_timer(st)
        close_other_category_workbook(other_wb)
        elapsed : float = time.time() - st
        per_row = elapsed / (num_rows - 1) if num_rows > 1 else 0.0
        ch = get_category_histogram() 
        m = (f"Task Complete: {time_taken} Mapped '{num_rows}' rows, to "
             f"'{len(ch)}' Categories, {per_row:6f} seconds per row, "
             f"'Other' category count: ({ch['Other']})({other_count})")
        logger.info(m)
        del transactions 
        return True, m
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        return False, m

def clear_worksheet(workbook, sheet_name):
    # Remove the existing worksheet
    workbook.remove(workbook[sheet_name])
    # Create a new worksheet with the same name
    workbook.create_sheet(sheet_name)
    workbook.active = workbook[sheet_name]

def copy_row_to_worksheet(source_row, dest_worksheet):
    # Get the next available row in the destination worksheet
    dest_row = dest_worksheet.max_row + 1

    for col, cell in enumerate(source_row, start=1):
        # Create a new cell in the destination worksheet
        dest_cell = dest_worksheet.cell(row=dest_row, column=col)
        
        # Copy the value
        dest_cell.value = cell.value
        
        # Copy the style
        if cell.has_style:
            dest_cell.font = cell.font.copy()
            dest_cell.border = cell.border.copy()
            dest_cell.fill = cell.fill.copy()
            dest_cell.number_format = cell.number_format
            dest_cell.protection = cell.protection.copy()
            dest_cell.alignment = cell.alignment.copy()

#
#endregion WORKFLOW_TASK_process_budget_category() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_categorize_transaction() function
def WORKFLOW_TASK_categorize_transaction(
        description : str, 
        txn_catalog:TXNCategoryCatalog,
        log_all : bool) -> Tuple[str,str]:
    """Use txn_catalog patterns to map description text to a category."""
    try:
        p3u.is_non_empty_str("description", description, raise_error=True)
        p3u.is_not_obj_of_type("txn_catalog", txn_catalog, TXNCategoryCatalog,
                               raise_error=True)
        txn_catalog.valid
        fi_key: str = txn_catalog.fi_key
        txn_category_collection = txn_catalog.txn_categories_workbook[WB_CATEGORY_COLLECTION]
        tcc_keys = list(txn_category_collection.keys())
        ccm: COMPLIED_CATEGORY_MAP_TYPE = txn_catalog.compiled_category_map
        ch = get_category_histogram()
        payee = ""
        rule_index = 0
        for rule_index, (pattern, category) in enumerate(ccm.items()):
            payee = ""
            m = pattern.search(description)
            if m:
                if category not in tcc_keys:
                    logger.warning(f"FI key '{fi_key}' rule_index: '{rule_index}', "
                                   f"Category '{category}', not found in "
                                   f"transaction category collection.")
                gc = len(m.groups())
                if gc == 0 or m[1] is None:
                    if log_all:
                        logger.debug(f"{P2}Matched rule_index: '{rule_index}', "
                                     f"pattern: '{pattern.pattern}' "
                                     f"category: '{category}', "
                                     f"groups: '{gc}', no payee found.")
                    return ch.count(category), payee, rule_index
                payee = m[1]
                # print(f"Matched pattern: '{pattern.pattern}' with payee: '{payee}'")
                if log_all:
                    logger.debug(f"{P2}Matched rule_index: '{rule_index}', "
                                 f"pattern: '{pattern.pattern}' "
                                 f"category: '{category}', "
                                 f"payee: '{payee}', groups: '{gc}'.")
                return ch.count(category), payee, rule_index
        logger.debug(f"{P2}No Match Description: [{description}]")
        return ch.count('Other'), 'Other', -1  # Default category if no match is found
    except re.PatternError as e:
        logger.error(p3u.exc_err_msg(e))
        logger.error(f'Pattern error: compiled_category_map dict: ' 
                     f'{{ \"{e.pattern}\": \"{category}\" }}')
        raise
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_TASK_categorize_transaction() function
# ---------------------------------------------------------------------------- +
#region open_other_category_workbook() function
def open_other_category_workbook(hdr: List[str],clear_content:bool=True) -> Tuple[Workbook, Worksheet]:
    """Open the 'Other' category workbook.

    Args:
        hdr (List[str]): The header row for the 'Other' category workbook.

    Returns:
        Optional[BDMWorkbook]: The 'Other' category workbook, or None if not found.
    """
    try:
        other_wb_abs_path_str = "C:/Users/ppain/OneDrive/budget/boa/Other.excel_txns.xlsx"
        other_wb_path: Path = Path(other_wb_abs_path_str)
        other_wb: Workbook = load_workbook(other_wb_path)
        other_ws: Worksheet = other_wb.active
        # Clear the other worksheet before processing.
        if clear_content:
            clear_worksheet(other_wb, BUDMAN_SHEET_NAME)
            other_ws = other_wb.active
            # Add the hdr row
            other_ws.append(hdr)
            confirm_hdr = [cell.value for cell in other_ws[1]] 
        return other_wb, other_ws
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise

def close_other_category_workbook(other_wb:Workbook) -> None:
    """Close the 'Other' category workbook.

    Args:
        other_wb (Workbook): The 'Other' category workbook to close.

    Returns:
        None
    """
    try:
        if p3u.is_not_obj_of_type("other_wb", other_wb, Workbook):
            raise TypeError(f"Expected 'other_wb' to be a Workbook, got {type(other_wb)}")
        other_wb_abs_path_str = "C:/Users/ppain/OneDrive/budget/boa/Other.excel_txns.xlsx"
        other_wb_path: Path = Path(other_wb_abs_path_str)
        other_wb.save(other_wb_path)  # Save the other workbook with 'Other' category rows.
        other_wb.close()  # Close the other workbook.
        return None
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion open_other_category_workbook() function
# ---------------------------------------------------------------------------- +
#region apply_check_register() function
def apply_check_register(cr_wb_content:BDM_CHECK_REGISTER_TYPE, 
                         trans_wb_ref:EXCEL_TXNS_WORKBOOK_TYPE) -> None:
    """Apply the check transactions to the worksheet.
    
    The sheet has banking transaction data in rows and columns. 
    The check_reg has a collection of checks. Scan and match the checks
    with the content in the worksheet. If found, modify the transactions.

    Args:
        sheet (openpyxl.worksheet): The worksheet to process.
        check_reg (dict): A dictionary of check transactions to apply.
    """
    try:
        # Validate the input parameters.
        p3u.is_not_obj_of_type("cr_wb_content", cr_wb_content, dict, raise_error=True)
        p3u.is_not_obj_of_type("trans_wb_ref",trans_wb_ref, Workbook, raise_error=True)
        # Validate the input parameters.
        cr = cr_wb_content
        sh = trans_wb_ref.active  # Get the active worksheet.
        if not WORKFLOW_TASK_check_sheet_columns(sh, add_columns=False):
            logger.error(f"Sheet '{sh.title}' cannot be mapped due to "
                         f"missing required columns.")
            return
        logger.info(f"Applying checks from check register to sheet: '{sh.title}' ")
        # transactions = WORKSHEET_data(sheet)
        # A row is a tuple of the Cell objects in the row. Tuples are 0-based
        # hdr is a list, also 0-based. So, using the index(name) will 
        # give the cell from a row tuple matching the column name in hdr.
        hdr = [cell.value for cell in sh[1]] 
        budget_cat = hdr.index(BUDGET_CATEGORY_COL_NAME)
        orig_desc = hdr.index(TRANSACTION_DESCRIPTION_COL_NAME)

        # For each check, with the check number and the Budget Category
        # 'Banking.Checks to Categorize', find the row in the worksheet to modify.
        target_cat = 'Banking.Checks to Categorize'
        check_pat = re.compile(r'^.*Check\s*x*(\d{1,6})\b.*$')  

        trans_match = []
        for row in sh.iter_rows(min_row=2):
            # row is a 'tuple' of Cell objects, 0-based index
            budget_cat_cell = row[budget_cat]
            if budget_cat_cell.value and target_cat in budget_cat_cell.value:
                orig_desc_cell = row[orig_desc]
                desc = row[orig_desc].value if orig_desc != -1 else ""
                m = check_pat.match(desc)
                if m:
                    check_key = m.group(1)
                    if check_key in cr:
                        # Modify the transaction with the check number.
                        pay_to = cr[check_key]['Pay-To']
                        new_cat = check_register_map[pay_to] if pay_to in check_register_map else 'Unknown'
                        new_desc = f"{pay_to} Check: {check_key}"
                        row[budget_cat].value = new_cat
                        row[orig_desc].value = new_desc
                        logger.info(f"Modified transaction for check: '{check_key}' "
                                    f"new_desc: '{new_desc}' new_cat: '{new_cat}'")
                logger.info(f"'{sh.title}' Match '{target_cat}' desc={desc}")
        logger.info(f"Found {len(trans_match)} transactions in sheet "
                    f"'{sh.title}' with category '{target_cat}' to modify.")
        return None
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise    
#endregion apply_check_register() function
# ---------------------------------------------------------------------------- +
