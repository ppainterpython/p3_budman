# ---------------------------------------------------------------------------- +
#region workflow_command_services.py module
""" workflow_command_services.py implements functions for BudMan app
workflow-related commands.

In general, services should return either data objects or command result objects.
Leave it to the caller, such as a View, or ViewMModel to handle additional
services, such as a pipeline, or to perform output functions.
"""
#endregion workflow_command_services.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import logging, shutil, datetime, re
from pathlib import Path
from typing import List, Type, Optional, Dict, Tuple, Any, Callable
from treelib import Tree
from datetime import datetime as dt

# third-party modules and packages
import p3_utils as p3u, pyjson5, p3logging as p3l, p3_mvvm as p3m
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# local modules and packages
import budman_settings as bdms
import budman_namespace as bdm
from budman_namespace import P2, P4, P8
from budman_namespace import BDMWorkbook
from budget_domain_model import BudgetDomainModel
from budman_data_context import BudManAppDataContext_Base
from budman_workflow_services import *
from budman_workflow_services.workflow_namespace import BUDMAN_TXNS_WORKBOOK_COL_NAMES
from budget_storage_model import (BSMFile, BSMFileTree) 
from budget_storage_model import (csv_DATA_LIST_url_copy, bsm_BDMWorkbook_copy)
from budman_command_services import *
from .budman_cp_namespace import * 
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_transfer() function
def WORKFLOW_CMD_transfer(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_TRANSFER_subcmd: Transfer data files and workflow workbooks
    between between workflow folders.

    Tasks required to transfer files to a workflow for a specified purpose.
    This is how raw input files from a file_list or workbooks from a work_book
    list are transformed into workbooks in the model. Processing requirements 
    vary based on the specific workflow, purpose and files and workbook types. 

    Arguments:
        cmd (p3m.Command): The Command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            CK_CMDLINE_FI_KEY - financial institution key for the file or 
            workbook being transferred.
            CK_TRANSFER_FILES - bool, if True, transfer files from a file_list.
            CK_TRANSFER_WORKBOOKS - bool, if True, transfer workbooks from a workbook_list.
            CK_FILE_LIST - list of src file_index values from the file_list
            CK_SRC_WF_KEY - wf_folder wf_key of destination.
            CK_SRC_WF_PURPOSE - wf_folder wf_purpose of destination.
            CK_WB_TYPE - workbook type of destination.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing. 
        The result_content upon succss is a List[int] of the new wb_index values 
        for the workbooks created from the transfer operation.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.

    """
    prev_fi_key: str | None = None
    model: BudgetDomainModel | None = None
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_WORKFLOW_TRANSFER_SUBCMD_KEY
        )
        transfer_files: bool = cmd_args.get(CK_TRANSFER_FILES, False)
        transfer_workbooks: bool = cmd_args.get(CK_TRANSFER_WORKBOOKS, False)
        if transfer_workbooks and not transfer_files:
            # This cmd is for transferring files, not workbooks.
            return WORKFLOW_CMD_transfer_workbooks(cmd, bdm_DC, level=level)
        if not transfer_workbooks and transfer_files:
            # This cmd is for transferring files, not workbooks.
            return WORKFLOW_CMD_transfer_files(cmd, bdm_DC, level=level)
        
        # Some invalid combination of parameters.
        msg = (f"{pad(level)}Unsupported combination of input parameters: "
                f"'{CK_TRANSFER_FILES}:{transfer_files}' "
                f" '{CK_TRANSFER_WORKBOOKS}:{transfer_workbooks}'")
        p3m.cp_user_error_message(msg)
        return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
    except p3m.CMDValidationException as e:
        logger.error(e.msg)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_transfer: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.cp_CMD_RESULT_ERROR_create(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
    finally:
        # Restore previous DC Values if they were modified.
        if (prev_fi_key is not None and 
            model is not None and
            model.bdm_FI_KEY_validate(prev_fi_key)):
            bdm_DC.dc_FI_KEY = prev_fi_key
    
#endregion WORKFLOW_CMD_transfer() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_transfer_workbooks() function
def WORKFLOW_CMD_transfer_workbooks(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_TRANSFER_subcmd: Transfer data workbooks between workflows.

    Tasks required to transfer workbooks to a workflow for a specified purpose.
    Processing requirements vary based on the specific workflow, purpose and 
    workbook types.

    Arguments:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            CK_WB_LIST - list of src wb_index values from the DC workbook_data_collection
            CK_WF_KEY - wf_folder wf_key
            CK_WF_PURPOSE - wf_folder wf_purpose
            CK_WB_TYPE - dst workflow workbook type

    Workbooks are files are in wf_folders. A list of wb_index values references 
    workbooks to transfer from a source (src) workflow folder to a destination (dst) 
    wf_folder. Depending on the dst wf_key and wf_purpose, naming conventions 
    applied to the workbooks being transferred. The dst wb_type also impacts the 
    dst workbook naming and possible conversion, or transformation tasks 
    applied during the transfer operation.

    At present, the workflow processes are fixed: intake, categorize_transactions,
    and budget. Each workflow may have up to three wf_folders configured for
    the wf_purpose values of: wf_input, wf_working, and wf_output. Also, a 
    wf_prefix value may be applied to workbooks arriving anew in a wf_folder.

    Hence, a dst full_file_name could have the following structure:
        <wf_prefix><filename><wb_type>.<extension>

    The src_filename is parsed for the four components, wf_prefix may be None and
    wb_type may be None for a file that is not yet a workbook.

    The dst_filename is constructed considering the dst_wf_prefix and dst_wb_type.
    The extension is determined by the wb_type.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.
    """
    try:
        #region Initialization and validation
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_transfer_workbooks.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_WORKFLOW_TRANSFER_SUBCMD_KEY
        )
        # Initializations
        transfer_files: bool = cmd_args.get(CK_TRANSFER_FILES, False)
        transfer_workbooks: bool = cmd_args.get(CK_TRANSFER_WORKBOOKS, False)
        if not transfer_workbooks or transfer_files:
            msg = (f"{pad(level)}Invalid combination: "
                    f"'{CK_TRANSFER_FILES}:{transfer_files}' "
                    f"'{CK_TRANSFER_WORKBOOKS}:{transfer_workbooks}'")
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        use_symlink: bool = cmd_args.get(CK_SYMLINK, False)
        model: BudgetDomainModel = bdm_DC.model
        valid_prefixes: List[str] = model.bdm_valid_WF_PREFIX_values()
        result_new_wb_index_list: List[int] = []
        success: bool = False
        result: Any = None
        src_wb: BDMWorkbook = None
        dst_wb: BDMWorkbook = None
        msg: str = ""
        src_wb_type: str = ''
        wb_index: int = -1
        # Extract and validate required parameters from the command.
        # The workbook list is from the source fi_key, so process it first
        # before the CK_CMDLINE_FI_KEY parameter.
        src_fi_key: str = bdm_DC.dc_FI_KEY
        selected_bdm_wb_list : List[BDMWorkbook] = None
        selected_bdm_wb_list = process_selected_workbook_input(cmd, bdm_DC)
        # Process the CK_CMDLINE_FI_KEY parameter and other destination related
        # parameters after the workbook list is processed since the workbook 
        # list is based on the source fi_key, and the destination fi_key may be different.
        dst_fi_key: str = cmd_args.get(CK_CMDLINE_FI_KEY, None)
        if (p3u.str_empty(dst_fi_key) or not model.bdm_FI_KEY_validate(dst_fi_key)):
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, f"Invalid fi_key: '{dst_fi_key}'")
        dst_wf_key = cmd_args.get(CK_CMDLINE_WF_KEY)
        if (p3u.str_empty(dst_wf_key) or not model.bdm_WF_KEY_validate(dst_wf_key)):
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, f"Invalid wf_key: '{dst_wf_key}'")
        dst_wf_purpose = cmd_args.get(CK_CMDLINE_WF_PURPOSE)
        if (p3u.str_empty(dst_wf_purpose) or not model.bdm_WF_PURPOSE_validate(dst_wf_purpose)):
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, f"Invalid wf_purpose: '{dst_wf_purpose}'")
        dst_wb_type = cmd_args.get(CK_WB_TYPE)
        if (p3u.str_empty(dst_wb_type) or dst_wb_type not in bdm.VALID_WB_TYPE_VALUES):
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, f"Invalid wb_type: '{dst_wb_type}'")
        #endregion Initialization and validation

        # Process the intended workbooks.
        for src_wb in selected_bdm_wb_list:
            # Select the current workbook in the Data Context.
            bdm_DC.dc_WORKBOOK = src_wb
            wb_index: int = bdm_DC.dc_WB_INDEX    # src fi_key
            bdm_wb_abs_path = src_wb.abs_path()
            msg = f"{pad(level)}{P2}workbook: {str(bdm_DC.dc_WB_INDEX):>4} '{bdm_DC.dc_WB_ID:<40}'"
            p3m.cp_user_info_message(msg)
            bdm_wb_abs_path = src_wb.abs_path()
            if bdm_wb_abs_path is None:
                m = f"{pad(level)}Workbook path is not valid: {src_wb.wb_url}"
                logger.error(m)
                p3m.cp_user_error_message(m)
                continue
            # Transfer cmd needs loaded workbooks
            if not src_wb.wb_loaded:
                # Load the workbook content if it is not loaded.
                success, result = bdm_DC.dc_WORKBOOK_content_get(src_wb)
                if not success:
                    selected_bdm_wb_list.remove(src_wb)
                    msg = f"{pad(level)}Excluded workbook: '{src_wb.wb_id}', "
                    msg += f"failed to load: {result}"
                    logger.error(msg)
                    p3m.cp_user_error_message(msg)
                    continue
            src_wb_type = src_wb.wb_type
            # Supportted transfer cases.
            # ---------------------------------------------------------------- +
            #region wb_type bdm.WB_TYPE_CSV_TXNS transfer
            if src_wb_type == bdm.WB_TYPE_CSV_TXNS and dst_wb_type == bdm.WB_TYPE_EXCEL_TXNS:
                # Transfer a csv_txns workbook by converting to a excel_txns workbook.
                # Create a file_url for the new workbook being transferred.
                src_wb_bsm_file: BSMFile = BSMFile(
                    file_url=src_wb.wb_url,
                    valid_prefixes=valid_prefixes,
                    valid_wb_types=bdm.VALID_WB_TYPE_VALUES
                )
                success, result = WORKFLOW_TASK_construct_bdm_workbook(
                    src_filename=src_wb_bsm_file.filename,
                    wb_type=dst_wb_type,
                    fi_key=dst_fi_key,
                    wf_key=dst_wf_key,
                    wf_purpose=dst_wf_purpose,
                    bdm_DC=bdm_DC
                )
                if not success:
                    msg = (f"{pad(level)}Failed to construct file URL for workbook: "
                            f"'{wb_index:2}:{src_wb.wb_name}' "
                            f" Error: {result}")
                    logger.error(msg)
                    p3m.cp_user_error_message(msg)
                    continue
                dst_wb = result
                WORKFLOW_TASK_convert_csv_txns_to_excel_txns(src_wb, dst_wb)
                # Add the new workbook to the wdc.
                wdc = bdm_DC.dc_WORKBOOK_DATA_COLLECTION
                wb_id = dst_wb.wb_id
                wdc[wb_id] = dst_wb
                wb_index = bdm_DC.dc_WORKBOOK_index(wb_id)
                result_new_wb_index_list.append(wb_index)
                msg = (f"{pad(level)}Added new workbook: '{wb_index:03}:{dst_wb.wb_name}' ")
                p3m.cp_user_info_message(msg)
            #endregion wb_type bdm.WB_TYPE_CSV_TXNS transfer
            # ---------------------------------------------------------------- +
            #region wb_type bdm.WB_TYPE_EXCEL_TXNS transfer
            elif src_wb_type == bdm.WB_TYPE_EXCEL_TXNS and dst_wb_type == bdm.WB_TYPE_EXCEL_TXNS:
                # Transfer a excel_txns workbook to another workflow wf_folder.
                # Create a file_url for the new workbook file being transferred.
                src_wb_bsm_file: BSMFile = BSMFile(
                    file_url=src_wb.wb_url,
                    valid_prefixes=valid_prefixes,
                    valid_wb_types=bdm.VALID_WB_TYPE_VALUES
                )
                # Task: Create a BDMWorkbook for the dst workbook.
                success, result = WORKFLOW_TASK_construct_bdm_workbook(
                    src_filename=src_wb_bsm_file.filename,
                    wb_type=dst_wb_type,
                    fi_key=dst_fi_key,
                    wf_key=dst_wf_key,
                    wf_purpose=dst_wf_purpose,
                    bdm_DC=bdm_DC
                )
                if not success:
                    msg = (f"Failed to create workbook: "
                            f"'{wb_index:2}:{src_wb.wb_name}' "
                            f" Error: {result}")
                    logger.error(msg)
                    p3m.cp_user_error_message(msg)
                    continue
                dst_wb = result
                if dst_wb.check_wb_url():
                    msg = (f"{pad(level)}Destination Workbook already exists for URL: "
                            f"'{dst_wb.wb_url}'")
                    p3m.cp_user_warning_message(msg)
                    msg = (f"{pad(level)}The Workbook will be overwritten.")
                    p3m.cp_user_warning_message(msg)
                # Task: Copy the workbook content from the src workbook to the dst workbook.
                bsm_BDMWorkbook_copy(src_wb, dst_wb, use_symlink)
                # Add the new workbook to the wdc.
                wdc = bdm_DC.dc_WORKBOOK_DATA_COLLECTION
                wb_id = dst_wb.wb_id
                wdc[wb_id] = dst_wb
                wb_index = bdm_DC.dc_WORKBOOK_index(wb_id)
                result_new_wb_index_list.append(wb_index)
                msg = (f"{pad(level)}Added new workbook: '{wb_index:03}:{dst_wb.wb_name}' ")
                p3m.cp_user_info_message(msg)
            #endregion wb_type bdm.WB_TYPE_EXCEL_TXNS transfer
            # ---------------------------------------------------------------- +
            else:
                msg = (f"Unsupported transfer from src wb_type '{src_wb_type}' "
                     f"to dst wb_type '{dst_wb_type}'")
                logger.error(msg)
                p3m.cp_user_error_message(msg)
                continue
        model.bdm_save_model()
        model.bdm_refresh_trees()
        p3m.cp_user_info_message(m + "End: ...")
        # End: --------------------------------------------------------------- +
        return p3m.cp_CMD_RESULT_create(
            cmd=cmd,
            status=True,
            type=CV_CMD_LIST_OUTPUT,
            content=result_new_wb_index_list
        )
    except p3m.CMDValidationException as e:
        logger.error(e.msg)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_transfer: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.cp_CMD_RESULT_ERROR_create(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
#endregion WORKFLOW_CMD_transfer_workbooks() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_transfer_files() function
def WORKFLOW_CMD_transfer_files(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_TRANSFER_subcmd: Transfer data files into workflow workbooks.

    Tasks required to transfer files to a workflow for a specified purpose.
    This is how raw input files from a file_list are transformed into workbooks 
    in the model. Processing requirements vary based on the specific workflow, 
    purpose and workbook types. This task uses file_indexes to indetify input
    files. Files are not yet workbooks and have no wb_index, just a file_index.

    Arguments:
        cmd (p3m.Command): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            CK_CMDLINE_FI_KEY (str): dst_fi_key for the dst workbook.
            CK_TRANSFER_FILES (bool): True to transfer files from a file_list.
            CK_TRANSFER_WORKBOOKS (bool): False always.
            CK_SYMLINK (bool): False, N/A here.
            CK_FILE_LIST (list): list of src file_index values from the file_list
            CK_SRC_WF_KEY (str): wf_folder wf_key of destination.
            CK_SRC_WF_PURPOSE (str): wf_folder wf_purpose of destination.
            CK_WB_TYPE (str): workbook type of destination.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing. 
        The result_content upon success is a List[int] of the new wb_index values 
        for the workbooks created from the transfer operation.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.

    Notes:
    ------
    
    Files are in wf_folders. The scope of the file_index values is the entire
    BDM, across multiple FIs. A list of file_index values references files to 
    transfer from their current (src) workflow folder to specified a destination
    (dst) wf_folder. Depending on the dst wf_key and wf_purpose, naming 
    conventions are applied to the files being transferred. The dst wb_type also
    impacts the dst filename and possible conversion, or transformation tasks 
    applied during the transfer operation.

    At present, the workflow processes are fixed: intake, categorize_transactions,
    and budget. Each workflow may have up to three wf_folders configured for
    the wf_purpose values of: wf_input, wf_working, and wf_output. Also, a 
    wf_prefix value may be applied to files arriving anew in a wf_folder.

    Hence, a dst full_file_name could have the following structure:
        <wf_prefix><filename><wb_type>.<extension>

    The src_filename is parsed for the four components, wf_prefix may be None and
    wb_type may be None for a file that is not yet a workbook.

    The dst_filename is constructed considering the dst_wf_prefix and dst_wb_type.
    The extension is determined by the wb_type.

    This cmd is how initial transaction files are orchestrated into the process.
    Based on the dst_fi_key, an INTAKE_TASK is used to transform different schemas 
    for incoming data files into common formats used within BudMan workflows. 
    .csv files downloaded from different financial institutions have different 
    schemas, column headings, mapped to the common workbook schema.

     This cmd does not modify the content of the input .csv file. It will 
     contain all of the columns from the original download. Each financial 
     institution has its own schema. Modifying this schema occurs in 
     INTAKE_TASKs.
    """
    try:
        #region Initialization and validation
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_transfer_files.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_WORKFLOW_TRANSFER_SUBCMD_KEY
        )
        # Initialization
        transfer_files: bool = cmd_args.get(CK_TRANSFER_FILES, False)
        transfer_workbooks: bool = cmd_args.get(CK_TRANSFER_WORKBOOKS, False)
        if not transfer_files or transfer_workbooks:
            msg = (f"{pad(level)}Invalid combination: "
                    f"'{CK_TRANSFER_FILES}:{transfer_files}' "
                    f"'{CK_TRANSFER_WORKBOOKS}:{transfer_workbooks}'")
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        model: BudgetDomainModel = bdm_DC.model
        bsm_file_tree : BSMFileTree = model.bsm_file_tree
        result_new_wb_index_list: List[int] = []
        msg: str = ""
        success: bool = False
        result: str = ""
        src_bsm_file: BSMFile = None
        dst_csv_wb: BDMWorkbook = None
        dst_wb_index: int = -1
        # Extract and validate required parameters from the command.
        # The source files for a transfer files subcmd is the bsm_file_tree 
        # which is scoped to the entire BDM and BSM, spanning all FIs.
        # For now, just retain the src_fi_key which is the current
        # fi_key in the DC.
        src_fi_key: str = bdm_DC.dc_FI_KEY
        # Process the transfer destination related
        # parameters.
        dst_fi_key: str = cmd_args.get(CK_CMDLINE_FI_KEY, None)
        if (p3u.str_empty(dst_fi_key) or not model.bdm_FI_KEY_validate(dst_fi_key)):
            msg = f"{pad(level)}Invalid destination fi_key: '{dst_fi_key}'"
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        dst_wf_key : str = cmd_args.get(CK_CMDLINE_WF_KEY)
        if (p3u.str_empty(dst_wf_key) or not model.bdm_WF_KEY_validate(dst_wf_key)):
            msg = f"{pad(level)}Invalid destination wf_key: '{dst_wf_key}'"
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        dst_wf_purpose: str = cmd_args.get(CK_CMDLINE_WF_PURPOSE)
        if (p3u.str_empty(dst_wf_purpose) or not model.bdm_WF_PURPOSE_validate(dst_wf_purpose)):
            msg = f"{pad(level)}Invalid destination wf_purpose: '{dst_wf_purpose}'"
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        dst_wb_type: str = cmd_args.get(CK_WB_TYPE)
        if (p3u.str_empty(dst_wb_type) or dst_wb_type not in bdm.VALID_WB_TYPE_VALUES):
            msg = f"{pad(level)}Invalid destination wb_type: '{dst_wb_type}'"
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        # Can be 1 or a list of file_indexes provided.
        src_file_index_list : List[int] = cmd_args.get(CK_FILE_LIST)
        error_file_index_list : List[int] = []
        src_bsm_files: List[BSMFile] = bsm_file_tree.validate_file_list(src_file_index_list)
        #endregion Initialization and validation

        # Supported cases:
        # 1. Transfer .csv files from file_list to a .csv_txns workbooks.
        #    These input .csv files are assumed to be raw_input from a
        #    financial institution with its own schema. The transfer process will
        #    not modify the content of the .csv file, but will create a .csv_txns
        #    workbook using naming conventions for the new wowrkbook files.
        #    
        for src_bsm_file in src_bsm_files:
            # Process for supported transfer dst wb_types.
            if dst_wb_type != bdm.WB_TYPE_CSV_TXNS:
                # Unsupported dst wb_type for transfer.
                msg = (f"{pad(level)}Unsupported destination workbook type for transfer: {dst_wb_type}")
                p3m.cp_user_warning_message(msg)
                error_file_index_list.append(src_bsm_file.file_index)
                continue
            # Dest wb_type is csv_txns workbook.
            # Input file must have .csv extension.
            if src_bsm_file.extension != bdm.WB_FILETYPE_CSV:
                # Unsupported file type for transfer.
                msg = (f"{pad(level)}Unsupported source file type file "
                        f"'{src_bsm_file.file_index:2}:{src_bsm_file.full_filename}'"
                        f" must be .csv file.")
                p3m.cp_user_error_message(msg)
                error_file_index_list.append(src_bsm_file.file_index)
                continue
            # Task: Create a BDMWorkbook for the new file being transferred.
            success, result = WORKFLOW_TASK_construct_bdm_workbook(
                src_filename=src_bsm_file.filename,
                wb_type=dst_wb_type,
                fi_key=dst_fi_key,
                wf_key=dst_wf_key,
                wf_purpose=dst_wf_purpose,
                bdm_DC=bdm_DC
            )
            if not success:
                msg = (f"{pad(level)}Failed to construct file URL for file: "
                        f"'{src_bsm_file.file_index:2}:{src_bsm_file.full_filename}'"
                        f" Error: {result}")
                p3m.cp_user_error_message(msg)
                error_file_index_list.append(src_bsm_file.file_index)
                continue
            dst_csv_wb = result
            if dst_csv_wb.check_wb_url():
                msg = (f"{pad(level)}Destination Workbook already exists for URL: "
                        f"'{dst_csv_wb.wb_url}'")
                p3m.cp_user_warning_message(msg)
                msg = (f"{pad(level)}The Workbook will be overwritten.")
                p3m.cp_user_warning_message(msg)
            # Transfer a WB_FILETYPE_CSV(.csv) file content to a 
            # WB_TYPE_CSV_TXNS(.csv) workbook, a new file in the dst workflow/purpose folder.
            success, result = WORKFLOW_TASK_transfer_csv_file_to_workbook(
                src_file_url=src_bsm_file._file_url,
                dst_wb=dst_csv_wb,
                wb_type=dst_wb_type
            )
            if not success:
                msg = (f"{pad(level)}Failed to transfer file: "
                        f"'{src_bsm_file.file_index:2}:{src_bsm_file.full_filename}' "
                        f" to .csv_txns workbook. Error: {result}")
                p3m.cp_user_warning_message(msg)
                error_file_index_list.append(src_bsm_file.file_index)
                continue

            # Add the new workbook to the wdc.
            bdm_DC.dc_WORKBOOK_DATA_COLLECTION_add(dst_csv_wb)
            dst_wb_index = bdm_DC.dc_WORKBOOK_index(dst_csv_wb.wb_id)
            result_new_wb_index_list.append(dst_wb_index)

            # Validate the schema from the header row etc. depending on the 
            # dst_csv_wb dst_fi_key. 
            # This is important to do before adding the new workbook to the DC 
            # and model, because if the schema is not correct, then the later 
            # steps of the workflow will likely fail and it is better to catch 
            # this early and not add a workbook that will just cause problems later.
            # Adjust the incoming .csv file schema to the workbook standard.
            success, result = INTAKE_TASK_convert_csv_txns_schema(dst_csv_wb, bdm_DC)
            if not success:
                msg = (f"{pad(level)}Failed to convert .csv file schema for workbook: "
                        f"'{dst_csv_wb.wb_id}'. Error: {result}")
                p3m.cp_user_warning_message(msg)
                error_file_index_list.append(src_bsm_file.file_index)
                continue
            # Successful completion
            msg = (f"{pad(level)}Added new workbook: '{dst_wb_index:03}:{dst_csv_wb.wb_name}' ")
            p3m.cp_user_info_message(msg)
        model.bdm_save_model()
        model.bdm_refresh_trees()
        # End: --------------------------------------------------------------- +
        p3m.cp_user_info_message(m + "End: ...")
        return p3m.cp_CMD_RESULT_create(
            status=True,
            type=CV_CMD_LIST_OUTPUT,
            content=result_new_wb_index_list,
            cmd=cmd
        )
    except p3m.CMDValidationException as e:
        logger.error(e.msg)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_transfer: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.cp_CMD_RESULT_ERROR_create(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
    
#endregion WORKFLOW_CMD_transfer_files() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_process() function
def WORKFLOW_CMD_process(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_PROCESS_subcmd: Execute a workflow process for input files or workbooks.

    NOTE: The process is hard coded for now as: 
    "intake" -> "categorize_transactions" -> "budget"
    The process can be made more dynamic in the future, but for now,
    
    Process new .csv file all the way through the
    workflow process applying CMD and TASK functions as necessary.

    The process must transfer input .csv files into .csv_txns workbooks, 
    then to .excel_txns workbooks, and then to categorize the transactions in 
    the .excel_txns workbook. The "budget" phase in not yet implemented.
    
    Arguments:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            CK_FILE_LIST - list of src file_index values.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.
    """
    try:
        #region Initialization and validation
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_process.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_PROCESS_SUBCMD_KEY
        )
        # Initializations
        msg: str = ""
        what_if: bool = cmd_args.get(CK_WHAT_IF, False)
        wif: str = "[yellow]What-If Mode[/yellow]" if what_if else ""
        wb_index_list: List[int] = []
        result : p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_create(
            status=True,
            type=CV_CMD_STRING_OUTPUT,
            content=wif,
            cmd=cmd
        )
        fi_key: str = cmd_args.get(CK_CMDLINE_FI_KEY, None)
        prev_fi_key: str = bdm_DC.dc_FI_KEY
        bdm_DC.dc_FI_KEY = fi_key if fi_key else prev_fi_key
        #endregion Initialization and validation

        #region Command 1: WORKFLOW_CMD_transfer_files()
        #    Transfer WB_FILETYPE_CSV(.csv) files from file_list to a 
        #    WB_TYPE_CSV_TXNS(.csv) workbooks.
        # Build a CMD object for WORKFLOW_CMD_transfer_files()
        new_cmd : p3m.CMD_OBJECT_TYPE = p3m.cp_CMD_OBJECT_create(
            cmd_name=CV_WORKFLOW_CMD_NAME,
            cmd_key=CV_WORKFLOW_CMD_KEY,
            subcmd_name=CV_TRANSFER_SUBCMD_NAME,
            subcmd_key=CV_WORKFLOW_TRANSFER_SUBCMD_KEY)
        new_cmd[CK_TRANSFER_FILES] = True
        new_cmd[CK_TRANSFER_WORKBOOKS] = False
        new_cmd[CK_FILE_LIST] = cmd_args.get(CK_FILE_LIST)
        new_cmd[CK_CMDLINE_FI_KEY] = fi_key
        new_cmd[CK_WF_KEY] = "intake"
        new_cmd[CK_WF_PURPOSE] =  bdm.WF_WORKING
        new_cmd[CK_WB_TYPE] = bdm.WB_TYPE_CSV_TXNS
        # Run CMD: WORKFLOW_CMD_transfer_files() and capture the result.
        if what_if:
            msg = f"{pad(level)}[yellow]What-If:[/yellow] Executing: {str(new_cmd)}"
            p3m.cp_user_info_message(msg)
        else:
            result = WORKFLOW_CMD_transfer_files(new_cmd, bdm_DC, level) 
            if not result[p3m.CK_CMD_RESULT_STATUS]:
                m = f"Failed to transfer files to {bdm.WB_TYPE_CSV_TXNS} workbooks."
                logger.error(m)
                return p3m.cp_CMD_RESULT_ERROR_create(cmd, m)
            wb_index_list = result[p3m.CK_CMD_RESULT_CONTENT]
        #endregion Command 1: WORKFLOW_CMD_transfer_files()

        # Check/modify schema of the intake/new .csv_txns workbook.

        #region Command 2: WORKFLOW_CMD_transfer_workbooks()
        #    Transfer WB_TYPE_CSV_TXNS(.csv) workbook from wb_index_list to a 
        #    WB_TYPE_EXCEL_TXNS(.xlsx) workbooks.
        # Build a CMD object for WORKFLOW_CMD_transfer_workbooks()
        new_cmd = None
        new_cmd : p3m.CMD_OBJECT_TYPE = p3m.cp_CMD_OBJECT_create(
            cmd_name=CV_WORKFLOW_CMD_NAME,
            cmd_key=CV_WORKFLOW_CMD_KEY,
            subcmd_name=CV_TRANSFER_SUBCMD_NAME,
            subcmd_key=CV_WORKFLOW_TRANSFER_SUBCMD_KEY)
        new_cmd[CK_TRANSFER_FILES] = False
        new_cmd[CK_TRANSFER_WORKBOOKS] = True
        new_cmd[CK_WB_LIST] = wb_index_list
        new_cmd[CK_WF_KEY] = "categorize_transactions"
        new_cmd[CK_WF_PURPOSE] =  bdm.WF_WORKING
        new_cmd[CK_WB_TYPE] = bdm.WB_TYPE_EXCEL_TXNS
        # Run CMD: WORKFLOW_CMD_transfer_workbooks() and capture the result.
        if what_if:
            msg = f"{pad(level)}[yellow]What-If:[/yellow] Executing: {str(new_cmd)}"
            p3m.cp_user_info_message(msg)
        else:
            result = WORKFLOW_CMD_transfer_workbooks(new_cmd, bdm_DC, level) 
            if not result[p3m.CK_CMD_RESULT_STATUS]:
                msg  = f"Failed to transfer {bdm.WB_TYPE_CSV_TXNS} workbooks to {bdm.WB_TYPE_EXCEL_TXNS} workbooks."
                logger.error(msg)
                return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
            wb_index_list = result[p3m.CK_CMD_RESULT_CONTENT]
        #endregion Command 2: WORKFLOW_CMD_transfer_workbooks()

        #region Command 3: WORKFLOW_TASK_check_workbooks()
        # Check the new WB_TYPE_EXCEL_TXNS workbooks with to ensure they are 
        # ready for the next step in the workflow process.
        new_cmd = None
        new_cmd : p3m.CMD_OBJECT_TYPE = p3m.cp_CMD_OBJECT_create(
            cmd_name=CV_WORKFLOW_CMD_NAME,
            cmd_key=CV_WORKFLOW_CMD_KEY,
            subcmd_name=CV_CHECK_SUBCMD_NAME,
            subcmd_key=CV_CHECK_SUBCMD_KEY)
        new_cmd[CK_LOAD_WORKBOOK_SWITCH] = True
        new_cmd[CK_FIX_SWITCH] = True
        new_cmd[CK_VALIDATE_CATEGORIES] = False
        new_cmd[CK_WB_LIST] = wb_index_list
        # Run CMD: WORKFLOW_CMD_check_workbooks() and capture the result.
        if what_if:
            m = f"{pad(level)}[yellow]What-If:[/yellow] Executing: {str(new_cmd)}"
            p3m.cp_user_info_message(m)
        else:
            result = WORKFLOW_CMD_check_workbooks(new_cmd, bdm_DC, level) 
            if not result[p3m.CK_CMD_RESULT_STATUS]:
                msg = f"Failed to check {bdm.WB_TYPE_EXCEL_TXNS} workbooks."
                logger.error(msg)
                return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        #endregion Command 3: Check the new WB_TYPE_EXCEL_TXNS workbooks.

        #region Command 4: WORKFLOW_TASK_categorize_transactions()
        #    Categorize the transactions in the new WB_TYPE_EXCEL_TXNS workbooks.
        new_cmd = None
        new_cmd : p3m.CMD_OBJECT_TYPE = p3m.cp_CMD_OBJECT_create(
            cmd_name=CV_WORKFLOW_CMD_NAME,
            cmd_key=CV_WORKFLOW_CMD_KEY,
            subcmd_name=CV_CATEGORIZATION_SUBCMD_NAME,
            subcmd_key=CV_CATEGORIZATION_SUBCMD_KEY)
        new_cmd[CK_LOAD_WORKBOOK_SWITCH] = True
        new_cmd[CK_LOG_ALL] = True
        new_cmd[CK_CLEAR_OTHER] = True
        new_cmd[CK_WB_LIST] = wb_index_list
        # Run CMD: WORKFLOW_TASK_categorize_transactions() and capture the result.
        if what_if:
            m = f"{pad(level)}[yellow]What-If:[/yellow] Executing: {str(new_cmd)}"
            p3m.cp_user_info_message(m)
        else:
            result = WORKFLOW_CMD_categorize_transactions(new_cmd, bdm_DC, level) 
            if not result[p3m.CK_CMD_RESULT_STATUS]:
                msg = f"Failed to categorize transactions in {bdm.WB_TYPE_EXCEL_TXNS} workbooks."
                logger.error(msg)
                return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        #endregion Command 4: WORKFLOW_TASK_categorize_transactions()

        level -= 1
        p3m.cp_user_info_message(m + "End: ...")
        # End: --------------------------------------------------------------- +
        return p3m.cp_CMD_RESULT_create(
            status=True,
            type=CV_CMD_DICT_OUTPUT,
            content=result,
            cmd=cmd
        )
    except p3m.CMDValidationException as e:
        logger.error(e.msg)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_CMD_process: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.cp_CMD_RESULT_ERROR_create(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
    finally:
        # Restore previous DC Values if they were modified.
        bdm_DC.dc_FI_KEY = prev_fi_key
#endregion WORKFLOW_CMD_process() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_categorize_transactions() execution method
def WORKFLOW_CMD_categorize_transactions(
        cmd: p3m.Command,
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0) -> p3m.CMD_RESULT_TYPE:
    """Apply categorization workflow to one or more WORKBOOKS in the DC.

    As a workflow process, the WORKFLOW_categorization_cmd method has the
    job to marshall the necessary input objects, implied by the command
    arguments and perhaps Data Context, and then invoke the appropriate 
    function or method to conduct the requested process, tasks, etc. It also
    catches the return status and result to return.

    Arguments:
        cmd (Dict): A valid BudMan View Model Command object. 

    Required cmd object attributes:
        cmd_key: 'workflow_cmd' 
    Optional cmd object attributes:
        cmd_name: CV_WORKFLOW_CMD
        Valid subcommands:
            subcmd_key: 'workflow_cmd_categorization'
            subcmd_name: CV_DELETE_SUBCMD
        subcmd_key: 'app_cmd_reload'
        subcmd_name: CV_RELOAD_SUBCMD
        subcmd_key: 'app_cmd_log'
        subcmd_name: CV_LOG_SUBCMD

    Returns:
        Tuple[success : bool, result : Any]: The outcome of the command 
        execution. If success is True, result contains result of the 
        command, if False, a description of the error.
        
    Raises:
        RuntimeError: A description of the
        root error is contained in the exception message.
    """
    try:
        #region Initialization and validation
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts}{WORKFLOW_CMD_categorize_transactions.__name__}()"
        p3m.cp_user_info_message(m + " Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_CATEGORIZATION_SUBCMD_KEY)
        # Extract and validate required parameters from the command.
        success: bool = False
        msg: str = ""
        model: BudgetDomainModel = bdm_DC.model
        selected_bdm_wb_list : List[BDMWorkbook] = None
        selected_bdm_wb_list = process_selected_workbook_input(
            cmd, 
            bdm_DC,
            validate_url=True)
        # Process the selected workbooks.
        if len(selected_bdm_wb_list) == 0:
            msg = f"{pad(level)}No workbooks selected to check."
            p3m.cp_user_warning_message(msg)
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        elif len(selected_bdm_wb_list) > 1:
            msg = f"{pad(level)}'{len(selected_bdm_wb_list)}' workbooks selected to check."
            p3m.cp_user_info_message(msg)
        else :
            msg = f"{pad(level)}A single workbook selected to check."
            p3m.cp_user_info_message(msg)
        p3m.cp_user_info_message(f"{pad(level)}Categorizing {len(selected_bdm_wb_list)} workbooks:")
        success : bool = False
        r : str = ""
        msg : str = ""
        log_all : bool = cmd_args.get(CK_LOG_ALL, False)
        clear_other: bool = cmd_args.get(CK_CLEAR_OTHER, False)
        cleared_other_now: bool = clear_other
        #endregion Initialization and validation

        # Process the intended workbooks.
        for bdm_wb in selected_bdm_wb_list:
            # Select the current workbook in the Data Context.
            bdm_DC.dc_WORKBOOK = bdm_wb
            bdm_wb_abs_path = bdm_wb.abs_path()
            p3m.cp_user_info_message(f"{pad(level)}Workbook: {str(bdm_DC.dc_WB_INDEX):>4} '{bdm_DC.dc_WB_ID:<40}'")
            bdm_wb_abs_path = bdm_wb.abs_path()
            if bdm_wb_abs_path is None:
                msg = f"Workbook path is not valid: {bdm_wb.wb_url}"
                p3m.cp_user_error_message(f"{pad(level + 1)}Error: {msg}")
                continue
            # Check cmd needs loaded workbooks to check
            if not bdm_wb.wb_loaded:
                # Load the workbook content if it is not loaded.
                success, result = bdm_DC.dc_WORKBOOK_content_get(bdm_wb)
                if not success:
                    selected_bdm_wb_list.remove(bdm_wb)
                    msg = f"{pad(level+1)}Excluded workbook: '{bdm_wb.wb_id}', "
                    msg += f"failed to load: {result}"
                    logger.error(msg)
                    p3m.cp_user_error_message(msg)
                    continue
            # Now we have a valid bdm_wb to process.
            if bdm_wb.wb_type == bdm.WB_TYPE_EXCEL_TXNS:
                task = "process_budget_category()"
                msg = (f"{pad(level+1)}Task: {task:30} {str(bdm_DC.dc_WB_INDEX):>4} "
                        f"'{bdm_DC.dc_WB_ID:<40}'")
                p3m.cp_user_info_message(msg)
                success, r = WORKFLOW_TASK_process_budget_category(bdm_wb, bdm_DC, 
                                                        log_all, cleared_other_now)
                cleared_other_now = False # Only clear_other for first workbook
                if not success:
                    r = (f"{pad(level + 1)}Task Failed: process_budget_category() Workbook: "
                            f"'{bdm_DC.dc_WB_ID}'\n{pad(level + 2)}Result: {r}")
                    p3m.cp_user_info_message(f"{pad(level + 1)}{r}")
                    continue
                p3m.cp_user_info_message(f"{pad(level + 1)}Result: {r}")
                task = "dc_WORKBOOK_save()"
                msg = (f"{pad(level+1)}Task: {task:30} {str(bdm_DC.dc_WB_INDEX):>4} "
                        f"'{bdm_DC.dc_WB_ID:<40}'")
                p3m.cp_user_info_message(msg)
                success, r = bdm_DC.dc_WORKBOOK_save(bdm_wb)
                if not success:
                    msg = (f"{pad(level + 1)}Task Failed: {task:30} Workbook: "
                            f"'{bdm_DC.dc_WB_ID}'\n{pad(level + 2)}Result: {r}")
                    p3m.cp_user_error_message(msg)
                    continue
                p3m.cp_user_info_message(f"{pad(level + 1)}Result: {r}")
                task = "dc_WORKBOOK_close()"
                msg = (f"{pad(level+1)}Task: {task:30} {str(bdm_DC.dc_WB_INDEX):>4} "
                        f"'{bdm_DC.dc_WB_ID:<40}'")
                p3m.cp_user_info_message(msg)
                success, r = bdm_DC.dc_WORKBOOK_close(bdm_wb)
                if not success:
                    msg = (f"{pad(level + 1)}Task Failed: {task:30} Workbook: "
                            f"'{bdm_DC.dc_WB_ID}'\n{pad(level + 1)}Result: {r}")
                    p3m.cp_user_error_message(msg)
                    continue
                p3m.cp_user_info_message(f"{pad(level + 1)}Result: {r}")
        p3m.cp_user_info_message(f"{m} Complete: ...")
        return p3m.cp_CMD_RESULT_create(True, p3m.CV_CMD_STRING_OUTPUT, "Complete", cmd)
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion WORKFLOW_TASK_categorize_transactions() execution method
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_delete_workbooks() function
def WORKFLOW_CMD_delete_workbooks(
        cmd: p3m.Command,
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_DELETE_subcmd: Delete data workbooks based on wb_index."""
    try:
        #region Initialization and validation
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_delete_workbooks.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_WORKFLOW_DELETE_SUBCMD_KEY
        )
        # Initializations
        # Extract and validate required parameters from the command.
        model: BudgetDomainModel = bdm_DC.model
        fi_key: str = bdm_DC.dc_FI_KEY
        fr: str = ""
        msg: str = ""
        # no_save: bool = cmd_args.get(CK_NO_SAVE)
        selected_bdm_wb_list : List[BDMWorkbook] = None
        selected_bdm_wb_list = process_selected_workbook_input(
            cmd, 
            bdm_DC,
            validate_url=False)
        #endregion Initialization and validation

        # Process the intended workbooks.
        src_wb: BDMWorkbook = None
        for src_wb in selected_bdm_wb_list:
            # Select the current workbook in the Data Context.
            bdm_DC.dc_WORKBOOK = src_wb
            wb_index: int = bdm_DC.dc_WB_INDEX
            # Remove the workbook from the DC and model first.
            del_wb: BDMWorkbook = model.bdm_WORKBOOK_delete(src_wb)
            if del_wb is None:
                msg = f"Failed to remove workbook: '{src_wb.wb_id}'."
                p3m.cp_user_error_message(msg)
                continue
            # Successfully removed from DC, BDM and BSM.
            p3m.cp_user_info_message(f"{pad(level)}Deleted workbook: {str(wb_index):>4} "
                                     f"'{del_wb.wb_id:<40}' '{del_wb.wb_url}'")
            src_wb = del_wb= None
        # Refresh the trees
        model.bdm_FILE_TREE_refresh()
        model.bdm_WORKBOOK_TREE_refresh()
        p3m.cp_user_info_message(m + "End: ...")
        # End: --------------------------------------------------------------- +
        return p3m.cp_CMD_RESULT_create(
            cmd=cmd,
            status=True,
            type=p3m.CV_CMD_STRING_OUTPUT,
            content="Complete"
        )
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_CMD_delete_workbooks() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_check_workbooks() function
def WORKFLOW_CMD_check_workbooks(
        cmd: p3m.Command,
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_CMD_check_workbooks: Check data workbooks."""
    try:
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_check_workbooks.__name__}()"
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY, 
            expected_subcmd_key=CV_CHECK_SUBCMD_KEY
        )
        # Initializations
        msg: str = ""
        src_wb: BDMWorkbook = None
        success: bool = False
        result: str = ''
        model: BudgetDomainModel = bdm_DC.model
        fi_key: str = bdm_DC.dc_FI_KEY
        selected_bdm_wb_list : List[BDMWorkbook] = None
        selected_bdm_wb_list = process_selected_workbook_input(
            cmd, 
            bdm_DC,
            validate_url=True)
        # Process the selected workbooks.
        if len(selected_bdm_wb_list) == 0:
            msg = f"{pad(level)}No workbooks selected to check."
            p3m.cp_user_warning_message(msg)
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        elif len(selected_bdm_wb_list) > 1:
            msg = f"{pad(level)}'{len(selected_bdm_wb_list)}' workbooks selected to check."
            p3m.cp_user_info_message(msg)
        else :
            msg = f"{pad(level)}A single workbook selected to check."
            p3m.cp_user_info_message(msg)
            # TODO: Notes. Command_Services should map from Commands with 
            # arguments to other services APIs. The APIs should be based on
            # BDMWorkbooks or BSMFiles with content, not exposing excel 
            # particulars like here. Clean up the API surface design in
            # budman_workflow_services package modules. When a file or workbook
            # are modified, the API function should save it by default.
        for src_wb in selected_bdm_wb_list:
            # Select the current workbook in the Data Context.
            bdm_DC.dc_WORKBOOK = src_wb
            bdm_wb_abs_path = src_wb.abs_path()
            msg = f"{pad(level)}workbook: {str(bdm_DC.dc_WB_INDEX):>4} '{src_wb.wb_id:<40}'"
            p3m.cp_user_info_message(msg)
            # Check cmd needs loaded workbooks to check
            if not src_wb.wb_loaded:
                msg = f"{pad(level)}wb_name '{src_wb.wb_name}' is not loaded, no action taken."
                p3m.cp_user_error_message(msg)
                continue
            # By default, check the sheet schema. But other cli switches
            # can added to check something else.
            if cmd.cmd_parms[CK_REMOVE_EXTRA_COLUMNS]:
                # TODO: Should be calling WORKFLOW_TASK_remove_extra_columns() 
                # instead with src_wb instead of worksheet
                task = "excel_WORKSHEET_remove_extra_columns()"
                ws = src_wb.wb_content.active
                success, result = excel_WORKSHEET_remove_extra_columns(ws,BUDMAN_TXNS_WORKBOOK_COL_NAMES)
                msg = (f"{pad(level)}Task: {task:40} {str(bdm_DC.dc_WB_INDEX):>4} "
                       f"'{src_wb.wb_id:<40}' - {'Success' if success else 'Failed'} - "
                       f"Result: {result}")
                p3m.cp_user_info_message(msg)
                if success: 
                    src_wb.wb_content.save(bdm_wb_abs_path)
            if cmd.cmd_parms[CK_VALIDATE_CATEGORIES]:
                # Validate the categories in the workbook.
                task = "validate_budget_categories()"
                msg = (f"{pad(level)}Task: {task:40} {str(bdm_DC.dc_WB_INDEX):>4} "
                      f"'{src_wb.wb_id:<40}' - {'Success' if success else 'Failed'} - "
                      f"Result: {result}")
                success, result = validate_budget_categories(src_wb, bdm_DC, P4)
                p3m.cp_user_debug_message(msg)
                continue
            success = check_sheet_schema(src_wb.wb_content)
            task = "check_sheet_schema()"
            msg = (f"{pad(level)}Task: {task:40} {str(bdm_DC.dc_WB_INDEX):>4} "
                    f"'{src_wb.wb_id:<40}' - {'Success' if success else 'Failed'}")
            p3m.cp_user_info_message(msg)
            if success:
                continue
            if cmd.cmd_parms[CK_FIX_SWITCH]:
                task = "check_sheet_schema()"
                ws = src_wb.wb_content.active
                success = WORKFLOW_TASK_check_sheet_columns(ws, add_columns=True)
                msg = (f"{pad(level)}Task: {task:40} {str(bdm_DC.dc_WB_INDEX):>4} "
                        f"'{src_wb.wb_id:<40}' --fix_switch - {'Success' if success else 'Failed'}")
                p3m.cp_user_info_message(msg)
                if success: 
                    src_wb.wb_content.save(bdm_wb_abs_path)
            continue
        # End: --------------------------------------------------------------- +
        p3m.cp_user_info_message(m + "End: ...")
        return p3m.cp_CMD_RESULT_create(success, p3m.CV_CMD_STRING_OUTPUT, 
                                            f"{P2}Complete", cmd)
    except Exception as e:
        p3m.cp_user_error_message(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_CMD_check_workbooks() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_update_catalog_map() function
def WORKFLOW_CMD_update_catalog_map(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_CMD_update_catalog_map: Update the catalog map workbook for
       a specified fi_key.

    Tasks required to update the catalog map workbook for an fi_key in the
    catalog manager.

    Arguments:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        Required cmd arguments:
            CK_FI_KEY - fi_key of a valid financial institution.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing.

    Raises:
        p3m.CMDValidationException: For unrecoverable errors.
    """
    try:
        #region Initialization and validation
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_update_catalog_map.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.

        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_WORKFLOW_UPDATE_SUBCMD_KEY
        )
        # Extract and validate required parameters from the command.
        settings: bdms.BudManSettings = bdms.BudManSettings()
        catman: BDMTXNCategoryManager = BDMTXNCategoryManager(settings)
        fi_key: str = cmd_args.get(CK_CMDLINE_FI_KEY, None)
        prev_fi_key: str = bdm_DC.dc_FI_KEY
        bdm_DC.dc_FI_KEY = fi_key if fi_key else prev_fi_key
        #endregion Initialization and validation

        # Update the catalog map workbook for the fi_key if it is already
        # loaded, else load it.
        if fi_key in catman.catalogs:   
            result_msg: str = catman.FI_TXN_CATEGORIES_WORKBOOK_update(fi_key)
            p3m.cp_user_info_message(f"{pad(level)}Updated catalog map workbook for fi_key: '{fi_key}'")
        else:
            result_msg: str = catman.FI_TXN_CATEGORIES_WORKBOOK_load(fi_key)
            p3m.cp_user_info_message(f"{pad(level)}Loaded catalog map workbook for fi_key: '{fi_key}'")
        level -= 1
        p3m.cp_user_info_message(m + "End: ...")
        # End: ------------------------------------------------------- +
        return p3m.cp_CMD_RESULT_create(
            cmd=cmd,
            status=True,
            type=p3m.CV_CMD_STRING_OUTPUT,
            content=result_msg
        )
    except p3m.CMDValidationException as e:
        logger.error(e.msg)
        raise
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_transfer: {m}")
        logger.error(err_msg)
        cmd_result_error = p3m.cp_CMD_RESULT_ERROR_create(cmd, err_msg)
        raise p3m.CMDValidationException(cmd=cmd, 
                                         msg=err_msg,
                                         cmd_result_error=cmd_result_error)
    finally:
        # Restore previous DC Values if they were modified.
        bdm_DC.dc_FI_KEY = prev_fi_key
#endregion WORKFLOW_CMD_update_catalog_map() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_set_value() function
def WORKFLOW_CMD_set_value(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """WORKFLOW_CMD_set_value: Set values in the DC for workflow tasks.

    This function sets values in the data context for workflow tasks based on the
    provided command object.

    Args:
        cmd (p3m.Command): A valid Command object.
        bdm_DC (BudManAppDataContext_Base): The data context for the 
            BudMan application.
        Associate cmd object keys for the set subcmd:
            CK_CMDLINE_WF_KEY
            CK_CMDLINE_WF_PURPOSE
    """
    try:
        # Assuming the cmd parameters have been validated before reaching this point.
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_set_value.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_SET_SUBCMD_KEY
        )
        wf_key: str = cmd_args.get(CK_CMDLINE_WF_KEY, None)
        wf_purpose: str = cmd_args.get(CK_CMDLINE_WF_PURPOSE, None)
        msg: str = ""
        if wf_key:
            bdm_DC.dc_WF_KEY = wf_key
            msg += f"Workflow key set to: '{wf_key}'"
        sep = ", " if wf_key else ""
        if wf_purpose:
            bdm_DC.dc_WF_PURPOSE = wf_purpose
            msg += f"{sep}Workflow purpose set to: '{wf_purpose}' "
        return p3m.cp_CMD_RESULT_create(
            cmd=cmd,
            status=True,
            type=p3m.CV_CMD_STRING_OUTPUT,
            content=msg
        )
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_CMD_set_value() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_task() function
def WORKFLOW_CMD_task(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """Model-Aware: Sync the WORKBOOK_DATA_COLLECTION for the current FI_KEY.

    For the current DC FI_KEY, synchronize the WORKBOOK_DATA_COLLECTION with 
    files in storage. Use the BSM functions.

    Args:
        bdm_DC (BudManAppDataContext_Base): The data context for the 
            BudMan application.

    Returns:
        treelib.Tree: The folder tree for the specified workflow.
    """
    try:
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_task.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_TASK_SUBCMD_KEY
        )
        # Initializations
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        msg: str = ''
        fi_key: str = bdm_DC.dc_FI_KEY
        if not fi_key:
            msg = "No FI_KEY set in the DC."
            logger.error(msg)
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        model:BudgetDomainModel = bdm_DC.model
        if not model:
            msg = "No BudgetDomainModel binding in the DC."
            logger.error(msg)
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
        task_name: str = cmd_args.get(CK_TASK_NAME, "Unnamed Task")
        if task_name == CV_SYNC:
            msg = f"DEPRECATED: Syncing WORKBOOK_DATA_COLLECTION for FI_KEY: '{fi_key}'"
            p3m.cp_user_warning_message(msg)
                # r_msg: str = ""
                # discovered_wdc: bdm.WORKBOOK_DATA_COLLECTION_TYPE = None
                # discovered_wdc, r_msg = model.bsm_FI_WORKBOOK_DATA_COLLECTION_resolve(fi_key)
            p3m.cp_user_info_message(m + "End: ...")
            # End: ----------------------------------------------------------- +
            return p3m.cp_CMD_RESULT_create(
                cmd=cmd,
                status=True,
                type=p3m.CV_CMD_STRING_OUTPUT,
                content=msg
            )
        else:
            msg = f"Unknown task name: '{task_name}'"
            return p3m.cp_CMD_RESULT_ERROR_create(cmd, msg)
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_CMD_task() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_CMD_apply() function
def WORKFLOW_CMD_apply(
        cmd: p3m.Command, 
        bdm_DC: BudManAppDataContext_Base,
        level: int = 0
        ) -> p3m.CMD_RESULT_TYPE:
    """Apply workflow tasks to WORKBOOKS.

        A WORKFLOW_apply_cmd command will use the wb_ref value in the cmd. 
        Value is a number or a wb_name.

        Arguments:
            cmd (Dict): A valid BudMan View Model Command object. For this
            command, must contain workflow_cmd = 'apply' resulting in
            a full command key of 'workflow_cmd_apply'.

        Returns:
            Tuple[success : bool, result : Any]: The outcome of the command 
            execution. If success is True, result contains result of the 
            command, if False, a description of the error.
            
        Raises:
            RuntimeError: A description of the
            root error is contained in the exception message.
    """
    try:
        level += 1
        ts: str = "[bold dark_orange]CMD: [/bold dark_orange]"
        m: str = f"{pad(level)}{ts} {WORKFLOW_CMD_task.__name__}() "
        p3m.cp_user_info_message(m + "Start: ...")
        level += 1
        # Start: ------------------------------------------------------------- +
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = cmd.validate_command(
            expected_cmd_key=CV_WORKFLOW_CMD_KEY,
            expected_subcmd_key=CV_TASK_SUBCMD_KEY
        )
        # Initializations
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        p3m.cp_msg_svc.user_info_message("workflow apply cmd starting..")
        p3m.cp_msg_svc.user_warning_message("workflow apply cmd Warning..")
        p3m.cp_msg_svc.user_error_message("workflow apply cmd Error..")
        p3m.cp_msg_svc.user_debug_message("workflow apply cmd Debug..")
        return p3m.cp_CMD_RESULT_create(True, p3m.CV_CMD_STRING_OUTPUT,
                            "Applied category_map to txn_categories.", cmd)   
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion WORKFLOW_CMD_apply() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_convert_csv_txns_to_excel_txns() function
def WORKFLOW_TASK_convert_csv_txns_to_excel_txns(csv_wb: BDMWorkbook, 
                                                 excel_wb: BDMWorkbook ) -> bdm.BUDMAN_RESULT_TYPE:
    """Convert CSV transactions to Excel transactions.

        Args:
        csv_wb (BDMWorkbook): The source CSV workbook.
        excel_wb (BDMWorkbook): The destination Excel workbook.
    """
    try:
        st = p3u.start_timer()
        fr: str = "WORKFLOW_TASK_convert_csv_txns_to_excel_txns:"
        logger.debug(f"{fr} ")
        csv_txns = csv_wb.wb_content
        headers = list(csv_txns[0].keys())  # Get headers from the first row
        
        excel_wb.wb_content = Workbook() # Create a new Excel workbook.
        ws: Worksheet = excel_wb.wb_content.active
        #TODO: get the worksheet title from settings or config.
        ws.title = "TransactionData" # Set the name of the first worksheet.

        ws.append(headers)  # Write headers to the first row
        # Need to convert date strings to datetime objects, and
        # convert amount strings to float.
        for row in csv_txns[1:]:
            for key, value in row.items():
                if key.lower() == "date" and isinstance(value, str):
                    value = datetime.datetime.strptime(value, "%m/%d/%Y").date()
                elif key.lower() == "amount":
                    if not isinstance(value, float):
                        cleaned = re.sub(r'[^\d.-]', '', value)  # Remove non-numeric characters
                        value = float(cleaned)
                row[key] = value
            ws.append(list(row.values()))
        # Task 3: Save the excel_txns workbook
        excel_wb.wb_content.save(excel_wb.abs_path())
        logger.debug(f"Saved excel_txns to {excel_wb.abs_path() }")
        return True, f"Converted CSV to Excel in {p3u.stop_timer(st)} "
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
                    # row[key].number_format = '#,##0.00'  # Set currency format in Excel
                    # row[key].number_format = 'DD/MM/YYYY'  # Set date format in Excel
#endregion WORKFLOW_TASK_convert_csv_txns_to_excel_txns() function
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_transfer_csv_file()
def WORKFLOW_TASK_transfer_csv_file_to_workbook(src_file_url: str,
                                    dst_wb: BDMWorkbook,
                                    wb_type: str) -> bdm.BUDMAN_RESULT_TYPE:
    """Transfer a CSV file to the specified wb_type.

    Args:
        src_file_url (str): The source file URL.
        dst_wb (BDMWorkbook): The destination workbook object.
        wb_type (str): The desired workbook type for the destination workbook.

    Returns:
        bdm.BUDMAN_RESULT_TYPE: The result of the transfer operation.
    """
    try:
        if wb_type == bdm.WB_TYPE_CSV_TXNS:
            # bsm_file must be a .csv file. Transfer to .csv_txns workbook
            # is just a copy to the new file_url.
            csv_DATA_LIST_url_copy(src_file_url, dst_wb.wb_url)
            return True, f"Transferred .csv to .csv_txns workbook: {dst_wb.wb_url}"
        else:
            err_msg = f"Unsupported wb_type for CSV transfer: {wb_type}"
            logger.error(err_msg)
            return False, err_msg
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception transfer_csv_file_to_workbook(): {m}")
        logger.error(err_msg)
        return False, err_msg
#endregion WORKFLOW_TASK_transfer_csv_file()
# ---------------------------------------------------------------------------- +
#region WORKFLOW_TASK_contstruct_wb_file_url() function
def WORKFLOW_TASK_construct_bdm_workbook(src_filename: str,
                                       wb_type: str,
                                       fi_key: str,
                                       wf_key: str,
                                       wf_purpose: str,
                                       bdm_DC: BudManAppDataContext_Base) -> bdm.BUDMAN_RESULT_TYPE:
    """Construct a workbook file URL based on the provided parameters.
    This function targets a workbook file url for for filename using:
        fi_key, wf_key, wf_purpose and wb_type.

    Args:
        filename (str): The base filename without extension.
        wb_type (str): The workbook type (e.g., 'excel_txns').
        fi_key (str): The financial institution key.
        wf_key (str): The workflow key.
        wf_purpose (str): The workflow purpose.

    Returns:
        bdm.BUDMAN_RESULT_TYPE: A tuple containing boolean indicating success, 
        and the constructed workbook file URL or an error message.
    """
    try:
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        model:BudgetDomainModel = bdm_DC.model
        if not model:
            m = "No BudgetDomainModel binding in the DC."
            logger.error(m)
            return False, m
        # Validate the provided parameters.
        if p3u.str_empty(src_filename):
            m = "Empty filename provided."
            logger.error(m)
            return False, m
        if (p3u.str_empty(fi_key) or not model.bdm_FI_KEY_validate(fi_key)):
            m = f"Invalid fi_key: '{fi_key}'"
            logger.error(m)
            return False, m
        if (p3u.str_empty(wf_key) or not model.bdm_WF_KEY_validate(wf_key)):
            m = f"Invalid wf_key: '{wf_key}'"
            logger.error(m)
            return False, m
        if (p3u.str_empty(wf_purpose) or not model.bdm_WF_PURPOSE_validate(wf_purpose)):
            m = f"Invalid wf_purpose: '{wf_purpose}'"
            logger.error(m)
            return False, m
        if (p3u.str_empty(wb_type) or wb_type not in bdm.VALID_WB_TYPE_VALUES):
            m = f"Invalid wb_type: '{wb_type}'"
            logger.error(m)
            return False, m
        # Construct the workbook file URL.
        wb_extension: str = bdm.WB_FILETYPE_MAP[wb_type]
        wb_prefix: str = model.bdm_WF_FOLDER_CONFIG_ATTRIBUTE(
            fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose,
            attribute=bdm.WF_PREFIX, raise_errors=True)
        wf_folder_url: str = model.bdm_WF_FOLDER_CONFIG_ATTRIBUTE(
            fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose,
            attribute=bdm.WF_FOLDER_URL, raise_errors=True)
        wf_folder: str = model.bdm_WF_FOLDER_CONFIG_ATTRIBUTE(
            fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose,
            attribute=bdm.WF_FOLDER, raise_errors=True)
        filename = f"{wb_prefix}{src_filename}{wb_type}"
        full_filename: str = f"{filename}{wb_extension}"
        folder_path = Path.from_uri(wf_folder_url) 
        file_path = folder_path / full_filename
        file_url: str = file_path.as_uri()
        new_wb: BDMWorkbook = BDMWorkbook(
            wb_name=full_filename,
            wb_filename=filename,
            wb_filetype=wb_extension,
            wb_type=wb_type,
            wb_url=file_url,    
            fi_key=fi_key,
            wf_key=wf_key,
            wf_purpose=wf_purpose,
            wf_folder_url=wf_folder_url,
            wf_folder=wf_folder
        )
        return True, new_wb
    except Exception as e:
        m = p3u.exc_err_msg(e)
        err_msg = (f"Exception during WORKFLOW_TASK_construct_wb_file_url: {m}")
        logger.error(err_msg)
        return False, err_msg
#endregion WORKFLOW_TASK_contstruct_wb_file_url() function
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region Workflow Command Services helper functions
# ---------------------------------------------------------------------------- +
#region pad()
def pad(level: int) -> str:
    """Pad level times P2 (2 spaces).

    Args:
        level (int): The level to pad.

    Returns:
        str: P2 times level.
`    """
    return P2 * level
#endregion pad()
# ---------------------------------------------------------------------------- +
#endregion Workflow Command Services helper functions
# ---------------------------------------------------------------------------- +
