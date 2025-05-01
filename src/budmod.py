# ---------------------------------------------------------------------------- +
#region PyExeclBudget - read Excel files from a bank and analyze budget.
""" p3ExcelBudget - read Excel files from a bank and analyze budget.

    This module is the main entry point for the p3ExcelBudget application.
    It reads Excel files from a bank and analyzes the budget.
"""
#endregion PyExeclBudget - read Excel files from a bank and analyze budget.
# ---------------------------------------------------------------------------- +
#region Imports
# python standard libraries packages and modules 
import atexit, pathlib, logging, inspect, logging.config  #, logging.handlers

# third-party  packages and module libraries
from openpyxl import Workbook
import inspect, pyjson5
import p3logging as p3l, p3_utils as p3u

# local packages and module libraries
from p3_excel_budget_constants import  *
import data.p3_budget_model as p3bm

logger = logging.getLogger(THIS_APP_NAME)
logger.propagate = True
#endregion Imports
# ---------------------------------------------------------------------------- +
#region log_workbook_info() function
def log_workbook_info(file_name : str = "unknown",wb : Workbook = None) -> None:  
    if wb is None or not isinstance(wb, Workbook):
        logger.error(f"Workbook({file_name}): None or not a Workbook.")
        return
    logger.info(f"Workbook({file_name}): with sheets: {str(wb.sheetnames)}")
    # sheet = wb.active
    # r = sheet.max_row
    # c = sheet.max_column
    logger.info(f"  Active sheet({wb.active.title}): " 
                f"({wb.active.max_row} x {wb.active.max_column})")
#endregion log_workbook_info() function
# ---------------------------------------------------------------------------- +
#region configure_logging() function
def configure_logging(logger_name : str = THIS_APP_NAME, logtest : bool = False) -> None:
    """Setup the application logger."""
    try:
        # Configure logging
        log_config_file = "budget_model_logging_config.jsonc"
        # Set the log filename for this application.
        # filenames = {"file": "logs/p3ExcelBudget.log"}
        _ = p3l.setup_logging(
            logger_name = logger_name,
            config_file = log_config_file
            )
        p3l.set_log_flag(p3l.LOG_FLAG_PRINT_CONFIG_ERRORS, True)
        logger = logging.getLogger(THIS_APP_NAME)
        logger.setLevel(logging.DEBUG)
        logger.info("+ ----------------------------------------------------- +")
        logger.info(f"+ Running {THIS_APP_NAME}...")
        logger.info("+ ----------------------------------------------------- +")
        if(logtest): 
            p3l.quick_logging_test(THIS_APP_NAME, log_config_file, reload = False)
    except Exception as e:
        logger.error(p3u.exc_msg(configure_logging, e))
        raise
#endregion configure_logging() function
# ---------------------------------------------------------------------------- +
#region budmod() function
def budmod():
    """Main function to run PyExcelBudget application."""
    try:
        p3bm.tryout_budget_model_template()

        # trans_file = "BOAChecking2025.xlsx"
        # # Initalize the p3_budget_model package.
        # bm = p3bm.BudgetModel()
        # bm.inititailize()
        # p3bm.execute_worklow_categorization(bm, "boa", p3bm.BM_WF_CATEGORIZATION)

        # wb = p3bm.load_fi_transactions(trans_file)
        # log_workbook_info(trans_file, wb)
        # sheet = wb.active
        # # Check for budget category column, add it if not present.
        # p3bm.check_budget_category(sheet)
        # # BOA specific: Map the 'Original Description' column to the 'Budget Category' column.
        # p3bm.map_budget_category(sheet, "Original Description", BUDGET_CATEGORY_COL)
        # p3bm.save_fi_transactions(wb, trans_file)
        _ = "pause"
    except Exception as e:
        m = p3u.exc_msg(budmod, e)
        logger.error(m)
        raise
#endregion budmod() function
# ---------------------------------------------------------------------------- +
#region Local __main__ stand-alone
if __name__ == "__main__":
    try:
        configure_logging(THIS_APP_NAME)
        logger.setLevel(logging.DEBUG)
        # bm = p3fi.init_budget_model()  # How to load the budget model config?
        budmod() # Application Main()
    except Exception as e:
        m = p3u.exc_msg("__main__", e)
        logger.error(m)
    logger.info(f"Exiting {THIS_APP_NAME}...")
    exit(1)
#endregion Local __main__ stand-alone
# ---------------------------------------------------------------------------- +
#region tryit() function
def tryit() -> None:
    """Try something."""
    try:
        result = inspect.stack()[1][3]
        print(f"result: '{result}'")
    except Exception as e:
        print(f"Error: tryit(): {str(e)}")
#endregion tryit() function
# ---------------------------------------------------------------------------- +
