# ---------------------------------------------------------------------------- +
#region budget_domain_model.py module
""" budget_domain_model.py implements the class BudgetModel.

    Following a rough MVVM pattern, the BudgetModel class is acting as 
    both the cohesive Model, representing a budget model domain in
    memory as object API. 

    TODO: rename BudgetModel to BudgetDomainModel.

    At present, the BudgetModel class is a singleton class that manages the 
    mapping of excel "workbooks" stored in the filesystem to the budget model.
    The BudgetModel class presents properties and methods to the outside world.
    Methods are separated into DomainModel-ish methods for the 
    Budget Domain Model (BDM), which is the in-memory data model,
    and StorageModel-ish methods for the Budget Storage Model (BSM), 
    which is the filesystem model structure.

    In the Budget Domain Model, a data pipeline pattern is used, anticipating 
    "raw data" will be introduced from financial institutions (FI) and and 
    processed through a series of workflow tasks. Raw data is a "workbook", 
    often an excel file, or a .cvs file.
    
    A "folder" concept is aligned with the stages of a pipeline as a 
    series of "workflows" applied to the data. Roughly, workflow task works 
    on data in its associated "work" folder, processing data inplace locally 
    or as modifications or output to workbooks in other folders. Configuration 
    is used to map the BDM folders to the BSM folders. But, a simple workflow 
    scheme is for a task to have a designated input, work and output folder in 
    BDM. Incoming data appears in the input folder. Tasks own the responsibility
    to move it into their work folder, and if and when to move it to the
    output folder. The workflow pipeline is roughly:

    Workflow: a function grouping of tasks to be applied to input data.
    workflowTaskFunc(InputFolder, Workfolder) -> OutputFolder

    Workflows are functional units of work with clearly defined concerns applied
    to input data and producing work and outout data. 

    Key Concepts:
    -------------
    - Folders: containers of workbooks associated with a workflow stage.
    - Workflows: a defined set of process functions applied to data.
    - Workbooks: a container of financial transaction data, e.g., excel file.
    - Raw Data: original data from a financial institution (FI), read-only.
    - Financial Budget: the finalized output, composed of workbooks,
        representing time-series transactions categorized by payments and
        deposits (debits and credits).
    - Financial Institution (FI): a bank or brokerage financial institution.
        
    Perhaps the primary domain should be "Financial Budget" instead of
    "Budget Model". The FB Domain and the "Storage Domain" or Sub-Domain. 

    The Budget Domain Model has the concern of presenting an API that is 
    independent of where the workbook data is sourced and stored. The 
    Budget Storage Model has the concern of managing sourcing and storing 
    workbooks in a storage system. Currently, it just maps the BDM structure to 
    filesystem folders and files.

    Assumptions:
    - The FI transaction data resides in a folder specified in the budget_config.
    - Banking transaction files are typically excel spreadsheets, referred to 
      as workbooks. 
    - A workbook's data content starts in cell A1.
    - Row 1 contains column headers. All subsequent rows are data with no blank
      rows until the end of the data.
"""
#endregion budget_domain_model.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import logging, os, getpass, time, copy
from pathlib import Path
from typing import List, Type, Generator, Dict, Tuple, Any

# third-party modules and packages
from config import settings
import p3_utils as p3u, pyjson5, p3logging as p3l
from openpyxl import Workbook, load_workbook

# local modules and packages
from .budget_model_constants import *
from .budget_domain_model_identity import BudgetDomainModelIdentity
# from .budget_storage_model import bsm_BDM_URL_load, bsm_BDM_URL_save
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(settings[APP_NAME])
# ---------------------------------------------------------------------------- +
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +

def dscr(_inst) -> str:
    try:
        if _inst is None: return "None"
        _id = hex(id(_inst))
        _cn = _inst.__class__.__name__
        return f"<instance '{_cn}':{_id}>"
    except Exception as e:
        return f"{type(e).__name__}()"
class SingletonMeta(type):
    """Metaclass for implementing the Singleton pattern for subclasses."""
    _instances = {}

    # As a metaclass, __call__() runs when an instance is created as an
    # override to the normal __new__ method which is called by the type() 
    # metaclass for all python classes. So, call super() to use the normal
    # class behavior in addition to what SingletonMeta is doing.
    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            # Runs only for the first instance of the cls (class).
            # Invokes cls.__new__(), then cls.__init__(), which constructs 
            # a new cls instance. By default, the type() metaclass does this.
            _new_inst = super().__call__(*args, **kwargs)
            # Apply the singleton pattern, one instance per class.
            cls._instances[cls] = _new_inst
            # Save the cls so code knows what subclass was instantiated.
            _new_inst._subclassname = cls.__name__
            logger.debug(f"Created first {dscr(_new_inst)}")
        logger.debug(f"Return {dscr(cls._instances[cls])}")
        return cls._instances[cls]
# ---------------------------------------------------------------------------- +
class BudgetModel(metaclass=SingletonMeta):
    # ======================================================================== +
    #region BudgetModel class intrinsics
    # ======================================================================== +
    """BudgetModel class implementing Singleton pattern.
    
        A singleton class to manage the budget model for the application.
        This class is used to store and manage the budget data, including
        the budget folder, institutions, and options. Uses only dict-friendly
        items.

        Properties:
        -----------
        - budget_folder_path_str: A pathname to a parent budget folder, 
            e.g., ~/OneDrive/budget.
        - budget_folder_abs_path: A Path object representing the absolute path
            to the budget folder.
        - institutions: A dictionary to store the financial institutions.
        - options: A dictionary to store the options for the budget model.
        - budget_data: A dictionary to store non-persistent the budget data.
    """
    # ------------------------------------------------------------------------ +
    #region class variables
    # A config Template should be a BudgetModel object, or a sub
    _default_config_object : object = None #
    #endregion class variables
    # ------------------------------------------------------------------------ +
    #region    BudgetModel class constructor __init__()
    def __init__(self, config_object : object = None) -> None:
        """Constructor for the BudgetModel class."""
        # Note: _subclassname set by SingletonMeta after __init__() completes.
        logger.debug("Start: ...")

        # Private attributes initialization, basic stuff only.
        # for serialization ease, always persist dates as str type.
        bdm_id = BudgetDomainModelIdentity()
        setattr(self, BDM_ID, bdm_id.uid)  # BudgetDomainModelIdentity
        setattr(self, BDM_CONFIG_OBJECT,config_object)
        setattr(self, BDM_INITIALIZED, False)
        setattr(self, BDM_FOLDER, None)  # budget folder path
        setattr(self, BDM_URL, bdm_id.bdm_store_abs_path().as_uri())  # path for budget model store
        setattr(self, BDM_FI_COLLECTION, {})  # financial institutions
        setattr(self, BDM_WF_COLLECTION, {}) 
        setattr(self, BDM_OPTIONS, {})  # budget model options
        setattr(self, BDM_OPTIONS, p3u.now_iso_date_string()) 
        setattr(self, BDM_LAST_MODIFIED_DATE, self._created_date)
        setattr(self, BDM_LAST_MODIFIED_BY, getpass.getuser())
        setattr(self, BDM_WORKING_DATA, {})  # wd - budget model working data
        logger.debug("Complete: BudgetModel().__init__() ...")
    #endregion BudgetModel class constructor __init__()
    # ------------------------------------------------------------------------ +
    #region    BudgetModel internal class methods
    def to_dict(self):
        '''Return BudgetModelTemplate dictionary object. Used for serialization.'''
        ret = {
            BDM_ID: self.bdm_id,
            BDM_CONFIG_OBJECT: self.bdm_config_object,
            BDM_INITIALIZED: self.bm_initialized,
            BDM_FOLDER: self.bm_folder,
            BDM_FI_COLLECTION: self.bdm_fi_collection,
            BDM_URL: self.bdm_url,
            BDM_WF_COLLECTION: self.bdm_wf_collection,
            BDM_OPTIONS: self.bm_options,
            BDM_OPTIONS: self.bm_created_date,
            BDM_LAST_MODIFIED_DATE: self.bm_last_modified_date,
            BDM_LAST_MODIFIED_BY: self.bm_last_modified_by,
            BDM_WORKING_DATA: self.bdm_working_data
        }
        return ret
    def __repr__(self) -> str:
        ''' Return a str representation of the BudgetModel object '''
        ret = f"{{ "
        ret += f"'{BDM_ID}': {self.bdm_id}, "
        ret += f"'{BDM_CONFIG_OBJECT}': {self.bdm_config_object}, "
        ret += f"'{BDM_INITIALIZED}': {self.bm_initialized}, "
        ret += f"'{BDM_FOLDER}': '{self.bm_folder}', "
        ret += f"'{BDM_FI_COLLECTION}': '{self.bdm_fi_collection}', "
        ret += f"'{BDM_URL}': '{self.bdm_url}', "
        ret += f"'{BDM_WF_COLLECTION}': '{self.bdm_wf_collection}', "
        ret += f"'{BDM_OPTIONS}': '{self.bm_options}', "
        ret += f"'{BDM_OPTIONS}': '{self.bm_created_date}', "
        ret += f"'{BDM_LAST_MODIFIED_DATE}': '{self.bm_last_modified_date}', "
        ret += f"'{BDM_LAST_MODIFIED_BY}': '{self.bm_last_modified_by}', "
        ret += f"'{BDM_WORKING_DATA}': '{self.bdm_working_data}' }} "
        return ret
    def __str__(self) -> str:
        ''' Return a str representation of the BudgetModel object '''
        ret = f"{self.__class__.__name__}({str(self.bm_folder)}): "
        ret += f"{BDM_ID} = {self.bdm_id}, "
        ret += f"{BDM_CONFIG_OBJECT} = {str(self.bdm_config_object)}, "
        ret += f"{BDM_INITIALIZED} = {str(self.bm_initialized)}, "
        ret += f"{BDM_FOLDER} = '{str(self.bm_folder)}', "
        ret += f"{BDM_FI_COLLECTION} = [{', '.join([repr(fi_key) for fi_key in self.bdm_fi_collection.keys()])}], "
        ret += f"{BDM_URL} = '{self.bdm_url}' "
        ret += f"{BDM_WF_COLLECTION} = '{self.bdm_wf_collection}' "
        ret += f"{BDM_OPTIONS} = '{self.bm_options}' "
        ret += f"{BDM_OPTIONS} = '{self.bm_created_date}', "
        ret += f"{BDM_LAST_MODIFIED_DATE} = '{self.bm_last_modified_date}', "
        ret += f"{BDM_LAST_MODIFIED_BY} = '{self.bm_last_modified_by}', "
        ret += f"{BDM_WORKING_DATA} = {self.bdm_working_data}"
        return ret
    #endregion BudgetModel internal class methods
    # ------------------------------------------------------------------------ +
    #region    BudgetModel public class properties
    @property
    def bdm_id(self) -> str:
        """The budget model ID."""
        return getattr(self, BDM_ID)
    @bdm_id.setter
    def bdm_id(self, value: str) -> None:
        """Set the budget model ID."""
        if not isinstance(value, str):
            raise ValueError(f"bm_id must be a string: {value}")
        setattr(self, BDM_ID, value)
    @property
    def bdm_config_object(self) -> object:
        """The budget model configuration object."""
        return getattr(self, BDM_CONFIG_OBJECT)
    @bdm_config_object.setter
    def bdm_config_object(self, value: object) -> None:
        """Set the budget model configuration object."""
        if not isinstance(value, object):
            raise ValueError(f"bm_config_object must be an object: {value}")
        setattr(self, BDM_CONFIG_OBJECT, value)
    @property
    def bm_initialized(self) -> bool:
        """The initialized value."""
        return self._initialized
    
    @bm_initialized.setter
    def bm_initialized(self, value )-> None:
        """Set the initialized value."""
        self._initialized = value

    @property
    def bm_folder(self) -> str:
        """The budget folder path is a string, e.g., '~/OneDrive/."""
        return self._budget_folder

    @bm_folder.setter
    def bm_folder(self, value: str) -> None:
        """Set the budget folder path."""
        self._budget_folder = value

    @property
    def bdm_url(self) -> str:
        """The budget manager store file url."""
        return self._bdm_url
    
    @bdm_url.setter
    def bdm_url(self, value: str) -> None:
        """Set the budget manager store file url."""
        self._bdm_url = value

    @property
    def bdm_fi_collection(self) -> dict:
        """The financial institutions collection."""
        return self._financial_institutions
    
    @bdm_fi_collection.setter
    def bdm_fi_collection(self, value: dict) -> None:
        """Set the financial institutions collection."""
        self._financial_institutions = value

    @property
    def bdm_wf_collection(self) -> dict:
        """The workflow collection."""
        return self._workflows
    
    @bdm_wf_collection.setter
    def bdm_wf_collection(self, value: dict) -> None:
        """Set the workflows collection."""
        self._workflows = value

    @property
    def bm_options(self) -> dict:
        """The budget model options dictionary."""
        return self._options
    
    @bm_options.setter
    def bm_options(self, value: dict) -> None:
        """Set the budget model options dictionary."""
        self._options = value

    @property
    def bm_created_date(self) -> str:
        """The created date."""
        return self._created_date
    
    @bm_created_date.setter
    def bm_created_date(self, value: str) -> None:  
        """Set the created date."""
        self._created_date = value

    @property
    def bm_last_modified_date(self) -> str:
        """The last modified date."""
        return self._last_modified_date
    
    @bm_last_modified_date.setter
    def bm_last_modified_date(self, value: str) -> None:
        """Set the last modified date."""
        self._last_modified_date = value

    @property
    def bm_last_modified_by(self) -> str:
        """The last modified by."""
        return self._last_modified_by
    
    @bm_last_modified_by.setter
    def bm_last_modified_by(self, value: str) -> None:
        """Set the last modified by."""
        self._last_modified_by = value
    
    @property
    def bdm_working_data(self) -> dict:
        """The budget domain model working data."""
        self._wd = {} if self._wd is None else self._wd
        return self._wd
    
    @bdm_working_data.setter
    def bdm_working_data(self, value: dict) -> None:
        """Set the budget domain model working data."""
        self._wd = {} if self._wd is None else self._wd
        self._wd = value

    @property
    def data_context(self) -> DATA_CONTEXT:
        """The data context for the budget model."""
        return self.bdm_working_data 
       
    @data_context.setter
    def data_context(self, value: DATA_CONTEXT) -> None:
        """Set the data context for the budget model."""
        self.bdm_working_data = value

    # budget_model_working_data is a dictionary to store dynamic, non-property data.
    def set_BDM_WORKING_DATA(self, key, value) -> None:
        self.bdm_working_data[key] = value

    def get_BDM_WORKING_DATA(self, key) -> Any:
        return self.bdm_working_data.get(key, 0)

    #endregion BudgetModel public class properties
    # ------------------------------------------------------------------------ +
    #endregion BudgetModel class intrinsics
    # ======================================================================== +

    # ======================================================================== +
    #region    Budget model Domain Model (BDM) object methods
    """ Budget Model Domain Model (BDM) Documentation.

    Budget model Domain Model (BDM) is the conceptual model used to track and 
    analyze transactions from financial institutions (FI) overtime for the 
    purpose of following a budget. As a Domain Model, the implementation is 
    the python data structures used to represent the domain data.

    Class BudgetModel is the base class for the budget model domain. As a 
    convention, while designing the structure, I adopted the use of constants
    for names of the various types and fields used for th class properties,
    default values and methods. Also, I adopted a pattern to use the
    BudgetModelTemplate class, a subclass of BudgetModel, as a means
    for storing the template for a budget model used for configuration and
    other conveniences. Each BudgetModel created by a user, may have an 
    associated Budget Manager (BudMan) store. This is referred to as the 
    BUDMAN_STORE and can be used as the configuration template.

    All BDM data storage is the concern of the BSM. BSM works with Path 
    objects, filenames, folders, relative, absolute path names and actual 
    filesystem operations. In the configuration data, the BDM data is stored 
    as strings. Internally, Path objects are used.

    In the BDM, all data is associated with a financial institution (FI). Data 
    is processed by workflow (WF) tasks that are configured for an FI. 
    Each FI has a unique fi_key which the api uses to keep data separated.

    Throughout, the following identifiers are used in reference to parts of
    the BDM.

    BDM_FOLDER - refers to the root folder for the budget model. BDM_FOLDER is 
    a path element, which is a str that will end up as a part of a Path str, 
    which is the BSM concern. See bsm_BDM_FOLDER methods for that.

    FI_FOLDER - refers to the root folder configured for a specific 
    Financial Institutions (FI). FI_FOLDER is a path element, which is a 
    str that will end up as a part of a Path str, which is the BSM concern.
    See bsm_FI_FOLDER methods for that. There is only one FI_FOLDER for
    each FI. The FI_FOLDER is a subfolder of the BDM_FOLDER.

    WF_INPUT_FOLDER - refers to the input folder for a specific workflow, again, 
    always in the context of a specific FI through the fi_key. WF_INPUT_FOLDER is
    a Path with various str reprs.

    WF_WORKBOOKS_IN - refers to the collection of workbooks currently in the 
    WF_INPUT_FOLDER folder.

    WF_OUTPUT_FOLDER - refers to the output folder for a specific workflow. It is
    a Path with various str reprs.

    WF_WORKBOOKS_OUT - refers to the collection of workbooks currently in 
    the WF_OUTPUT_FOLDER folder.

    From the configuration data, the values of BDM_FOLDER, FI_FOLDER and
    WF_FOLDER are used to construct folder and file names for the file system
    and the BSM. The BDM methods do not use Path objects. The overlap seen in 
    the WF_WORKBOOKS_IN and WF_WORKBOOKS_OUT for a purpose. The workbooks are
    common in the two layers.
     
    The following naming conventions apply to the purpose of the BDM and BSM 
    methods to  separate the concerns for partial path names, full path names 
    and absolute pathname values..

    Naming Conventions:
    -------------------
    - bm_ - BudgetModel class properties, e.g., bm_folder, bm_fi, etc.
    - fi_ - Related to financial institution, e.g., fi_key, fi_name, etc.
    - wf_ - Related to workflow, e.g., wf_key, wf_name, etc.
    - bdm_ - BudgetModel Domain methods, concerning the in memory data model.
    - bsm_ - BudgetModel Storage Model methods, folders/files stored the filesystem.
    - _path_str - is the simplest string for a path name or component of. These
    methods return str values and do not manipulate with Path objects. Some
    folder values in the BDM, for example, are simply the name of a folder, 
    not a complete path name.
    - _path - constructs a Path object using _path_str values, 
    invokes .expanduser(). 
    _abs_path - invokes .resolve() on _path results, return Path object.
    - Path objects are never serialized to .jsonc files, only _path_str values.

    Abbrevs used in method names: 
        bdm - Budget Model Domain, the domain model of the budget model.
        bms - Budget Storage Model, the filesystem storage model.
        bm - BudgetModel class instance, parent of the Budget Model data structure.
        bf - budget folder - attribute, root folder, used in path objects.
        fi - financial institution dictionary, attribute of bdm.
             fi_key = int_key, short name of FI, e.g., "boa", "chase", etc.
             fi_value = dict of info about the FI.
        wf - workflow
    """
    # ======================================================================== +
    #region    BDM_CONFIG_OBJECT bdm_resolve_config_object(self) -> Dict
    def bdm_resolve_config_object(self) -> Dict:
        """Resolve the configuration object to initialize BudgetModel.

        Best to provide a config_object to BudgetModel() at instantiation
        time. If not, the BudgetModelTemplate is used as the default.

        Returns:
            Dict: The resolved configuration object.
        """
        try:
            if self.bdm_config_object is None:
                # If bdm_config_object is None, use the BudgetModelTemplate.
                logger.debug("bm_config_object property is None, "
                             f"resolve with builtin BudgetModelTemplate.")
                # Default is left behind if BudgetModelTemplate() is instantiated.
                bdm_config = BudgetModel._default_config_object  
                if bdm_config is None:
                    # Last hope, obtain the BudgetModelTemplate without 
                    # a circular reference, since it is a subclass.
                    from src.budman_model.get_budget_model_template import __get_budget_model_template__
                    bdm_config = __get_budget_model_template__()
                self.bdm_config_object = bdm_config
            return self.bdm_config_object
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BDM_CONFIG_OBJECT bdm_resolve_config_object(self) -> Dict 
    # ------------------------------------------------------------------------ +
    #region    BDM_INITIALIZED BDM bdm_initialize(self, bsm_init, ...)
    def bdm_initialize(self, 
                 bsm_init : bool = True,
                 create_missing_folders : bool = True,
                 raise_errors : bool = True
                 ) -> "BudgetModel":
        """Initialize BDM, from a config_object.

        Currently, as a singleton class, BudgetModel is initialized just once.
        So, having a resolved config_object Dict is important, lest not much
        can be initialized. That is the concern of bdm_resolve_config_src().

        Args:
            bsm_init (bool): Initialize the BSM if True.
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
        """
        st = p3u.start_timer()
        logger.debug(f"Start: ...")
        bdm_config : Dict = None
        try:
            if not hasattr(self, "_subclassname"):
                # hmm, _subclassname should be set in SingletonMeta metaclass.
                logger.debug("No subclass name, setting to None")
                self._subclassname = None
            # Initialize from a BudgetModel-based object such as the BMT template.
            bdm_config = self.bdm_resolve_config_object()
            # Apply the configuration to the budget model (self)
            for k,v in bdm_config.items():
                if k in BSM_PERSISTED_PROPERTIES and hasattr(self, k):
                    setattr(self, k, v)
            if bsm_init:
                self.bsm_initialize(create_missing_folders, raise_errors)
            self.bdm_working_data = self.bdm_BDM_WORKING_DATA_initialize()
            self.bm_initialized = True
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return self
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BDM_INITIALIZED BDM bdm_initialize(self, bsm_init, ...) 
    # ------------------------------------------------------------------------ +
    #region    bdm_initialize_from_BDM_URL(self)
    def bdm_initialize_from_BDM_URL(self,bsm_init:bool=True) -> None:
        """Initialize the BudgetModel, dynamically, from BDM_CONFIG values.

        The current session state of the BudgetModel configuration can be stored
        using the Budget Storage Model based on the URI in the BDM_URL
        property. Load that and apply the values to the BudgetModel instance.

        Returns:
            None: on success, else raises an exception.
        """
        try:
            st = p3u.start_timer()
            logger.debug(f"Start: ...")
            # Initialize from the BDM_URL persisted configuration and data.
            # Load the BudgetModel Store values as a Dict with persisted
            # attributes.
            # Apply the configuration to the budget model (self)
            # BSM_PERSISTED_PROPERTIES defines the attributes to be applied.
            self.__dict__.update(copy.deepcopy(self.bdm_config_object))
            bdm_config : Dict = None #bsm_BDM_URL_load(self)
            for attr in BSM_PERSISTED_PROPERTIES:
                if attr in bdm_config and hasattr(self, attr):
                    setattr(self, attr, bdm_config[attr])
            if bsm_init:
                self.bsm_initialize()
            self.bdm_working_data = self.bdm_BDM_WORKING_DATA_initialize()
            self.bm_initialized = True
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return 
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdm_initialize_from_BDM_URL(self)
    # ------------------------------------------------------------------------ +
    #region    BDM bdm_BDM_WORKING_DATA_initialize() 
    def bdm_BDM_WORKING_DATA_initialize(self) -> dict:
        """Initialize BDWD, the budget domain working data.
        
        When BudgetModel is initialized, the BDM_WORKING_DATA property is
        set to an empty dictionary. This method adds additional initial
        values to that dictionary. 

        Note: use the get_BDM_WORKING_DATA() and 
        set_BDM_WORKING_DATA() methods to access the working data.
        """
        try:
            logger.debug("Start: ...")
            # If the BDM_WORKING_DATA is already initialized, return it.
            if (hasattr(self, BDM_WORKING_DATA) and
                isinstance(self.bdm_working_data, dict) and
                BDWD_INITIALIZED in self.bdm_working_data and
                self.bdm_working_data[BDWD_INITIALIZED]): 
                return self.bdm_working_data
            # Initialize the budget model working data.
            for wd_key in BDM_WORKING_DATA_VALID_ATTR_KEYS:
                if wd_key == BDWD_LOADED_WORKBOOKS:
                    self.set_BDM_WORKING_DATA(
                        BDWD_LOADED_WORKBOOKS, list())
            self.set_BDM_WORKING_DATA(BDWD_INITIALIZED, True)
            logger.debug("Complete:")
            return self.bdm_working_data
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion   BDM bdm_BDM_WORKING_DATA_initialize() 
    # ------------------------------------------------------------------------ +
    #region    BDM FI_OBJECT pseudo-Object properties
    def bdm_FI_OBJECT(self, fi_key:str) -> FI_OBJECT:
        """Return the FI_OBJECT for fi_key."""
        self.bdm_FI_KEY_validate(fi_key)
        return self.bdm_fi_collection[fi_key]
    
    def bdm_FI_KEY(self, fi_key:str) -> str:
        """Return the FI_KEY value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_KEY]

    def bdm_FI_NAME(self, fi_key:str) -> str:
        """Return the FI_NAME value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_NAME]
    
    def bdm_FI_TYPE(self, fi_key:str) -> str:
        """Return the FI_TYPE value of the FI_OBJECT for fi_key.."""
        return self.bdm_FI_OBJECT(fi_key)[FI_TYPE]

    def bdm_FI_FOLDER(self, fi_key:str) -> str:
        """Return the FI_FOLDER value of the FI_OBJECT for fi_key."""
        return self.bdm_FI_OBJECT(fi_key)[FI_FOLDER]
    
    def bdm_FI_DATA_COLLECTION(self, fi_key:str) -> DATA_COLLECTION:
        """Return the DATA_COLLECTION object of the FI_OBJECT for fi_key."""
        return self.bdm_FI_OBJECT(fi_key)[FI_DATA_COLLECTION]

    def bdm_FI_DATA_COLLECTION_count(self, fi_key:str) -> int:
        """Return a count of objects in the FI_DATA_COLLECTION."""
        fi_data_collection = self.bdm_FI_DATA_COLLECTION(fi_key)
        if fi_data_collection is None:
            return 0
        if not isinstance(fi_data_collection, dict):
            m = f"FI_DATA_COLLECTION is not a dict: {type(fi_data_collection)}"
            logger.error(m)
            raise ValueError(m)
        return len(fi_data_collection)

    def bdm_FI_KEY_validate(self, fi_key:str) -> bool:
        """Validate the financial institution key."""
        if fi_key not in self.bdm_fi_collection and fi_key != ALL_KEY:
            m = f"Financial Institution '{fi_key}' not 'all' or not found in "
            m += f"{self.bdm_fi_collection.keys()}"
            logger.error(m)
            raise ValueError(m)
        return True
    #endregion BDM FI_OBJECT pseudo-Object properties
    # ------------------------------------------------------------------------ +
    #region    BDM FI_WF_OBJECT DATA_OBJECT pseudo-Object properties
    def bdm_FI_WF_DATA_OBJECT(self, fi_key:str, wf_key : str) -> DATA_OBJECT:
        """Return the WF_DATA_OBJECT value for fi_key, wf_key."""
        if self.bdm_FI_DATA_COLLECTION_count(fi_key) == 0:
            return None
        if self.bdm_FI_DATA_COLLECTION(fi_key) is None:
            return None
        if wf_key not in self.bdm_FI_DATA_COLLECTION(fi_key).keys():
            return None
        return self.bdm_FI_DATA_COLLECTION(fi_key)[wf_key]
    
    def bdm_FI_WF_WORKBOOK_LIST(self, 
                                fi_key:str, wf_key:str, 
                                wb_type:str) -> WORKBOOK_LIST:
        """Return the WORKBOOK_LIST for the specified fi_key, wf_key, wb_type."""
        if not self.bdm_FI_KEY_validate(fi_key):
            m = f"Invalid financial institution key '{fi_key}'."
            logger.error(m)
            raise ValueError(m)
        if not self.bdm_WF_KEY_validate(wf_key):
            m = f"Invalid workflow key '{wf_key}'."
            logger.error(m)
            raise ValueError(m)
        if wb_type not in WF_WORKBOOK_TYPES:
            m = f"Invalid workbook type '{wb_type}' for workflow '{wf_key}'."
            logger.error(m)
            raise ValueError(m)
        wf_do = self.bdm_FI_WF_DATA_OBJECT(fi_key, wf_key)
        if wf_do is None:
            return None
        if wb_type not in wf_do.keys():
            return None
        wf_wbl = wf_do[wb_type]
        if wf_wbl is None:
            return None
        if not isinstance(wf_wbl, list):
            m = f"FI_DATA '{fi_key}' '{wf_key}' is not a WORKBOOK_LIST"
            m += f"{type(wf_wbl)}"
            logger.error(m)
            raise ValueError(m)
        return wf_wbl

    def bdm_FI_WF_WORKBOOK_LIST_count(self,
                                fi_key:str, wf_key:str, wb_type:str) -> int:
        """Return a count of WORKBOOK_LIST in the FI_DATA for fi_key, wf_key."""
        wf_wbl = self.bdm_FI_WF_WORKBOOK_LIST(fi_key, wf_key, wb_type)
        if wf_wbl is None:
            return 0
        if not isinstance(wf_wbl, list):
            m = f"FI_DATA '{fi_key}' '{wf_key}' is not a WORKBOOK_LIST"
            m += f"{type(wf_wbl)}"
            logger.error(m)
            raise ValueError(m)
        return len(wf_wbl)
    #endregion BDM FI_WF_OBJECT DATA_OBJECT pseudo-Object properties
    # ------------------------------------------------------------------------ +
    #region    BDM WF_OBJECT pseudo-Object properties
    def bdm_WF_OBJECT(self, wf_key:str) -> WF_OBJECT:
        """Return the WF_OBJECT specified wf_key."""
        self.bdm_WF_KEY_validate(wf_key)
        return self.bdm_wf_collection[wf_key]
    
    def bdm_WF_OBJECT_count(self) -> int:
        """Return a count of WF_OBJECTS in the BDM_WF_COLLECTION."""
        if self.bdm_wf_collection is None:
            return 0
        if not isinstance(self.bdm_wf_collection, dict):
            m = f"BDM_WF_COLLECTION is not a dict, "
            m+= "but type: {type(self.bdm_wf_collection)}"
            logger.error(m)
            raise ValueError(m)
        return len(self.bdm_wf_collection)
    
    def bdm_WF_KEY(self, wf_key:str) -> str:
        """Return the WF_KEY value the WF_OBJECT for wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_KEY]

    def bdm_WF_NAME(self, wf_key:str) -> str:
        """Return the WF_NAME value for wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_NAME]
    
    def bdm_WF_FOLDER(self, wf_key:str, folder_id : str) -> str:
        """Return the WF_FOLDER value for (wf_key, folder_id).
        Not a full path, just the configuration value."""
        if folder_id not in (WF_INPUT_FOLDER, WF_OUTPUT_FOLDER):
            m = f"Invalid folder_id '{folder_id}' for workflow '{wf_key}'."
            logger.error(m)
            raise ValueError(m)
        return self.bdm_WF_OBJECT(wf_key)[folder_id]
    
    def bdm_WF_PREFIX_IN(self, wf_key:str) -> str:
        """Return the WF_PREFIX_IN value for wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_PREFIX_IN]
    
    def bdm_WF_PREFIX_OUT(self, wf_key:str) -> str:
        """Return the WF_PREFIX_OUT value for wf_key."""
        return self.bdm_WF_OBJECT(wf_key)[WF_PREFIX_OUT]    
    
    def bdm_WF_WORKBOOK_MAP(self, wf_key:str, wb_type:str=None) -> str|dict:
        """Return the WF_WORKBOOK_MAP or specific mapped value for wb_type."""
        if wb_type is None:
            return self.bdm_WF_OBJECT(wf_key)[WF_WORKBOOK_MAP]
        return self.bdm_WF_OBJECT(wf_key)[WF_WORKBOOK_MAP][wb_type]

    def bdm_WF_KEY_validate(self, wf_key:str) -> bool:
        """Validate the workflow key."""
        supp_wf = self.bdm_wf_collection
        if supp_wf is None or (wf_key not in supp_wf and wf_key != ALL_KEY):
            m = f"Workflow('{wf_key}') is not 'all' and not found in "
            m += f"{str(self.bdm_wf_collection)}"
            logger.error(m)
            raise ValueError(m)
        return True
    #endregion BDM WF_OBJECT pseudo-Object properties
    # ------------------------------------------------------------------------ +
    #endregion BudgetModel Domain methods
    # ======================================================================== +

    # ======================================================================== +
    #region    BudgetDomainModel Storage Model (BSM) methods
    """ Budget model Storage Model (BSM) Documentation.

    All BDM data is stored in the filesystem by the BSM. BSM works with Path 
    objects, filenames, folders, relative, absolute path names and actual 
    filesystem operations. In the configuration data, the BDM data is stored 
    as strings. Internally, Path objects are used.

    In the BDM, all data is associated with a financial institution (FI). Data 
    is processed by workflows (WF) that are configured for an FI. Hence,
    in all cases the BSM requires an fi_key to identify which FI's data is being
    processed. The wf_key, identifying a particular workflow, is also used
    in the BSM methods, because naming conventions for folder and files are 
    based on the workflow settings.

    Throughout, the following identifiers are used as a convention:
    
    FI_FOLDER - refers to the root folder configured for a specific 
    Financial Institutions (FI). FI_FOLDER is a Path with various str reprs.

    WF_INPUT_FOLDER - refers to the input folder for a specific workflow, again, 
    always in the context of a specific FI through the fi_key. WF_INPUT_FOLDER is
    a Path with various str reprs.

    WF_WORKBOOKS_IN - refers to the collection of workbooks currently in the 
    WF_INPUT_FOLDER folder.

    WF_OUTPUT_FOLDER - refers to the output folder for a specific workflow. It is
    a Path with various str reprs.

    WF_WORKBOOKS_OUT - refers to the collection of workbooks currently in 
    the WF_OUTPUT_FOLDER folder.

    The configuration data and stored budget model data contains str values 
    that are utilized in path names for folders and files. The following 
    naming conventions apply to the purpose of the BDM and BSM methods to 
    separate the concerns for partial path names, full path names and absolute
    pathname values..

    Naming Conventions:
    -------------------
    - bm_ - BudgetModel class properties, e.g., bm_folder, bm_fi, etc.
    - fi_ - Related to financial institution, e.g., fi_key, fi_name, etc.
    - wf_ - Related to workflow, e.g., wf_key, wf_name, etc.
    - bdm_ - BudgetModel Domain methods, concerning the in memory data model.
    - bsm_ - BudgetModel Storage Model methods, folders/files stored the filesystem.
    - _path_str - is the simplest string for a path name or component of. These
    methods return str values and do not manipulate with Path objects. Some
    folder values in the BDM, for example, are simply the name of a folder, 
    not a complete path name.
    - _path - constructs a Path object using _path_str values, 
    invokes .expanduser(). 
    _abs_path - invokes .resolve() on _path results, return Path object.
    - Path objects are never serialized to .jsonc files, only _path_str values.

    Abbrevs used in method names: 
        bdm - Budget Model Domain, the domain model of the budget model.
        bms - Budget Storage Model, the filesystem storage model.
        bm - BudgetModel class instance, parent of the Budget Model data structure.
        bf - budget folder - attribute, root folder, used in path objects.
        fi - financial institution dictionary, attribute of bdm.
             fi_key = int_key, short name of FI, e.g., "boa", "chase", etc.
             fi_value = dict of info about the FI.
        wf - workflow
    """
    # ======================================================================== +
    #region bsm_initialize() method
    def bsm_initialize(self, 
                        create_missing_folders:bool=True,
                        raise_errors:bool=True) -> "BudgetModel":
        """Initialize BSM aspects of the BDM.
        
        Examine elements of self, the BudgetModel class, representing the BDM.
        Validate the mapping of the BDM data dependent on mapping to folders 
        and files in the filesystem. Flags control whether to create the
        filesystem structure if it is not present. Also scan the workflow 
        folders for the presence of workbook excel files and load their 
        references into the BDM as appropriate.

        Args:
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
        """
        st = p3u.start_timer()
        # Plan: validate filesystem folders: bf, fi, wf
        #       update the BDM workbook mappings to BSM 
        try:
            logger.debug("Start: ...")
            self.bsm_BDM_FOLDER_resolve(create_missing_folders, raise_errors)
            # Enumerate the financial institutions.

            _ = [self.bsm_FI_initialize(fi_key, create_missing_folders, raise_errors) 
                   for fi_key, fi_object in self.bdm_fi_collection.items()]                
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return self
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_initialize() method
    # ------------------------------------------------------------------------ +    
    #region bsm_FI_initialize() method
    def bsm_FI_initialize(self, fi_key : str, 
                        create_missing_folders:bool=True,
                        raise_errors:bool=True) -> bool:
        """Initialize BSM aspects for a financial institution.
        
        Examine elements of self for the financial institution indicated by
        fi_key. Validate the mapping of the BDM data dependent on mapping to 
        folders and files in the filesystem. Flags control whether to create the
        filesystem structure if it is not present. Also scan the workflow 
        folders for the presence of workbook excel files and load their 
        references into the BDM as appropriate.

        Args:
            fi_key (str): The financial institution key.
            create_missing_folders (bool): Create missing folders if True.
            raise_errors (bool): Raise errors if True.
            
        Returns:
            bool: True if successful, False otherwise.

        Raises:
            TypeError: If the financial institution key is not a string.
            ValueError: If the financial institution key is not valid.
            Exception: If there is an error during initialization.
        """
        st = p3u.start_timer()
        # Plan: validate filesystem folders: bf, fi, wf
        #       update the BDM workbook mappings to BSM 
        try:
            logger.debug("Start: ...")
            p3u.str_empty(fi_key, raise_error=True) # Raises TypeError, ValueError
            # Resolve FI_FOLDER path.
            self.bsm_FI_FOLDER_resolve(fi_key, 
                                        create_missing_folders, raise_errors)
            # Resolve WF_FOLDER paths for the workflows.
            self.bsm_WF_FOLDER_resolve(fi_key, 
                                        create_missing_folders, raise_errors)
            # Resolve the FI_DATA collection, refresh actual
            # data from folders in BSM.
            self.bsm_FI_DATA_COLLECTION_resolve(fi_key)
            logger.debug(f"Complete: {p3u.stop_timer(st)}")   
            return True
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bsm_FI_initialize() method
    # ------------------------------------------------------------------------ +    
    #region BDM_FOLDER Path methods
    def bsm_BDM_FOLDER_validate(self) -> bool:
        """Validate the bm_folder property setting.
        
        Raise a ValueError if the bm_folder property is not set or is not
        usable as part of a valid path string.
        """
        # TODO: expand and resolve flags?
        if self.bm_folder is None or len(self.bm_folder) == 0:
            m = f"Budget folder path is not set. "
            m += f"Set the BDM_FOLDER('{BDM_FOLDER}') property to valid path value."
            logger.error(m)
            raise ValueError(m)
        return True
    def bsm_BDM_FOLDER_path_str(self) -> str:
        """str version of the BDM_FOLDER value as a Path."""
        # In the BSM, the bm_folder property must be a valid setting that will
        # result in a valid Path. Raise a ValueError if not.
        self.bsm_BDM_FOLDER_validate() # Raises ValueError if not valid
        return str(Path(self.bm_folder))
    def bsm_BDM_FOLDER_path(self) -> Path:
        """Path of self.bsm_BDM_FOLDER_path_str().expanduser()."""
        return Path(self.bsm_BDM_FOLDER_path_str()).expanduser()
    def bsm_BDM_FOLDER_abs_path(self) -> Path:
        """Path of self.bsm_BDM_FOLDER_path().resolve()."""
        return self.bsm_BDM_FOLDER_path().resolve()
    def bsm_BDM_FOLDER_abs_path_str(self) -> str:
        """str of self.bsm_BDM_FOLDER_abs_path()."""
        return str(self.bsm_BDM_FOLDER_abs_path())
    
    def bsm_BDM_FOLDER_resolve(self, 
                              create_missing_folders : bool=True,
                              raise_errors : bool=True) -> None:
        """Resolve the BDM_FOLDER path and create it if it does not exist."""
        try:
            logger.info(f"Checking BDM_FOLDER path: '{self.bm_folder}'")
            if self.bm_folder is None:
                m = f"Budget folder path is not set. "
                m += f"Set the '{BDM_FOLDER}' property to a valid path."
                logger.error(m)
                raise ValueError(m)
            # Resolve the BDM_FOLDER path.
            bf_ap = self.bsm_BDM_FOLDER_abs_path()
            bsm_verify_folder(bf_ap, create_missing_folders, raise_errors)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BDM_FOLDER Path methods
    # ------------------------------------------------------------------------ +
    #region BDM_URL Path methods
    def bsm_BDM_URL_validate(self) -> bool:
        """Validate the BDM_URL property setting.
        
        Raise a ValueError if the BDM_URL property is not set or is not
        usable as part of a valid path string.
        """
        # TODO: expand and resolve flags?
        if self.bdm_url is None or not isinstance(self.bdm_url,str) or len(self.bdm_url) == 0:
            m = f"BDM_URL value is not set to a non-zero length str. "
            m += f"Set the BDM_URL('{BDM_URL}') property to valid path str."
            logger.error(m)
            raise ValueError(m)
        return True
    def bsm_BDM_URL_path_str(self) -> str:
        """str version of the BDM_URL validated as a Path."""
        # In the BSM, the BDM_URL property must be a valid setting that will
        # result in a valid Path. Raise a ValueError if not.
        self.bsm_BDM_URL_validate() # Raises ValueError if not valid
        return str(Path(self.bdm_url))
    def bsm_BDM_URL_path(self) -> Path:
        """Path of self.bsm_BDM_URL_path_str().expanduser()."""
        return Path(self.bsm_BDM_URL_path_str()).expanduser()
    def bsm_BDM_URL_abs_path(self) -> Path:
        """Path of self.bsm_BDM_URL_path().resolve()."""
        return self.bsm_BDM_URL_path().resolve()
    def bsm_BDM_URL_abs_path_str(self) -> str:
        """str of self.bsm_BDM_URL_abs_path()."""
        return str(self.bsm_BDM_URL_abs_path())
    
    def bsm_BDM_URL_resolve(self, 
                              create_missing_folders : bool=True,
                              raise_errors : bool=True) -> None:
        """Resolve the BDM_URL path and create it if it does not exist."""
        try:
            logger.info(f"Checking BDM_URL path: '{self.bdm_url}'")
            if self.bdm_url is None:
                m = f"Budget folder path is not set. "
                m += f"Set the '{BDM_URL}' property to a valid path."
                logger.error(m)
                raise ValueError(m)
            # Resolve the BDM_URL path.
            bf_ap = self.bsm_BDM_URL_abs_path()
            bsm_verify_folder(bf_ap, create_missing_folders, raise_errors)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion BDM_URL Path methods
    # ------------------------------------------------------------------------ +
    #region FI_DATA FI_FOLDER Path methods
    def bsm_FI_FOLDER_path_str(self, fi_key: str) -> str:
        """str version of the FI_FOLDER value as a Path.
        
        Raises ValueError for invalid settings for any of: bm_folder property,
        fi_key, of FI_FOLDER value of FI dictionary."""
        bf_p_s = self.bsm_BDM_FOLDER_path_str() # ValueError on BDM_FOLDER property
        self.bdm_FI_KEY_validate(fi_key) # ValueError fi_key
        fi_p_s = str(Path(self.bdm_FI_FOLDER(fi_key))) 
        if fi_p_s is None or len(fi_p_s) == 0:
            m = f"FI_FOLDER value is not set for FI_KEY('{fi_key}'). "
            m += f"In the BudgetModel configuration, correct FI_FOLDER setting."
            logger.error(m)
            raise ValueError(m)   
        return str(Path(bf_p_s) / fi_p_s)
    def bsm_FI_FOLDER_path(self, fi_key: str) -> Path:
        """Path of self.bsm_FI_FOLDER_path_str().expanduser()."""
        return Path(self.bsm_FI_FOLDER_path_str(fi_key)).expanduser()
    def bsm_FI_FOLDER_abs_path(self, fi_key: str) -> Path:
        """Path of self.bsm_FI_FOLDER_path().resolve()."""
        return Path(self.bsm_FI_FOLDER_path(fi_key)).resolve()
    def bsm_FI_FOLDER_abs_path_str(self, fi_key: str) -> str:
        """str of self.bsm_FI_FOLDER_abs_path()."""
        return str(self.bsm_FI_FOLDER_abs_path(fi_key))
    def bsm_FI_FOLDER_resolve(self, fi_key : str,
                              create_missing_folders : bool=True,
                                raise_errors : bool=True) -> None: 
        """Resolve the FI_FOLDER path and create it if it does not exist."""
        fi_ap = self.bsm_FI_FOLDER_abs_path(fi_key)
        logger.info(f"FI_KEY('{fi_key}') Checking FI_FOLDER('{fi_ap}')")
        bsm_verify_folder(fi_ap, create_missing_folders, raise_errors)
    #endregion FI_DATA FI_FOLDER Path methods
    # ------------------------------------------------------------------------ +   
    #region bsm_FI_DATA_COLLECTION_resolve() method
    def bsm_FI_DATA_COLLECTION_resolve(self, fi_key:str) -> None:
        """Resolve the FI_WORKFLOW_DATA for the specified fi_key and wf_key."""
        try:
            if self.bdm_FI_DATA_COLLECTION_count(fi_key) == 0:
                m = f"FI_KEY('{fi_key}') has no workflow data."
                logger.debug(m)
                return
            # Enumerate the FI_OBJECT WF_DATA_COLLECTION.
            wf_dc : WF_DATA_COLLECTION = self.bdm_FI_DATA_COLLECTION(fi_key)
            for wf_key, wf_data_object in wf_dc.items():
                # Resolve each WF_DATA_OBJECT in the collection.
                self.bsm_WF_DATA_OBJECT_resolve(wf_data_object,
                                                fi_key, wf_key)
        except Exception as e:
                m = p3u.exc_err_msg(e)
                logger.error(m)
                raise
    #endregion bsm_FI_DATA_COLLECTION_resolve() method
    # ------------------------------------------------------------------------ + #region WF_OBJECT WF_FOLDER Path methods
    #region WF_OBJECT WF_FOLDER Path methods
    """WF_FOLDER refers to a str element used to construct a folder path name.
    It is an actual Path value or a string that is used
    to initialize a Path object. These Path methods are for that purpose.
    """
    def bsm_WF_FOLDER_path_str(self, fi_key : str, wf_key : str,
                               folder_id : str) -> str:
        """str version of the WF_FOLDER value for fi_key/wf_key/folder_id."""
        if folder_id not in [WF_INPUT_FOLDER, WF_OUTPUT_FOLDER]:
            m = f"Invalid folder_id '{folder_id}' for FI_KEY('{fi_key}') "
            m += f"and WF_KEY('{wf_key}')"
            logger.error(m)
            raise ValueError(m)
        fi_p_s = self.bsm_FI_FOLDER_path_str(fi_key) # FI_FOLDER Path component
        wf_p_s = self.bdm_WF_FOLDER(wf_key, folder_id)
        return str(Path(fi_p_s) / wf_p_s) if wf_p_s is not None else None
    def bsm_WF_FOLDER_path(self, fi_key : str, wf_key : str, 
                           folder_id:str) -> Path:
        """Path of self.bsm_wf_folder_path_str().expanduser()."""
        p_s = self.bsm_WF_FOLDER_path_str(fi_key, wf_key, folder_id)
        return Path(p_s).expanduser() if p_s is not None else None
    def bsm_WF_FOLDER_abs_path(self, fi_key : str, wf_key : str,
                               folder_id : str) -> Path:
        """Path of self.bsm_wf_folder_path().resolve()."""
        p = self.bsm_WF_FOLDER_path(fi_key, wf_key, folder_id)
        return p.resolve() if p is not None else None
    def bsm_WF_FOLDER_abs_path_str(self, fi_key : str, wf_key : str,
                                   folder_id : str) -> str:
        """str of self.bsm_wf_folder_abs_path()."""
        p = self.bsm_WF_FOLDER_abs_path(fi_key, wf_key, folder_id)
        return str(p) if p is not None else None
    def bsm_WF_FOLDER_resolve(self, fi_key:str, 
                              create_missing_folders:bool=True, 
                              raise_errors:bool=True) -> None:
        """Resolve any WF-related folders for the fi_key, create if requested."""
        try:
            logger.debug(f"FI_KEY('{fi_key}') scan all WF_FOLDERs.")
            for wf_key, wf_object in self.bdm_wf_collection.items():
                # Resolve the WF_FOLDER path elements.
                for f_id in WF_FOLDER_PATH_ELEMENTS:
                    wf_in_p = self.bsm_WF_FOLDER_path(fi_key, wf_key, f_id)
                    if wf_in_p is not None:
                        logger.info(f"FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
                                    f"Checking WF_FOLDER('{f_id}')")
                        bsm_verify_folder(wf_in_p, create_missing_folders, 
                                         raise_errors)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion WF_OBJECT WF_FOLDER Path methods
    # ------------------------------------------------------------------------ + #region WF_OBJECT WF_FOLDER Path methods
    #region WF_DATA_OBJECT (WF_DO) pseudo-property methods
    """
    A WF_DATA_OBJECT(Dict) is a DATA_OBJECT(Dict) with key/value pairs specific
    to data for a workflow. A wf_do is retrieved with the 
    bdm_FI_WF_DATA_OBJECT() method which could return other types of 
    DATA_OBJECTs in the future. These methods are BSM-related.

    With the BSM, the wf_do properties relating to the storage model are 
    WF_INPUT_FOLDER, WF_WORKBOOKS_IN, WF_OUTPUT_FOLDER and WF_WORKBOOKS_OUT. These 
    concern actual access to data in the filesystem. The BDM methods are used to
    access the data in the BDM.

    One key overlap, is bsm_WF_WORKBOOKS_IN and bsm_WF_WORKBOOKS_OUT. These are 
    the same methods in both the BDM and BSM. Each are dictionaries with tuples
    of each workbook in the corresponding WF_INPUT_FOLDER and WF_OUTPUT_FOLDER folders.
    The key is the filename, or the workbook name, the value is the full path 
    to the file.
    """
    def bsm_WF_DATA_OBJECT_resolve(self, wf_do: WF_DATA_OBJECT,
                                   fi_key : str, wf_key : str):
        """Resolve the WF_DATA_OBJECT based on the keys and values present."""
        try:
            logger.debug(f"FI_KEY('{fi_key}') WF_KEY('{wf_key}')")
            if wf_do is None or len(wf_do) == 0:
                logger.debug(f"  WF_DATA_OBJECT is empty.")
                return
            logger.debug(f"  WF_DATA_OBJECT({len(wf_do)} keys): {str(list(wf_do.keys()))}")
            # Resolve all keys in the WF_DATA_OBJECT.
            did_workbooks = False
            for wf_do_key, wf_do_value in wf_do.items():
                if wf_do_key not in WF_DATA_OBJECT_VALID_ATTR_KEYS:
                    m = f"Invalid WF_DATA_OBJECT key '{wf_do_key}' "
                    m += f"for FI_KEY('{fi_key}') and WF_KEY('{wf_key}')"
                    logger.error(m)
                    raise ValueError(m)
                if wf_do_key in WF_WORKBOOK_TYPES:
                    if not did_workbooks:
                        # Resolve all the WORKBOOK_LIST for WF_FOLDERs, just once.
                        self.bsm_WF_DATA_OBJECT_WORKBOOKS_resolve(wf_do, fi_key, wf_key)
                        did_workbooks = True
                    else:
                        continue
                else:
                    raise NotImplementedError(f"WF_DATA_OBJECT key '{wf_do_key}' ")
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
 
    def bsm_WF_DATA_OBJECT_WORKBOOKS_resolve(self, wf_do: WF_DATA_OBJECT, 
                                 fi_key: str, wf_key : str):
        """Resolve all WORKBOOK_LISTs for WF_FOLDERS(fi_key, wf_key) in the WF_DATA_OBJECT. 

        Workbooks are contained in folders in the filesystem. The folder_id 
        determines the absolute path to the folder, as per the following 
        mapping:
        - WF_WORKBOOKS_IN -> WF_INPUT_FOLDER - input folder for the workflow.
        - WF_WORKBOOKS_OUT -> WF_OUTPUT_FOLDER - output folder for the workflow.

        Args:
            wf_do (WF_DATA_OBJECT): The workflow data object to resolve.
            fi_key (str): The key of the institution to get the workbooks for.
            wf_key (str): The workflow to get the workbooks for.

        Returns:
            dict { str: str, ... }: A dict of workbook file names and their paths.
        """
        try:
            # TODO: validate parameters.
            for wf_do_key in wf_do.keys():
                # Only handle WF_WORKBOOK scope for the workflow.
                # A wf_do may have other keys, but we only care about the workbooks.
                if wf_do_key not in WF_WORKBOOK_TYPES:
                    logger.debug(f"  Skipping WF_DATA_OBJECT key: '{wf_do_key}'")
                    continue
                # Only interested in the WF_WORKBOOK_TYPES keys. Each of them
                # is configured to one WF_FOLDER_PATH_ELEMENT. Get the
                # appropriate folder_id for the WF_WORKBOOK_TYPE.
                f_id = self.bdm_WF_WORKBOOK_MAP(wf_key, wf_do_key)
                if f_id is None:
                    logger.debug(f"  FI('{fi_key}') WF('{wf_key}') "
                                f"WF_DO['{wf_do_key}'] is None")
                    continue
                # WORKBOOKS are found in WF_FOLDERS, scan them all.
                # Resolve all folders in this workflow for fi_key, wf_key.
                wb_list = []
                wf_f_ap = self.bsm_WF_FOLDER_abs_path(fi_key, wf_key, f_id)
                wb_files = list(wf_f_ap.glob("*.xlsx"))
                for wb_p in wb_files:
                    wb_name = wb_p.name
                    wb_item = (wb_name, str(wb_p))
                    wb_list.append(wb_item)
                wf_do[wf_do_key] = wb_list # save wb_list to the wf_do
                logger.debug(f"  FI('{fi_key}') WF('{wf_key}') "
                            f"WF_DO['{wf_do_key}'] maps to FOLDER('{f_id}') "
                            f"with {len(wb_list)} workbooks)")
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion WF_DATA_OBJECT (WF_DO) pseudo-property methods
    # ------------------------------------------------------------------------ +   
    #region bsm_FI_WF Methods
    def bsm_FI_WF_WORKBOOKS_load(self,
                        wbl : WORKBOOK_LIST = None) -> LOADED_WORKBOOKS_LIST: 
        """Load WORKBOOK_LIST returning a LOADED_WORKBOOKS_LIST
        
        A WORKBOOK_LIST has pairs of wb_name and wb_abs_path. Iterate the
        list and load each one. This is BSM-scope only, loads from the 
        filesystem, no side effects.

        Args:
            wbl (WORKBOOK_LIST): A list of tuples containing the workbook name
            and the absolute path to the workbook file.
        
        Returns:
            LOADED_WORKBOOKS_LIST: a new list of tuples containing the file name
            and the loaded workbook object.

        Raises: exceptions from any errors.
        """
        try:
            st = p3u.start_timer()
            returned_wb_list = []
            if wbl is None:
                logger.warning("No workbooks to load, wbl arg was None.")
                return returned_wb_list
            logger.debug(f"Loading {len(wbl)} workbooks.")
            for wb_name, wb_path in wbl:
                wb = load_workbook(filename=wb_path)
                returned_wb_list.append((wb_name, wb))
            logger.debug(f"Complete: Loaded {len(returned_wb_list)} workbooks. "
                         f"{p3u.stop_timer(st)}")
            return returned_wb_list
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bsm_FI_WF_WORKBOOKS_save(self,
                fi_key : str, 
                wf_key : str, 
                wb_type : str): 
        """Save all workbooks for a workflow to storage
        
        For fi_key,wf_key,wb_type, save all workbooks of type wb_type.
        
        Returns:
            None: as success. Updates BDM_WORKING_DATA.BDWD_LOADED_WORKBOOKS 
            property with the loaded workbooks.
        Raises: exceptions from any errors.
        """
        try:
            wbl = self.bdwd_LOADED_WORKBOOKS_get(fi_key, wf_key, wb_type)
            if wbl is None:
                return
            for wb_name, wb in wbl:
                wb.save(wb_path)
                self.bdwd_LOADED_WORKBOOKS_add(wb_name, wb)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bsm_FI_WF_WORKBOOK_generate(self,
                fi_key : str, 
                wf_key : str, 
                wb_type : str) -> Generator[Tuple[str, Workbook], None, None]: 
        """Generate a list of loaded workbooks.
        
        For fi_key,wf_key,wb_type, yield (wb_name, loaded_Workbook).
        
        Yields:
            Tuple[str, Workbook]: a tuple containing the file name the 
            loaded workbook object.
        """
        try:
            wbl = self.bdm_FI_WF_WORKBOOK_LIST(fi_key, wf_key, wb_type)
            if wbl is None:
                return
            for wb_name, wb_path in wbl:
                wb = load_workbook(filename=wb_path)
                self.bdwd_LOADED_WORKBOOKS_add(wb_name, wb)
                yield (wb_name, wb)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bsm_FI_WF_WORKBOOK_save(self, wb : Workbook, wb_name: str, 
                               fi_key:str, wf_key:str, wb_type : str):
        """Save workbook output to storage associated (fi_key,wf_key,wb_type).
        
        Map the fi_key and wf_key to the appropriate WF_OUTPUT_FOLDER folder in 
        the filesystem.
        """
        try:
            f_id = self.bdm_WF_WORKBOOK_MAP(wf_key, wb_type)
            if f_id is None:
                m = f"Invalid folder_id '{f_id}' for FI_KEY('{fi_key}') "
                m += f"and WF_KEY('{wf_key}')"
                logger.error(m)
                raise ValueError(m)
            fi_wf_ap = self.bsm_WF_FOLDER_abs_path(fi_key, wf_key, f_id)
            # TODO: strip the in_prefix if it is there.
            # Prepend the out_prefix to the workbook name.
            wb_name = f"{self.bdm_WF_PREFIX_OUT(wf_key)}{wb_name}"
            wb_path = fi_wf_ap / wb_name
            wb.save(wb_path)
            logger.info(f"Saved workbook '{wb_name}' to '{str(fi_wf_ap)}'")
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bsm_FI_WF_WORKBOOK_load(self, fi_key:str, wf_key:str, 
                              workbook_name:str) -> Workbook:
        """Load a workbook transaction file for a Financial Institution Workflow.

        ViewModel: This is a ViewModel function, mapping budget domain model 
        to how budget model data is stored in filesystem.

        Args:
            fi_key (str): The key of the institution to load the transaction file for.
            workflow (str): The workflow to load the transaction file from.
            workbook_name (str): The name of the workbook file to load.

        Returns:
            Workbook: The loaded transaction workbook.
        """
        me = self.bsm_FI_WF_WORKBOOK_load
        try:
            # Budget Folder Financial Institution Workflow Folder absolute path
            bffiwfap = self.bsm_FI_FOLDER_wf_abs_path(fi_key, wf_key) 
            wbap = bffiwfap / workbook_name # workbook absolute path
            m = f"BDM: Loading FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
            m += f"workbook('{workbook_name}'): abs_path: '{str(wbap)}'"
            logger.debug(m)
            wb = self.bsm_load_workbook(wbap)
            return wb
        except Exception as e:
            logger.error(p3u.exc_err_msg(e))
            raise
    #endregion FI_WF Methods
    # ------------------------------------------------------------------------ +   
    #region bsm_load_workbook(self, workbook_path:Path) function
    def bsm_load_workbook(self, input_path:Path) -> Workbook:
        """Load a transaction file for a Financial Institution Workflow.

        Storage Model: This is a Model function, loading an excel workbook
        file into memory.

        Args:
            workbook_path (Path): The path of the workbook file to load.

        Returns:
            Workbook: The loaded transaction workbook.
        """
        me = self.bsm_load_workbook
        try:
            logger.debug(f"BSM: Loading workbook file: '{input_path}'")
            wb = load_workbook(filename=input_path)
            return wb
        except Exception as e:
            logger.error(p3u.exc_msg(me, e))
            raise
    #endregion bsm_load_workbook(self, fi_key:str, process_folder:str) function
    # ------------------------------------------------------------------------ +
    #region bsm_save_workbook() function
    def bsm_save_workbook(self, workbook : Workbook = None, output_path:str=None) -> None:
        """Save the FI transactions workbook to the filesystem.
        
        The file is assumed to be in the folder specified in the budget_config. 
        Use the folder specified in the budget_config.json file. 
        the budget_config may have an output_prefix specified which will be
        prepended to the output_path name.

        TODO: If output_path is not specified, then scan the folder for files. 
        Return a dictionary of workbooks with the file name as the key.

        Args:
            output_path (str): The absolute path of the transaction file to save.

        Returns:
            None

        """
        me = self.save_fi_transactions
        st = time.time()
        try:
            # TODO: add logic to for workbook open in excel, work around.
            if (budget_config["output_prefix"] is not None and 
                isinstance(budget_config["output_prefix"], str) and
                len(budget_config["output_prefix"]) > 0):
                file_path = budget_config["output_prefix"] + output_path
            else:
                file_path = output_path
            trans_path = Path(budget_config["bank_transactions_folder"]) / file_path
            logger.info("Saving FI transactions...")
            workbook.save(filename=trans_path)
            logger.info(f"Saved FI transactions to '{str(trans_path)}'")
            return
        except Exception as e:
            logger.error(p3u.exc_msg(me, e))
            raise    
    #endregion bsm_save_workbook() function
    # ------------------------------------------------------------------------ +
    #endregion BudgetDomainModel Storage Model methods
    # ======================================================================== +

    # ======================================================================== +
    #region    BDWD - Budget Domain Model Working Data methods
    """ Budget Domain Model Working Data (BDWD) methods.
    BDWD methods are used to access the working data in the BDM. The working
    data is used by client packages for ViewModel, View, UX, CLI etc.
    It is transient data derived from the BDM/BSM operations.

    Note: the bdwd_ prefixed methods only use the 
    get_BDM_WORKING_DATA() and set_BDM_WORKING_DATA() 
    methods to access the BDM_WORKING_DATA property.
    """
    # --------------------------------------------------------------------- +
    #region bdwd_LOADED_WORKBOOKS() methods
    def bdwd_LOADED_WORKBOOKS_count(self) -> int:
        """Return total count of BDWD_LOADED_WORKBOOKS dictionary."""
        self.bdwd_INITIALIZED()
        return len(self.bdwd_LOADED_WORKBOOKS_get())

    def bdwd_LOADED_WORKBOOKS_get(self) -> List[Tuple[str, Workbook]] | None:
        """Get the BDWD_LOADED_WORKBOOKS from the BDM_WORKING_DATA.

        Returns:
            List(Tuple(wb_name: Workbook object))
        """
        try:
            self.bdwd_INITIALIZED()
            if self.bdm_working_data is None:
                m = f"BDM_WORKING_DATA is not set. "
                logger.error(m)
                raise ValueError(m)
            return self.get_BDM_WORKING_DATA(BDWD_LOADED_WORKBOOKS)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bdwd_LOADED_WORKBOOKS_add(self,wb_name : str, wb : Workbook) -> None:
        """Add an entry to BDWD_LOADED_WORKBOOKS in the BDM_WORKING_DATA.

        Returns:
            None: on success.
        Raises:
            ValueError: if the BDM_WORKING_DATA is not set.
            TypeError: if wb_name is not a str.
            TypeError: if wb is not a Workbook object.
            ValueError: if wb_name is None or an empty str.
            ValueError: if wb is None.
        """
        try:
            self.bdwd_INITIALIZED()
            p3u.str_empty(wb_name, raise_error = True)
            p3u.is_obj_of_type("wb", wb, Workbook, raise_TypeError=True)
            lwbs_list = self.bdwd_LOADED_WORKBOOKS_get()
            lwbs_list.append((wb_name, wb))
            logger.debug(f"Added ('{wb_name}', '{str(wb)}') "
                         f"to BDWD_LOADED_WORKBOOKS.")
            return None
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdwd_LOADED_WORKBOOKS() methods
    # ------------------------------------------------------------------------ +
    #region bdwd_FI methods
    def bdwd_FI_WORKBOOKS_load(self, fi_key : str, wf_key : str,
                               wb_type : str) -> LOADED_WORKBOOKS_LIST:
        """Load wbs for fi_key,wf_key,wb_type, merge to BDWD_LOADED_WORKBOOKS.

        For a given fi_key, wf_key and wb_type, load the workbooks from the
        filesystem and merge them into the BDWD_LOADED_WORKBOOKS list.

        Args:
            fi_key (str): The financial institution key.
            wf_key (str): The workflow key.
            wb_type (str): The workbook type.

        Returns:
            LOADED_WORKBOOKS_LIST: A list of tuples containing the 
            workbook name and the loaded workbook object. Newly loaded
            workbooks are appended to the BDWD_LOADED_WORKBOOKS list.
        """
        try:
            self.bdwd_INITIALIZED()
            # Get the WORKBOOK_LIST for fi_key, wf_key and wb_type from the BDM.
            returned_wb_list = []
            wbl = self.bdm_FI_WF_WORKBOOK_LIST(fi_key, wf_key, wb_type)
            if wbl is None or len(wbl) == 0:
                logger.debug(f"No workbooks to load for FI_KEY('{fi_key}') "
                             f"WF_KEY('{wf_key}')")
                return returned_wb_list
            # Use the BSM to load the workbooks for fi_key, wf_key and wb_type.
            logger.debug(f"Loading {len(wbl)} workbooks for "
                         f"FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
                         f"WB_TYPE('{wb_type}')")
            new_lwbl : LOADED_WORKBOOKS_LIST = []
            new_lwbl = self.bsm_FI_WF_WORKBOOKS_load(wbl)
            new_count = len(new_lwbl) if new_lwbl is not None else 0
            logger.debug(f"Loaded {new_count} workbooks for "
                         f"FI_KEY('{fi_key}') WF_KEY('{wf_key}') "
                         f"WB_TYPE('{wb_type}')")
            if new_lwbl is None or len(new_lwbl) == 0: return []
            bdwd_lwbs = self.bdwd_LOADED_WORKBOOKS_get()
            bdwd_lwbs_count = len(bdwd_lwbs) if bdwd_lwbs is not None else 0
            # TODO: How to handle duplicates?
            bdwd_lwbs.extend(new_lwbl)
            logger.debug(f"Added {new_count} to "
                         f"BDWD_LOADED_WORKBOOKS({bdwd_lwbs_count}) "
                         f"total = {len(bdwd_lwbs)}")
            return bdwd_lwbs
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise

    def bdwd_FI_WORKBOOKS_save(self, fi_key : str, wf_key : str,
                               wb_type : str) -> None:
        """Save workbooks for an FI workflow."""
        try:
            self.bdwd_INITIALIZED()
            self.bsm_FI_WF_WORKBOOKS_save(fi_key, wf_key, wb_type)
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise            

    def bdwd_FI_set(self, fi_key : str) -> None:
        """Set the BDWD_FI in the BDM_WORKING_DATA.

        Args:
            fi_key (str): The financial institution key.
        """
        try:
            self.bdwd_INITIALIZED()
            # fi_key must be a valid key or 'all'.
            _ = p3u.str_empty(fi_key, raise_error=True) # Raises TypeError, ValueError
            _ = self.bdm_FI_KEY_validate(fi_key) # Raises ValueError fi_key
            self.set_BDM_WORKING_DATA(BDWD_FI, fi_key)
            logger.debug(f"Set BDWD_FI to '{fi_key}'")
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdwd_FI methods
    # ------------------------------------------------------------------------ +
    #region bdwd_INITIALIZED methods
    def bdwd_INITIALIZED(self) -> bool:
        """Test if BDM_WORKING_DATA was initialized. Raise RuntimeError if not.
        """
        try:
            if (hasattr(self, BDM_WORKING_DATA) and
                isinstance(self.bdm_working_data, dict) and
                BDWD_INITIALIZED in self.bdm_working_data and
                self.bdm_working_data[BDWD_INITIALIZED]): return True
            m = f"BDM_WORKING_DATA was not initialized. "
            m += f"Use the bdm_BDM_WORKING_DATA_initialize() method."
            logger.error(m)
            raise RuntimeError(m)
        except RuntimeError as e:
            raise
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    def bdwd_INITIALIZED_get(self) -> bool:
        """Get the BDWD_INITIALIZED from the BDM_WORKING_DATA.
        """
        try:
            self.bdwd_INITIALIZED()
            return self.get_BDM_WORKING_DATA(BDWD_INITIALIZED) 
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    def bdwd_INITIALIZED_set(self, value : bool) -> bool:
        """Set the BDWD_INITIALIZED in BDM_WORKING_DATA.
        """
        try:
            self.bdwd_INITIALIZED()
            return self.set_BDM_WORKING_DATA(BDWD_INITIALIZED, value) 
        except Exception as e:
            m = p3u.exc_err_msg(e)
            logger.error(m)
            raise
    #endregion bdwd_INITIALIZED methods
    # ------------------------------------------------------------------------ +
    #endregion BDWD - Budget Domain Model Working Data methods
    # ======================================================================== +

# ---------------------------------------------------------------------------- +
#region log_BDM_info() function
def log_BDM_info(bdm : BudgetModel) -> None:  
    try:
        if bdm is None:
            logger.warning("bm parameter is None.")
            return None
        logger.debug("BDM Content:")
        logger.debug(f"{P2}BDM_INITIALIZED['{BDM_INITIALIZED}']: "
                     f"{bdm.bm_initialized}")
        logger.debug(f"{P2}BDM_FOLDER['{BDM_FOLDER}']: '{bdm.bm_folder}'")
        logger.debug(f"{P2}BDM_URL['{BDM_URL}]: '{bdm.bdm_url}'")
        # Enumerate the financial institutions in the budget model
        c = len(bdm.bdm_fi_collection)
        logger.debug(
            f"{P2}BDM_FI_COLLECTION['{BDM_FI_COLLECTION}']({c}): "
            f"{str(list(bdm.bdm_fi_collection.keys()))}")
        for fi_key in bdm.bdm_fi_collection.keys():
            logger.debug(f"{P4}Financial Institution: "
                         f"{bdm.bdm_FI_KEY(fi_key)}:{bdm.bdm_FI_NAME(fi_key)}:"
                         f"{bdm.bdm_FI_TYPE(fi_key)}: '{bdm.bdm_FI_FOLDER(fi_key)}'")
        # Enumerate workflows in the budget model
        c = bdm.bdm_WF_OBJECT_count()
        logger.debug(
            f"{P2}BDM_WF_COLLECTION['{BDM_WF_COLLECTION}']({c}): "
            f"{str(list(bdm.bdm_wf_collection.keys()))}")
        for wf_key in bdm.bdm_wf_collection.keys():
            logger.debug(f"{P4}Workflow:({bdm.bdm_WF_KEY(wf_key)}:{bdm.bdm_WF_NAME(wf_key)}: ")
            logger.debug(f"{P6}WF_WORKBOOKS_IN: '{bdm.bdm_WF_FOLDER(wf_key,WF_INPUT_FOLDER)}'")
            logger.debug(f"{P6}WF_OUTPUT_FOLDER: '{bdm.bdm_WF_FOLDER(wf_key,WF_OUTPUT_FOLDER)}'")
            logger.debug(f"{P6}WF_PREFIX_IN: '{bdm.bdm_WF_PREFIX_IN(wf_key)}' "
                         f"WF_PREFIX_OUT: '{bdm.bdm_WF_PREFIX_OUT(wf_key)}'")
            logger.debug(f"{P6}WF_WORKBOOK_MAP: {str(bdm.bdm_WF_WORKBOOK_MAP(wf_key))}")
        # Enumerate Budget Model Options
        bmoc = len(bdm.bm_options)
        logger.debug(f"{P2}BDM_OPTION['{BDM_OPTIONS}']({bmoc})")
        for opt_key, opt in bdm.bm_options.items():
            logger.debug(f"{P4}Option('{opt_key}') = '{opt}'")

        # And the rest
        logger.debug(f"{P2}BDM_OPTIONS['{BDM_OPTIONS}']: "
                        f"{bdm.bm_created_date}")
        logger.debug(f"{P2}BDM_LAST_MODIFIED_DATE['{BDM_LAST_MODIFIED_DATE}']: "
                        f"{bdm.bm_last_modified_date}")
        logger.debug(f"{P2}BDM_LAST_MODIFIED_BY['{BDM_LAST_MODIFIED_BY}']: "
                        f"'{bdm.bm_last_modified_by}'")
        logger.debug(f"{P2}BDM_WORKING_DATA('{BDM_WORKING_DATA}'): "
                        f"'{bdm.bdm_working_data}'")
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion log_BDM_info() function
# ---------------------------------------------------------------------------- +
#region log_BSM_info() function
def log_BSM_info(bdm : BudgetModel) -> None:  
    try:
        if bdm is None:
            logger.warning("bm parameter is None.")
            return None
        logger.debug("BSM Content:")
        bf_p = bdm.bsm_BDM_FOLDER_path() # budget folder path
        bf_p_exists = "exists." if bf_p.exists() else "does not exist!"
        bf_ap = bdm.bsm_BDM_FOLDER_abs_path() # budget folder path
        bf_ap_exists = "exists." if bf_ap.exists() else "does not exist!"
        logger.debug(f"{P2}BDM_FOLDER['{BDM_FOLDER}']: '{bdm.bm_folder}' {bf_ap_exists}")
        logger.debug(f"{P4}bsm_BDM_FOLDER_path(): '{str(bf_p)}' {bf_p_exists}")
        logger.debug(f"{P4}bsm_BDM_FOLDER_abs_path(): '{str(bf_ap)}' {bf_ap_exists}")
        bmc_p = bf_p / BSM_DEFAULT_BUDGET_MODEL_FILE_NAME # bmc: BM config file
        bmc_p_exists = "exists." if bmc_p.exists() else "does not exist!"
        logger.debug(
            f"{P2}BDM_URL['{BDM_URL}]: '{bdm.bdm_url}' "
            f"{bmc_p_exists}")
        # Enumerate the financial institutions in the budget model
        c = len(bdm.bdm_fi_collection)
        logger.debug(
            f"{P2}BDM_FI_COLLECTION['{BDM_FI_COLLECTION}']({c}): "
            f"{str(list(bdm.bdm_fi_collection.keys()))}")
        for fi_key in bdm.bdm_fi_collection.keys():
            logger.debug(f"{P4}Financial Institution: "
                         f"{bdm.bdm_FI_KEY(fi_key)}:{bdm.bdm_FI_NAME(fi_key)}:"
                         f"{bdm.bdm_FI_TYPE(fi_key)}: '{bdm.bdm_FI_FOLDER(fi_key)}'")
            c = bdm.bdm_FI_DATA_COLLECTION_count(fi_key)
            fi_data = bdm.bdm_FI_DATA_COLLECTION(fi_key)
            m = str(list(fi_data.keys())) if fi_data is not None else "'None'"
            logger.debug(f"{P6}FI Data({c}): {m}")
        # Enumerate workflows in the budget model
        c = bdm.bdm_WF_OBJECT_count()
        logger.debug(
            f"{P2}BDM_WF_COLLECTION['{BDM_WF_COLLECTION}']({c}): "
            f"{str(list(bdm.bdm_wf_collection.keys()))}")
        for wf_key in bdm.bdm_wf_collection.keys():
            logger.debug(f"{P4}Workflow:({bdm.bdm_WF_KEY(wf_key)}:{bdm.bdm_WF_NAME(wf_key)}: ")
            logger.debug(f"{P6}WF_INPUT_FOLDER: '{bdm.bdm_WF_FOLDER(wf_key,WF_INPUT_FOLDER)}'")
            logger.debug(f"{P6}WF_OUTPUT_FOLDER: '{bdm.bdm_WF_FOLDER(wf_key,WF_OUTPUT_FOLDER)}'")
            logger.debug(f"{P6}WF_WORKBOOK_MAP: {str(bdm.bdm_WF_WORKBOOK_MAP(wf_key))}")
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion log_BSM_info() function
# ---------------------------------------------------------------------------- +
#region check_budget_model() -> bool
def check_budget_model() -> bool:   
    """Quick check of the budget model schema is valid.

    Returns:
        bool: True if the budget model schema is valid, raise otherwise.
    """
    me = check_budget_model
    global budget_config, budget_model, budget_config_expected_keys
    global institution_expected_keys, options_expected_keys
    try:
        if (budget_model is None or 
            not isinstance(budget_model, dict) or
            len(budget_model) == 0):
            logger.warning("Budget model is not initialized.")
            return False
        # Check if the budget model contains expected keys
        for key in budget_config_expected_keys:
            if key not in budget_model:
                m = f"Budget model missing key: '{key}'"
                logger.error(m)
                raise ValueError(m)
        # Check if the institutions are valid
        for institution_key in BudgetModel.valid_institutions_keys:
            if institution_key not in budget_config["institutions"].keys():
                m = f"Budget model has invalid institution key: '{institution_key}'"
                logger.error(m)
                raise ValueError(m)
            for key in institution_expected_keys:
                if key not in budget_model["institutions"][institution_key]:
                    m = f"Budget model institution[{institution_key}] missing key: '{key}'"
                    logger.error(m)
                    raise ValueError(m)
        for key in options_expected_keys:
            if key not in budget_model["options"]:
                m = f"Budget model missing options key: '{key}'"
                logger.error(m)
                raise ValueError(m)
        return True
    except Exception as e:
        logger.error(p3u.exc_msg(me, e))
        return False
#endregion check_budget_model() -> bool
# ---------------------------------------------------------------------------- +
#region verify_folder(ap: Path, create:bool=True, raise_errors:bool=True) -> bool
def bsm_verify_folder(ap: Path, create:bool=True, raise_errors:bool=True) -> bool:
    """Verify the folder exists, create it if it does not exist.

    Args:
        ap (Path): The absolute path to the folder to verify.
        create (bool): Create the folder if it does not exist.
        raise_errors (bool): Raise errors if True.
    """
    try:
        if not ap.is_absolute():
            m = f"Path is not absolute: '{str(ap)}'"
            logger.error(m)
            raise ValueError(m)
        if ap.exists() and ap.is_dir():
            logger.debug(f"Folder exists: '{str(ap)}'")
            return True
        if not ap.exists():
            m = f"Folder does not exist: '{str(ap)}'"
            logger.error(m)
            if create:
                logger.info(f"Creating folder: '{str(ap)}'")
                ap.mkdir(parents=True, exist_ok=True)
            else:
                raise ValueError(m)
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion verify_folder(ap: Path, create:bool=True, raise_errors:bool=True) -> bool
# ---------------------------------------------------------------------------- +
#region Local __main__ stand-alone
if __name__ == "__main__":
    try:
        # this_app_name = os.path.basename(__file__)
        # Configure logging
        logger_name = THIS_APP_NAME
        log_config = "budget_model_logging_config.jsonc"
        # Set the log filename for this application.
        # filenames = {"file": "logs/p3ExcelBudget.log"}
        _ = p3l.setup_logging(
            logger_name = logger_name,
            config_file = log_config
        )
        p3l.set_log_flag(p3l.LOG_FLAG_PRINT_CONFIG_ERRORS, True)
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)
        logger.info("+ ----------------------------------------------------- +")
        logger.info(f"+ Running {THIS_APP_NAME} ...")
        logger.info("+ ----------------------------------------------------- +")
        logger.debug(f"Start: {THIS_APP_NAME}...")

        bdm = BudgetModel()
        bdm.initialize()
        bms = str(bdm)
        bmr = repr(bdm)
        bdm = bdm.to_dict()

        logger.debug(f"Budget Model: str() = '{bms}'")
        logger.debug(f"Budget Model: repr() = '{bmr}'")
        logger.debug(f"Budget Model: to_dict() = '{bdm}'")
        logger.info(f"Budget Model: {str(bdm)}")
        _ = "pause"
    except Exception as e:
        m = p3u.exc_msg("__main__", e)
        logger.error(m)
    logger.info(f"Exiting {THIS_APP_NAME}...")
    exit(1)
#endregion
# ---------------------------------------------------------------------------- +
