# ---------------------------------------------------------------------------- +
#region budman_command_services.py module
""" budman_command_services.py implements functions for BudMan app commands.

    The budman_command_processor package provides modules with functions
    to execute validated commands within the BudMan application. In general,
    a CMD_OBJECT is dispatched to the appropriate command handler for execution.
    A handler will decompose a CMD_OBJECT into a series of CMD tasks to be
    executed.

    In general, services should return either data objects or command result 
    objects. Leave it to the caller, such as a View, or ViewMModel to handle 
    additional services, such as a pipeline, or to perform output functions.

    A key aspect of the BudMan application patterns is the treatement of 
    workbooks versus files. Workbooks returned in lists or tree objects are
    in the Budget Domain Model (BDM) abstraction, with the container structure
    being bdm_folder, bdm_FI_FOLDER, and wf_folders bases on wf_key and 
    wf_purpose. CMDs associated with workbooks use the "workbooks" parameter.

    Some CMDs are aimed at the files in the Budget Storage Model (BSM). Files
    returned in lists or tree objects are based on the configured storage model,
    but are the usual hierarchy of folders and files.
"""
#endregion budman_command_services.py module
# ---------------------------------------------------------------------------- +
#region Imports
# python standard library modules and packages
import logging, io, sys, getpass, time, copy, importlib
from pathlib import Path
from typing import List, Type, Optional, Dict, Tuple, Any, Callable
from datetime import datetime as dt

# third-party modules and packages
from arrow import now
import p3_utils as p3u, pyjson5, p3logging as p3l, p3_mvvm as p3m
from openpyxl import Workbook, load_workbook
from treelib import Tree, Node
from rich import print
from rich.markup import escape
from rich.text import Text
from rich.tree import Tree as RichTree
from rich.table import Table

# local modules and packages
from .budman_cp_namespace import * 
import budman_namespace as bdm
from budman_namespace import (BDMWorkbook, P2, P4)
import budman_settings as bdms
from budget_domain_model import BudgetDomainModel
from budman_data_context import BudManAppDataContext_Base
from budget_storage_model import (
    BSMFile, BSMFileTree,
    bsm_verify_folder, bsm_URL_verify_file_scheme,)
from budget_categorization import (BDMTXNCategoryManager, TXNCategoryMap)
import budman_workflows
#endregion Imports
# ---------------------------------------------------------------------------- +
#region Globals and Constants
logger = logging.getLogger(__name__)
#endregion Globals and Constants
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region BudMan Application CMD_ functions
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_process() function
def BUDMAN_CMD_process(cmd: p3m.CMD_OBJECT_TYPE,
                     bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Process a CMD_TASK command.

    Args:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the command.

    Returns:
        CMD_RESULT_TYPE: The result of the command processing.
    """
    try:
        # If bdm_DC is bad, just raise an error.
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        # Assuming the CMD_OBJECT has been validated before reaching this point.
        # Process the CMD_OBJECT based on its cmd_key and subcmd_key.
        # List command
        if cmd[p3m.CK_CMD_KEY] == CV_LIST_CMD_KEY:
            if cmd[p3m.CK_SUBCMD_KEY] == CV_LIST_WORKBOOKS_SUBCMD_KEY:
                # List workbooks
                return BUDMAN_CMD_list_workbooks(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_NAME] == CV_BDM_STORE_SUBCMD_NAME:
                # List BDM_STORE as JSON
                return BUDMAN_CMD_list_bdm_store_json(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_KEY] == CV_LIST_FILES_SUBCMD_KEY:
                # List files in the BDM store
                return BUDMAN_CMD_list_files(cmd, bdm_DC)
            else:
                return p3m.cp_CMD_RESULT_ERROR_unknown(cmd)
        # Show command
        elif cmd[p3m.CK_CMD_KEY] == CV_SHOW_CMD_KEY:
            if cmd[p3m.CK_SUBCMD_KEY] == CV_SHOW_DATA_CONTEXT_SUBCMD_KEY:
                return BUDMAN_CMD_show_DATA_CONTEXT(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_KEY] == CV_SHOW_BUDGET_CATEGORIES_SUBCMD_KEY:
                return BUDMAN_CMD_TASK_show_BUDGET_CATEGORIES(cmd, bdm_DC)
            else:
                return p3m.cp_CMD_RESULT_ERROR_unknown(cmd)
        # App command
        elif cmd[p3m.CK_CMD_KEY] == CV_APP_CMD_KEY:
            if cmd[p3m.CK_SUBCMD_KEY] == CV_SYNC_SUBCMD_KEY:
                # App Sync the BDM_STORE to the BSM.
                return BUDMAN_CMD_app_sync(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_KEY] == CV_LOG_SUBCMD_KEY:
                # Show the current log level.
                return BUDMAN_CMD_app_log(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_KEY] == CV_APP_REFRESH_SUBCMD_KEY:
                # App Refresh subcommand.
                return BUDMAN_CMD_app_refresh(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_KEY] == CV_RELOAD_SUBCMD_KEY:
                # Reload the application configuration and settings.
                return BUDMAN_CMD_app_reload(cmd, bdm_DC)
            elif cmd[p3m.CK_SUBCMD_KEY] == CV_FILE_DELETE_SUBCMD_KEY:
                # Delete workbooks from the BDM store and BSM.
                return BUDMAN_CMD_app_delete(cmd, bdm_DC)
            else:
                return p3m.cp_CMD_RESULT_ERROR_unknown(cmd)
        # Unknown command
        else:
            return p3m.cp_CMD_RESULT_ERROR_unknown(cmd)
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion BUDMAN_CMD_process() function
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_list_workbooks()
def BUDMAN_CMD_list_workbooks(cmd: p3m.CMD_OBJECT_TYPE,
        bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """List workbooks in the BudMan application data context.

    Args:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        CMD parameters:
            CK_BDM_TREE (bool): --bdm_tree | -t switch, dispaly result as
            a tree, else display a table.
    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command execution.
    """
    try:
        # Construct CMD_RESULT for return.
        cmd_result: p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_create(
            cmd_object=cmd
        )
        # check bdm_tree flag
        bdm_tree: bool = cmd.get(CK_BDM_TREE, False)
        if bdm_tree:
            # Return all workbooks in a RichTree
            result_tree: RichTree = BUDMAN_CMD_TASK_get_workbook_tree(bdm_DC)
            if result_tree is None:
                # Failed to construct the model_tree for workbooks
                cmd_result[p3m.CK_CMD_RESULT_CONTENT] = "Model workbooks tree not constructed."
                cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_ERROR_STRING_OUTPUT
                logger.error(cmd_result[p3m.CK_CMD_RESULT_CONTENT])
                return cmd_result
            # Success for workbook_tree
            cmd_result[p3m.CK_CMD_RESULT_STATUS] = True
            cmd_result[p3m.CK_CMD_RESULT_CONTENT] = result_tree
            cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_WORKBOOK_TREE_OBJECT
            return cmd_result
        else:
            # CMD_TASK_list_workbook_info_table()
            # List the workbooks selected by list command line arguments.
            selected_bdm_wb_list : List[BDMWorkbook] = None
            selected_bdm_wb_list = process_selected_workbook_input(cmd, bdm_DC)
            # Collect the wb info for workbooks in the selected_bdm_wb_list.
            # Construct the output dictionary result
            cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_WORKBOOK_INFO_TABLE
            cmd_result[p3m.CK_CMD_RESULT_CONTENT] = list()
            for wb in selected_bdm_wb_list:
                wb_index = bdm_DC.dc_WORKBOOK_index(wb.wb_id)
                cmd_result[p3m.CK_CMD_RESULT_CONTENT].append(wb.wb_info_dict(wb_index))
            if len(selected_bdm_wb_list) == 1:
                    bdm_DC.dc_WORKBOOK = wb
            # Success for workbook_tree
            cmd_result[p3m.CK_CMD_RESULT_STATUS] = True
            return cmd_result
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_list_workbooks()
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_list_bdm_store_json()
def BUDMAN_CMD_list_bdm_store_json(cmd: p3m.CMD_OBJECT_TYPE,
                                  bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """List the BDM store in JSON format.
    Args:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        CMD parameters:
            CK_BDM_TREE (bool): --bdm_tree | -t switch, dispaly result as
            a tree, else display a table.
    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command execution.
    """
    try:
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error=True)
        model: BudgetDomainModel = bdm_DC.model
        if not model:
            m = "No BudgetDomainModel binding in the DC."
            logger.error(m)
            return p3m.cp_CMD_RESULT_create(
                cmd_object=cmd,
                cmd_result_status=False,
                result_content_type=p3m.CV_CMD_STRING_OUTPUT,
                result_content=m
            )
        # Construct CMD_RESULT for return.
        cmd_result: p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_create(
            cmd_object=cmd
        )
        cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CMD_JSON_OUTPUT
        cmd_result[p3m.CK_CMD_RESULT_CONTENT] = bdm_DC.model.bdm_BDM_STORE_json()
        cmd_result[p3m.CK_CMD_RESULT_STATUS] = True
        return cmd_result
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_list_bdm_store_json()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_list_files()
def BUDMAN_CMD_list_files(cmd: p3m.CMD_OBJECT_TYPE,
                         bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """List the files in the BDM store.
    Args:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command execution.
    """
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = validate_cmd_arguments(
            cmd=cmd, 
            bdm_DC=bdm_DC,
            cmd_key=CV_LIST_CMD_KEY, 
            subcmd_key=CV_LIST_FILES_SUBCMD_KEY,
            required_args=[
                CK_ALL_FILES,
                CK_SRC_WF_FOLDER,
                CK_SRC_WF_KEY,
                CK_SRC_WF_PURPOSE,
                CK_RAW_FORMAT
            ]
        )
        cmd_result: p3m.CMD_RESULT_TYPE = validate_model_binding(bdm_DC)
        # Validate DC is Model-aware with a model binding.
        if not cmd_result[p3m.CK_CMD_RESULT_STATUS]:
            return cmd_result
        model: BudgetDomainModel = bdm_DC.model
        bsm_file_tree : BSMFileTree = model.bsm_file_tree
        file_tree: Tree = None
        cmd_err: bool = False
        err_msg:str = ""
        # Construct CMD_RESULT for return.
        cmd_result = p3m.cp_CMD_RESULT_create(
            cmd_object=cmd
        )
        # List files all_files | wf_folder
        all_files: bool = cmd_args.get(CK_ALL_FILES, False)
        src_wf_folder: bool = cmd_args.get(CK_SRC_WF_FOLDER, False)
        raw_format: bool = cmd_args.get(CK_RAW_FORMAT, False)
        # If src_wf_folder, then will need wf_key and wf_purpose.
        if src_wf_folder:
            wf_key: str = cmd_args.get(CK_SRC_WF_KEY, None)
            if not wf_key:
                # If wf_key is not present in cmd, try to get it from the data context
                # No wf_key in cmdline, try DC
                wf_key = bdm_DC.dc_WF_KEY
                if not wf_key:
                    # No wf_key to work with
                    cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_ERROR_STRING_OUTPUT
                    cmd_result[p3m.CK_CMD_RESULT_CONTENT] = "No wf_key from cmd args or DC."
                    logger.error(cmd_result[p3m.CK_CMD_RESULT_CONTENT])
                    return cmd_result
            wf_purpose: str = cmd_args.get(CK_SRC_WF_PURPOSE, None)
            if not wf_purpose:
                # If wf_purpose is not present in cmd, try to get it from the data context
                # No wf_purpose in cmdline, try DC
                wf_purpose = bdm_DC.dc_WF_PURPOSE
                if not wf_purpose:
                    # No wf_purpose to work with
                    cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_ERROR_STRING_OUTPUT
                    cmd_result[p3m.CK_CMD_RESULT_CONTENT] = "No wf_purpose from cmd args or DC."
                    logger.error(cmd_result[p3m.CK_CMD_RESULT_CONTENT])
                    return cmd_result
        if all_files:
            # List all files in the BDM store.
            file_tree = bsm_file_tree.file_tree
            if file_tree and raw_format:
                # Return a raw list of file URLs
                file_list = bsm_file_tree.output_all_files()
                cmd_result[p3m.CK_CMD_RESULT_STATUS] = True
                cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_STRING_OUTPUT
                cmd_result[p3m.CK_CMD_RESULT_CONTENT] = file_list
                return cmd_result
            if not file_tree:
                err_msg = "No file_tree available from the model."
                cmd_err = True
        if src_wf_folder and not all_files:
            fi_key: str = bdm_DC.dc_FI_KEY
            # From the Model, get the wf_folder_url for an fi_key, wf_key, wf_purpose
            fi_wf_folder_url: str = model.bdm_WF_FOLDER_CONFIG_ATTRIBUTE(
                fi_key=fi_key, wf_key=wf_key, wf_purpose=wf_purpose, 
                attribute=bdm.WF_FOLDER_URL, raise_errors=False)
            # Retrieve the sub tree for that folder_url
            if fi_wf_folder_url:
                file_tree: Tree = bsm_file_tree.get_sub_file_tree(fi_wf_folder_url)
                if not file_tree:
                    err_msg = (f"No wf_folder found for FI_KEY: "
                               f"'{fi_key}', WF_KEY: '{wf_key}', WF_PURPOSE: '{wf_purpose}'")
                    cmd_err = True
        if cmd_err:
            cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_ERROR_STRING_OUTPUT
            cmd_result[p3m.CK_CMD_RESULT_CONTENT] = err_msg
            logger.error(cmd_result[p3m.CK_CMD_RESULT_CONTENT])
            return cmd_result
        cmd_result[p3m.CK_CMD_RESULT_STATUS] = True
        cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_FILE_TREE_OBJECT
        cmd_result[p3m.CK_CMD_RESULT_CONTENT] = file_tree
        return cmd_result
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_list_files()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_show_DATA_CONTEXT()
def BUDMAN_CMD_show_DATA_CONTEXT(cmd: p3m.CMD_OBJECT_TYPE,
                                bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Show the data context for the command."""
    try:
        # Gather the current content of the DATA_CONTEXT.
        bs = bdm_DC.dc_BDM_STORE
        bs_str = p3u.first_n(str(bs))
        # Prepare the Command output result
        result = f"Budget Manager Data Context:"
        result += f"{P2}{bdm.DC_INITIALIZED}: {bdm_DC.dc_INITIALIZED}\n"
        result += f"{P2}{bdm.DC_BDM_STORE}: {bs_str}\n"
        result += "Current Workflow Location:"
        result += ( f"{P2}{bdm.FI_KEY}: {bdm_DC.dc_FI_KEY}"
                    f"{P2}{bdm.WF_KEY}: {bdm_DC.dc_WF_KEY}"
                    f"{P2}{bdm.WF_PURPOSE}: {bdm_DC.dc_WF_PURPOSE}\n")
        result += "Current Workbook:"
        wb:BDMWorkbook = bdm_DC.dc_WORKBOOK
        if wb:
            result += (f"{P2}{bdm.WB_ID}: {bdm_DC.dc_WB_ID}"
                        f"{P2}{bdm.WB_INDEX}: {bdm_DC.dc_WB_INDEX}"
                        f"{P2}{bdm.WB_NAME}: {bdm_DC.dc_WB_NAME}"
                        f"{P2}{bdm.WB_TYPE}: {bdm_DC.dc_WB_TYPE}\n")
        else:
            result += (f"{P2}{bdm.WB_ID}: ..."
                        f"{P2}{bdm.WB_INDEX}: ..."
                        f"{P2}{bdm.WB_NAME}: ..."
                        f"{P2}{bdm.WB_TYPE}: ...\n")
        _, wdc_result = get_workbook_data_collection_info_str(bdm_DC)
        result += wdc_result
        return p3m.cp_CMD_RESULT_create(
            cmd_result_status=True,
            result_content=result,
            result_content_type=p3m.CV_CMD_STRING_OUTPUT,
            cmd_object=cmd
        )
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_show_DATA_CONTEXT()
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_app_sync()
def BUDMAN_CMD_app_sync(cmd: p3m.CMD_OBJECT_TYPE,
                         bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Sync the BDM store to the BSM.
    
        Compare the files in model.bsm_file_tree to the workbooks in 
        model overall (all FI's). Prepare collections of BDMWorkbooks found in
        BSM but are missing from BDM, and BDMWorkbooks found in BDM but
        missing from BSM. Depending on the cmd args, apply changes or not.
    Args:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        CMD Args:
            CK_SAVE (bool): --save | -s switch, if present, save the BDM_STORE.
            CK_FIX_SWITCH (bool): --fix_switch | -f switch, if present,  fix
            discrepancies between BDM and BSM.
    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command execution.
    """
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = validate_cmd_arguments(
            cmd=cmd, 
            bdm_DC=bdm_DC,
            cmd_key=CV_APP_CMD_KEY, 
            subcmd_key=CV_SYNC_SUBCMD_KEY,
            required_args=[CK_SAVE, CK_FIX_SWITCH]
        )
        cmd_result: p3m.CMD_RESULT_TYPE = validate_model_binding(bdm_DC)
        # Validate DC is Model-aware with a model binding.
        if not cmd_result[p3m.CK_CMD_RESULT_STATUS]:
            return cmd_result
        model: BudgetDomainModel = bdm_DC.model
        bsm_file_tree : BSMFileTree = model.bsm_file_tree
        # Sync the BDM_STORE to the BSM.
        wb: BDMWorkbook = None
        for bsm_file in bsm_file_tree.all_workbooks():
            logger.debug(f"BSM file_url: '{bsm_file._file_url}'")
            wb = model.bdm_WORKBOOK_DATA_COLLECTION_find(bdm.WB_URL,
                                                          bsm_file._file_url)
            if wb:
                logger.debug(f"Found matching BDMWorkbook: '{wb.wb_id}'")
        # Success
        return p3m.cp_CMD_RESULT_create(
            cmd_object=None,
            cmd_result_status=True,
            result_content_type=p3m.CV_CMD_STRING_OUTPUT,
            result_content="foo"
        )
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_app_sync()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_app_log()
def BUDMAN_CMD_app_log(cmd: p3m.CMD_OBJECT_TYPE,
                      bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Show the current log level."""
    try:
        # Construct CMD_RESULT for return.
        cmd_result: p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_create(
            cmd_object=cmd
        )
        log_level_name: str = logging.getLevelName(logger.getEffectiveLevel())
        cmd_result[p3m.CK_CMD_RESULT_STATUS] = True
        cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_STRING_OUTPUT
        cmd_result[p3m.CK_CMD_RESULT_CONTENT] = f"Current log level: '{log_level_name}'"
        return cmd_result
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_app_log()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_app_refresh()
def BUDMAN_CMD_app_refresh(cmd: p3m.CMD_OBJECT_TYPE,
                      bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Show the current log level."""
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = validate_cmd_arguments(
            cmd=cmd, 
            bdm_DC=bdm_DC,
            cmd_key=CV_APP_CMD_KEY, 
            subcmd_key=CV_APP_REFRESH_SUBCMD_KEY,
            model_aware=True,
            required_args=[]
        )
        fr: str = "App Refresh Start: ..."
        r: str = ""
        logger.info(fr)
        model: BudgetDomainModel = bdm_DC.model
        r = f"{P2}Save the BDM_STORE model."
        fr += f"\n{r}"
        logger.debug(r)
        model.bdm_save_model()
        r = f"{P2}Refresh the BSM file tree and BDM Workbook tree."
        fr += f"\n{r}"
        logger.debug(r)
        model.bdm_refresh_trees()
        return p3m.cp_CMD_RESULT_create(True, p3m.CV_CMD_STRING_OUTPUT, fr, cmd)
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_app_refresh()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_app_reload()
def BUDMAN_CMD_app_reload(cmd: p3m.CMD_OBJECT_TYPE,
                      bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Reload the application configuration and settings."""
    try:
        r_msg: str = "Start: ...{P2}"
        reload_target = cmd.get(CK_RELOAD_TARGET, None)
        if reload_target is None:
            m = f"reload_target is None, no action taken."
            logger.error(m)
            return p3m.cp_CMD_RESULT_create(False, p3m.CV_CMD_STRING_OUTPUT, m, cmd)
        if reload_target == CV_CATEGORY_MAP:
            catman: BDMTXNCategoryManager = BDMTXNCategoryManager() #self.WF_CATEGORY_MANAGER
            category_catalog: TXNCategoryMap = None
            if catman :
                category_catalog = catman.catalogs[bdm_DC.dc_FI_KEY]
                category_catalog.CATEGORY_MAP_WORKBOOK_import()
                mod = category_catalog.category_map_module
                if mod:
                    cat_count = len(category_catalog.category_collection)
                    rule_count = len(category_catalog.category_map)
                    task = "CATEGORY_MAP_WORKBOOK_import()"
                    m = (f"{P2}Task: {task:30} {rule_count:>3} "
                            f"rules, {cat_count:>3} categories.")
                    logger.debug(m)
                    return p3m.cp_CMD_RESULT_create(
                        True, p3m.CV_CMD_STRING_OUTPUT, m, cmd)
                else:
                    return p3m.cp_CMD_RESULT_create(
                        False, p3m.CV_CMD_STRING_OUTPUT, 
                        "Failed to reload category_map_module", cmd)
        if reload_target == CV_FI_WORKBOOK_DATA_COLLECTION:
            m = "deprecated"
            # wdc: WORKBOOK_DATA_COLLECTION_TYPE = None
            # wdc, m = self.model.bsm_FI_WORKBOOK_DATA_COLLECTION_resolve(self.dc_FI_KEY)
            return p3m.cp_CMD_RESULT_create(
                True, p3m.CV_CMD_STRING_OUTPUT, m, cmd)
        if reload_target == CV_WORKFLOWS_MODULE:
            importlib.reload(budman_workflows.workflow_utils)
            return p3m.cp_CMD_RESULT_create(
                True, p3m.CV_CMD_STRING_OUTPUT, 
                "Reloaded workflows module.", cmd)
        return p3m.cp_CMD_RESULT_create(
            True, p3m.CV_CMD_STRING_OUTPUT, r_msg, cmd)
    except Exception as e:
        m = f"Error reloading target: {reload_target}: {p3u.exc_err_msg(e)}"
        logger.error(m)
        return p3m.cp_CMD_RESULT_create(
            False, p3m.CV_CMD_STRING_OUTPUT, m, cmd)
#endregion BUDMAN_CMD_app_reload()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_app_delete()
def BUDMAN_CMD_app_delete(cmd: p3m.CMD_OBJECT_TYPE,
                        bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Delete files from the BSM, based on file_index."""
    # Delete a workbook from the DC or a file from BSM.
    try:
        # Validate the cmd argsuments.
        cmd_args: p3m.CMD_ARGS_TYPE = validate_cmd_arguments(
            cmd=cmd, 
            bdm_DC=bdm_DC,
            cmd_key=CV_APP_CMD_KEY, 
            subcmd_key=CV_FILE_DELETE_SUBCMD_KEY,
            model_aware=True,
            required_args=[
                CK_FILE_LIST, 
                CK_ALL_FILES
            ]
        )
        fr: str = "Delete files start: ..."
        m: str = ""
        logger.info(fr)
        model: BudgetDomainModel = bdm_DC.model
        bsm_file_tree : BSMFileTree = model.bsm_file_tree
        all_files: bool = cmd.get(CK_ALL_FILES)
        selected_file_list: List[str] = cmd_args.get(CK_FILE_LIST)
        bsm_files: List[BSMFile] = None
        bsm_files = bsm_file_tree.validate_file_list(selected_file_list)
        bsm_file: BSMFile = None
        for bsm_file in bsm_files:
            logger.debug(f"{P2}Deleting BSM file_url: '{bsm_file._file_url}'")
            if bsm_file.in_bdm:
                m = f"{P2}BSM file_url: '{bsm_file._file_url}' is flagged in BDM."
                logger.warning(m)
                fr += f"\n{m}"
            tree_node: Node = bsm_file_tree.file_tree.get_node(bsm_file._file_url)
            # Remove the bsm_file node from the file_tree
            if tree_node:
                bsm_file_tree.file_tree.remove_node(tree_node.identifier)
            # Remove the physical file from storage
            bsm_file.delete_file()
            m = f"BizEVENT: Deleted file: '{bsm_file._file_url}'"
            logger.info(m)
            fr += f"\n{m}"
            # Delete the BSMFile object.
            del bsm_file

        # Refresh the trees
        model.bdm_refresh_trees()
        return p3m.cp_CMD_RESULT_create(True, p3m.CV_CMD_STRING_OUTPUT, fr, cmd)
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
# ---------------------------------------------------------------------------- +    
#endregion BudMan Application CMD_ functions
# ---------------------------------------------------------------------------- +    
#endregion BudMan Application CMD_ functions
# --------------------------------------------------------------------------- +

# --------------------------------------------------------------------------- +
#region BudMan Application CMD_TASK_ functions
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_TASK_sync_BDM_STORE_to_BSM()
def BUDMAN_CMD_TASK_sync_BDM_STORE_to_BSM(bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:        
    """Sync the BDM_STORE to the BSM.

    Args:
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command execution.
    """
    try:
        # Validate DC is Model-aware with a model binding.
        cmd_result: p3m.CMD_RESULT_TYPE = validate_model_binding(bdm_DC)
        if not cmd_result[p3m.CK_CMD_RESULT_STATUS]:
            return cmd_result
        model: BudgetDomainModel = bdm_DC.model
        # Sync the BDM_STORE to the BSM.
        sync_result: Tuple[bool, str] = model.bdm_sync_BDM_STORE_to_BSM()
        if not sync_result[0]:
            # Sync failed
            return p3m.cp_CMD_RESULT_create(
                cmd_object=None,
                cmd_result_status=False,
                result_content_type=p3m.CV_CMD_STRING_OUTPUT,
                result_content=sync_result[1]
            )
        # Success
        return p3m.cp_CMD_RESULT_create(
            cmd_object=None,
            cmd_result_status=True,
            result_content_type=p3m.CV_CMD_STRING_OUTPUT,
            result_content=sync_result[1]
        )
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(None, e)
#endregion BUDMAN_CMD_TASK_sync_BDM_STORE_to_BSM()
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_TASK_get_workbook_tree() method
def BUDMAN_CMD_TASK_get_workbook_tree(bdm_DC: BudManAppDataContext_Base) -> RichTree:
    """Return a workbook_tree (RichTree) for the BDM workbooks in bdm_DC."""
    try:
        st = p3u.start_timer()
        logger.debug(f"Start.")
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        model:BudgetDomainModel = bdm_DC.model
        now_str = dt.now().strftime("%Y-%m-%d %I:%M:%S %p")
        if not model:
            return None
        settings = bdms.BudManSettings()
        p3u.is_not_obj_of_type("settings", settings, bdms.BudManSettings, raise_error=True)
        bdm_store_full_filename = (settings[bdms.BDM_STORE_FILENAME] +
                                   settings[bdms.BDM_STORE_FILETYPE])
        bdm_folder = settings[bdms.BDM_FOLDER] + '/'
        purpose_icon = {
            "wf_input": ":inbox_tray:",
            "wf_output": ":outbox_tray:",
            "wf_working": ":gear:"
        }
        p_str: str = f"model['{bdm_folder + bdm_store_full_filename}'] {now_str}" 
        workbook_tree = RichTree(
            f":open_file_folder: {escape(p_str)}",
            guide_style="bold white"
        )
        wdc : bdm.WORKBOOK_DATA_COLLECTION_TYPE = None
        # For all Financial Institutions (FI) in the model.
        for fi_key, fi_obj in model.bdm_fi_collection.items():
            fi_folder = model.bdm_FI_FOLDER(fi_key)
            fi_name = model.bdm_FI_NAME(fi_key)
            # fi_key node.
            wdc = model.bdm_FI_WORKBOOK_DATA_COLLECTION(fi_key)
            count = len(wdc) if wdc else 0
            fi_key_branch = workbook_tree.add(
                f"[bold cyan]:open_file_folder: "
                f"'{fi_folder}' :bank: workbook count: '{count}'",
                guide_style="bold white"
            )
            if count == 0: 
                continue
            # Put the wdc in wb_index order, which is sorted by wb_id key.
            wdc = dict(sorted(wdc.items(), key=lambda item: item[0]))
            for wf_key, wf_folder_config_list in model.bdm_FI_WF_FOLDER_CONFIG_COLLECTION(fi_key).items():
                x_key = f"{fi_key}_{wf_key}"
                wf_key_branch = fi_key_branch.add(
                    f"[bold yellow]:repeat: {wf_key} workflow",
                    guide_style="bold white"
                )
                for wf_folder_config in wf_folder_config_list:
                    # wf_purpose node.
                    wf_purpose = wf_folder_config[bdm.WF_PURPOSE]
                    wf_folder = wf_folder_config[bdm.WF_FOLDER]
                    y_key = f"{fi_key}_{wf_key}_{wf_purpose}" # node key
                    wf_purpose_branch = wf_key_branch.add(
                        f"[bold gold3]:open_file_folder: '{wf_folder}' "
                        f"{purpose_icon.get(wf_purpose, '')}",
                        guide_style="bold white"
                    )
                    # wf_purpose workbook names.
                    wb_names = workbook_names(wdc, wf_key, wf_purpose)
                    logger.debug(f"FI: '{fi_key}', WF: '{wf_key}', WF_PURPOSE: '{wf_purpose}', workbooks({len(wb_names)}): {wb_names}")
                    if len(wb_names) > 0:
                        for wb_name in wb_names:
                            wf_purpose_branch.add(
                                f"[bold green]:page_facing_up: '{wb_name}' (wb_name)",
                            )
        logger.debug(f"Complete.")
        return workbook_tree
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        return None
#endregion BUDMAN_CMD_TASK_get_workbook_tree
# ---------------------------------------------------------------------------- +    
#region BUDMAN_CMD_TASK_show_BUDGET_CATEGORIES()
def BUDMAN_CMD_TASK_show_BUDGET_CATEGORIES(cmd: p3m.CMD_OBJECT_TYPE,
                                      bdm_DC: BudManAppDataContext_Base) -> p3m.CMD_RESULT_TYPE:
    """Show the budget categories for the command."""
    try:
        result: str = "no result"
        # Show the budget categories.
        cat_list = cmd.get(CK_CAT_LIST, [])
        tree_level = cmd.get(CK_LEVEL, 2)
        fi_key: str = bdm_DC.dc_FI_KEY
        catman : BDMTXNCategoryManager = bdm_DC.WF_CATEGORY_MANAGER
        fi_catmap : TXNCategoryMap = catman.catalogs[fi_key]
        result: str = fi_catmap.output_category_tree(level=tree_level, cat_list=cat_list)
        return p3m.cp_CMD_RESULT_create(
            cmd_result_status=True,
            result_content=result,
            result_content_type=p3m.CV_CMD_STRING_OUTPUT,
            cmd_object=cmd
        )
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_TASK_show_BUDGET_CATEGORIES()
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_TASK_construct_dst_file_url()
def BUDMAN_CMD_TASK_construct_dst_file_url(cmd: p3m.CMD_OBJECT_TYPE,
                                            bdm_DC: BudManAppDataContext_Base,
                                            bsm_file: "BSMFile") -> p3m.CMD_RESULT_TYPE:
    """Construct the destination file URL from the BSMFile.

    Arguments:
        cmd (CMD_OBJECT_TYPE): The command object to process.
        bdm_DC (BudManAppDataContext_Base): The data context for the BudMan application.
        bsm_file (BSMFile): The BSMFile object to construct the URL from.

    Returns:
        p3m.CMD_RESULT_TYPE: The result of the command processing.
            .result_content: The constructed destination file URL.

    """
    try:
        # Construct the destination file URL.
        # nedd dst_prefix, dst_wb_type, dst_folder_abs_path
        dst_file_url = f"{bsm_file.wf_prefix}{bsm_file.wb_name}" \
                       f"{bsm_file.wb_type}.{bsm_file.extension}"

        return p3m.cp_CMD_RESULT_create(
            cmd_result_status=True,
            result_content=dst_file_url,
            result_content_type=p3m.CV_CMD_STRING_OUTPUT,
            cmd_object=cmd
        )
    except Exception as e:
        return p3m.cp_CMD_RESULT_EXCEPTION_create(cmd, e)
#endregion BUDMAN_CMD_TASK_construct_dst_file_url()
# ---------------------------------------------------------------------------- +
#region extract_bdm_tree() function
def extract_bdm_tree(bdm_DC: BudManAppDataContext_Base) -> Tree:
    """Return a tree structure of the BDM in the provided DataContext."""
    try:
        st = p3u.start_timer()
        logger.debug(f"Start.")
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error= True)
        model:BudgetDomainModel = bdm_DC.model
        now_str = dt.now().strftime("%Y-%m-%d %I:%M:%S %p")
        if not model:
            m = "No BudgetDomainModel binding in the DC."
            logger.error(m)
            return False, m
        settings = bdms.BudManSettings()
        p3u.is_not_obj_of_type("settings", settings, bdms.BudManSettings, raise_error=True)
        bdm_store_full_filename = (settings[bdms.BDM_STORE_FILENAME] +
                                   settings[bdms.BDM_STORE_FILETYPE])
        bdm_folder = settings[bdms.BDM_FOLDER] + '/'
        p_str: str = f"BDM_STORE['{bdm_folder + bdm_store_full_filename}'] {now_str}" 
        tree = Tree()
        tree.create_node(f"model: {p_str}", "root")  # root node
        wdc : bdm.WORKBOOK_DATA_COLLECTION_TYPE = None
        # For all Financial Institutions (FI) in the model.
        for fi_key, fi_obj in model.bdm_fi_collection.items():
            fi_folder = model.bdm_FI_FOLDER(fi_key)
            fi_name = model.bdm_FI_NAME(fi_key)
            # fi_key node.
            wdc = model.bdm_FI_WORKBOOK_DATA_COLLECTION(fi_key)
            if wdc is None or len(wdc) == 0:
                tag = f"'{fi_folder}' (fi_folder) workbook count: '0'"
                tree.create_node(tag, f"{fi_key}", parent="root")  # fi_key node
                continue
            tag = f"'{fi_folder}' (fi_folder) workbook count: '{len(wdc)}'"
            tree.create_node(tag, f"{fi_key}", parent="root") # fi_key node
            # Put the wdc in wb_index order, which is sorted by wb_id key.
            wdc = dict(sorted(wdc.items(), key=lambda item: item[0]))
            for wf_key, wf_folder_config_list in model.bdm_FI_WF_FOLDER_CONFIG_COLLECTION(fi_key).items():
                x_key = f"{fi_key}_{wf_key}"
                tree.create_node(f"{wf_key} workflow (wf_key)", x_key, parent=f"{fi_key}") # wf_key node
                for wf_folder_config in wf_folder_config_list:
                    # wf_purpose node.
                    wf_purpose = wf_folder_config[bdm.WF_PURPOSE]
                    wf_folder = wf_folder_config[bdm.WF_FOLDER]
                    y_key = f"{fi_key}_{wf_key}_{wf_purpose}" # node key
                    tree.create_node(f"'{wf_folder}' {wf_purpose} (wf_folder)", 
                                     y_key, parent=x_key) # wf_purpose node
                    # wf_purpose workbook names.
                    wb_names = workbook_names(wdc, wf_key, wf_purpose)
                    logger.debug(f"FI: '{fi_key}', WF: '{wf_key}', WF_PURPOSE: '{wf_purpose}', workbooks({len(wb_names)}): {wb_names}")
                    if len(wb_names) > 0:
                        for wb_name in wb_names:
                            tree.create_node(f"'{wb_name}' (wb_name)", # workbook node
                                             f"{y_key}_{wb_name}", 
                                             parent=y_key)
        print
        logger.debug(f"Complete.")
        return tree
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion extract_bdm_tree() function
# ---------------------------------------------------------------------------- +
#endregion BudMan Application CMD_TASK_ functions
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region BudMan Application Command File Services
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_FILE_SERVICE_get_BSMFile()
def BUDMAN_CMD_FILE_SERVICE_get_BSMFile(bdm_DC: BudManAppDataContext_Base, file_index: int) -> Optional[BSMFile]:
    """Get the BSMFile object from the file tree for a given file_index."""
    try:
        # Return the bsm_file_tree from the model via the DC.
        bsm_file_tree: BSMFileTree = bdm_DC.model.bsm_file_tree
        return bsm_file_tree.get_BSMFile(file_index)
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion BUDMAN_CMD_FILE_SERVICE_get_BSMFile()
# ---------------------------------------------------------------------------- +
#region BUDMAN_CMD_FILE_SERVICE_get_full_filename()
def BUDMAN_CMD_FILE_SERVICE_get_full_filename(file_tree: Tree, file_index: int) -> Optional[str]:
    """Get the filename from the file tree for a given file_index."""
    try:
        p3u.is_not_obj_of_type("file_tree", file_tree, Tree, raise_error=True)
        for node_id in file_tree.expand_tree():
            file_node: Node = file_tree.get_node(node_id)
            if file_node.is_leaf(): # only look at file nodes, which are leafs
                if (file_node.data and isinstance(file_node.data, BSMFile)):
                    bsm_file: BSMFile = file_node.data
                    this_index: int = bsm_file.file_index
                    if this_index == file_index:
                        return bsm_file.full_filename
        return None
    except Exception as e:
        logger.error(p3u.exc_err_msg(e))
        raise
#endregion BUDMAN_CMD_FILE_SERVICE_get_full_filename()
# ---------------------------------------------------------------------------- +
#endregion BudMan Application Command File Services
# ---------------------------------------------------------------------------- +

# ---------------------------------------------------------------------------- +
#region BudMan Application Command Helper functions
# ---------------------------------------------------------------------------- +
#region workbook_names() function
def workbook_names(wdc: bdm.WORKBOOK_DATA_COLLECTION_TYPE, wf_key: str, wf_purpose: str) -> List[str]:
    """Return a list of workbook names for the given wf_key and wf_purpose."""
    try:
        wb_name_list: List[str] = []
        if wdc is None or len(wdc) == 0:
            return wb_name_list
        wb_id_list: List[str] = list(wdc.keys())
        for wb_id,bdm_wb in wdc.items():
            if not isinstance(bdm_wb, bdm.BDMWorkbook):
                continue
            if bdm_wb.wf_key != wf_key:
                continue
            if bdm_wb.wf_purpose != wf_purpose:
                continue
            wb_index = wb_id_list.index(wb_id)
            name_str: str = f"{str(wb_index):>4} {bdm_wb.wb_name}"
            wb_name_list.append(name_str)
        return wb_name_list
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        return []
#endregion workbook_names() function
# ---------------------------------------------------------------------------- +
#region process_selected_workbook_input()
def process_selected_workbook_input(cmd: p3m.CMD_OBJECT_TYPE,
                                    bdm_DC: BudManAppDataContext_Base,
                                    validate_url:bool = True) -> List[BDMWorkbook]:
    """Process the workbook input from the command, return a list of 
    BDMWorkbooks, which may be empty.

    Arguments:
        cmd (Dict): A p3m.CMD_OBJECT.

    Returns:
        List[BDMWorkbook]: A list of BDMWorkbook objects.
    """
    try:
        # Extract common command attributes to select workbooks for task action.
        wb_list : List[int] = cmd.get(CK_WB_LIST, [])
        all_wbs : bool = cmd.get(CK_ALL_WBS, bdm_DC.dc_ALL_WBS)
        selected_bdm_wb_list : List[BDMWorkbook] = []
        load_workbook:bool = cmd.get(CK_LOAD_WORKBOOK, False)
        if all_wbs:
            # If all_wbs is True, process all workbooks in the data context.
            selected_bdm_wb_list = list(bdm_DC.dc_WORKBOOK_DATA_COLLECTION.values())
        elif len(wb_list) > 0:
            for wb_index in wb_list:
                bdm_wb = bdm_DC.dc_WORKBOOK_by_index(wb_index)
                selected_bdm_wb_list.append(bdm_wb)
        else:
            # No workbooks selected by the command parameters.
            return selected_bdm_wb_list
        if validate_url:
            for bdm_wb in selected_bdm_wb_list:
                bdm_wb_abs_path = bdm_wb.abs_path()
                if bdm_wb_abs_path is None:
                    selected_bdm_wb_list.remove(bdm_wb)
                    m = f"Excluded workbook: '{bdm_wb.wb_id}', "
                    m += f"workbook url is not valid: {bdm_wb.wb_url}"   
                    logger.error(m)
                    continue
                if load_workbook and not bdm_wb.wb_loaded:
                    # Load the workbook content if it is not loaded.
                    success, result = bdm_DC.dc_WORKBOOK_content_get(bdm_wb)
                    if not success:
                        selected_bdm_wb_list.remove(bdm_wb)
                        m = f"Excluded workbook: '{bdm_wb.wb_id}', "
                        m += f"failed to load: {result}"
                        logger.error(m)
                        continue
                    if not bdm_wb.wb_loaded:
                        logger.warning(f"Workbook '{bdm_wb.wb_id}' wb_loaded was False!")
                        bdm_wb.wb_loaded = True
                    # self.dc_LOADED_WORKBOOKS[bdm_wb.wb_id] = wb_content
        return selected_bdm_wb_list
    except Exception as e:
        m = f"Error processing workbook input: {p3u.exc_err_msg(e)}"
        logger.error(m)
        raise RuntimeError(m)
#endregion process_selected_workbook_input()
# ------------------------------------------------------------------------ +    
#region get_workbook_data_collection_info_str() method
def get_workbook_data_collection_info_str(bdm_DC: BudManAppDataContext_Base) -> bdm.BUDMAN_RESULT_TYPE:
    """Construct an outout string with information about the WORKBOOKS."""
    try:
        logger.debug(f"Start: ...")
        # Be workbook-centric is this view of the DC
        wdc = bdm_DC.dc_WORKBOOK_DATA_COLLECTION
        wdc_count = len(wdc) if wdc else 0
        lwbc = bdm_DC.dc_LOADED_WORKBOOKS

        # Prepare the output result
        result = f"{P2}{bdm.FI_WORKBOOK_DATA_COLLECTION}: {wdc_count}\n"
        result += f"{P4}{bdm.WB_INDEX:6}{P2}{bdm.WB_ID:55}{P2}"
        result += f"{bdm.WB_TYPE:15}{P2}{bdm.WB_CONTENT:30}"
        result += "\n"
        bdm_wb : BDMWorkbook = None
        if wdc_count > 0:
            for i, bdm_wb in enumerate(wdc.values()):
                r = f"{bdm_wb.wb_index_display_str(i)}"
                result += r + "\n"
        logger.info(f"Complete:")
        return True, result
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        return False, m
#endregion get_workbook_data_collection_info_str() method
# ------------------------------------------------------------------------ +
#region verify_cmd_key()
def verify_cmd_key( cmd: p3m.CMD_OBJECT_TYPE, 
                   expected_cmd_key: str) -> p3m.CMD_RESULT_TYPE:
    """Verify the command key in the command object.

    Args:
        cmd (Dict): A p3m.CMD_OBJECT.
        expected_cmd_key (str): The expected command key.

    Returns:
        CMD_RESULT_TYPE: True if the command key matches, False otherwise,
        with a returnable CMD_RESULT_TYPE.
    """
    cmd_result: p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_create(
        cmd_result_status=True,
        result_content_type=p3m.CV_CMD_STRING_OUTPUT,
        result_content=f"Expected cmd_key: {expected_cmd_key} is valid.",
        cmd_object=cmd
    )
    actual_cmd_key = cmd.get(p3m.CK_CMD_KEY)
    if actual_cmd_key != expected_cmd_key:
        # Invalid cmd_key
        m = (f"Invalid cmd_key: {cmd[p3m.CK_CMD_KEY]} "
             f"expected: {expected_cmd_key}")
        logger.error(m)
        cmd_result[p3m.CK_CMD_RESULT_STATUS] = False
        cmd_result[p3m.CK_CMD_RESULT_CONTENT] = m
        cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_ERROR_STRING_OUTPUT
    return cmd_result
#endregion verify_cmd_key()
# ---------------------------------------------------------------------------- +
#region verify_subcmd_key()
def verify_subcmd_key( cmd: p3m.CMD_OBJECT_TYPE, 
                   expected_subcmd_key: str) -> p3m.CMD_RESULT_TYPE:
    """Verify the subcommand key in the command object.

    Args:
        cmd (Dict): A p3m.CMD_OBJECT.
        expected_subcmd_key (str): The expected subcommand key.

    Returns:
        CMD_RESULT_TYPE: True if the command key matches, False otherwise,
        with a returnable CMD_RESULT_TYPE.
    """
    cmd_result: p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_create(
        cmd_result_status=True,
        result_content_type=p3m.CV_CMD_STRING_OUTPUT,
        result_content=f"Expected subcmd_key: {expected_subcmd_key} is valid.",
        cmd_object=cmd
    )
    actual_subcmd_key = cmd.get(p3m.CK_SUBCMD_KEY)
    if actual_subcmd_key != expected_subcmd_key:
        # Invalid subcmd_key
        m = (f"Invalid subcmd_key: {cmd[p3m.CK_SUBCMD_KEY]} "
             f"expected: {expected_subcmd_key}")
        logger.error(m)
        cmd_result[p3m.CK_CMD_RESULT_STATUS] = False
        cmd_result[p3m.CK_CMD_RESULT_CONTENT] = m
        cmd_result[p3m.CK_CMD_RESULT_CONTENT_TYPE] = p3m.CV_CMD_ERROR_STRING_OUTPUT
    return cmd_result
#endregion verify_cmd_key()
# ---------------------------------------------------------------------------- +
#region validate_cmd_arguments()
def validate_cmd_arguments(cmd: p3m.CMD_OBJECT_TYPE,
                            bdm_DC: BudManAppDataContext_Base,
                            cmd_key: str,
                            subcmd_key: str,
                            required_args: List[str],
                            model_aware:bool=True) -> p3m.CMD_ARGS_TYPE:
    """Validate the components of the command object.

    Args:
        cmd (CMD_OBJECT_TYPE): The command object to validate.
        bdm_DC (BudManAppDataContext_Base): The data context for the command.
        required_args (List[str]): A list of required argument cmd keys (CK_).

    Returns:
        cmd_args: p3m.CMD_ARGS_TYPE: The resulting cmd arg from the validation.
    """
    try:
        try:
            p3u.is_not_obj_of_type("cmd", cmd, dict, raise_error=True)
            p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                                raise_error=True)
        except Exception as e:
            msg = f"Error validating command components: {str(e)}"
            cmd_result: p3m.CMD_RESULT_TYPE = p3m.cp_CMD_RESULT_ERROR_create(
                cmd=cmd, msg=msg
            )
            logger.error(msg)
            raise p3m.CMDValidationException(cmd=cmd, 
                                             msg=msg,
                                             cmd_result_error=cmd_result)
        cmd_result : p3m.CMD_RESULT_TYPE = None
        # Should be indicated cmd_key.
        cmd_result = verify_cmd_key(cmd, cmd_key)
        if not cmd_result[p3m.CK_CMD_RESULT_STATUS]: 
            raise p3m.CMDValidationException(cmd=cmd, 
                                             msg=cmd_result[p3m.CK_CMD_RESULT_CONTENT],
                                             cmd_result_error=cmd_result)
        # Should be indicated subcmd_key.
        cmd_result = verify_subcmd_key(cmd, subcmd_key)
        if not cmd_result[p3m.CK_CMD_RESULT_STATUS]: 
            raise p3m.CMDValidationException(cmd=cmd, 
                                             msg=cmd_result[p3m.CK_CMD_RESULT_CONTENT],
                                             cmd_result_error=cmd_result)
        if model_aware:
            cmd_result = validate_model_binding(bdm_DC)
            # Validate DC is Model-aware with a model binding.
            if not cmd_result[p3m.CK_CMD_RESULT_STATUS]:
                return cmd_result
        model: BudgetDomainModel = bdm_DC.model
        # Check for required command key arguments (CK_) in the command object
        cmd_args: p3m.CMD_ARGS_TYPE = {}
        cmd_args[p3m.CK_CMD_KEY] = cmd[p3m.CK_CMD_KEY]
        cmd_args[p3m.CK_SUBCMD_KEY] = cmd[p3m.CK_SUBCMD_KEY]
        missing_arg_error_msg: str = "Required arguments validation:"
        missing_arg_list = []
        missing_arg_count = 0
        for key in required_args:
            if key not in cmd:
                m = f"Missing required command argument key: {key}"
                logger.error(m)
                missing_arg_error_msg += f"{m}\n"
                missing_arg_count += 1
                missing_arg_list.append(key)
            else:
                cmd_args[key] = cmd[key]
        if missing_arg_count > 0:
            m = (f"Missing {missing_arg_count} required arguments: "
                 f"{missing_arg_list}")
            logger.error(m)
            cmd_result = p3m.cp_CMD_RESULT_create(
                cmd_result_status=False,
                result_content_type=p3m.CV_CMD_ERROR_STRING_OUTPUT,
                result_content=missing_arg_error_msg,
                cmd_object=cmd
                )
            raise p3m.CMDValidationException(cmd=cmd,
                                             msg=cmd_result[p3m.CK_CMD_RESULT_CONTENT],
                                             cmd_result_error=cmd_result)
        return cmd_args
    except p3m.CMDValidationException as ve:
        raise
    except Exception as e:
        logger.error(f"Error validating command components: {e}")
        return p3m.cp_CMD_RESULT_create(
            cmd_result_status=False,
            result_content_type=p3m.CV_CMD_ERROR_STRING_OUTPUT,
            result_content=str(e),
            cmd_object=cmd
        )
#endregion validate_cmd_arguments()
# ---------------------------------------------------------------------------- +
#region validate_model_binding()
def validate_model_binding(bdm_DC: BudManAppDataContext_Base,
                           raise_error: bool = False) -> p3m.CMD_RESULT_TYPE:
    """Validate that the BudManAppDataContext has a valid BudgetDomainModel binding.

    Args:
        bdm_DC (BudManAppDataContext_Base): The data context to validate.

    Returns:
        CMD_RESULT_TYPE: The result of the validation.
            .result_content: True if the model is valid, False otherwise.
    """
    try:
        p3u.is_not_obj_of_type("bdm_DC", bdm_DC, BudManAppDataContext_Base,
                               raise_error=True)
        model: BudgetDomainModel = bdm_DC.model
        if not model:
            m = "No BudgetDomainModel binding in the DC."
            logger.error(m)
            if raise_error:
                raise ValueError(m)
            return p3m.cp_CMD_RESULT_create(
                cmd_result_status=False,
                result_content=m,
                result_content_type=p3m.CV_CMD_ERROR_STRING_OUTPUT,
                cmd_object=None
            )
        if not isinstance(model, BudgetDomainModel):
            m = f"Invalid BudgetDomainModel binding in the DC: {type(model)}"
            logger.error(m)
            if raise_error:
                raise ValueError(m)
            return p3m.cp_CMD_RESULT_create(
                cmd_result_status=False,
                result_content=m,
                result_content_type=p3m.CV_CMD_ERROR_STRING_OUTPUT,
                cmd_object=None
            )
        # Model is valid
        return p3m.cp_CMD_RESULT_create(
            cmd_result_status=True,
            result_content=model,
            result_content_type=p3m.CV_CMD_BDM_MODEL_OBJECT,
            cmd_object=None
        )
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        if raise_error:
            raise
        return p3m.cp_CMD_RESULT_create(
            cmd_result_status=False,
            result_content=m,
            result_content_type=p3m.CV_CMD_ERROR_STRING_OUTPUT,
            cmd_object=None
        )
#ednregion validate_model_binding()
# ---------------------------------------------------------------------------- +
#endregion BudMan Application Command Helper functions
# ---------------------------------------------------------------------------- +
#endregion BudMan Application Command Helper functions
# ---------------------------------------------------------------------------- +
